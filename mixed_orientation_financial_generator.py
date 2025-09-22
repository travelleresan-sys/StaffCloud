#!/usr/bin/env python3
"""
混合レイアウト財務諸表Excel生成機能
- 貸借対照表、損益計算書、キャッシュフロー: A4縦
- 株主資本等変動計算書、附属明細書: A4横
"""

from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from datetime import datetime


class MixedOrientationFinancialGenerator:
    """混合レイアウト財務諸表Excel生成クラス"""

    def __init__(self):
        # フォント・スタイル設定
        self.title_font = Font(name='ＭＳ ゴシック', size=12, bold=True)
        self.header_font = Font(name='ＭＳ ゴシック', size=10, bold=True)
        self.normal_font = Font(name='ＭＳ ゴシック', size=9)
        self.small_font = Font(name='ＭＳ ゴシック', size=8)

        # 横向け用（少し大きめ）
        self.title_font_landscape = Font(name='ＭＳ ゴシック', size=14, bold=True)
        self.header_font_landscape = Font(name='ＭＳ ゴシック', size=11, bold=True)
        self.normal_font_landscape = Font(name='ＭＳ ゴシック', size=10)
        self.small_font_landscape = Font(name='ＭＳ ゴシック', size=9)

        # アライメント
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')
        self.right_align = Alignment(horizontal='right', vertical='center')

        # 罫線
        self.thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        self.double_bottom_border = Border(bottom=Side(style='double'))

        # 背景色
        self.header_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
        self.subtotal_fill = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')

    def setup_a4_portrait(self, ws):
        """A4縦向けページ設定"""
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins = PageMargins(
            left=0.5, right=0.5, top=0.5, bottom=0.5,
            header=0.3, footer=0.3
        )
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.print_options.horizontalCentered = True
        ws.page_setup.scale = 85

    def setup_a4_landscape(self, ws):
        """A4横向けページ設定"""
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins = PageMargins(
            left=0.5, right=0.5, top=0.5, bottom=0.5,
            header=0.3, footer=0.3
        )
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.print_options.horizontalCentered = True
        ws.page_setup.scale = 90  # 横向けは少し大きめに

    def generate_all_statements(self, company_name="株式会社サンプル", fiscal_year=2025):
        """全ての財務諸表を生成（混合レイアウト）"""
        wb = openpyxl.Workbook()

        # 貸借対照表（A4縦）
        ws_bs = wb.active
        ws_bs.title = "貸借対照表"
        self.create_balance_sheet_portrait(ws_bs, company_name, fiscal_year)

        # 損益計算書（A4縦）
        ws_pl = wb.create_sheet("損益計算書")
        self.create_income_statement_portrait(ws_pl, company_name, fiscal_year)

        # キャッシュフロー計算書（A4縦）
        ws_cf = wb.create_sheet("キャッシュフロー計算書")
        self.create_cash_flow_statement_portrait(ws_cf, company_name, fiscal_year)

        # 株主資本等変動計算書（A4横）
        ws_eq = wb.create_sheet("株主資本等変動計算書")
        self.create_equity_statement_landscape(ws_eq, company_name, fiscal_year)

        # 附属明細書（A4横）
        ws_notes = wb.create_sheet("附属明細書")
        self.create_notes_landscape(ws_notes, company_name, fiscal_year)

        return wb

    def create_balance_sheet_portrait(self, ws, company_name, fiscal_year):
        """A4縦向け貸借対照表作成"""
        self.setup_a4_portrait(ws)

        # 列幅設定
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 2
        ws.column_dimensions['E'].width = 22
        ws.column_dimensions['F'].width = 12

        # タイトル部分
        ws.merge_cells('A1:F1')
        ws['A1'] = "貸借対照表"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = self.center_align

        ws.merge_cells('A2:F2')
        ws['A2'] = f"{company_name}　{fiscal_year}年3月31日現在　（単位：円）"
        ws['A2'].font = self.small_font
        ws['A2'].alignment = self.center_align

        # ヘッダー行
        row = 4
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = "資産の部"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].alignment = self.center_align
        ws[f'A{row}'].fill = self.header_fill

        ws.merge_cells(f'D{row}:F{row}')
        ws[f'D{row}'] = "負債及び純資産の部"
        ws[f'D{row}'].font = self.header_font
        ws[f'D{row}'].alignment = self.center_align
        ws[f'D{row}'].fill = self.header_fill

        # データ（従来と同じ）
        row = 5
        assets_data = [
            ("Ⅰ 流動資産", None, True),
            ("　現金及び預金", 5000000, False),
            ("　売掛金", 8000000, False),
            ("　商品", 3000000, False),
            ("　その他", 2500000, False),
            ("　流動資産合計", 18500000, True),
            ("", None, False),
            ("Ⅱ 固定資産", None, True),
            ("　建物（純額）", 20000000, False),
            ("　機械装置（純額）", 7000000, False),
            ("　土地", 30000000, False),
            ("　ソフトウェア", 2000000, False),
            ("　投資有価証券", 5000000, False),
            ("　その他", 1000000, False),
            ("　固定資産合計", 65000000, True),
            ("", None, False),
            ("資産合計", 83500000, True)
        ]

        liabilities_data = [
            ("Ⅰ 流動負債", None, True),
            ("　買掛金", 4000000, False),
            ("　短期借入金", 3000000, False),
            ("　未払金等", 4500000, False),
            ("　流動負債合計", 11500000, True),
            ("", None, False),
            ("Ⅱ 固定負債", None, True),
            ("　長期借入金", 20000000, False),
            ("　退職給付引当金", 3000000, False),
            ("　固定負債合計", 23000000, True),
            ("負債合計", 34500000, True),
            ("", None, False),
            ("Ⅲ 純資産の部", None, True),
            ("　資本金", 25000000, False),
            ("　資本剰余金", 5000000, False),
            ("　利益剰余金", 19000000, False),
            ("　純資産合計", 49000000, True),
            ("負債及び純資産合計", 83500000, True)
        ]

        # データ出力
        for i, (item, amount, is_total) in enumerate(assets_data):
            current_row = row + i
            ws[f'B{current_row}'] = item
            ws[f'B{current_row}'].font = self.header_font if is_total else self.normal_font
            ws[f'B{current_row}'].alignment = self.left_align
            if is_total:
                ws[f'B{current_row}'].fill = self.subtotal_fill

            if amount is not None:
                ws[f'C{current_row}'] = f"{amount:,}"
                ws[f'C{current_row}'].font = self.header_font if is_total else self.normal_font
                ws[f'C{current_row}'].alignment = self.right_align
                if is_total:
                    ws[f'C{current_row}'].fill = self.subtotal_fill
                if item == "資産合計":
                    ws[f'C{current_row}'].border = self.double_bottom_border

        for i, (item, amount, is_total) in enumerate(liabilities_data):
            current_row = row + i
            ws[f'E{current_row}'] = item
            ws[f'E{current_row}'].font = self.header_font if is_total else self.normal_font
            ws[f'E{current_row}'].alignment = self.left_align
            if is_total:
                ws[f'E{current_row}'].fill = self.subtotal_fill

            if amount is not None:
                ws[f'F{current_row}'] = f"{amount:,}"
                ws[f'F{current_row}'].font = self.header_font if is_total else self.normal_font
                ws[f'F{current_row}'].alignment = self.right_align
                if is_total:
                    ws[f'F{current_row}'].fill = self.subtotal_fill
                if item == "負債及び純資産合計":
                    ws[f'F{current_row}'].border = self.double_bottom_border

        # 罫線適用
        max_row = row + len(assets_data) - 1
        for r in range(4, max_row + 1):
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                if ws[f'{col}{r}'].value is not None or r == 4:
                    ws[f'{col}{r}'].border = self.thin_border

    def create_income_statement_portrait(self, ws, company_name, fiscal_year):
        """A4縦向け損益計算書作成"""
        self.setup_a4_portrait(ws)

        # 列幅設定
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 26
        ws.column_dimensions['C'].width = 13
        ws.column_dimensions['D'].width = 13

        # タイトル部分
        ws.merge_cells('A1:D1')
        ws['A1'] = "損益計算書"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = self.center_align

        ws.merge_cells('A2:D2')
        ws['A2'] = f"{company_name}　{fiscal_year}年3月期　（単位：円）"
        ws['A2'].font = self.small_font
        ws['A2'].alignment = self.center_align

        # ヘッダー行
        row = 4
        ws['B4'] = "科目"
        ws['B4'].font = self.header_font
        ws['B4'].alignment = self.center_align
        ws['B4'].fill = self.header_fill

        ws['C4'] = "当期"
        ws['C4'].font = self.header_font
        ws['C4'].alignment = self.center_align
        ws['C4'].fill = self.header_fill

        ws['D4'] = "前期"
        ws['D4'].font = self.header_font
        ws['D4'].alignment = self.center_align
        ws['D4'].fill = self.header_fill

        # データ（従来と同じ）
        pl_data = [
            ("Ⅰ 売上高", 120000000, 110000000, True),
            ("", None, None, False),
            ("Ⅱ 売上原価", 74500000, 67500000, True),
            ("売上総利益", 45500000, 42500000, True),
            ("", None, None, False),
            ("Ⅲ 販売費及び一般管理費", None, None, True),
            ("　役員報酬", 12000000, 12000000, False),
            ("　給料手当", 18000000, 16000000, False),
            ("　法定福利費", 2500000, 2300000, False),
            ("　賃借料", 3600000, 3600000, False),
            ("　減価償却費", 2800000, 2900000, False),
            ("　その他", 4100000, 3800000, False),
            ("　販管費合計", 43000000, 40600000, True),
            ("", None, None, False),
            ("営業利益", 2500000, 1900000, True),
            ("", None, None, False),
            ("Ⅳ 営業外収益", 250000, 195000, True),
            ("Ⅴ 営業外費用", 800000, 900000, True),
            ("経常利益", 1950000, 1195000, True),
            ("", None, None, False),
            ("Ⅵ 特別利益", 0, 500000, True),
            ("Ⅶ 特別損失", 200000, 0, True),
            ("税引前当期純利益", 1750000, 1695000, True),
            ("", None, None, False),
            ("法人税等", 525000, 509000, False),
            ("当期純利益", 1225000, 1186000, True)
        ]

        # データ出力
        row = 5
        for i, (item, current, previous, is_total) in enumerate(pl_data):
            current_row = row + i

            ws[f'B{current_row}'] = item
            ws[f'B{current_row}'].font = self.header_font if is_total else self.normal_font
            ws[f'B{current_row}'].alignment = self.left_align
            if is_total:
                ws[f'B{current_row}'].fill = self.subtotal_fill

            if current is not None:
                ws[f'C{current_row}'] = f"{current:,}"
                ws[f'C{current_row}'].font = self.header_font if is_total else self.normal_font
                ws[f'C{current_row}'].alignment = self.right_align
                if is_total:
                    ws[f'C{current_row}'].fill = self.subtotal_fill
                if item == "当期純利益":
                    ws[f'C{current_row}'].border = self.double_bottom_border

            if previous is not None:
                ws[f'D{current_row}'] = f"{previous:,}"
                ws[f'D{current_row}'].font = self.header_font if is_total else self.normal_font
                ws[f'D{current_row}'].alignment = self.right_align
                if is_total:
                    ws[f'D{current_row}'].fill = self.subtotal_fill
                if item == "当期純利益":
                    ws[f'D{current_row}'].border = self.double_bottom_border

        # 罫線適用
        max_row = row + len(pl_data) - 1
        for r in range(4, max_row + 1):
            for col in ['A', 'B', 'C', 'D']:
                if ws[f'{col}{r}'].value is not None or r == 4:
                    ws[f'{col}{r}'].border = self.thin_border

    def create_cash_flow_statement_portrait(self, ws, company_name, fiscal_year):
        """A4縦向けキャッシュフロー計算書作成"""
        self.setup_a4_portrait(ws)

        # 列幅設定
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 32
        ws.column_dimensions['C'].width = 14

        # タイトル部分
        ws.merge_cells('A1:C1')
        ws['A1'] = "キャッシュ・フロー計算書"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = self.center_align

        ws.merge_cells('A2:C2')
        ws['A2'] = f"{company_name}　{fiscal_year}年3月期　（単位：円）"
        ws['A2'].font = self.small_font
        ws['A2'].alignment = self.center_align

        # ヘッダー行
        row = 4
        ws['B4'] = "科目"
        ws['B4'].font = self.header_font
        ws['B4'].alignment = self.center_align
        ws['B4'].fill = self.header_fill

        ws['C4'] = "金額"
        ws['C4'].font = self.header_font
        ws['C4'].alignment = self.center_align
        ws['C4'].fill = self.header_fill

        # データ（従来と同じ）
        cf_data = [
            ("Ⅰ 営業活動によるキャッシュ・フロー", None, True),
            ("　税引前当期純利益", 1750000, False),
            ("　減価償却費", 2800000, False),
            ("　引当金の増減額", 500000, False),
            ("　受取利息及び受取配当金", -250000, False),
            ("　支払利息", 800000, False),
            ("　売上債権の増減額", -1500000, False),
            ("　たな卸資産の増減額", -500000, False),
            ("　仕入債務の増減額", 800000, False),
            ("　その他", -200000, False),
            ("　小計", 4200000, True),
            ("　利息及び配当金の受取額", 250000, False),
            ("　利息の支払額", -800000, False),
            ("　法人税等の支払額", -500000, False),
            ("　営業活動CF", 3150000, True),
            ("", None, False),
            ("Ⅱ 投資活動によるキャッシュ・フロー", None, True),
            ("　有形固定資産の取得", -5000000, False),
            ("　無形固定資産の取得", -800000, False),
            ("　投資有価証券の取得", -1000000, False),
            ("　その他", -200000, False),
            ("　投資活動CF", -7000000, True),
            ("", None, False),
            ("Ⅲ 財務活動によるキャッシュ・フロー", None, True),
            ("　短期借入金の純増減額", 1000000, False),
            ("　長期借入れによる収入", 5000000, False),
            ("　長期借入金の返済", -2000000, False),
            ("　配当金の支払額", -500000, False),
            ("　財務活動CF", 3500000, True),
            ("", None, False),
            ("現金及び現金同等物の増減額", -350000, True),
            ("現金及び現金同等物の期首残高", 4850000, True),
            ("現金及び現金同等物の期末残高", 4500000, True)
        ]

        # データ出力
        row = 5
        for i, (item, amount, is_total) in enumerate(cf_data):
            current_row = row + i

            ws[f'B{current_row}'] = item
            ws[f'B{current_row}'].font = self.header_font if is_total else self.normal_font
            ws[f'B{current_row}'].alignment = self.left_align
            if is_total:
                ws[f'B{current_row}'].fill = self.subtotal_fill

            if amount is not None:
                ws[f'C{current_row}'] = f"{amount:,}"
                ws[f'C{current_row}'].font = self.header_font if is_total else self.normal_font
                ws[f'C{current_row}'].alignment = self.right_align
                if is_total:
                    ws[f'C{current_row}'].fill = self.subtotal_fill
                if "期末残高" in item:
                    ws[f'C{current_row}'].border = self.double_bottom_border

        # 罫線適用
        max_row = row + len(cf_data) - 1
        for r in range(4, max_row + 1):
            for col in ['A', 'B', 'C']:
                if ws[f'{col}{r}'].value is not None or r == 4:
                    ws[f'{col}{r}'].border = self.thin_border

    def create_equity_statement_landscape(self, ws, company_name, fiscal_year):
        """A4横向け株主資本等変動計算書作成"""
        self.setup_a4_landscape(ws)

        # 列幅設定（横向けで余裕を持って）
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 15

        # タイトル部分
        ws.merge_cells('A1:H1')
        ws['A1'] = "株主資本等変動計算書"
        ws['A1'].font = self.title_font_landscape
        ws['A1'].alignment = self.center_align

        ws.merge_cells('A2:H2')
        ws['A2'] = f"{company_name}　{fiscal_year}年3月期　（単位：円）"
        ws['A2'].font = self.small_font_landscape
        ws['A2'].alignment = self.center_align

        # ヘッダー行（横向けで拡張）
        row = 4
        ws.merge_cells(f'B{row}:F{row}')
        ws[f'B{row}'] = "株主資本"
        ws[f'B{row}'].font = self.header_font_landscape
        ws[f'B{row}'].alignment = self.center_align
        ws[f'B{row}'].fill = self.header_fill

        ws[f'G{row}'] = "評価・換算差額等"
        ws[f'G{row}'].font = self.header_font_landscape
        ws[f'G{row}'].alignment = self.center_align
        ws[f'G{row}'].fill = self.header_fill

        ws[f'H{row}'] = "純資産合計"
        ws[f'H{row}'].font = self.header_font_landscape
        ws[f'H{row}'].alignment = self.center_align
        ws[f'H{row}'].fill = self.header_fill

        # サブヘッダー
        row = 5
        headers = ["", "資本金", "資本剰余金", "利益剰余金", "自己株式", "株主資本合計", "その他有価証券評価差額金", "合計"]
        for i, header in enumerate(headers):
            col = chr(ord('A') + i)
            ws[f'{col}{row}'] = header
            ws[f'{col}{row}'].font = self.header_font_landscape
            ws[f'{col}{row}'].alignment = self.center_align
            ws[f'{col}{row}'].fill = self.subtotal_fill

        # データ（拡張版）
        equity_data = [
            ("当期首残高", 25000000, 5000000, 18275000, 0, 48275000, 0, 48275000),
            ("当期変動額", None, None, None, None, None, None, None),
            ("　剰余金の配当", 0, 0, -500000, 0, -500000, 0, -500000),
            ("　当期純利益", 0, 0, 1225000, 0, 1225000, 0, 1225000),
            ("　株主資本以外の項目の当期変動額（純額）", 0, 0, 0, 0, 0, 0, 0),
            ("当期変動額合計", 0, 0, 725000, 0, 725000, 0, 725000),
            ("当期末残高", 25000000, 5000000, 19000000, 0, 49000000, 0, 49000000)
        ]

        # データ出力
        row = 6
        for i, data_row in enumerate(equity_data):
            current_row = row + i
            item = data_row[0]

            ws[f'A{current_row}'] = item
            ws[f'A{current_row}'].font = self.header_font_landscape if "残高" in item or "合計" in item else self.normal_font_landscape
            ws[f'A{current_row}'].alignment = self.left_align

            for j, value in enumerate(data_row[1:], 1):
                col = chr(ord('A') + j)
                if value is not None:
                    ws[f'{col}{current_row}'] = f"{value:,}"
                    ws[f'{col}{current_row}'].font = self.header_font_landscape if "残高" in item or "合計" in item else self.normal_font_landscape
                    ws[f'{col}{current_row}'].alignment = self.right_align
                    if item == "当期末残高":
                        ws[f'{col}{current_row}'].border = self.double_bottom_border
                else:
                    ws[f'{col}{current_row}'] = ""

            if "残高" in item or "合計" in item:
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                    ws[f'{col}{current_row}'].fill = self.subtotal_fill

        # 罫線適用
        max_row = row + len(equity_data) - 1
        for r in range(4, max_row + 1):
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                if ws[f'{col}{r}'].value is not None or r in [4, 5]:
                    ws[f'{col}{r}'].border = self.thin_border

    def create_notes_landscape(self, ws, company_name, fiscal_year):
        """A4横向け附属明細書作成"""
        self.setup_a4_landscape(ws)

        # 列幅設定（横向けで拡張）
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15

        # タイトル部分
        ws.merge_cells('A1:H1')
        ws['A1'] = "附属明細書"
        ws['A1'].font = self.title_font_landscape
        ws['A1'].alignment = self.center_align

        ws.merge_cells('A2:H2')
        ws['A2'] = f"{company_name}　{fiscal_year}年3月期"
        ws['A2'].font = self.small_font_landscape
        ws['A2'].alignment = self.center_align

        # 主な会計方針
        row = 4
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = "１．主な会計方針"
        ws[f'A{row}'].font = self.header_font_landscape
        ws[f'A{row}'].alignment = self.left_align
        ws[f'A{row}'].fill = self.header_fill

        # 会計方針（横向けで詳細版）
        accounting_policies = [
            "（１）有価証券の評価基準及び評価方法",
            "　　満期保有目的の債券：償却原価法（定額法）",
            "　　その他有価証券：時価法（評価差額は全部純資産直入法により処理し、売却原価は移動平均法により算定）",
            "",
            "（２）たな卸資産の評価基準及び評価方法",
            "　　商品：先入先出法による原価法（貸借対照表価額については収益性の低下に基づく簿価切下げの方法）",
            "",
            "（３）固定資産の減価償却の方法",
            "　　有形固定資産：定率法（ただし、建物については定額法）",
            "　　無形固定資産：定額法（自社利用のソフトウェアについては、社内における利用可能期間5年に基づく定額法）",
            "",
            "（４）引当金の計上基準",
            "　　賞与引当金：従業員への賞与支給に備えるため、支給見込額のうち当期負担額を計上",
            "　　退職給付引当金：従業員の退職給付に備えるため、期末における退職給付債務の見込額に基づき、当期末において発生していると認められる額を計上"
        ]

        row += 1
        for policy in accounting_policies:
            ws.merge_cells(f'A{row}:H{row}')
            ws[f'A{row}'] = policy
            ws[f'A{row}'].font = self.normal_font_landscape
            ws[f'A{row}'].alignment = self.left_align
            row += 1

        # 有形固定資産明細（横向けで詳細版）
        row += 1
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = "２．有形固定資産等明細書"
        ws[f'A{row}'].font = self.header_font_landscape
        ws[f'A{row}'].alignment = self.left_align
        ws[f'A{row}'].fill = self.header_fill

        row += 1
        # 詳細テーブルヘッダー
        fixed_asset_headers = ["資産の種類", "期首帳簿価額", "当期増加額", "当期減少額", "期末帳簿価額", "減価償却累計額", "期末取得価額", "摘要"]
        for i, header in enumerate(fixed_asset_headers):
            col = chr(ord('A') + i)
            ws[f'{col}{row}'] = header
            ws[f'{col}{row}'].font = self.header_font_landscape
            ws[f'{col}{row}'].alignment = self.center_align
            ws[f'{col}{row}'].fill = self.subtotal_fill

        # 詳細データ
        fixed_assets_data = [
            ("建物", 22000000, 0, 0, 20000000, 5000000, 25000000, "本社ビル"),
            ("構築物", 2000000, 0, 0, 1800000, 400000, 2200000, "駐車場設備"),
            ("機械装置", 8500000, 2000000, 0, 7000000, 8000000, 15000000, "製造設備"),
            ("車両運搬具", 1500000, 500000, 800000, 1200000, 1000000, 2200000, "営業車両"),
            ("工具器具備品", 800000, 300000, 0, 700000, 400000, 1100000, "事務用機器"),
            ("土地", 30000000, 0, 0, 30000000, 0, 30000000, "本社用地"),
            ("建設仮勘定", 0, 1000000, 1000000, 0, 0, 0, "設備改修"),
            ("合計", 64800000, 3800000, 1800000, 60700000, 14800000, 75500000, "")
        ]

        row += 1
        for asset_data in fixed_assets_data:
            for i, value in enumerate(asset_data):
                col = chr(ord('A') + i)
                if isinstance(value, int) and value != 0:
                    ws[f'{col}{row}'] = f"{value:,}"
                    ws[f'{col}{row}'].alignment = self.right_align
                else:
                    ws[f'{col}{row}'] = str(value)
                    ws[f'{col}{row}'].alignment = self.left_align if i in [0, 7] else self.right_align

                ws[f'{col}{row}'].font = self.header_font_landscape if asset_data[0] == "合計" else self.normal_font_landscape
                if asset_data[0] == "合計":
                    ws[f'{col}{row}'].fill = self.subtotal_fill
                    ws[f'{col}{row}'].border = self.double_bottom_border
                else:
                    ws[f'{col}{row}'].border = self.thin_border
            row += 1

        # 借入金明細（横向けで詳細版）
        row += 1
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = "３．借入金明細書"
        ws[f'A{row}'].font = self.header_font_landscape
        ws[f'A{row}'].alignment = self.left_align
        ws[f'A{row}'].fill = self.header_fill

        row += 1
        loan_headers = ["借入先", "期首残高", "当期増減", "期末残高", "平均利率", "返済期限", "担保の内容", "摘要"]
        for i, header in enumerate(loan_headers):
            col = chr(ord('A') + i)
            ws[f'{col}{row}'] = header
            ws[f'{col}{row}'].font = self.header_font_landscape
            ws[f'{col}{row}'].alignment = self.center_align
            ws[f'{col}{row}'].fill = self.subtotal_fill

        loans_data = [
            ("○○銀行", 15000000, 8000000, 23000000, "1.5%", "2030年3月", "本社ビル", "設備資金"),
            ("△△信用金庫", 3000000, -3000000, 0, "2.0%", "返済済み", "-", "運転資金"),
            ("合計", 18000000, 5000000, 23000000, "-", "-", "-", "")
        ]

        row += 1
        for loan_data in loans_data:
            for i, value in enumerate(loan_data):
                col = chr(ord('A') + i)
                if isinstance(value, int) and value != 0:
                    ws[f'{col}{row}'] = f"{value:,}"
                    ws[f'{col}{row}'].alignment = self.right_align
                else:
                    ws[f'{col}{row}'] = str(value)
                    ws[f'{col}{row}'].alignment = self.left_align if i in [0, 4, 5, 6, 7] else self.right_align

                ws[f'{col}{row}'].font = self.header_font_landscape if loan_data[0] == "合計" else self.normal_font_landscape
                if loan_data[0] == "合計":
                    ws[f'{col}{row}'].fill = self.subtotal_fill
                else:
                    ws[f'{col}{row}'].border = self.thin_border
            row += 1

        # 従業員数等
        row += 1
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = "４．従業員数等"
        ws[f'A{row}'].font = self.header_font_landscape
        ws[f'A{row}'].alignment = self.left_align
        ws[f'A{row}'].fill = self.header_fill

        row += 1
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = "従業員数：25名（男性15名、女性10名）　平均年齢：42歳　平均勤続年数：8年　役員報酬総額：12,000,000円　平均給与額（年額）：4,800,000円"
        ws[f'A{row}'].font = self.normal_font_landscape
        ws[f'A{row}'].alignment = self.left_align


def generate_mixed_orientation_financial_statements(company_name="株式会社サンプル", fiscal_year=2025):
    """混合レイアウト財務諸表Excel生成のメイン関数"""
    generator = MixedOrientationFinancialGenerator()
    wb = generator.generate_all_statements(company_name, fiscal_year)

    # BytesIOオブジェクトに保存
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


if __name__ == "__main__":
    # テスト実行
    print("=== 混合レイアウト財務諸表Excel生成テスト ===")

    try:
        output = generate_mixed_orientation_financial_statements()

        # ファイル保存
        filename = "/home/esan/employee-db/mixed_orientation_financial_statements.xlsx"
        with open(filename, 'wb') as f:
            f.write(output.getvalue())

        import os
        file_size = os.path.getsize(filename)
        print(f"✅ 混合レイアウト財務諸表Excel生成成功！")
        print(f"📁 ファイル: {filename}")
        print(f"📊 サイズ: {file_size:,} bytes")
        print(f"📋 レイアウト:")
        print(f"  📄 A4縦向け:")
        print(f"    ✓ 貸借対照表")
        print(f"    ✓ 損益計算書")
        print(f"    ✓ キャッシュフロー計算書")
        print(f"  📄 A4横向け:")
        print(f"    ✓ 株主資本等変動計算書（詳細版）")
        print(f"    ✓ 附属明細書（詳細版）")

    except Exception as e:
        print(f"❌ エラー: {e}")
        import traceback
        traceback.print_exc()