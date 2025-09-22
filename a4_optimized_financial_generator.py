#!/usr/bin/env python3
"""
A4縦サイズ最適化された日本の財務諸表Excel生成機能
"""

from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from datetime import datetime


class A4OptimizedFinancialGenerator:
    """A4縦サイズに最適化された財務諸表Excel生成クラス"""

    def __init__(self):
        # A4縦向けに最適化されたフォント・スタイル設定
        self.title_font = Font(name='ＭＳ ゴシック', size=12, bold=True)
        self.header_font = Font(name='ＭＳ ゴシック', size=10, bold=True)
        self.normal_font = Font(name='ＭＳ ゴシック', size=9)
        self.small_font = Font(name='ＭＳ ゴシック', size=8)

        # アライメント
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')
        self.right_align = Alignment(horizontal='right', vertical='center')

        # 罫線（細めに調整）
        self.thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        self.double_bottom_border = Border(bottom=Side(style='double'))

        # 背景色（薄めに調整）
        self.header_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
        self.subtotal_fill = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')

    def setup_a4_page(self, ws):
        """A4縦向けページ設定"""
        # ページ設定
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        # 余白設定（狭く設定）
        ws.page_margins = PageMargins(
            left=0.5, right=0.5, top=0.5, bottom=0.5,
            header=0.3, footer=0.3
        )

        # 印刷設定
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.print_options.horizontalCentered = True
        ws.page_setup.scale = 85  # 85%に縮小

    def generate_all_statements(self, company_name="株式会社サンプル", fiscal_year=2025):
        """全ての財務諸表を生成"""
        wb = openpyxl.Workbook()

        # 貸借対照表
        ws_bs = wb.active
        ws_bs.title = "貸借対照表"
        self.create_balance_sheet_a4(ws_bs, company_name, fiscal_year)

        # 損益計算書
        ws_pl = wb.create_sheet("損益計算書")
        self.create_income_statement_a4(ws_pl, company_name, fiscal_year)

        # キャッシュフロー計算書
        ws_cf = wb.create_sheet("キャッシュフロー計算書")
        self.create_cash_flow_statement_a4(ws_cf, company_name, fiscal_year)

        # 株主資本等変動計算書
        ws_eq = wb.create_sheet("株主資本等変動計算書")
        self.create_equity_statement_a4(ws_eq, company_name, fiscal_year)

        # 附属明細書
        ws_notes = wb.create_sheet("附属明細書")
        self.create_notes_a4(ws_notes, company_name, fiscal_year)

        return wb

    def create_balance_sheet_a4(self, ws, company_name, fiscal_year):
        """A4縦向け貸借対照表作成"""
        self.setup_a4_page(ws)

        # 列幅設定（A4に最適化）
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 2
        ws.column_dimensions['E'].width = 22
        ws.column_dimensions['F'].width = 12

        # タイトル部分（コンパクト化）
        ws.merge_cells('A1:F1')
        ws['A1'] = "貸借対照表"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = self.center_align

        ws.merge_cells('A2:F2')
        ws['A2'] = f"{company_name}　{fiscal_year}年3月31日現在　（単位：円）"
        ws['A2'].font = self.small_font
        ws['A2'].alignment = self.center_align

        # ヘッダー行（3行目から開始）
        row = 4
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = "資産の部"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].alignment = self.center_align
        ws[f'A{row}'].fill = self.header_fill

        ws.merge_cells(f'D{row}:F{row}')
        ws[f'D{row}'] = "負債及び純資産の部"
        ws[f'D{row}'].font = self.header_font
        ws[f'D{row}'].alignment = self.center_align
        ws[f'D{row}'].fill = self.header_fill

        # 資産の部データ（コンパクト化）
        row = 5
        assets_data = [
            ("Ⅰ 流動資産", None, True),
            ("　現金及び預金", 5000000, False),
            ("　売掛金", 8000000, False),
            ("　商品", 3000000, False),
            ("　その他", 2500000, False),
            ("　流動資産合計", 18500000, True),
            ("", None, False),
            ("Ⅱ 固定資産", None, True),
            ("　建物（純額）", 20000000, False),
            ("　機械装置（純額）", 7000000, False),
            ("　土地", 30000000, False),
            ("　ソフトウェア", 2000000, False),
            ("　投資有価証券", 5000000, False),
            ("　その他", 1000000, False),
            ("　固定資産合計", 65000000, True),
            ("", None, False),
            ("資産合計", 83500000, True)
        ]

        # 負債及び純資産の部データ（コンパクト化）
        liabilities_data = [
            ("Ⅰ 流動負債", None, True),
            ("　買掛金", 4000000, False),
            ("　短期借入金", 3000000, False),
            ("　未払金等", 4500000, False),
            ("　流動負債合計", 11500000, True),
            ("", None, False),
            ("Ⅱ 固定負債", None, True),
            ("　長期借入金", 20000000, False),
            ("　退職給付引当金", 3000000, False),
            ("　固定負債合計", 23000000, True),
            ("負債合計", 34500000, True),
            ("", None, False),
            ("Ⅲ 純資産の部", None, True),
            ("　資本金", 25000000, False),
            ("　資本剰余金", 5000000, False),
            ("　利益剰余金", 19000000, False),
            ("　純資産合計", 49000000, True),
            ("負債及び純資産合計", 83500000, True)
        ]

        # データ出力
        for i, (item, amount, is_total) in enumerate(assets_data):
            current_row = row + i
            ws[f'B{current_row}'] = item
            ws[f'B{current_row}'].font = self.header_font if is_total else self.normal_font
            ws[f'B{current_row}'].alignment = self.left_align
            if is_total:
                ws[f'B{current_row}'].fill = self.subtotal_fill

            if amount is not None:
                ws[f'C{current_row}'] = f"{amount:,}"
                ws[f'C{current_row}'].font = self.header_font if is_total else self.normal_font
                ws[f'C{current_row}'].alignment = self.right_align
                if is_total:
                    ws[f'C{current_row}'].fill = self.subtotal_fill
                if item == "資産合計":
                    ws[f'C{current_row}'].border = self.double_bottom_border

        # 負債及び純資産の部を出力
        for i, (item, amount, is_total) in enumerate(liabilities_data):
            current_row = row + i
            ws[f'E{current_row}'] = item
            ws[f'E{current_row}'].font = self.header_font if is_total else self.normal_font
            ws[f'E{current_row}'].alignment = self.left_align
            if is_total:
                ws[f'E{current_row}'].fill = self.subtotal_fill

            if amount is not None:
                ws[f'F{current_row}'] = f"{amount:,}"
                ws[f'F{current_row}'].font = self.header_font if is_total else self.normal_font
                ws[f'F{current_row}'].alignment = self.right_align
                if is_total:
                    ws[f'F{current_row}'].fill = self.subtotal_fill
                if item == "負債及び純資産合計":
                    ws[f'F{current_row}'].border = self.double_bottom_border

        # 罫線適用
        max_row = row + len(assets_data) - 1
        for r in range(4, max_row + 1):
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                if ws[f'{col}{r}'].value is not None or r == 4:
                    ws[f'{col}{r}'].border = self.thin_border

    def create_income_statement_a4(self, ws, company_name, fiscal_year):
        """A4縦向け損益計算書作成"""
        self.setup_a4_page(ws)

        # 列幅設定
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 26
        ws.column_dimensions['C'].width = 13
        ws.column_dimensions['D'].width = 13

        # タイトル部分
        ws.merge_cells('A1:D1')
        ws['A1'] = "損益計算書"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = self.center_align

        ws.merge_cells('A2:D2')
        ws['A2'] = f"{company_name}　{fiscal_year}年3月期　（単位：円）"
        ws['A2'].font = self.small_font
        ws['A2'].alignment = self.center_align

        # ヘッダー行
        row = 4
        ws['B4'] = "科目"
        ws['B4'].font = self.header_font
        ws['B4'].alignment = self.center_align
        ws['B4'].fill = self.header_fill

        ws['C4'] = "当期"
        ws['C4'].font = self.header_font
        ws['C4'].alignment = self.center_align
        ws['C4'].fill = self.header_fill

        ws['D4'] = "前期"
        ws['D4'].font = self.header_font
        ws['D4'].alignment = self.center_align
        ws['D4'].fill = self.header_fill

        # 損益計算書データ（A4に収まるよう簡略化）
        pl_data = [
            ("Ⅰ 売上高", 120000000, 110000000, True),
            ("", None, None, False),
            ("Ⅱ 売上原価", 74500000, 67500000, True),
            ("売上総利益", 45500000, 42500000, True),
            ("", None, None, False),
            ("Ⅲ 販売費及び一般管理費", None, None, True),
            ("　役員報酬", 12000000, 12000000, False),
            ("　給料手当", 18000000, 16000000, False),
            ("　法定福利費", 2500000, 2300000, False),
            ("　賃借料", 3600000, 3600000, False),
            ("　減価償却費", 2800000, 2900000, False),
            ("　その他", 4100000, 3800000, False),
            ("　販管費合計", 43000000, 40600000, True),
            ("", None, None, False),
            ("営業利益", 2500000, 1900000, True),
            ("", None, None, False),
            ("Ⅳ 営業外収益", 250000, 195000, True),
            ("Ⅴ 営業外費用", 800000, 900000, True),
            ("経常利益", 1950000, 1195000, True),
            ("", None, None, False),
            ("Ⅵ 特別利益", 0, 500000, True),
            ("Ⅶ 特別損失", 200000, 0, True),
            ("税引前当期純利益", 1750000, 1695000, True),
            ("", None, None, False),
            ("法人税等", 525000, 509000, False),
            ("当期純利益", 1225000, 1186000, True)
        ]

        # データ出力
        row = 5
        for i, (item, current, previous, is_total) in enumerate(pl_data):
            current_row = row + i

            ws[f'B{current_row}'] = item
            ws[f'B{current_row}'].font = self.header_font if is_total else self.normal_font
            ws[f'B{current_row}'].alignment = self.left_align
            if is_total:
                ws[f'B{current_row}'].fill = self.subtotal_fill

            if current is not None:
                ws[f'C{current_row}'] = f"{current:,}"
                ws[f'C{current_row}'].font = self.header_font if is_total else self.normal_font
                ws[f'C{current_row}'].alignment = self.right_align
                if is_total:
                    ws[f'C{current_row}'].fill = self.subtotal_fill
                if item == "当期純利益":
                    ws[f'C{current_row}'].border = self.double_bottom_border

            if previous is not None:
                ws[f'D{current_row}'] = f"{previous:,}"
                ws[f'D{current_row}'].font = self.header_font if is_total else self.normal_font
                ws[f'D{current_row}'].alignment = self.right_align
                if is_total:
                    ws[f'D{current_row}'].fill = self.subtotal_fill
                if item == "当期純利益":
                    ws[f'D{current_row}'].border = self.double_bottom_border

        # 罫線適用
        max_row = row + len(pl_data) - 1
        for r in range(4, max_row + 1):
            for col in ['A', 'B', 'C', 'D']:
                if ws[f'{col}{r}'].value is not None or r == 4:
                    ws[f'{col}{r}'].border = self.thin_border

    def create_cash_flow_statement_a4(self, ws, company_name, fiscal_year):
        """A4縦向けキャッシュフロー計算書作成"""
        self.setup_a4_page(ws)

        # 列幅設定
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 32
        ws.column_dimensions['C'].width = 14

        # タイトル部分
        ws.merge_cells('A1:C1')
        ws['A1'] = "キャッシュ・フロー計算書"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = self.center_align

        ws.merge_cells('A2:C2')
        ws['A2'] = f"{company_name}　{fiscal_year}年3月期　（単位：円）"
        ws['A2'].font = self.small_font
        ws['A2'].alignment = self.center_align

        # ヘッダー行
        row = 4
        ws['B4'] = "科目"
        ws['B4'].font = self.header_font
        ws['B4'].alignment = self.center_align
        ws['B4'].fill = self.header_fill

        ws['C4'] = "金額"
        ws['C4'].font = self.header_font
        ws['C4'].alignment = self.center_align
        ws['C4'].fill = self.header_fill

        # キャッシュフローデータ（A4に収まるよう簡略化）
        cf_data = [
            ("Ⅰ 営業活動によるキャッシュ・フロー", None, True),
            ("　税引前当期純利益", 1750000, False),
            ("　減価償却費", 2800000, False),
            ("　引当金の増減額", 500000, False),
            ("　受取利息及び受取配当金", -250000, False),
            ("　支払利息", 800000, False),
            ("　売上債権の増減額", -1500000, False),
            ("　たな卸資産の増減額", -500000, False),
            ("　仕入債務の増減額", 800000, False),
            ("　その他", -200000, False),
            ("　小計", 4200000, True),
            ("　利息及び配当金の受取額", 250000, False),
            ("　利息の支払額", -800000, False),
            ("　法人税等の支払額", -500000, False),
            ("　営業活動CF", 3150000, True),
            ("", None, False),
            ("Ⅱ 投資活動によるキャッシュ・フロー", None, True),
            ("　有形固定資産の取得", -5000000, False),
            ("　無形固定資産の取得", -800000, False),
            ("　投資有価証券の取得", -1000000, False),
            ("　その他", -200000, False),
            ("　投資活動CF", -7000000, True),
            ("", None, False),
            ("Ⅲ 財務活動によるキャッシュ・フロー", None, True),
            ("　短期借入金の純増減額", 1000000, False),
            ("　長期借入れによる収入", 5000000, False),
            ("　長期借入金の返済", -2000000, False),
            ("　配当金の支払額", -500000, False),
            ("　財務活動CF", 3500000, True),
            ("", None, False),
            ("現金及び現金同等物の増減額", -350000, True),
            ("現金及び現金同等物の期首残高", 4850000, True),
            ("現金及び現金同等物の期末残高", 4500000, True)
        ]

        # データ出力
        row = 5
        for i, (item, amount, is_total) in enumerate(cf_data):
            current_row = row + i

            ws[f'B{current_row}'] = item
            ws[f'B{current_row}'].font = self.header_font if is_total else self.normal_font
            ws[f'B{current_row}'].alignment = self.left_align
            if is_total:
                ws[f'B{current_row}'].fill = self.subtotal_fill

            if amount is not None:
                ws[f'C{current_row}'] = f"{amount:,}"
                ws[f'C{current_row}'].font = self.header_font if is_total else self.normal_font
                ws[f'C{current_row}'].alignment = self.right_align
                if is_total:
                    ws[f'C{current_row}'].fill = self.subtotal_fill
                if "期末残高" in item:
                    ws[f'C{current_row}'].border = self.double_bottom_border

        # 罫線適用
        max_row = row + len(cf_data) - 1
        for r in range(4, max_row + 1):
            for col in ['A', 'B', 'C']:
                if ws[f'{col}{r}'].value is not None or r == 4:
                    ws[f'{col}{r}'].border = self.thin_border

    def create_equity_statement_a4(self, ws, company_name, fiscal_year):
        """A4縦向け株主資本等変動計算書作成"""
        self.setup_a4_page(ws)

        # 列幅設定
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12

        # タイトル部分
        ws.merge_cells('A1:F1')
        ws['A1'] = "株主資本等変動計算書"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = self.center_align

        ws.merge_cells('A2:F2')
        ws['A2'] = f"{company_name}　{fiscal_year}年3月期　（単位：円）"
        ws['A2'].font = self.small_font
        ws['A2'].alignment = self.center_align

        # ヘッダー行
        row = 4
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'B{row}'] = "株主資本"
        ws[f'B{row}'].font = self.header_font
        ws[f'B{row}'].alignment = self.center_align
        ws[f'B{row}'].fill = self.header_fill

        ws[f'F{row}'] = "合計"
        ws[f'F{row}'].font = self.header_font
        ws[f'F{row}'].alignment = self.center_align
        ws[f'F{row}'].fill = self.header_fill

        # サブヘッダー
        row = 5
        headers = ["", "資本金", "資本剰余金", "利益剰余金", "自己株式", "合計"]
        for i, header in enumerate(headers):
            col = chr(ord('A') + i)
            ws[f'{col}{row}'] = header
            ws[f'{col}{row}'].font = self.header_font
            ws[f'{col}{row}'].alignment = self.center_align
            ws[f'{col}{row}'].fill = self.subtotal_fill

        # データ（簡略化）
        equity_data = [
            ("当期首残高", 25000000, 5000000, 18275000, 0, 48275000),
            ("当期変動額", None, None, None, None, None),
            ("　剰余金の配当", 0, 0, -500000, 0, -500000),
            ("　当期純利益", 0, 0, 1225000, 0, 1225000),
            ("当期変動額合計", 0, 0, 725000, 0, 725000),
            ("当期末残高", 25000000, 5000000, 19000000, 0, 49000000)
        ]

        # データ出力
        row = 6
        for i, (item, capital, surplus, retained, treasury, total) in enumerate(equity_data):
            current_row = row + i

            ws[f'A{current_row}'] = item
            ws[f'A{current_row}'].font = self.header_font if "残高" in item or "合計" in item else self.normal_font
            ws[f'A{current_row}'].alignment = self.left_align

            if capital is not None:
                ws[f'B{current_row}'] = f"{capital:,}"
                ws[f'B{current_row}'].font = self.header_font if "残高" in item or "合計" in item else self.normal_font
                ws[f'B{current_row}'].alignment = self.right_align
            else:
                ws[f'B{current_row}'] = ""

            if surplus is not None:
                ws[f'C{current_row}'] = f"{surplus:,}"
                ws[f'C{current_row}'].font = self.header_font if "残高" in item or "合計" in item else self.normal_font
                ws[f'C{current_row}'].alignment = self.right_align
            else:
                ws[f'C{current_row}'] = ""

            if retained is not None:
                ws[f'D{current_row}'] = f"{retained:,}"
                ws[f'D{current_row}'].font = self.header_font if "残高" in item or "合計" in item else self.normal_font
                ws[f'D{current_row}'].alignment = self.right_align
            else:
                ws[f'D{current_row}'] = ""

            if treasury is not None:
                ws[f'E{current_row}'] = f"{treasury:,}"
                ws[f'E{current_row}'].font = self.header_font if "残高" in item or "合計" in item else self.normal_font
                ws[f'E{current_row}'].alignment = self.right_align
            else:
                ws[f'E{current_row}'] = ""

            if total is not None:
                ws[f'F{current_row}'] = f"{total:,}"
                ws[f'F{current_row}'].font = self.header_font if "残高" in item or "合計" in item else self.normal_font
                ws[f'F{current_row}'].alignment = self.right_align
                if item == "当期末残高":
                    ws[f'F{current_row}'].border = self.double_bottom_border
            else:
                ws[f'F{current_row}'] = ""

            if "残高" in item or "合計" in item:
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    ws[f'{col}{current_row}'].fill = self.subtotal_fill

        # 罫線適用
        max_row = row + len(equity_data) - 1
        for r in range(4, max_row + 1):
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                if ws[f'{col}{r}'].value is not None or r in [4, 5]:
                    ws[f'{col}{r}'].border = self.thin_border

    def create_notes_a4(self, ws, company_name, fiscal_year):
        """A4縦向け附属明細書作成"""
        self.setup_a4_page(ws)

        # 列幅設定
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 44

        # タイトル部分
        ws.merge_cells('A1:B1')
        ws['A1'] = "附属明細書"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = self.center_align

        ws.merge_cells('A2:B2')
        ws['A2'] = f"{company_name}　{fiscal_year}年3月期"
        ws['A2'].font = self.small_font
        ws['A2'].alignment = self.center_align

        # 主な会計方針（A4に収まるよう簡略化）
        row = 4
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = "１．主な会計方針"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].alignment = self.left_align
        ws[f'A{row}'].fill = self.header_fill

        accounting_policies = [
            "（１）有価証券の評価基準及び評価方法",
            "　満期保有目的の債券：償却原価法",
            "　その他有価証券：時価法（評価差額は全部純資産直入法）",
            "",
            "（２）たな卸資産の評価基準及び評価方法",
            "　商品：先入先出法による原価法",
            "",
            "（３）固定資産の減価償却の方法",
            "　有形固定資産：定率法、無形固定資産：定額法",
            "",
            "（４）引当金の計上基準",
            "　賞与引当金：従業員への賞与支給に備えるため、支給見込額に基づき計上",
            "　退職給付引当金：従業員の退職給付に備えるため、期末における退職給付債務の見込額に基づき計上"
        ]

        row += 1
        for policy in accounting_policies:
            ws[f'B{row}'] = policy
            ws[f'B{row}'].font = self.normal_font
            ws[f'B{row}'].alignment = self.left_align
            row += 1

        # 固定資産明細（コンパクト化）
        row += 1
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = "２．有形固定資産の明細"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].alignment = self.left_align
        ws[f'A{row}'].fill = self.header_fill

        row += 1
        # テーブル形式（A4に収まるよう列を調整）
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10

        fixed_asset_headers = ["資産の種類", "期首帳簿価額", "当期増加額", "当期減少額", "期末帳簿価額"]
        for i, header in enumerate(fixed_asset_headers):
            col = chr(ord('A') + i)
            ws[f'{col}{row}'] = header
            ws[f'{col}{row}'].font = self.header_font
            ws[f'{col}{row}'].alignment = self.center_align
            ws[f'{col}{row}'].fill = self.subtotal_fill

        fixed_assets_data = [
            ("建物", 22000000, 0, 0, 20000000),
            ("機械装置", 8500000, 2000000, 0, 7000000),
            ("土地", 30000000, 0, 0, 30000000),
            ("合計", 60500000, 2000000, 0, 57000000)
        ]

        row += 1
        for asset_name, beginning, increase, decrease, ending in fixed_assets_data:
            ws[f'A{row}'] = asset_name
            ws[f'B{row}'] = f"{beginning:,}"
            ws[f'C{row}'] = f"{increase:,}"
            ws[f'D{row}'] = f"{decrease:,}"
            ws[f'E{row}'] = f"{ending:,}"

            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{row}'].font = self.header_font if asset_name == "合計" else self.normal_font
                ws[f'{col}{row}'].alignment = self.center_align if col == 'A' else self.right_align
                if asset_name == "合計":
                    ws[f'{col}{row}'].fill = self.subtotal_fill
                    ws[f'{col}{row}'].border = self.double_bottom_border
            row += 1

        # 従業員数等（簡略化）
        row += 1
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = "３．従業員数等"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].alignment = self.left_align
        ws[f'A{row}'].fill = self.header_fill

        row += 1
        ws[f'B{row}'] = "従業員数：25名　役員報酬総額：12,000,000円　平均給与額（年額）：4,800,000円"
        ws[f'B{row}'].font = self.normal_font
        ws[f'B{row}'].alignment = self.left_align


def generate_a4_financial_statements(company_name="株式会社サンプル", fiscal_year=2025):
    """A4縦向け最適化された財務諸表Excel生成のメイン関数"""
    generator = A4OptimizedFinancialGenerator()
    wb = generator.generate_all_statements(company_name, fiscal_year)

    # BytesIOオブジェクトに保存
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


if __name__ == "__main__":
    # テスト実行
    print("=== A4縦向け最適化財務諸表Excel生成テスト ===")

    try:
        output = generate_a4_financial_statements()

        # ファイル保存
        filename = "/home/esan/employee-db/a4_optimized_financial_statements.xlsx"
        with open(filename, 'wb') as f:
            f.write(output.getvalue())

        import os
        file_size = os.path.getsize(filename)
        print(f"✅ A4最適化財務諸表Excel生成成功！")
        print(f"📁 ファイル: {filename}")
        print(f"📊 サイズ: {file_size:,} bytes")
        print(f"📋 最適化内容:")
        print(f"  ✓ A4縦向けページ設定")
        print(f"  ✓ 印刷倍率85%設定")
        print(f"  ✓ 適切な余白設定")
        print(f"  ✓ コンパクト化されたレイアウト")
        print(f"  ✓ フォントサイズ最適化")
        print(f"  ✓ 列幅調整")

    except Exception as e:
        print(f"❌ エラー: {e}")
        import traceback
        traceback.print_exc()