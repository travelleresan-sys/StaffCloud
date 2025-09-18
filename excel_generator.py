#!/usr/bin/env python3
"""
従業員情報Excel生成機能
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import date

def create_employee_excel(employee, leave_data=None):
    """
    従業員情報をExcel形式で生成
    
    Args:
        employee: Employee オブジェクト
        leave_data: 年休データ辞書（オプション）
        
    Returns:
        BytesIO: Excel ファイルのバイナリデータ
    """
    
    # 新しいワークブック作成
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"従業員情報_{employee.name}"
    
    # スタイル設定
    header_font = Font(name='メイリオ', size=14, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    title_font = Font(name='メイリオ', size=16, bold=True)
    normal_font = Font(name='メイリオ', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # タイトル行
    ws['A1'] = f"従業員情報 - {employee.name}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:B1')
    
    # 発行日
    ws['A2'] = f"発行日: {date.today().strftime('%Y年%m月%d日')}"
    ws['A2'].font = normal_font
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:B2')
    
    # 空行
    current_row = 4
    
    # 基本情報セクション
    ws[f'A{current_row}'] = "基本情報"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
    ws[f'B{current_row}'].fill = header_fill
    ws.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 1
    
    # 基本情報データ
    basic_info = [
        ('氏名', employee.name),
        ('生年月日', employee.birth_date.strftime('%Y年%m月%d日') if employee.birth_date else '未設定'),
        ('性別', employee.gender or '未設定'),
        ('入社年月日', employee.join_date.strftime('%Y年%m月%d日') if employee.join_date else '未設定'),
        ('電話番号', employee.phone_number or '未設定'),
        ('住所', employee.address or '未設定'),
        ('国籍', employee.nationality or '未設定'),
        ('在留カード期限', employee.residence_card_expiry.strftime('%Y年%m月%d日') if employee.residence_card_expiry else '未設定'),
        ('自動車保険満了日', employee.car_insurance_expiry.strftime('%Y年%m月%d日') if employee.car_insurance_expiry else '未設定'),
        ('在籍状況', employee.status or '未設定'),
    ]
    
    for label, value in basic_info:
        ws[f'A{current_row}'] = label
        ws[f'B{current_row}'] = value
        
        # スタイル適用
        ws[f'A{current_row}'].font = Font(name='メイリオ', size=11, bold=True)
        ws[f'B{current_row}'].font = normal_font
        ws[f'A{current_row}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        # ボーダー
        ws[f'A{current_row}'].border = border
        ws[f'B{current_row}'].border = border
        
        current_row += 1
    
    # 空行
    current_row += 1
    
    # 年次有給休暇情報
    if leave_data:
        ws[f'A{current_row}'] = "年次有給休暇情報"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].fill = PatternFill(start_color='28a745', end_color='28a745', fill_type='solid')
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        ws[f'B{current_row}'].fill = PatternFill(start_color='28a745', end_color='28a745', fill_type='solid')
        ws.merge_cells(f'A{current_row}:B{current_row}')
        current_row += 1
        
        leave_info = [
            ('付与日数合計', f"{leave_data.get('total_credited', 0)}日"),
            ('取得日数合計', f"{leave_data.get('total_taken', 0)}日"),
            ('残日数', f"{leave_data.get('remaining_leave', 0)}日"),
            ('法定付与日数', f"{leave_data.get('legal_leave_days', 0)}日"),
        ]
        
        for label, value in leave_info:
            ws[f'A{current_row}'] = label
            ws[f'B{current_row}'] = value
            
            # スタイル適用
            ws[f'A{current_row}'].font = Font(name='メイリオ', size=11, bold=True)
            ws[f'B{current_row}'].font = normal_font
            ws[f'A{current_row}'].fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
            
            # ボーダー
            ws[f'A{current_row}'].border = border
            ws[f'B{current_row}'].border = border
            
            current_row += 1
        
        # 年休付与履歴
        if leave_data.get('leave_credits'):
            current_row += 1
            ws[f'A{current_row}'] = "年休付与履歴"
            ws[f'A{current_row}'].font = header_font
            ws[f'A{current_row}'].fill = PatternFill(start_color='8e44ad', end_color='8e44ad', fill_type='solid')
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
            ws[f'B{current_row}'].fill = PatternFill(start_color='8e44ad', end_color='8e44ad', fill_type='solid')
            ws.merge_cells(f'A{current_row}:B{current_row}')
            current_row += 1
            
            # ヘッダー
            ws[f'A{current_row}'] = "付与日"
            ws[f'B{current_row}'] = "付与日数"
            ws[f'A{current_row}'].font = Font(name='メイリオ', size=11, bold=True)
            ws[f'B{current_row}'].font = Font(name='メイリオ', size=11, bold=True)
            ws[f'A{current_row}'].fill = PatternFill(start_color='D2C4E8', end_color='D2C4E8', fill_type='solid')
            ws[f'B{current_row}'].fill = PatternFill(start_color='D2C4E8', end_color='D2C4E8', fill_type='solid')
            ws[f'A{current_row}'].border = border
            ws[f'B{current_row}'].border = border
            current_row += 1
            
            # データ
            for credit in leave_data['leave_credits'][:10]:  # 最大10件
                ws[f'A{current_row}'] = credit.date_credited.strftime('%Y年%m月%d日')
                ws[f'B{current_row}'] = f"{credit.days_credited}日"
                ws[f'A{current_row}'].font = normal_font
                ws[f'B{current_row}'].font = normal_font
                ws[f'A{current_row}'].border = border
                ws[f'B{current_row}'].border = border
                current_row += 1
        
        # 年休取得履歴
        if leave_data.get('leave_records'):
            current_row += 1
            ws[f'A{current_row}'] = "年休取得履歴"
            ws[f'A{current_row}'].font = header_font
            ws[f'A{current_row}'].fill = PatternFill(start_color='8e44ad', end_color='8e44ad', fill_type='solid')
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
            ws[f'B{current_row}'].fill = PatternFill(start_color='8e44ad', end_color='8e44ad', fill_type='solid')
            ws.merge_cells(f'A{current_row}:B{current_row}')
            current_row += 1
            
            # ヘッダー
            ws[f'A{current_row}'] = "取得日"
            ws[f'B{current_row}'] = "取得日数"
            ws[f'A{current_row}'].font = Font(name='メイリオ', size=11, bold=True)
            ws[f'B{current_row}'].font = Font(name='メイリオ', size=11, bold=True)
            ws[f'A{current_row}'].fill = PatternFill(start_color='D2C4E8', end_color='D2C4E8', fill_type='solid')
            ws[f'B{current_row}'].fill = PatternFill(start_color='D2C4E8', end_color='D2C4E8', fill_type='solid')
            ws[f'A{current_row}'].border = border
            ws[f'B{current_row}'].border = border
            current_row += 1
            
            # データ
            for record in leave_data['leave_records'][:10]:  # 最大10件
                ws[f'A{current_row}'] = record.date_taken.strftime('%Y年%m月%d日')
                ws[f'B{current_row}'] = f"{record.days_taken}日"
                ws[f'A{current_row}'].font = normal_font
                ws[f'B{current_row}'].font = normal_font
                ws[f'A{current_row}'].border = border
                ws[f'B{current_row}'].border = border
                current_row += 1
    
    # 列幅調整
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    
    # BytesIOに保存
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer

def test_excel_generation():
    """Excel生成機能をテスト"""
    from app import app, Employee, LeaveCredit, LeaveRecord, db
    
    print("=== Excel生成テスト ===")
    
    with app.app_context():
        employees = Employee.query.all()
        
        for employee in employees:
            try:
                print(f"従業員: {employee.name}")
                
                # 年休データを取得
                from sqlalchemy import func
                total_credited = db.session.query(func.sum(LeaveCredit.days_credited))\
                    .filter_by(employee_id=employee.id).scalar() or 0
                total_taken = db.session.query(func.sum(LeaveRecord.days_taken))\
                    .filter_by(employee_id=employee.id).scalar() or 0
                
                leave_credits = LeaveCredit.query.filter_by(employee_id=employee.id)\
                    .order_by(LeaveCredit.date_credited.desc()).all()
                leave_records = LeaveRecord.query.filter_by(employee_id=employee.id)\
                    .order_by(LeaveRecord.date_taken.desc()).all()
                
                leave_data = {
                    'total_credited': total_credited,
                    'total_taken': total_taken,
                    'remaining_leave': total_credited - total_taken,
                    'legal_leave_days': 20,  # 簡易計算
                    'leave_credits': leave_credits,
                    'leave_records': leave_records
                }
                
                # Excel生成
                excel_buffer = create_employee_excel(employee, leave_data)
                
                # ファイルに保存
                filename = f"employee_{employee.id}.xlsx"
                with open(filename, 'wb') as f:
                    f.write(excel_buffer.read())
                
                import os
                file_size = os.path.getsize(filename)
                print(f"  ✓ Excel生成成功: {filename} ({file_size} bytes)")
                
            except Exception as e:
                print(f"  ✗ Excel生成失敗: {e}")
                import traceback
                traceback.print_exc()

if __name__ == "__main__":
    test_excel_generation()