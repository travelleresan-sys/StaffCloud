from flask import Flask, render_template, redirect, url_for, request, flash, make_response, jsonify
import urllib.parse
# Flask-Login: User session management for Flask
from flask_login import LoginManager, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import os
import uuid
import io
from io import BytesIO
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage, PageBreak
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.utils import ImageReader
from weasyprint import HTML, CSS
from models import (db, User, Employee, LeaveCredit, LeaveRecord, 
                    CompanyCalendar, CalendarSettings, LeaveRequest, PersonalInfoRequest, PerformanceEvaluation,
                    WorkingTimeRecord, PayrollCalculation, CompanySettings, LaborStandardsSettings, LegalHolidaySettings, Agreement36History, Agreement36,
                    PayrollSlip, EmployeePayrollSettings, AccountingAccount, JournalEntry, JournalEntryDetail, GeneralLedger, TransactionPattern, BusinessPartner,
                    AccountingPeriod, OpeningBalance)
from payroll_slip_pdf_generator import create_payroll_slip_pdf
from timecard_pdf_generator import create_timecard_pdf
from datetime import date, datetime, timedelta
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your_secret_key'  # 必ず後で変更してください
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///employees.db'

# 画像アップロード用の設定
UPLOAD_FOLDER = 'static/uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
ALLOWED_RESIDENCE_CARD_EXTENSIONS = {'png', 'jpg', 'jpeg', 'pdf'}
ALLOWED_CAR_INSURANCE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'pdf'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# アップロードフォルダが存在しない場合は作成
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# データベース初期化
db.init_app(app)

# Login Manager初期化
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# 年次有給休暇の自動付与ロジック
def calculate_annual_leave_days(join_date, current_date=None):
    """
    日本の労働基準法に基づく年次有給休暇の付与日数を計算
    
    労働基準法第39条に基づく：
    - 6ヶ月継続勤務：10日付与
    - 1年6ヶ月継続勤務：11日付与
    - 2年6ヶ月継続勤務：12日付与
    - 以降1年ごとに1日追加（最大20日）
    """
    if current_date is None:
        current_date = date.today()
    
    # 入社日から現在までの勤続期間を計算
    years_of_service = (current_date - join_date).days / 365.25
    
    if years_of_service < 0.5:  # 6ヶ月未満
        return 0
    elif years_of_service < 1.5:  # 6ヶ月〜1年6ヶ月未満
        return 10
    elif years_of_service < 2.5:  # 1年6ヶ月〜2年6ヶ月未満
        return 11
    else:  # 2年6ヶ月以降
        # 2年6ヶ月以降は1年ごとに1日追加（最大20日）
        additional_days = min(int(years_of_service - 2.5), 8)  # 最大8日追加
        return 12 + additional_days

def should_auto_grant_leave(employee, current_date=None):
    """
    年次有給休暇の自動付与が必要かどうかを判定
    入社から6ヶ月経過時点から自動付与可能
    """
    if current_date is None:
        current_date = date.today()
    
    # 入社日がNoneの場合は付与しない
    if not employee.join_date:
        return False
    
    # 入社日から6ヶ月経過しているかチェック
    days_since_join = (current_date - employee.join_date).days
    
    if days_since_join < 182:  # 約6ヶ月 (182日)
        return False
    
    # 前回の自動付与日をチェック
    last_auto_grant = db.session.query(LeaveCredit)\
        .filter(LeaveCredit.employee_id == employee.id)\
        .order_by(LeaveCredit.date_credited.desc())\
        .first()
    
    # 初回付与（6ヶ月経過時点）の場合
    if not last_auto_grant:
        return days_since_join >= 182  # 6ヶ月経過していれば付与
    
    # 2回目以降の付与（前回から1年経過）の場合
    days_since_last_grant = (current_date - last_auto_grant.date_credited).days
    if days_since_last_grant >= 365:  # 1年経過していれば付与
        return True
    
    return False

@app.route('/auto_grant_annual_leave', methods=['GET', 'POST'])
@login_required
def auto_grant_annual_leave():
    # ... 関数の処理 ...
    if current_user.role != 'admin':
        flash('アクセス権がありません。')
        return redirect(url_for('leave_management'))
    
    try:
        current_date = date.today()
        granted_count = 0
        
        # 全従業員をチェック
        employees = Employee.query.filter_by(status='在籍中').all()
        
        for employee in employees:
            if should_auto_grant_leave(employee, current_date):
                # 法律に基づく付与日数を計算
                days_to_grant = calculate_annual_leave_days(employee.join_date, current_date)
                
                if days_to_grant > 0:
                    # 自動付与を記録
                    new_leave_credit = LeaveCredit(
                        employee_id=employee.id,
                        days_credited=days_to_grant,
                        date_credited=current_date
                    )
                    
                    db.session.add(new_leave_credit)
                    granted_count += 1
        
        if granted_count > 0:
            db.session.commit()
            flash(f'{granted_count}名の従業員に年次有給休暇を自動付与しました。')
        else:
            flash('自動付与対象の従業員はいませんでした。')
            
    except Exception:
        db.session.rollback()
        flash('自動付与処理中にエラーが発生しました。')
    
    return redirect(url_for('leave_management'))


def allowed_file(filename):
    """アップロードされたファイルの拡張子をチェック"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def save_photo(file):
    """アップロードされた画像ファイルを安全に保存"""
    if file and file.filename and allowed_file(file.filename):
        # 安全なファイル名を生成（UUID + 元の拡張子）
        file_extension = file.filename.rsplit('.', 1)[1].lower()
        safe_filename = f"{uuid.uuid4().hex}.{file_extension}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], safe_filename)
        file.save(file_path)
        return safe_filename
    return None

def allowed_residence_card_file(filename):
    """在留カードファイルの拡張子をチェック"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_RESIDENCE_CARD_EXTENSIONS

def save_residence_card_file(file):
    """アップロードされた在留カードファイルを安全に保存"""
    if file and file.filename and allowed_residence_card_file(file.filename):
        # 安全なファイル名を生成（UUID + 元の拡張子）
        file_extension = file.filename.rsplit('.', 1)[1].lower()
        safe_filename = f"residence_card_{uuid.uuid4().hex}.{file_extension}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], safe_filename)
        file.save(file_path)
        return safe_filename
    return None

def allowed_car_insurance_file(filename):
    """自動車保険証ファイルの拡張子をチェック"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_CAR_INSURANCE_EXTENSIONS

def save_car_insurance_file(file):
    """アップロードされた自動車保険証ファイルを安全に保存"""
    if file and file.filename and allowed_car_insurance_file(file.filename):
        # 安全なファイル名を生成（UUID + 元の拡張子）
        file_extension = file.filename.rsplit('.', 1)[1].lower()
        safe_filename = f"car_insurance_{uuid.uuid4().hex}.{file_extension}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], safe_filename)
        file.save(file_path)
        return safe_filename
    return None

def get_calendar_setting(key, default_value=None):
    """カレンダー設定を取得"""
    setting = CalendarSettings.query.filter_by(setting_key=key).first()
    return setting.setting_value if setting else default_value

def set_calendar_setting(key, value, description=None):
    """カレンダー設定を保存"""
    setting = CalendarSettings.query.filter_by(setting_key=key).first()
    if setting:
        setting.setting_value = value
        setting.updated_at = date.today()
        if description:
            setting.description = description
    else:
        setting = CalendarSettings(
            setting_key=key,
            setting_value=value,
            description=description
        )
        db.session.add(setting)
    db.session.commit()

def setup_japanese_font():
    """日本語フォント設定の共通関数 - CIDフォント優先版"""
    
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    
    # まずCIDフォントを試行（最も確実）
    cid_fonts = [
        ('HeiseiKakuGo-W5', 'HeiseiKakuGo-W5'),  # 日本語ゴシック
        ('HeiseiMin-W3', 'HeiseiMin-W3'),        # 日本語明朝
    ]
    
    for regular_name, bold_name in cid_fonts:
        try:
            pdfmetrics.registerFont(UnicodeCIDFont(regular_name))
            print(f"✓ setup_japanese_font: CIDフォント {regular_name} 登録成功")
            return regular_name, bold_name
        except Exception as e:
            print(f"✗ setup_japanese_font: CIDフォント {regular_name} 登録失敗: {e}")
            continue
    
    # CIDフォント失敗時のTTFフォント試行
    font_configs = [
        # DejaVu Sans（システム標準）
        {
            'paths': ['/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'],
            'regular': 'DejaVuSans-Regular',
            'bold': 'DejaVuSans-Bold'
        },
        
        # Liberation Sans
        {
            'paths': ['/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf'],
            'regular': 'LiberationSans-Regular', 
            'bold': 'LiberationSans-Bold'
        }
    ]
    
    # フォントを順番に試行
    for config in font_configs:
        for font_path in config['paths']:
            try:
                # レギュラーフォント登録
                pdfmetrics.registerFont(TTFont(config['regular'], font_path))
                
                # ボールドフォント登録（ボールド版が存在する場合）
                bold_path = font_path.replace('Regular', 'Bold').replace('regular', 'bold')
                try:
                    if os.path.exists(bold_path):
                        pdfmetrics.registerFont(TTFont(config['bold'], bold_path))
                    else:
                        # ボールド版がない場合はレギュラーを使用
                        config['bold'] = config['regular']
                except:
                    config['bold'] = config['regular']
                
                print(f"日本語フォント設定成功: {config['regular']}, {config['bold']}")
                return config['regular'], config['bold']
                
            except Exception as e:
                continue
    
    # フォールバック1: DejaVu Sans（基本的な文字をサポート）
    try:
        pdfmetrics.registerFont(TTFont('DejaVuSans', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
        pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
        print("フォールバック: DejaVuSans使用")
        return 'DejaVuSans', 'DejaVuSans-Bold'
    except:
        pass
    
    # フォールバック2: Ubuntu フォント（一部日本語文字サポート）
    try:
        # 実在するUbuntuフォントファイル
        pdfmetrics.registerFont(TTFont('Ubuntu', '/usr/share/fonts/truetype/ubuntu/Ubuntu[wdth,wght].ttf'))
        # ボールドも同じファイルから取得（Variable Font）
        pdfmetrics.registerFont(TTFont('Ubuntu-Bold', '/usr/share/fonts/truetype/ubuntu/Ubuntu[wdth,wght].ttf'))
        print("フォールバック: Ubuntuフォント使用（日本語の一部をサポート）")
        return 'Ubuntu', 'Ubuntu-Bold'
    except Exception as e:
        print(f"Ubuntuフォント読み込み失敗: {e}")
        pass
    
    # フォールバック3: Liberation フォント
    try:
        pdfmetrics.registerFont(TTFont('LiberationSans', '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf'))
        pdfmetrics.registerFont(TTFont('LiberationSans-Bold', '/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf'))
        print("フォールバック: Liberationフォント使用")
        return 'LiberationSans', 'LiberationSans-Bold'
    except:
        pass
    
    # 最終フォールバック: ReportLab内蔵のTimes-Roman
    try:
        print("最終フォールバック: Times-Roman使用（日本語は表示されない可能性があります）")
        return 'Times-Roman', 'Times-Bold'
    except:
        # 究極のフォールバック
        print("警告: フォント設定に失敗。Helveticaを使用します。")
        return 'Helvetica', 'Helvetica-Bold'

def create_employee_pdf(employee):
    """従業員情報のPDFを生成 - CIDフォント版（確実な日本語表示）"""
    buffer = BytesIO()
    
    # CIDフォント設定（日本語専用フォント）
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    
    japanese_font = None
    japanese_font_bold = None
    
    # CIDフォント候補を順番に試行
    cid_fonts = [
        ('HeiseiKakuGo-W5', 'HeiseiKakuGo-W5'),  # 日本語ゴシック
        ('HeiseiMin-W3', 'HeiseiMin-W3'),        # 日本語明朝
    ]
    
    for regular_name, bold_name in cid_fonts:
        try:
            pdfmetrics.registerFont(UnicodeCIDFont(regular_name))
            japanese_font = regular_name
            japanese_font_bold = bold_name  # CIDフォントでは同じ名前
            print(f"✓ CIDフォント {regular_name} 登録成功")
            break
        except Exception as e:
            print(f"✗ CIDフォント {regular_name} 登録失敗: {e}")
            continue
    
    # CIDフォント失敗時のTTFフォールバック
    if japanese_font is None:
        font_path = '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'
        if os.path.exists(font_path):
            try:
                pdfmetrics.registerFont(TTFont('JapaneseFont', font_path))
                japanese_font = 'JapaneseFont'
                japanese_font_bold = 'JapaneseFont'
                print("TTFフォールバック: DejaVu Sans使用")
            except:
                japanese_font = 'Helvetica'
                japanese_font_bold = 'Helvetica-Bold'
        else:
            japanese_font = 'Helvetica'
            japanese_font_bold = 'Helvetica-Bold'
    
    # 年休データを計算
    total_credited = db.session.query(db.func.sum(LeaveCredit.days_credited))\
                               .filter_by(employee_id=employee.id).scalar() or 0
    total_taken = db.session.query(db.func.sum(LeaveRecord.days_taken))\
                            .filter_by(employee_id=employee.id).scalar() or 0
    remaining_leave = total_credited - total_taken
    
    # 法定付与日数の計算
    years_employed = (date.today() - employee.join_date).days / 365.25 if employee.join_date else 0
    if years_employed < 0.5:
        legal_leave_days = 0
    elif years_employed < 1.5:
        legal_leave_days = 10
    elif years_employed < 2.5:
        legal_leave_days = 11
    elif years_employed < 3.5:
        legal_leave_days = 12
    elif years_employed < 4.5:
        legal_leave_days = 14
    elif years_employed < 5.5:
        legal_leave_days = 16
    elif years_employed < 6.5:
        legal_leave_days = 18
    else:
        legal_leave_days = 20
    
    # 顔写真情報を保存（後で右上に配置）
    photo_data = None
    if employee.photo_filename:
        try:
            photo_path = os.path.join('static', 'uploads', employee.photo_filename)
            if os.path.exists(photo_path):
                from reportlab.lib.utils import ImageReader
                
                # 画像を読み込み
                img = ImageReader(photo_path)
                img_width, img_height = img.getSize()
                
                # 写真のサイズを調整（140x140ピクセル相当の1.2倍サイズ）
                base_photo_size = 140 * 72 / 96  # ピクセルをポイントに変換
                photo_size = base_photo_size * 1.2  # 1.2倍に拡大
                aspect_ratio = img_width / img_height
                if aspect_ratio > 1:
                    width = photo_size
                    height = photo_size / aspect_ratio
                else:
                    width = photo_size * aspect_ratio
                    height = photo_size
                
                # 写真データを保存
                photo_data = {
                    'path': photo_path,
                    'width': width,
                    'height': height
                }
        except Exception as e:
            print(f"顔写真の読み込みエラー: {e}")
    
    # PDFドキュメントを作成（マージンを最小限に設定）
    from reportlab.platypus import PageTemplate, Frame
    from reportlab.lib.pagesizes import A4
    
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20, bottomMargin=30, leftMargin=40, rightMargin=40)
    
    # ページテンプレート用のフッター・ヘッダー関数（クロージャーでphoto_dataを参照）
    def add_footer_and_photo(canvas, doc):
        width, height = A4
        
        # 発行日を右下に固定配置
        canvas.setFont(japanese_font, 10)
        canvas.setFillColor(colors.black)
        canvas.drawRightString(width - 40, 20, f"発行日: {date.today().strftime('%Y年%m月%d日')}")
        
        # 顔写真を基本情報テーブルの上部に合わせて配置
        if photo_data:
            try:
                # 写真を5px左、10px下に移動
                # 1ポイント ≈ 1.33ピクセルなので、5px ≈ 3.75pt、10px ≈ 7.5pt
                photo_x = width - photo_data['width'] - 53.75  # 右端から53.75pt内側（元の50ptから3.75pt左に移動）
                # 基本情報テーブルの上部に合わせるため、タイトル+スペーサー分を考慮し、さらに10px下に移動
                # タイトル(16pt) + spaceAfter(5pt) + Spacer(8pt) + セクションタイトル(12pt) + spaceAfter(3pt) + Spacer(3pt) = 約47pt + 7.5pt(10px下)
                photo_y = height - photo_data['height'] - 105  # 上端から105pt内側（元の97.5ptからさらに7.5pt下に移動）
                
                canvas.drawImage(
                    photo_data['path'], 
                    photo_x, photo_y,
                    width=photo_data['width'], 
                    height=photo_data['height']
                )
            except Exception as e:
                print(f"写真描画エラー: {e}")
    
    # カスタムページテンプレートを作成（顔写真分の右マージンを追加）
    if photo_data:
        # 顔写真がある場合は右マージンを広くする（タイトルは全幅、その他は写真を避ける）
        frame = Frame(40, 30, A4[0]-80-(photo_data['width']+20), A4[1]-50, id='normal')
    else:
        # 顔写真がない場合は通常のマージン
        frame = Frame(40, 30, A4[0]-80, A4[1]-50, id='normal')
    
    template = PageTemplate(id='main', frames=[frame], onPage=add_footer_and_photo)
    doc.addPageTemplates([template])
    
    styles = getSampleStyleSheet()
    story = []
    
    # タイトル（完全中央揃え）
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=5,
        alignment=1,  # 中央揃え
        fontName=japanese_font_bold,
        textColor=colors.Color(0.1, 0.3, 0.7)  # ダークブルー
    )
    
    # タイトルを完全中央に配置するため、写真の有無に関わらず全幅を使用
    from reportlab.platypus import KeepTogether
    title = Paragraph("スタッフ情報", title_style)
    story.append(title)
    story.append(Spacer(1, 8))
    
    # 顔写真は右上固定配置のため、ここでは何もしない
    
    # 基本情報セクション
    basic_title_style = ParagraphStyle(
        'BasicTitle',
        parent=styles['Heading2'],
        fontSize=12,
        fontName=japanese_font_bold,
        textColor=colors.Color(0.1, 0.4, 0.7),
        spaceAfter=3
    )
    basic_title = Paragraph("基本情報", basic_title_style)
    story.append(basic_title)
    story.append(Spacer(1, 3))
    
    # 基本情報テーブル（圧縮版）
    data = [
        ['項目', '内容'],
        ['氏名', employee.name],
        ['生年月日 / 性別', f"{employee.birth_date.strftime('%Y年%m月%d日') if employee.birth_date else '未設定'} / {employee.gender or '未設定'}"],
        ['入社年月日', employee.join_date.strftime('%Y年%m月%d日') if employee.join_date else '未設定'],
        ['電話番号', employee.phone_number or '未設定'],
        ['国籍', employee.nationality or '未設定'],
        ['在留カード期限', employee.residence_card_expiry.strftime('%Y年%m月%d日') if employee.residence_card_expiry else '未設定'],
        ['在籍状況', employee.status or '未設定'],
    ]
    
    # 写真がある場合はテーブル幅を調整
    if photo_data:
        # 写真分のスペースを考慮してテーブル幅を縮小
        table = Table(data, colWidths=[2*inch, 3*inch])
    else:
        table = Table(data, colWidths=[2.5*inch, 4*inch])
    
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.2, 0.4, 0.8)),  # ブルーヘッダー
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), japanese_font_bold),
        ('FONTNAME', (0, 1), (-1, -1), japanese_font),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
        ('TOPPADDING', (0, 0), (-1, 0), 5),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BACKGROUND', (0, 1), (-1, -1), colors.Color(0.95, 0.97, 1.0)),  # 薄いブルー
        ('GRID', (0, 0), (-1, -1), 1.5, colors.Color(0.3, 0.5, 0.9)),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.97, 1.0)]),
    ]))
    table.hAlign = 'LEFT'  # テーブル全体を左詰め
    
    story.append(table)
    story.append(Spacer(1, 3))
    
    # 人事情報セクション
    if employee.position or employee.department or employee.hire_type or employee.salary_grade or employee.work_history or employee.skills or employee.qualifications:
        hr_title_style = ParagraphStyle(
            'HRTitle',
            parent=styles['Heading2'],
            fontSize=12,
            fontName=japanese_font_bold,
            textColor=colors.Color(0.1, 0.4, 0.7),
            spaceAfter=3
        )
        hr_title = Paragraph("人事情報", hr_title_style)
        story.append(hr_title)
        story.append(Spacer(1, 3))
        
        # HR情報を2列に配置
        hr_data = []
        if employee.position or employee.department:
            position_dept = []
            if employee.position:
                position_dept.append(f"役職: {employee.position}")
            if employee.department:
                position_dept.append(f"部署: {employee.department}")
            hr_data.append(['役職/部署', ' / '.join(position_dept)])
        
        if employee.hire_type or employee.salary_grade:
            hire_salary = []
            if employee.hire_type:
                hire_salary.append(employee.hire_type)
            if employee.salary_grade:
                hire_salary.append(f"等級: {employee.salary_grade}")
            hr_data.append(['雇用形態/等級', ' / '.join(hire_salary)])
        
        if employee.work_history:
            # 長いテキストは適切に短縮
            work_history = employee.work_history[:50] + "..." if len(employee.work_history) > 50 else employee.work_history
            hr_data.append(['経歴', work_history])
        
        if employee.skills:
            skills = employee.skills[:50] + "..." if len(employee.skills) > 50 else employee.skills
            hr_data.append(['スキル', skills])
        
        if employee.qualifications:
            qualifications = employee.qualifications[:50] + "..." if len(employee.qualifications) > 50 else employee.qualifications
            hr_data.append(['資格', qualifications])
        
        if hr_data:
            hr_data.insert(0, ['項目', '内容'])  # ヘッダーを追加
            
            # 人事情報テーブル幅調整：項目列を狭くし、内容列を広くする
            hr_table = Table(hr_data, colWidths=[2.0*inch, 4.5*inch])
            hr_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.2, 0.4, 0.8)),  # 基本情報と同じブルーヘッダー
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), japanese_font_bold),
                ('FONTNAME', (0, 1), (-1, -1), japanese_font),
                ('FONTSIZE', (0, 0), (-1, 0), 11),  # 基本情報と統一
                ('FONTSIZE', (0, 1), (-1, -1), 9),  # 基本情報と統一
                ('BOTTOMPADDING', (0, 0), (-1, 0), 5),  # 基本情報と統一
                ('TOPPADDING', (0, 0), (-1, 0), 5),     # 基本情報と統一
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4), # 基本情報と統一
                ('TOPPADDING', (0, 1), (-1, -1), 4),    # 基本情報と統一
                ('BACKGROUND', (0, 1), (-1, -1), colors.Color(0.95, 0.97, 1.0)),  # 基本情報と同じ薄いブルー
                ('GRID', (0, 0), (-1, -1), 1.5, colors.Color(0.3, 0.5, 0.9)),     # 基本情報と統一
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.97, 1.0)]),  # 基本情報と統一
            ]))
            hr_table.hAlign = 'LEFT'
            
            story.append(hr_table)
            story.append(Spacer(1, 3))
    
    # 休暇情報セクション
    leave_title_style = ParagraphStyle(
        'LeaveTitle',
        parent=styles['Heading2'],
        fontSize=12,
        fontName=japanese_font_bold,
        textColor=colors.Color(0.1, 0.4, 0.7),
        spaceAfter=3
    )
    leave_title = Paragraph("休暇情報", leave_title_style)
    story.append(leave_title)
    story.append(Spacer(1, 3))
    
    # 年次有給休暇データ
    annual_leave_data = [
        ['年次有給休暇', '日数'],
        ['付与日数合計', f'{total_credited}日'],
        ['取得日数合計', f'{total_taken}日'],
        ['残日数', f'{remaining_leave}日'],
        ['法定付与日数', f'{legal_leave_days}日'],
    ]
    
    # その他休暇データ（サンプル）
    other_leave_data = [
        ['その他休暇', '日数'],
        ['特別休暇', '0日'],
        ['慶弔休暇', '0日'],
        ['産前産後休暇', '0日'],
        ['育児休暇', '0日'],
    ]
    
    # 休暇情報を横並びで表示
    vacation_table_data = [['年次有給休暇', 'その他休暇']]
    
    # 年次有給休暇の文字列を作成
    annual_text_list = []
    for row in annual_leave_data[1:]:  # ヘッダーを除く
        annual_text_list.append(f"{row[0]}: {row[1]}")
    annual_text = "\n".join(annual_text_list)
    
    # その他休暇の文字列を作成
    other_text_list = []
    for row in other_leave_data[1:]:  # ヘッダーを除く
        other_text_list.append(f"{row[0]}: {row[1]}")
    other_text = "\n".join(other_text_list)
    
    vacation_table_data.append([annual_text, other_text])
    
    leave_table = Table(vacation_table_data, colWidths=[3.25*inch, 3.25*inch])
    leave_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.2, 0.4, 0.8)),  # 基本情報と同じブルーヘッダー
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), japanese_font_bold),
        ('FONTNAME', (0, 1), (-1, -1), japanese_font),
        ('FONTSIZE', (0, 0), (-1, 0), 11),  # 基本情報と統一
        ('FONTSIZE', (0, 1), (-1, -1), 9),  # 基本情報と統一
        ('BOTTOMPADDING', (0, 0), (-1, 0), 5),  # 基本情報と統一
        ('TOPPADDING', (0, 0), (-1, 0), 5),     # 基本情報と統一
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4), # 基本情報と統一
        ('TOPPADDING', (0, 1), (-1, -1), 4),    # 基本情報と統一
        ('BACKGROUND', (0, 1), (-1, -1), colors.Color(0.95, 0.97, 1.0)),  # 基本情報と同じ薄いブルー背景
        ('GRID', (0, 0), (-1, -1), 1.5, colors.Color(0.3, 0.5, 0.9)),     # 基本情報と統一
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.97, 1.0)]),  # 基本情報と統一
    ]))
    leave_table.hAlign = 'LEFT'  # テーブル全体を左詰め
    
    story.append(leave_table)
    
    # 年休履歴（付与と取得を並列表示して省スペース化）
    leave_credits = LeaveCredit.query.filter_by(employee_id=employee.id).order_by(LeaveCredit.date_credited.desc()).limit(2).all()
    leave_records = LeaveRecord.query.filter_by(employee_id=employee.id).order_by(LeaveRecord.date_taken.desc()).limit(2).all()
    
    if leave_credits or leave_records:
        story.append(Spacer(1, 3))
        history_title_style = ParagraphStyle('HistoryTitle', parent=styles['Heading3'], fontName=japanese_font_bold, fontSize=10, textColor=colors.Color(0.2, 0.4, 0.7))
        history_title = Paragraph("休暇履歴（直近2件）", history_title_style)
        story.append(history_title)
        story.append(Spacer(1, 2))
        
        # 履歴を横並びで表示（年休と他の休暇を分けて表示）
        history_data = [['年次有給休暇履歴', 'その他休暇履歴']]
        
        # 年次有給休暇履歴の文字列を作成
        annual_history_list = []
        
        # 付与履歴
        if leave_credits:
            annual_history_list.append("【付与履歴】")
            for credit in leave_credits:
                annual_history_list.append(f"{credit.date_credited.strftime('%Y/%m/%d')} +{credit.days_credited}日")
        
        # 取得履歴
        if leave_records:
            if annual_history_list:  # 付与履歴がある場合は改行を追加
                annual_history_list.append("")
            annual_history_list.append("【取得履歴】")
            for record in leave_records:
                annual_history_list.append(f"{record.date_taken.strftime('%Y/%m/%d')} -{record.days_taken}日")
        
        annual_history_text = "\n".join(annual_history_list) if annual_history_list else "なし"
        
        # その他休暇履歴（現在はサンプル）
        other_history_text = "【特別休暇】\nなし\n\n【慶弔休暇】\nなし"
        
        history_data.append([annual_history_text, other_history_text])
        
        history_table = Table(history_data, colWidths=[3.25*inch, 3.25*inch])
        history_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.2, 0.4, 0.8)),  # 基本情報と同じブルーヘッダー
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, 0), japanese_font_bold),
            ('FONTNAME', (0, 1), (-1, -1), japanese_font),
            ('FONTSIZE', (0, 0), (-1, 0), 11),  # 基本情報と統一
            ('FONTSIZE', (0, 1), (-1, -1), 9),  # 基本情報と統一
            ('BOTTOMPADDING', (0, 0), (-1, 0), 5),  # 基本情報と統一
            ('TOPPADDING', (0, 0), (-1, 0), 5),     # 基本情報と統一
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4), # 基本情報と統一
            ('TOPPADDING', (0, 1), (-1, -1), 4),    # 基本情報と統一
            ('BACKGROUND', (0, 1), (-1, -1), colors.Color(0.95, 0.97, 1.0)),  # 基本情報と同じ薄いブルー
            ('GRID', (0, 0), (-1, -1), 1.5, colors.Color(0.3, 0.5, 0.9)),     # 基本情報と統一
        ]))
        history_table.hAlign = 'LEFT'
        story.append(history_table)
    
    # PDFを生成
    doc.build(story)
    buffer.seek(0)
    return buffer

def create_calendar_pdf():
    """会社カレンダーのPDFを升目デザインで生成 - CIDフォント版（確実な日本語表示）"""
    buffer = BytesIO()
    
    # CIDフォント設定（日本語専用フォント）
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    
    japanese_font = None
    japanese_font_bold = None
    
    # CIDフォント候補を順番に試行
    cid_fonts = [
        ('HeiseiKakuGo-W5', 'HeiseiKakuGo-W5'),  # 日本語ゴシック
        ('HeiseiMin-W3', 'HeiseiMin-W3'),        # 日本語明朝
    ]
    
    for regular_name, bold_name in cid_fonts:
        try:
            pdfmetrics.registerFont(UnicodeCIDFont(regular_name))
            japanese_font = regular_name
            japanese_font_bold = bold_name  # CIDフォントでは同じ名前
            print(f"✓ カレンダーPDF用CIDフォント {regular_name} 登録成功")
            break
        except Exception as e:
            print(f"✗ CIDフォント {regular_name} 登録失敗: {e}")
            continue
    
    # CIDフォント失敗時のTTFフォールバック
    if japanese_font is None:
        font_path = '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'
        if os.path.exists(font_path):
            try:
                pdfmetrics.registerFont(TTFont('JapaneseFont', font_path))
                japanese_font = 'JapaneseFont'
                japanese_font_bold = 'JapaneseFont'
                print("TTFフォールバック: DejaVu Sans使用")
            except:
                japanese_font = 'Helvetica'
                japanese_font_bold = 'Helvetica-Bold'
        else:
            japanese_font = 'Helvetica'
            japanese_font_bold = 'Helvetica-Bold'
    
    # PDFドキュメントを作成（縦向きA4）
    doc = SimpleDocTemplate(buffer, pagesize=A4, 
                           topMargin=15*mm, bottomMargin=15*mm,
                           leftMargin=15*mm, rightMargin=15*mm)
    styles = getSampleStyleSheet()
    story = []
    
    # タイトル（コンパクトに調整）
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=japanese_font_bold,
        fontSize=14,  # サイズを小さく
        spaceAfter=3,  # 下余白を大幅削減
        alignment=1  # 中央揃え
    )
    
    current_year = datetime.now().year
    title = Paragraph(f"{current_year}年 会社カレンダー", title_style)
    story.append(title)
    story.append(Spacer(1, 2))  # タイトル後のスペースも削減
    
    # カレンダー開始月設定を取得
    start_month = int(get_calendar_setting('start_month', '1'))
    
    # 年間カレンダーデータを生成
    calendar_data = generate_year_calendar(current_year, start_month)
    
    # 12ヶ月を確実に1ページに収める3x4レイアウト（上に詰める）
    # ページサイズとマージンの計算
    page_width = A4[0]  # 210mm
    page_height = A4[1]  # 297mm
    
    # マージンを最小限に設定（確実に1Pに収める）
    top_margin = 2*mm  # 上余白を最小に
    bottom_margin = 2*mm  
    left_margin = 5*mm
    right_margin = 5*mm
    title_space = 3*mm  # タイトル用スペースを極小に
    
    usable_width = page_width - left_margin - right_margin  # 200mm
    usable_height = page_height - top_margin - bottom_margin - title_space  # 290mm
    
    # 月間の間隔設定（密着しない最小限の間隔）
    month_h_spacing = 2.5*mm  # 横間隔（2.5mm = 約7px）
    month_v_spacing = 1.5*mm  # 縦間隔（1.5mm = 約4.2px）
    
    # 各月に割り当てられるスペース計算
    available_width_per_month = (usable_width - 2 * month_h_spacing) / 3  # (200-5)/3 = 65mm
    available_height_per_month = (usable_height - 3 * month_v_spacing) / 4  # (290-4.5)/4 = 71.4mm
    
    # 月タイトル用スペース
    month_title_height = 3.5*mm
    
    # カレンダーのセルサイズ計算（縦横最低値を取って正方形に近づける）
    max_cell_width = available_width_per_month / 7  # 9.29mm
    max_cell_height = (available_height_per_month - month_title_height) / 7  # 9.70mm
    
    # 正方形に近づけるため小さい方を採用
    cell_size = min(max_cell_width, max_cell_height, 9*mm)  # 最大9mmで制限
    
    # 実際の月カレンダーサイズ（セル高さ調整を反映）
    actual_month_width = cell_size * 7
    actual_month_height = cell_size * 0.5 + cell_size * 1.1 * 6 + month_title_height  # ヘッダー0.5 + 本文1.1×6
    
    # デバッグ情報表示
    total_needed_height = actual_month_height * 4 + month_v_spacing * 3
    print(f"Debug: cell_size={cell_size/mm:.2f}mm, month_size={actual_month_width/mm:.1f}x{actual_month_height/mm:.1f}mm")
    print(f"Debug: total_needed_height={total_needed_height/mm:.1f}mm, available={usable_height/mm:.1f}mm")
    
    for row_idx in range(4):  # 4行
        row_elements = []
        
        for col_idx in range(3):  # 3列
            month_idx = row_idx * 3 + col_idx
            if month_idx >= len(calendar_data):
                # 空セルで埋める
                row_elements.append(Spacer(actual_month_width, actual_month_height))
                continue
            
            month_data = calendar_data[month_idx]
            
            # 月のテーブルデータを作成
            month_table_data = []
            
            # 曜日ヘッダー
            month_table_data.append(['日', '月', '火', '水', '木', '金', '土'])
            
            # 各週のデータ
            for week in month_data['weeks']:
                week_row = []
                for day_data in week:
                    if day_data['day']:
                        # 日付とイベントを分離して配置（日付：上部、イベント：中央から下部）
                        day_number = str(day_data['day'])
                        
                        if day_data['events']:
                            # 複数イベントがある場合は最も重要なものを選択
                            event = day_data['events'][0]
                            title = event.title
                            
                            # イベント名の表示を改善（改行なしで表示）
                            if len(title) <= 4:
                                # 4文字以下はそのまま表示
                                display_title = title
                            elif len(title) <= 8:
                                # 8文字以下はそのまま表示
                                display_title = title
                            else:
                                # 9文字以上は省略して表示
                                display_title = title[:7] + '..'
                            
                            # 日付とイベント名を分離して配置
                            day_content = day_number + '\n' + display_title
                        else:
                            day_content = day_number
                    else:
                        day_content = ''
                    week_row.append(day_content)
                month_table_data.append(week_row)
            
            # 週数を6週に固定
            while len(month_table_data) - 1 < 6:
                month_table_data.append(['', '', '', '', '', '', ''])
            
            # カレンダーテーブル作成（セル高さを調整して日付とイベントを分離表示）
            month_table = Table(month_table_data, 
                               colWidths=[cell_size] * 7,
                               rowHeights=[cell_size * 0.5] + [cell_size * 1.1] * 6)  # ヘッダー行を少し高くして見切れ防止
            
            # テーブルスタイル
            table_styles = [
                ('GRID', (0, 0), (-1, -1), 0.3, colors.black),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('FONTNAME', (0, 0), (-1, 0), japanese_font_bold),
                ('FONTSIZE', (0, 0), (-1, 0), 4),  # ヘッダーフォントサイズを適度に調整
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                ('FONTNAME', (0, 1), (-1, -1), japanese_font),
                ('FONTSIZE', (0, 1), (-1, -1), 3),  # 本文フォントサイズを少し見やすく
                ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0.5),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0.5),
                ('TOPPADDING', (0, 0), (-1, -1), 0.5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0.5),
            ]
            
            # 背景色設定
            for week_idx, week in enumerate(month_data['weeks']):
                for day_idx, day_data in enumerate(week):
                    row = week_idx + 1
                    col = day_idx
                    
                    if day_data['day']:
                        # 日曜日・土曜日の基本色
                        if day_idx == 0:  # 日曜日
                            table_styles.append(('TEXTCOLOR', (col, row), (col, row), colors.red))
                            table_styles.append(('BACKGROUND', (col, row), (col, row), colors.mistyrose))
                        elif day_idx == 6:  # 土曜日
                            table_styles.append(('TEXTCOLOR', (col, row), (col, row), colors.blue))
                            table_styles.append(('BACKGROUND', (col, row), (col, row), colors.lightcyan))
                        
                        # イベントがある場合は背景色を上書き
                        if day_data['events']:
                            for event in day_data['events']:
                                if event.event_type in ['holiday', 'company_holiday']:
                                    table_styles.append(('BACKGROUND', (col, row), (col, row), colors.mistyrose))
                                    table_styles.append(('TEXTCOLOR', (col, row), (col, row), colors.red))
                                elif event.event_type == 'vacation':
                                    table_styles.append(('BACKGROUND', (col, row), (col, row), colors.lightgreen))
                                    table_styles.append(('TEXTCOLOR', (col, row), (col, row), colors.darkgreen))
                                else:
                                    table_styles.append(('BACKGROUND', (col, row), (col, row), colors.lightyellow))
            
            month_table.setStyle(TableStyle(table_styles))
            
            # 月タイトルを作成（小さめのフォント）
            month_title_style = ParagraphStyle(
                'MonthTitle',
                parent=styles['Heading3'],
                fontName=japanese_font_bold,
                fontSize=5,  # 固定で小さめ
                spaceAfter=0,
                spaceBefore=0,
                alignment=1
            )
            month_title = Paragraph(f"{month_data['display_year']}年{month_data['month']}月", month_title_style)
            
            # 月タイトルとカレンダーを組み合わせ（コンパクトに）
            month_content = [
                [month_title],
                [month_table]
            ]
            month_wrapper = Table(month_content, 
                                colWidths=[actual_month_width],
                                rowHeights=[month_title_height, actual_month_height - month_title_height])
            month_wrapper.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ]))
            
            row_elements.append(month_wrapper)
        
        # 行テーブルを作成（間隔を確実に設ける）
        if len(row_elements) == 1:
            # 1個のみの場合（通常はありえない）
            row_elements.extend([Spacer(month_h_spacing, actual_month_height), 
                               Spacer(actual_month_width, actual_month_height),
                               Spacer(month_h_spacing, actual_month_height),
                               Spacer(actual_month_width, actual_month_height)])
        elif len(row_elements) == 2:
            # 2個の場合
            row_elements.extend([Spacer(month_h_spacing, actual_month_height), 
                               Spacer(actual_month_width, actual_month_height)])
        elif len(row_elements) == 3:
            # 3個の場合（正常） - 間にスペーサーを挿入
            new_row_elements = [
                row_elements[0],
                Spacer(month_h_spacing, actual_month_height),
                row_elements[1], 
                Spacer(month_h_spacing, actual_month_height),
                row_elements[2]
            ]
            row_elements = new_row_elements
        
        # 各要素の幅を計算
        col_widths = []
        for i, element in enumerate(row_elements):
            if i % 2 == 0:  # カレンダー要素
                col_widths.append(actual_month_width)
            else:  # スペーサー要素
                col_widths.append(month_h_spacing)
        
        row_table = Table([row_elements], 
                         colWidths=col_widths,
                         rowHeights=[actual_month_height])
        row_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        
        story.append(row_table)
        
        # 行間スペース（最後の行以外）
        if row_idx < 3:
            story.append(Spacer(1, month_v_spacing))
    
    # 改ページして2ページ目へ
    story.append(PageBreak())
    
    # イベント一覧を追加
    events = CompanyCalendar.query.filter(
        db.or_(
            db.extract('year', CompanyCalendar.event_date) == current_year,
            CompanyCalendar.is_recurring == True
        )
    ).order_by(CompanyCalendar.event_date.asc()).all()
    
    if events:
        event_title_style = ParagraphStyle(
            'EventTitle',
            parent=styles['Heading2'],
            fontName=japanese_font_bold,
            fontSize=16,
            spaceAfter=10
        )
        event_title = Paragraph("年間イベント一覧", event_title_style)
        story.append(event_title)
        story.append(Spacer(1, 10))
        
        # 休日とイベントを分ける
        holidays = []
        company_events = []
        
        for event in events:
            if event.is_recurring:
                event_date = date(current_year, event.event_date.month, event.event_date.day)
            else:
                event_date = event.event_date
                
            if event_date.year == current_year:
                if event.event_type == 'holiday':
                    holidays.append((event_date, event))
                else:
                    company_events.append((event_date, event))
        
        # 休日セクション
        if holidays:
            holiday_title_style = ParagraphStyle(
                'HolidayTitle',
                parent=styles['Heading3'],
                fontName=japanese_font_bold,
                fontSize=12
            )
            holiday_title = Paragraph("休日", holiday_title_style)
            story.append(holiday_title)
            story.append(Spacer(1, 5))
            
            holiday_data = [['月日', '名称', '備考']]
            for event_date, event in sorted(holidays):
                holiday_data.append([
                    event_date.strftime('%m/%d'),
                    event.title,
                    event.description[:20] + '...' if event.description and len(event.description) > 20 else (event.description or '')
                ])
            
            holiday_table = Table(holiday_data, colWidths=[25*mm, 60*mm, 85*mm])
            holiday_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.red),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), japanese_font_bold),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTNAME', (0, 1), (-1, -1), japanese_font),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('LEFTPADDING', (0, 0), (-1, -1), 5),
                ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ]))
            story.append(holiday_table)
            story.append(Spacer(1, 10))
        
        # イベントセクション
        if company_events:
            event_section_title_style = ParagraphStyle(
                'EventSectionTitle',
                parent=styles['Heading3'],
                fontName=japanese_font_bold,
                fontSize=12
            )
            event_section_title = Paragraph("会社行事・イベント", event_section_title_style)
            story.append(event_section_title)
            story.append(Spacer(1, 5))
            
            event_data = [['月日', '名称', '詳細']]
            for event_date, event in sorted(company_events):
                event_data.append([
                    event_date.strftime('%m/%d'),
                    event.title,
                    event.description[:20] + '...' if event.description and len(event.description) > 20 else (event.description or '')
                ])
            
            event_table = Table(event_data, colWidths=[25*mm, 60*mm, 85*mm])
            event_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), japanese_font_bold),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTNAME', (0, 1), (-1, -1), japanese_font),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('LEFTPADDING', (0, 0), (-1, -1), 5),
                ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ]))
            story.append(event_table)
    
    # 生成日時を追加（日本語フォント使用）
    story.append(Spacer(1, 20))
    generation_style = ParagraphStyle(
        'GenerationInfo',
        parent=styles['Normal'],
        fontName=japanese_font,
        fontSize=9,
        textColor=colors.grey,
        alignment=2  # 右揃え
    )
    generation_info = Paragraph(
        f"生成日時: {datetime.now().strftime('%Y年%m月%d日 %H時%M分')}", 
        generation_style
    )
    story.append(generation_info)
    
    # PDF生成（マージンを最小限に）
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, 
                           topMargin=top_margin, bottomMargin=bottom_margin, 
                           leftMargin=left_margin, rightMargin=right_margin)
    doc.build(story)
    buffer.seek(0)
    return buffer

# データベースとアプリケーションを連携は既に上で完了

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# --- ここから下が各ページの処理 ---

@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return render_template('index.html')


@app.route('/save_company_settings', methods=['POST'])
def save_company_settings():
    try:
        company_name = request.form.get('company_name')
        if not company_name:
            flash('会社名は必須項目です。')
            return redirect(url_for('general_affairs_36agreement'))
        
        # 既存の設定を取得または新規作成
        company_settings = CompanySettings.query.first()
        if not company_settings:
            company_settings = CompanySettings(
                company_name=company_name,
                created_at=datetime.now(),
                updated_at=datetime.now()
            )
            db.session.add(company_settings)
        else:
            company_settings.updated_at = datetime.now()
        
        # フォームデータを設定
        company_settings.company_name = company_name
        company_settings.company_address = request.form.get('company_address')
        company_settings.company_phone = request.form.get('company_phone')
        company_settings.representative_name = request.form.get('representative_name')
        company_settings.business_type = request.form.get('business_type')
        
        # 日付フィールドの処理
        establishment_date_str = request.form.get('establishment_date')
        if establishment_date_str:
            company_settings.establishment_date = datetime.strptime(establishment_date_str, '%Y-%m-%d').date()
        
        # 数値フィールドの処理
        employee_count_str = request.form.get('employee_count')
        if employee_count_str:
            company_settings.employee_count = int(employee_count_str)
        
        db.session.commit()
        flash('企業情報が正常に保存されました。')
        
    except Exception as e:
        db.session.rollback()
        flash('企業情報の保存中にエラーが発生しました。')
    
    return redirect(url_for('general_affairs_36agreement'))

@app.route('/save_labor_settings', methods=['POST'])
def save_labor_settings():
    try:
        # 既存の設定を取得または新規作成
        labor_settings = LaborStandardsSettings.query.first()
        if not labor_settings:
            labor_settings = LaborStandardsSettings(
                created_at=datetime.now(),
                updated_at=datetime.now()
            )
            db.session.add(labor_settings)
        else:
            labor_settings.updated_at = datetime.now()
        
        # フォームデータを設定
        labor_settings.labor_office_name = request.form.get('labor_office_name')
        labor_settings.labor_office_address = request.form.get('labor_office_address')
        labor_settings.agreement_36_filed = 'agreement_36_filed' in request.form
        labor_settings.work_rules_filed = 'work_rules_filed' in request.form
        
        # 日付フィールドの処理
        for field in ['agreement_36_submission_date', 'agreement_36_period_start', 'agreement_36_period_end', 'agreement_36_expiry_date', 'work_rules_last_updated']:
            date_str = request.form.get(field)
            if date_str:
                setattr(labor_settings, field, datetime.strptime(date_str, '%Y-%m-%d').date())
        
        # 数値フィールドの処理
        for field in ['max_overtime_hours_monthly', 'max_overtime_hours_yearly']:
            value_str = request.form.get(field)
            if value_str:
                setattr(labor_settings, field, int(value_str))
        
        db.session.commit()
        flash('労働基準監督署情報が正常に保存されました。')
        
    except Exception as e:
        db.session.rollback()
        flash('労働基準監督署情報の保存中にエラーが発生しました。')
    
    return redirect(url_for('general_affairs_36agreement'))

@app.route('/save_holiday_settings', methods=['POST'])
def save_holiday_settings():
    try:
        # 既存の設定を取得または新規作成
        holiday_settings = LegalHolidaySettings.query.first()
        if not holiday_settings:
            holiday_settings = LegalHolidaySettings(
                # 初期値は全てFalse（チェックなし）
                monday_legal_holiday=False,
                tuesday_legal_holiday=False,
                wednesday_legal_holiday=False,
                thursday_legal_holiday=False,
                friday_legal_holiday=False,
                saturday_legal_holiday=False,
                sunday_legal_holiday=False,
                created_at=datetime.now(),
                updated_at=datetime.now()
            )
            db.session.add(holiday_settings)
        else:
            holiday_settings.updated_at = datetime.now()
        
        # フォームデータを設定（曜日別法定休日）
        holiday_settings.monday_legal_holiday = 'monday_legal_holiday' in request.form
        holiday_settings.tuesday_legal_holiday = 'tuesday_legal_holiday' in request.form
        holiday_settings.wednesday_legal_holiday = 'wednesday_legal_holiday' in request.form
        holiday_settings.thursday_legal_holiday = 'thursday_legal_holiday' in request.form
        holiday_settings.friday_legal_holiday = 'friday_legal_holiday' in request.form
        holiday_settings.saturday_legal_holiday = 'saturday_legal_holiday' in request.form
        holiday_settings.sunday_legal_holiday = 'sunday_legal_holiday' in request.form
        
        # 週の起算日
        week_start_day_str = request.form.get('week_start_day')
        if week_start_day_str:
            holiday_settings.week_start_day = int(week_start_day_str)
        
        db.session.commit()
        flash('法定休日設定が正常に保存されました。')
        
    except Exception as e:
        db.session.rollback()
        flash('法定休日設定の保存中にエラーが発生しました。')
    
    return redirect(url_for('general_affairs_36agreement'))

@app.route('/submit_agreement_36', methods=['POST'])
def submit_agreement_36():
    """36協定の新規提出を記録"""
    try:
        # フォームデータを取得
        submission_date_str = request.form.get('submission_date')
        period_start_str = request.form.get('period_start')
        period_end_str = request.form.get('period_end')
        expiry_date_str = request.form.get('expiry_date')
        
        # 必須項目チェック
        if not all([submission_date_str, period_start_str, period_end_str, expiry_date_str]):
            flash('提出日、協定期間、期限切れ日は必須項目です。')
            return redirect(url_for('general_affairs_36agreement'))
        
        # 日付変換
        submission_date = datetime.strptime(submission_date_str, '%Y-%m-%d').date()
        period_start = datetime.strptime(period_start_str, '%Y-%m-%d').date()
        period_end = datetime.strptime(period_end_str, '%Y-%m-%d').date()
        expiry_date = datetime.strptime(expiry_date_str, '%Y-%m-%d').date()
        
        # 数値項目の取得
        max_overtime_monthly = int(request.form.get('max_overtime_hours_monthly', 45))
        max_overtime_yearly = int(request.form.get('max_overtime_hours_yearly', 360))
        
        # 既存の有効な36協定を無効化
        Agreement36History.query.filter_by(is_active=True).update({'is_active': False, 'status': 'superseded'})
        
        # 新しい36協定履歴を作成
        new_agreement = Agreement36History(
            submission_date=submission_date,
            period_start=period_start,
            period_end=period_end,
            expiry_date=expiry_date,
            max_overtime_hours_monthly=max_overtime_monthly,
            max_overtime_hours_yearly=max_overtime_yearly,
            labor_office_name=request.form.get('labor_office_name', ''),
            submission_method=request.form.get('submission_method', ''),
            document_number=request.form.get('document_number', ''),
            representative_name=request.form.get('representative_name', ''),
            notes=request.form.get('notes', ''),
            is_active=True,
            status='active',
            created_by=current_user.id if current_user.is_authenticated else None
        )
        
        db.session.add(new_agreement)
        db.session.commit()
        
        flash(f'36協定の提出が正常に記録されました。（提出日: {submission_date}）')
        
    except Exception as e:
        db.session.rollback()
        flash('36協定提出の記録中にエラーが発生しました。')
        print(f"Error: {e}")  # デバッグ用
    
    return redirect(url_for('general_affairs_36agreement'))

# --- 詳細な36協定管理 ---
@app.route('/create_36agreement', methods=['GET', 'POST'])
@login_required
def create_36agreement():
    """36協定の新規作成処理"""
    if current_user.role not in ['admin', 'general_affairs']:
        flash('アクセス権がありません。')
        return redirect(url_for('dashboard'))
    
    if request.method == 'GET':
        # GET リクエスト: フォーム表示
        # 企業情報を取得
        company_settings = CompanySettings.query.first()
        
        # 従業員一覧を取得（労働者代表・使用者選択用）
        employees = Employee.query.filter_by(status='在籍中').order_by(Employee.name).all()
        
        return render_template('create_36agreement.html',
                             company_settings=company_settings,
                             employees=employees)
    
    # POST リクエスト: フォーム送信処理
    try:
        # 必須項目のバリデーション
        required_fields = ['business_type', 'business_name', 'business_postal_code', 
                          'business_address', 'business_phone', 'agreement_start_date', 
                          'agreement_end_date', 'agreement_conclusion_date', 
                          'overtime_reason_1', 'overtime_business_type_1', 'overtime_employee_count',
                          'overtime_hours_daily', 'overtime_hours_monthly',
                          'worker_representative_employee_id', 'employer_employee_id']
        
        # 必須フィールドのチェック
        for field in required_fields:
            if not request.form.get(field):
                flash(f'必須項目「{field}」が入力されていません。')
                return redirect(url_for('create_36agreement'))
        
        # 日付フィールドの変換
        agreement_start_date = datetime.strptime(request.form['agreement_start_date'], '%Y-%m-%d').date()
        agreement_end_date = datetime.strptime(request.form['agreement_end_date'], '%Y-%m-%d').date()
        agreement_conclusion_date = datetime.strptime(request.form['agreement_conclusion_date'], '%Y-%m-%d').date()
        
        # 時刻フィールドの変換（任意）
        holiday_work_start_time = None
        holiday_work_end_time = None
        if request.form.get('holiday_work_start_time'):
            holiday_work_start_time = datetime.strptime(request.form['holiday_work_start_time'], '%H:%M').time()
        if request.form.get('holiday_work_end_time'):
            holiday_work_end_time = datetime.strptime(request.form['holiday_work_end_time'], '%H:%M').time()
        
        # 労働者代表・使用者の従業員情報を取得
        worker_representative = Employee.query.get(int(request.form['worker_representative_employee_id']))
        employer = Employee.query.get(int(request.form['employer_employee_id']))
        
        if not worker_representative or not employer:
            flash('労働者代表または使用者の従業員情報が見つかりません。')
            return redirect(url_for('create_36agreement'))
        
        # 既存の有効な36協定を無効化
        Agreement36.query.filter_by(is_active=True).update({'is_active': False, 'status': 'superseded'})
        
        # 新しい36協定を作成
        new_agreement = Agreement36(
            # 基本情報
            business_type=request.form['business_type'],
            business_name=request.form['business_name'],
            business_postal_code=request.form['business_postal_code'],
            business_address=request.form['business_address'],
            business_phone=request.form['business_phone'],
            
            # 協定期間
            agreement_start_date=agreement_start_date,
            agreement_end_date=agreement_end_date,
            
            # 時間外労働（複数事由対応）
            overtime_reason_1=request.form.get('overtime_reason_1'),
            overtime_reason_2=request.form.get('overtime_reason_2'),
            overtime_reason_3=request.form.get('overtime_reason_3'),
            overtime_reason_4=request.form.get('overtime_reason_4'),
            overtime_business_type_1=request.form.get('overtime_business_type_1'),
            overtime_business_type_2=request.form.get('overtime_business_type_2'),
            overtime_business_type_3=request.form.get('overtime_business_type_3'),
            overtime_business_type_4=request.form.get('overtime_business_type_4'),
            overtime_employee_count=int(request.form['overtime_employee_count']),
            overtime_hours_daily=int(request.form['overtime_hours_daily']),
            overtime_hours_monthly=int(request.form['overtime_hours_monthly']),
            
            # 休日労働（複数事由対応）
            holiday_work_reason_1=request.form.get('holiday_work_reason_1'),
            holiday_work_reason_2=request.form.get('holiday_work_reason_2'),
            holiday_work_reason_3=request.form.get('holiday_work_reason_3'),
            holiday_work_reason_4=request.form.get('holiday_work_reason_4'),
            holiday_work_business_type_1=request.form.get('holiday_work_business_type_1'),
            holiday_work_business_type_2=request.form.get('holiday_work_business_type_2'),
            holiday_work_business_type_3=request.form.get('holiday_work_business_type_3'),
            holiday_work_business_type_4=request.form.get('holiday_work_business_type_4'),
            holiday_work_employee_count=int(request.form.get('holiday_work_employee_count', 0)) if request.form.get('holiday_work_employee_count') else None,
            legal_holiday_days_count=int(request.form.get('legal_holiday_days_count', 0)) if request.form.get('legal_holiday_days_count') else None,
            holiday_work_start_time=holiday_work_start_time,
            holiday_work_end_time=holiday_work_end_time,
            
            # 協定締結情報
            agreement_conclusion_date=agreement_conclusion_date,
            worker_representative_selection_method=request.form.get('worker_representative_selection_method'),
            
            # 労働者代表
            worker_representative_employee_id=worker_representative.id,
            worker_representative_name=worker_representative.name,
            worker_representative_position=request.form.get('worker_representative_position', ''),
            
            # 使用者
            employer_employee_id=employer.id,
            employer_name=employer.name,
            employer_position=request.form.get('employer_position', ''),
            
            # 提出情報
            submission_date=None,  # まだ提出していない
            labor_office_name=None,
            submission_method=None,
            document_number=None,
            
            # 状態
            is_active=True,
            status='draft',
            
            # 備考
            notes=request.form.get('notes'),
            
            # メタ情報
            created_by=current_user.id
        )
        
        db.session.add(new_agreement)
        db.session.commit()
        
        flash('36協定が正常に作成されました。')
        return redirect(url_for('view_36agreement', id=new_agreement.id))
        
    except ValueError as e:
        db.session.rollback()
        flash('日付または数値の入力形式が正しくありません。')
        return redirect(url_for('create_36agreement'))
    except Exception as e:
        db.session.rollback()
        flash('36協定作成中にエラーが発生しました。')
        print(f"Error: {e}")  # デバッグ用
        return redirect(url_for('create_36agreement'))

@app.route('/view_36agreement/<int:id>')
@login_required
def view_36agreement(id):
    """36協定の詳細表示"""
    if current_user.role not in ['admin', 'general_affairs']:
        flash('アクセス権がありません。')
        return redirect(url_for('dashboard'))
    
    agreement = Agreement36.query.get_or_404(id)
    return render_template('view_36agreement.html', agreement=agreement)

@app.route('/list_36agreements')
@login_required
def list_36agreements():
    """36協定の一覧表示"""
    if current_user.role not in ['admin', 'general_affairs']:
        flash('アクセス権がありません。')
        return redirect(url_for('dashboard'))
    
    agreements = Agreement36.query.order_by(Agreement36.created_at.desc()).all()
    return render_template('list_36agreements.html', agreements=agreements)

# --- 総務事務専用36協定管理 ---
@app.route('/general_affairs_36agreement')
@login_required
def general_affairs_36agreement():
    if current_user.role not in ['admin', 'general_affairs']:
        flash('アクセス権がありません。')
        return redirect(url_for('dashboard'))
    
    # 現在有効な36協定を取得
    current_agreement = Agreement36History.query.filter_by(is_active=True).first()
    
    # 36協定履歴を取得
    agreement_history = Agreement36History.query.order_by(Agreement36History.submission_date.desc()).all()
    
    # 企業情報を取得
    company_settings = CompanySettings.query.first()
    
    # 法定休日設定を取得
    holiday_settings = LegalHolidaySettings.query.first()
    
    # 労働基準監督署情報を取得
    labor_settings = LaborStandardsSettings.query.first()
    
    return render_template('general_affairs_36agreement.html',
                         current_agreement=current_agreement,
                         agreement_history=agreement_history,
                         company_settings=company_settings,
                         holiday_settings=holiday_settings,
                         labor_settings=labor_settings)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
        
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        user = User.query.filter_by(email=email).first()
        
        if user and check_password_hash(user.password, password):
            login_user(user)
            return redirect(url_for('dashboard'))
        else:
            flash('メールアドレスまたはパスワードが正しくありません。')

    return render_template('login.html')

@app.route('/employee_login', methods=['GET', 'POST'])
def employee_login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
        
    if request.method == 'POST':
        employee_name = request.form.get('employee_name')
        employee_id = request.form.get('employee_id')
        
        if employee_name and employee_id:
            try:
                employee_id_int = int(employee_id)
                # 従業員IDと氏名で検索
                employee = Employee.query.filter_by(id=employee_id_int, name=employee_name).first()
                
                if employee:
                    # 該当する従業員ユーザーを検索
                    user = User.query.filter_by(employee_id=employee.id, role='employee').first()
                    if user:
                        login_user(user)
                        return redirect(url_for('dashboard'))
                    else:
                        flash('ユーザーアカウントが見つかりません。管理者にお問い合わせください。')
                else:
                    flash('従業員IDまたは氏名が正しくありません。')
            except ValueError:
                flash('従業員IDは数字で入力してください。')
        else:
            flash('従業員IDと氏名の両方を入力してください。')

    return render_template('employee_login.html')

@app.route('/admin_login', methods=['GET', 'POST'])
def admin_login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
        
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        # 総務事務ログイン、人事事務ログイン、従来のadminロールをサポート
        user = User.query.filter(User.email == email, User.role.in_(['admin', 'general_affairs', 'hr_affairs'])).first()
        
        if user and check_password_hash(user.password, password):
            login_user(user)
            return redirect(url_for('dashboard'))
        else:
            flash('メールアドレスまたはパスワードが正しくありません。')

    return render_template('admin_login.html')

@app.route('/accounting_login', methods=['GET', 'POST'])
def accounting_login():
    if current_user.is_authenticated:
        return redirect(url_for('accounting_dashboard'))
        
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        user = User.query.filter_by(email=email, role='accounting').first()
        
        if user and check_password_hash(user.password, password):
            login_user(user)
            return redirect(url_for('accounting_dashboard'))
        else:
            flash('メールアドレスまたはパスワードが正しくありません。')

    return render_template('accounting_login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))

# 従業員向け有給休暇申請機能
@app.route('/request_leave', methods=['GET', 'POST'])
@login_required
def request_leave():
    if current_user.role != 'employee':
        flash('アクセス権がありません。')
        return redirect(url_for('dashboard'))
        
    employee = Employee.query.filter_by(id=current_user.employee_id).first()
    if not employee:
        flash('従業員情報が見つかりません。')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        leave_type = request.form.get('leave_type', 'annual_leave')
        start_date_str = request.form.get('start_date')
        end_date_str = request.form.get('end_date')
        reason = request.form.get('reason')
        
        if start_date_str and end_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                
                if start_date > end_date:
                    flash('終了日は開始日以降を選択してください。')
                    return render_template('leave_request.html', employee=employee)
                
                # 営業日数を計算（土日を除く）
                days_requested = 0
                current_date = start_date
                while current_date <= end_date:
                    if current_date.weekday() < 5:  # 月〜金（0〜4）
                        days_requested += 1
                    current_date += timedelta(days=1)
                
                # 年次有給休暇の場合のみ残日数チェック
                if leave_type == 'annual_leave':
                    total_credited = sum([credit.days_credited for credit in employee.leave_credits])
                    total_taken = sum([record.days_taken for record in employee.leave_records])
                    remaining_leave = total_credited - total_taken
                    
                    if days_requested > remaining_leave:
                        flash(f'申請日数({days_requested}日)が残り有給日数({remaining_leave}日)を超えています。')
                        return render_template('leave_request.html', employee=employee)
                
                leave_request = LeaveRequest(
                    employee_id=employee.id,
                    leave_type=leave_type,
                    start_date=start_date,
                    end_date=end_date,
                    days_requested=days_requested,
                    reason=reason
                )
                
                db.session.add(leave_request)
                db.session.commit()
                
                flash('有給休暇申請を提出しました。管理者の承認をお待ちください。')
                return redirect(url_for('dashboard'))
                
            except ValueError:
                flash('日付の形式が正しくありません。')
                return render_template('leave_request.html', employee=employee)
        else:
            flash('開始日と終了日は必須です。')
    
    return render_template('leave_request.html', employee=employee)

@app.route('/leave_history')
@login_required
def leave_history():
    if current_user.role != 'employee':
        flash('アクセス権がありません。')
        return redirect(url_for('dashboard'))
        
    employee = Employee.query.filter_by(id=current_user.employee_id).first()
    if not employee:
        flash('従業員情報が見つかりません。')
        return redirect(url_for('dashboard'))
    
    leave_requests = LeaveRequest.query.filter_by(employee_id=employee.id).order_by(LeaveRequest.created_at.desc()).all()
    
    return render_template('leave_history.html', employee=employee, leave_requests=leave_requests)

# 従業員向け個人情報更新申請機能
@app.route('/request_personal_info_update', methods=['GET', 'POST'])
@login_required
def request_personal_info_update():
    if current_user.role != 'employee':
        flash('アクセス権がありません。')
        return redirect(url_for('dashboard'))
        
    employee = Employee.query.filter_by(id=current_user.employee_id).first()
    if not employee:
        flash('従業員情報が見つかりません。')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        
        request_type = request.form.get('request_type')
        new_value = request.form.get('new_value')
        new_file = request.files.get('new_file')
        reason = request.form.get('reason')
        
        # ファイルアップロードの場合とテキスト入力の場合を分ける
        uploaded_filename = None
        if request_type == 'residence_card_file':
            if new_file and new_file.filename:
                uploaded_filename = save_residence_card_file(new_file)
                if not uploaded_filename:
                    flash('ファイル形式が正しくありません。PNG、JPG、JPEG、PDF形式のみ対応しています。')
                    return render_template('personal_info_request.html', employee=employee)
                new_value = uploaded_filename  # ファイル名を新しい値として保存
            else:
                flash('ファイルを選択してください。')
                return render_template('personal_info_request.html', employee=employee)
        elif request_type == 'car_insurance_file':
            if new_file and new_file.filename:
                uploaded_filename = save_car_insurance_file(new_file)
                if not uploaded_filename:
                    flash('ファイル形式が正しくありません。PNG、JPG、JPEG、PDF形式のみ対応しています。')
                    return render_template('personal_info_request.html', employee=employee)
                new_value = uploaded_filename  # ファイル名を新しい値として保存
            else:
                flash('ファイルを選択してください。')
                return render_template('personal_info_request.html', employee=employee)
        
        # バリデーション：テキスト入力またはファイル入力のどちらかが必要
        if request_type in ['residence_card_file', 'car_insurance_file']:
            validation_ok = uploaded_filename is not None
        else:
            validation_ok = new_value and new_value.strip()
        
        if request_type and validation_ok:
            # 現在の値を取得
            current_value = ''
            if request_type == 'address':
                current_value = employee.address or ''
            elif request_type == 'phone':
                current_value = employee.phone_number or ''
            elif request_type == 'residence_card_expiry':
                current_value = employee.residence_card_expiry.strftime('%Y-%m-%d') if employee.residence_card_expiry else ''
            elif request_type == 'residence_card_file':
                current_value = employee.residence_card_filename or ''
            elif request_type == 'car_insurance_expiry':
                current_value = employee.car_insurance_expiry.strftime('%Y-%m-%d') if employee.car_insurance_expiry else ''
            elif request_type == 'car_insurance_file':
                current_value = employee.car_insurance_filename or ''
            
            try:
                personal_info_request = PersonalInfoRequest(
                    employee_id=employee.id,
                    request_type=request_type,
                    current_value=current_value,
                    new_value=new_value,
                    uploaded_filename=uploaded_filename,
                    reason=reason
                )
                
                db.session.add(personal_info_request)
                db.session.commit()
                
                flash('個人情報更新申請を提出しました。管理者の承認をお待ちください。')
                return redirect(url_for('dashboard'))
            except Exception as e:
                db.session.rollback()
                flash('申請の提出中にエラーが発生しました。もう一度お試しください。')
                return render_template('personal_info_request.html', employee=employee)
        else:
            if not request_type:
                flash('変更項目を選択してください。')
            elif request_type == 'residence_card_file':
                flash('ファイルを選択してください。')
            else:
                flash('新しい値を入力してください。')
    
    return render_template('personal_info_request.html', employee=employee)

# 管理者向け申請管理機能
@app.route('/admin_requests')
@login_required
def admin_requests():
    if current_user.role not in ['admin', 'hr_affairs']:
        flash('アクセス権がありません。')
        return redirect(url_for('dashboard'))
    
    
    # 承認待ちの申請を取得
    pending_leave_requests = LeaveRequest.query.filter_by(status='pending').order_by(LeaveRequest.created_at.desc()).all()
    pending_personal_info_requests = PersonalInfoRequest.query.filter_by(status='pending').order_by(PersonalInfoRequest.created_at.desc()).all()
    
    # 最近処理された申請も表示（直近10件）
    recent_leave_requests = LeaveRequest.query.filter(LeaveRequest.status.in_(['approved', 'rejected'])).order_by(LeaveRequest.reviewed_at.desc()).limit(10).all()
    recent_personal_info_requests = PersonalInfoRequest.query.filter(PersonalInfoRequest.status.in_(['approved', 'rejected'])).order_by(PersonalInfoRequest.reviewed_at.desc()).limit(10).all()
    
    return render_template('admin_requests.html',
                         pending_leave_requests=pending_leave_requests,
                         pending_personal_info_requests=pending_personal_info_requests,
                         recent_leave_requests=recent_leave_requests,
                         recent_personal_info_requests=recent_personal_info_requests)

@app.route('/review_leave_request/<int:request_id>/<action>')
@login_required
def review_leave_request(request_id, action):
    if current_user.role not in ['admin', 'hr_affairs']:
        flash('アクセス権がありません。')
        return redirect(url_for('dashboard'))
    
    
    leave_request = LeaveRequest.query.get_or_404(request_id)
    
    if action == 'approve':
        leave_request.status = 'approved'
        leave_request.reviewed_at = datetime.now()
        leave_request.reviewed_by = current_user.id
        
        # 有給休暇記録を追加
        leave_record = LeaveRecord(
            employee_id=leave_request.employee_id,
            date_taken=leave_request.start_date,
            days_taken=leave_request.days_requested
        )
        db.session.add(leave_record)
        
        flash(f'{leave_request.employee.name}さんの有給休暇申請を承認しました。')
        
    elif action == 'reject':
        leave_request.status = 'rejected'
        leave_request.reviewed_at = datetime.now()
        leave_request.reviewed_by = current_user.id
        
        flash(f'{leave_request.employee.name}さんの有給休暇申請を却下しました。')
    
    db.session.commit()
    return redirect(url_for('admin_requests'))

@app.route('/review_personal_info_request/<int:request_id>/<action>')
@login_required
def review_personal_info_request(request_id, action):
    if current_user.role not in ['admin', 'hr_affairs']:
        flash('アクセス権がありません。')
        return redirect(url_for('dashboard'))
    
    
    personal_info_request = PersonalInfoRequest.query.get_or_404(request_id)
    
    if action == 'approve':
        personal_info_request.status = 'approved'
        personal_info_request.reviewed_at = datetime.now()
        personal_info_request.reviewed_by = current_user.id
        
        # 実際に従業員情報を更新
        employee = personal_info_request.employee
        if personal_info_request.request_type == 'address':
            employee.address = personal_info_request.new_value
        elif personal_info_request.request_type == 'phone':
            employee.phone_number = personal_info_request.new_value
        elif personal_info_request.request_type == 'residence_card_expiry':
            employee.residence_card_expiry = datetime.strptime(personal_info_request.new_value, '%Y-%m-%d').date()
        elif personal_info_request.request_type == 'residence_card_file':
            employee.residence_card_filename = personal_info_request.uploaded_filename
        elif personal_info_request.request_type == 'car_insurance_expiry':
            employee.car_insurance_expiry = datetime.strptime(personal_info_request.new_value, '%Y-%m-%d').date()
        elif personal_info_request.request_type == 'car_insurance_file':
            employee.car_insurance_filename = personal_info_request.uploaded_filename
        
        flash(f'{employee.name}さんの個人情報更新申請を承認し、情報を更新しました。')
        
    elif action == 'reject':
        personal_info_request.status = 'rejected'
        personal_info_request.reviewed_at = datetime.now()
        personal_info_request.reviewed_by = current_user.id
        
        flash(f'{personal_info_request.employee.name}さんの個人情報更新申請を却下しました。')
    
    db.session.commit()
    return redirect(url_for('admin_requests'))

@app.route('/dashboard')
@login_required
def dashboard():
    if current_user.role in ['admin', 'general_affairs', 'hr_affairs']:
        return admin_dashboard()
    elif current_user.role == 'employee':
        return employee_dashboard()
    elif current_user.role == 'system_admin':
        return redirect(url_for('system_admin_dashboard'))
    elif current_user.role == 'accounting':
        return redirect(url_for('accounting_dashboard'))
    else:
        flash('アクセス権がありません。')
        return redirect(url_for('login'))

def admin_dashboard():
    employees = Employee.query.all()
    
    # 期限切れの警告フラグを計算（2か月/1か月の二段階）
    from datetime import date, timedelta
    today = date.today()
    threshold_60 = today + timedelta(days=60)
    threshold_30 = today + timedelta(days=30)

    def categorize(expiry):
        if not expiry:
            return 'none'
        if expiry <= today:
            return 'expired'
        if expiry <= threshold_30:
            return 'warn30'
        if expiry <= threshold_60:
            return 'warn60'
        return 'normal'

    for employee in employees:
        # 各レベルを算出
        employee.residence_card_level = categorize(employee.residence_card_expiry)
        employee.car_insurance_level = categorize(employee.car_insurance_expiry)

        # 後方互換（既存テンプレの参照があれば）
        level_map_to_old = {
            'none': 'none',
            'normal': 'normal',
            'warn60': 'warning',
            'warn30': 'warning',
            'expired': 'expired',
        }
        employee.residence_card_warning = level_map_to_old[employee.residence_card_level]
        employee.car_insurance_warning = level_map_to_old[employee.car_insurance_level]

        # 行背景は warn30 と expired のときのみ
        employee.is_bg_row = (
            employee.residence_card_level in ['warn30', 'expired'] or
            employee.car_insurance_level in ['warn30', 'expired']
        )
        # アイコンは warn30 と expired のときのみ
        employee.show_icon = (
            employee.residence_card_level in ['warn30', 'expired'] or
            employee.car_insurance_level in ['warn30', 'expired']
        )

    return render_template('dashboard.html', employees=employees)

def employee_dashboard():
    
    # 従業員情報の取得
    employee = Employee.query.filter_by(id=current_user.employee_id).first()
    if not employee:
        flash('従業員情報が見つかりません。')
        return redirect(url_for('login'))
    
    # 期限警告レベルの計算
    today = date.today()
    threshold_30 = today + timedelta(days=30)
    threshold_60 = today + timedelta(days=60)
    
    def categorize(expiry):
        if not expiry:
            return 'none'
        if expiry <= today:
            return 'expired'
        if expiry <= threshold_30:
            return 'warn30'
        if expiry <= threshold_60:
            return 'warn60'
        return 'normal'
    
    # 各レベルを算出
    employee.residence_card_level = categorize(employee.residence_card_expiry)
    employee.car_insurance_level = categorize(employee.car_insurance_expiry)
    
    # 有給休暇の残日数計算
    total_credited = sum([credit.days_credited for credit in employee.leave_credits])
    total_taken = sum([record.days_taken for record in employee.leave_records])
    remaining_leave = total_credited - total_taken
    
    # 申請履歴の取得
    leave_requests = LeaveRequest.query.filter_by(employee_id=employee.id).order_by(LeaveRequest.created_at.desc()).limit(5).all()
    personal_info_requests = PersonalInfoRequest.query.filter_by(employee_id=employee.id).order_by(PersonalInfoRequest.created_at.desc()).limit(5).all()
    
    return render_template('employee_dashboard.html', 
                         employee=employee, 
                         remaining_leave=remaining_leave,
                         leave_requests=leave_requests,
                         personal_info_requests=personal_info_requests)

@app.route('/add', methods=['POST'])
@login_required
def add_employee():
    if current_user.role != 'admin':
        flash('アクセス権がありません。')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        name = request.form.get('name')
        birth_date_str = request.form.get('birth_date')
        gender = request.form.get('gender')
        join_date_str = request.form.get('join_date')
        phone_number = request.form.get('phone_number')
        address = request.form.get('address')
        status = request.form.get('status')
        photo = request.files.get('photo')
        residence_card_file = request.files.get('residence_card_file')
        car_insurance_file = request.files.get('car_insurance_file')
        nationality = request.form.get('nationality')
        residence_card_expiry_str = request.form.get('residence_card_expiry')
        car_insurance_expiry_str = request.form.get('car_insurance_expiry')
        
        if name and join_date_str and status:
            try:
                # 日付文字列をDateオブジェクトに変換
                join_date = datetime.strptime(join_date_str, '%Y-%m-%d').date()
                
                # 生年月日の処理
                birth_date = None
                if birth_date_str:
                    birth_date = datetime.strptime(birth_date_str, '%Y-%m-%d').date()
                
                # 在留カード期限の処理
                residence_card_expiry = None
                if residence_card_expiry_str:
                    residence_card_expiry = datetime.strptime(residence_card_expiry_str, '%Y-%m-%d').date()
                
                # 自動車保険満了日の処理
                car_insurance_expiry = None
                if car_insurance_expiry_str:
                    car_insurance_expiry = datetime.strptime(car_insurance_expiry_str, '%Y-%m-%d').date()
                
                # 画像ファイルの処理
                photo_filename = None
                if photo:
                    photo_filename = save_photo(photo)
                    if not photo_filename:
                        flash('画像ファイルの形式が正しくありません。PNG、JPG、JPEG、GIF形式のみ対応しています。')
                        return redirect(url_for('dashboard'))
                
                # 在留カードファイルの処理
                residence_card_filename = None
                if residence_card_file:
                    residence_card_filename = save_residence_card_file(residence_card_file)
                    if not residence_card_filename:
                        flash('在留カードファイルの形式が正しくありません。PNG、JPG、JPEG、PDF形式のみ対応しています。')
                        return redirect(url_for('dashboard'))
                
                # 自動車保険証ファイルの処理
                car_insurance_filename = None
                if car_insurance_file:
                    car_insurance_filename = save_car_insurance_file(car_insurance_file)
                    if not car_insurance_filename:
                        flash('自動車保険証ファイルの形式が正しくありません。PNG、JPG、JPEG、PDF形式のみ対応しています。')
                        return redirect(url_for('dashboard'))
                
                # 新しい従業員を作成
                # 新しいフィールドの取得
                position = request.form.get('position')
                department = request.form.get('department')
                hire_type = request.form.get('hire_type')
                work_history = request.form.get('work_history')
                skills = request.form.get('skills')
                qualifications = request.form.get('qualifications')
                salary_grade = request.form.get('salary_grade')
                
                new_employee = Employee(
                    name=name,
                    birth_date=birth_date,
                    gender=gender if gender else None,
                    join_date=join_date,
                    phone_number=phone_number if phone_number else None,
                    address=address if address else None,
                    photo_filename=photo_filename,
                    nationality=nationality if nationality else None,
                    residence_card_expiry=residence_card_expiry,
                    residence_card_filename=residence_card_filename,
                    car_insurance_expiry=car_insurance_expiry,
                    car_insurance_filename=car_insurance_filename,
                    status=status,
                    # 新しい管理者専用フィールド
                    position=position if position else None,
                    department=department if department else None,
                    hire_type=hire_type if hire_type else None,
                    work_history=work_history if work_history else None,
                    skills=skills if skills else None,
                    qualifications=qualifications if qualifications else None,
                    salary_grade=salary_grade if salary_grade else None
                )
                
                # データベースに保存
                db.session.add(new_employee)
                db.session.commit()
                
                flash('従業員が正常に追加されました。')
                return redirect(url_for('dashboard'))
                
            except ValueError:
                flash('日付の形式が正しくありません。')
                return redirect(url_for('dashboard'))
        else:
            flash('必須項目（氏名、入社年月日、在籍状況）を入力してください。')
            return redirect(url_for('dashboard'))
    
    return redirect(url_for('dashboard'))

@app.route('/edit/<int:employee_id>', methods=['GET', 'POST'])
@login_required
def edit_employee(employee_id):
    if current_user.role != 'admin':
        flash('アクセス権がありません。')
        return redirect(url_for('login'))
    
    employee = Employee.query.get_or_404(employee_id)
    
    if request.method == 'POST':
        name = request.form.get('name')
        birth_date_str = request.form.get('birth_date')
        gender = request.form.get('gender')
        join_date_str = request.form.get('join_date')
        phone_number = request.form.get('phone_number')
        address = request.form.get('address')
        status = request.form.get('status')
        photo = request.files.get('photo')
        residence_card_file = request.files.get('residence_card_file')
        car_insurance_file = request.files.get('car_insurance_file')
        nationality = request.form.get('nationality')
        residence_card_expiry_str = request.form.get('residence_card_expiry')
        car_insurance_expiry_str = request.form.get('car_insurance_expiry')
        
        # 新しいHRフィールドを取得
        position = request.form.get('position')
        department = request.form.get('department')
        hire_type = request.form.get('hire_type')
        work_history = request.form.get('work_history')
        skills = request.form.get('skills')
        qualifications = request.form.get('qualifications')
        salary_grade = request.form.get('salary_grade')
        
        if name and join_date_str and status:
            try:
                # 日付文字列をDateオブジェクトに変換
                join_date = datetime.strptime(join_date_str, '%Y-%m-%d').date()
                
                # 生年月日の処理
                birth_date = None
                if birth_date_str:
                    birth_date = datetime.strptime(birth_date_str, '%Y-%m-%d').date()
                
                # 在留カード期限の処理
                residence_card_expiry = None
                if residence_card_expiry_str:
                    residence_card_expiry = datetime.strptime(residence_card_expiry_str, '%Y-%m-%d').date()
                
                # 自動車保険満了日の処理
                car_insurance_expiry = None
                if car_insurance_expiry_str:
                    car_insurance_expiry = datetime.strptime(car_insurance_expiry_str, '%Y-%m-%d').date()
                
                # 画像ファイルの処理
                if photo:
                    photo_filename = save_photo(photo)
                    if not photo_filename:
                        flash('画像ファイルの形式が正しくありません。PNG、JPG、JPEG、GIF形式のみ対応しています。')
                        return redirect(url_for('edit_employee', employee_id=employee_id))
                    employee.photo_filename = photo_filename
                
                # 在留カードファイルの処理
                if residence_card_file:
                    residence_card_filename = save_residence_card_file(residence_card_file)
                    if not residence_card_filename:
                        flash('在留カードファイルの形式が正しくありません。PNG、JPG、JPEG、PDF形式のみ対応しています。')
                        return redirect(url_for('edit_employee', employee_id=employee_id))
                    employee.residence_card_filename = residence_card_filename
                
                # 自動車保険証ファイルの処理
                if car_insurance_file:
                    car_insurance_filename = save_car_insurance_file(car_insurance_file)
                    if not car_insurance_filename:
                        flash('自動車保険証ファイルの形式が正しくありません。PNG、JPG、JPEG、PDF形式のみ対応しています。')
                        return redirect(url_for('edit_employee', employee_id=employee_id))
                    employee.car_insurance_filename = car_insurance_filename
                
                # 従業員情報を更新
                employee.name = name
                employee.birth_date = birth_date
                employee.gender = gender if gender else None
                employee.join_date = join_date
                employee.phone_number = phone_number if phone_number else None
                employee.address = address if address else None
                employee.nationality = nationality if nationality else None
                employee.residence_card_expiry = residence_card_expiry
                employee.car_insurance_expiry = car_insurance_expiry
                employee.status = status
                
                # 新しいHRフィールドを更新
                employee.position = position if position else None
                employee.department = department if department else None
                employee.hire_type = hire_type if hire_type else None
                employee.work_history = work_history if work_history else None
                employee.skills = skills if skills else None
                employee.qualifications = qualifications if qualifications else None
                employee.salary_grade = salary_grade if salary_grade else None
                
                # データベースに保存
                db.session.commit()
                
                flash('従業員情報が正常に更新されました。')
                return redirect(url_for('dashboard'))
                
            except ValueError:
                flash('日付の形式が正しくありません。')
        else:
            flash('必須項目（氏名、入社年月日、在籍状況）を入力してください。')
    
    return render_template('edit_employee.html', employee=employee)

@app.route('/delete/<int:employee_id>')
@login_required
def delete_employee(employee_id):
    if current_user.role != 'admin':
        flash('アクセス権がありません。')
        return redirect(url_for('login'))
    
    employee = Employee.query.get_or_404(employee_id)
    
    try:
        # 従業員を削除
        db.session.delete(employee)
        db.session.commit()
        
        flash('従業員が正常に削除されました。')
    except Exception:
        flash('従業員の削除中にエラーが発生しました。')
        db.session.rollback()
    
    return redirect(url_for('dashboard'))

@app.route('/leave_management')
@login_required
def leave_management():
    if current_user.role not in ['admin', 'hr_affairs']:
        flash('アクセス権がありません。')
        return redirect(url_for('login'))
    
    # 全従業員を取得
    employees = Employee.query.all()
    
    # 各従業員の年休付与合計、取得合計、残日数を計算
    for employee in employees:
        # 年休付与合計
        total_credited = db.session.query(db.func.sum(LeaveCredit.days_credited))\
            .filter(LeaveCredit.employee_id == employee.id)\
            .scalar() or 0
        employee.total_leave_credited = total_credited
        
        # 年休取得合計
        total_taken = db.session.query(db.func.sum(LeaveRecord.days_taken))\
            .filter(LeaveRecord.employee_id == employee.id)\
            .scalar() or 0
        employee.total_leave_taken = total_taken
        
        # 残日数
        employee.remaining_leave = total_credited - total_taken
        
        # 法律に基づく付与日数（join_dateがNoneの場合は0）
        if employee.join_date:
            employee.legal_leave_days = calculate_annual_leave_days(employee.join_date)
        else:
            employee.legal_leave_days = 0
        
        # 次回自動付与予定日を計算
        if employee.status == '在籍中' and employee.join_date:
            # 前回の自動付与日を取得
            last_auto_grant = db.session.query(LeaveCredit)\
                .filter(LeaveCredit.employee_id == employee.id)\
                .order_by(LeaveCredit.date_credited.desc())\
                .first()
            
            if last_auto_grant:
                # 前回の付与日から1年後
                from datetime import timedelta
                employee.next_auto_grant_date = last_auto_grant.date_credited + timedelta(days=365)
            else:
                # 入社日から1年後
                from datetime import timedelta
                employee.next_auto_grant_date = employee.join_date + timedelta(days=365)
        else:
            employee.next_auto_grant_date = None
    
    return render_template('leave_management.html', employees=employees)

@app.route('/add_leave_credit', methods=['POST'])
@login_required
def add_leave_credit():
    if current_user.role != 'admin':
        flash('アクセス権がありません。')
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        employee_id = request.form.get('employee_id')
        days_credited = request.form.get('days_credited')
        
        if employee_id and days_credited:
            try:
                employee_id = int(employee_id)
                days_credited = int(days_credited)
                
                # 従業員が存在するかチェック
                employee = Employee.query.get(employee_id)
                if not employee:
                    flash('指定された従業員が見つかりません。')
                    return redirect(url_for('leave_management'))
                
                # 年休付与を記録
                new_leave_credit = LeaveCredit(
                    employee_id=employee_id,
                    days_credited=days_credited,
                    date_credited=date.today()
                )
                
                db.session.add(new_leave_credit)
                db.session.commit()
                
                flash(f'{employee.name}に{days_credited}日分の年休を付与しました。')
                return redirect(url_for('leave_management'))
                
            except ValueError:
                flash('日数は正しい数値を入力してください。')
                return redirect(url_for('leave_management'))
        else:
            flash('従業員と日数を選択してください。')
            return redirect(url_for('leave_management'))
    
    return redirect(url_for('leave_management'))

@app.route('/add_leave_record', methods=['POST'])
@login_required
def add_leave_record():
    if current_user.role != 'admin':
        flash('アクセス権がありません。')
        return redirect(url_for('leave_management'))
    
    if request.method == 'POST':
        employee_id = request.form.get('employee_id')
        date_taken_str = request.form.get('date_taken')
        days_taken = request.form.get('days_taken')
        
        if employee_id and date_taken_str and days_taken:
            try:
                employee_id = int(employee_id)
                days_taken = int(days_taken)
                
                # 日付文字列をDateオブジェクトに変換
                date_taken = datetime.strptime(date_taken_str, '%Y-%m-%d').date()
                
                # 従業員が存在するかチェック
                employee = Employee.query.get(employee_id)
                if not employee:
                    flash('指定された従業員が見つかりません。')
                    return redirect(url_for('leave_management'))
                
                # 年休取得を記録
                new_leave_record = LeaveRecord(
                    employee_id=employee_id,
                    date_taken=date_taken,
                    days_taken=days_taken
                )
                
                db.session.add(new_leave_record)
                db.session.commit()
                
                flash(f'{employee.name}の年休取得（{days_taken}日）を記録しました。')
                return redirect(url_for('leave_management'))
                
            except ValueError:
                flash('日数は正しい数値を入力してください。')
                return redirect(url_for('leave_management'))
        else:
            flash('従業員、取得日、日数を入力してください。')
            return redirect(url_for('leave_management'))
    
    return redirect(url_for('leave_management'))

@app.route('/company_calendar')
@login_required
def company_calendar():
    if current_user.role not in ['admin', 'general_affairs']:
        flash('アクセス権がありません。')
        return redirect(url_for('login'))
    
    # イベントを日付順で取得
    calendar_events = CompanyCalendar.query.order_by(CompanyCalendar.event_date.asc()).all()
    
    # 現在の年の休日数を計算（祝日 + 土日）
    current_year = datetime.now().year
    holiday_count = 0
    
    # 祝日・会社休日をカウント
    for event in calendar_events:
        if event.event_type in ['holiday', 'company_holiday']:
            if event.is_recurring or event.event_date.year == current_year:
                holiday_count += 1
    
    # 土曜日・日曜日をカウント
    from calendar import monthrange
    weekend_count = 0
    for month in range(1, 13):
        days_in_month = monthrange(current_year, month)[1]
        for day in range(1, days_in_month + 1):
            weekday = date(current_year, month, day).weekday()
            if weekday in [5, 6]:  # 土曜日=5, 日曜日=6
                weekend_count += 1
    
    holiday_count += weekend_count
    
    # カレンダー開始月設定を取得
    start_month = int(get_calendar_setting('start_month', '1'))
    
    return render_template('company_calendar.html', 
                         calendar_events=calendar_events,
                         holiday_count=holiday_count,
                         current_year=current_year,
                         start_month=start_month)

@app.route('/calendar_view')
@login_required
def calendar_view():
    if current_user.role not in ['admin', 'general_affairs', 'employee']:
        flash('アクセス権がありません。')
        return redirect(url_for('login'))
    
    # 年を指定できるようにする（デフォルトは現在年）
    year = request.args.get('year', datetime.now().year, type=int)
    
    # カレンダー開始月設定を取得
    start_month = int(get_calendar_setting('start_month', '1'))
    
    # 年間カレンダーデータを生成
    calendar_data = generate_year_calendar(year, start_month)
    
    return render_template('calendar_view.html', calendar_data=calendar_data, year=year, start_month=start_month)

def generate_year_calendar(year, start_month=1):
    """指定された年の12ヶ月カレンダーデータを生成（開始月指定可能）"""
    import calendar
    calendar_data = []
    
    # その年のイベントを取得
    events = CompanyCalendar.query.filter(
        db.or_(
            db.extract('year', CompanyCalendar.event_date) == year,
            CompanyCalendar.is_recurring == True
        )
    ).all()
    
    # イベントを日付別に整理
    events_by_date = {}
    for event in events:
        if event.is_recurring:
            # 毎年繰り返しの場合、指定年の同じ月日に設定
            event_date = date(year, event.event_date.month, event.event_date.day)
        else:
            event_date = event.event_date
        
        if event_date.year == year:
            date_str = event_date.strftime('%Y-%m-%d')
            if date_str not in events_by_date:
                events_by_date[date_str] = []
            events_by_date[date_str].append(event)
    
    # 各月のカレンダーを生成（開始月から順番に）
    cal = calendar.Calendar(firstweekday=6)  # 日曜日を週の最初の日とする
    
    # 開始月から12ヶ月分生成
    for i in range(12):
        month = ((start_month - 1 + i) % 12) + 1
        display_year = year
        
        # 年をまたぐ場合の調整
        if start_month > 1 and month < start_month:
            display_year = year + 1
        
        month_data = {
            'month': month,
            'month_name': f'{month}月',
            'display_year': display_year,
            'weeks': []
        }
        
        # その月のカレンダーを取得
        month_calendar = cal.monthdayscalendar(display_year, month)
        
        for week in month_calendar:
            week_data = []
            for day in week:
                if day == 0:
                    # 月の範囲外の日
                    week_data.append({
                        'day': '',
                        'date': None,
                        'events': [],
                        'is_today': False,
                        'is_weekend': False,
                        'is_other_month': True
                    })
                else:
                    # 該当月の日
                    current_date = date(display_year, month, day)
                    date_str = current_date.strftime('%Y-%m-%d')
                    weekday = current_date.weekday()  # 0=月曜日, 6=日曜日
                    
                    # 土日判定（土曜日=5、日曜日=6）
                    is_weekend = weekday >= 5
                    
                    # 今日かどうか判定
                    is_today = current_date == date.today()
                    
                    week_data.append({
                        'day': day,
                        'date': current_date,
                        'events': events_by_date.get(date_str, []),
                        'is_today': is_today,
                        'is_weekend': is_weekend,
                        'is_other_month': False
                    })
            
            month_data['weeks'].append(week_data)
        
        # 週数を5週に固定（不足分は空行で補完）
        while len(month_data['weeks']) < 5:
            empty_week = []
            for i in range(7):
                empty_week.append({
                    'day': '',
                    'date': None,
                    'events': [],
                    'is_today': False,
                    'is_weekend': False,
                    'is_other_month': True
                })
            month_data['weeks'].append(empty_week)
        
        # 6週の場合は5週に制限
        if len(month_data['weeks']) > 5:
            month_data['weeks'] = month_data['weeks'][:5]
        
        calendar_data.append(month_data)
    
    return calendar_data

@app.route('/add_calendar_event', methods=['POST'])
@login_required
def add_calendar_event():
    if current_user.role != 'admin':
        flash('アクセス権がありません。')
        return redirect(url_for('company_calendar'))
    
    if request.method == 'POST':
        title = request.form.get('title')
        event_date_str = request.form.get('event_date')
        event_type = request.form.get('event_type')
        description = request.form.get('description')
        is_recurring = request.form.get('is_recurring') == 'on'
        
        if title and event_date_str and event_type:
            try:
                event_date = datetime.strptime(event_date_str, '%Y-%m-%d').date()
                
                new_event = CompanyCalendar(
                    title=title,
                    event_date=event_date,
                    event_type=event_type,
                    description=description if description else None,
                    is_recurring=is_recurring
                )
                
                db.session.add(new_event)
                db.session.commit()
                
                flash(f'{"休日" if event_type == "holiday" else "イベント"}「{title}」を追加しました。')
                return redirect(url_for('company_calendar'))
                
            except ValueError:
                flash('日付の形式が正しくありません。')
                return redirect(url_for('company_calendar'))
        else:
            flash('タイトル、日付、種別を入力してください。')
            return redirect(url_for('company_calendar'))
    
    return redirect(url_for('company_calendar'))

@app.route('/edit_calendar_event/<int:event_id>', methods=['GET', 'POST'])
@login_required
def edit_calendar_event(event_id):
    if current_user.role != 'admin':
        flash('アクセス権がありません。')
        return redirect(url_for('company_calendar'))
    
    event = CompanyCalendar.query.get_or_404(event_id)
    
    if request.method == 'POST':
        title = request.form.get('title')
        event_date_str = request.form.get('event_date')
        event_type = request.form.get('event_type')
        description = request.form.get('description')
        is_recurring = request.form.get('is_recurring') == 'on'
        
        if title and event_date_str and event_type:
            try:
                event_date = datetime.strptime(event_date_str, '%Y-%m-%d').date()
                
                event.title = title
                event.event_date = event_date
                event.event_type = event_type
                event.description = description if description else None
                event.is_recurring = is_recurring
                event.updated_at = date.today()
                
                db.session.commit()
                
                flash(f'{"休日" if event_type == "holiday" else "イベント"}「{title}」を更新しました。')
                return redirect(url_for('company_calendar'))
                
            except ValueError:
                flash('日付の形式が正しくありません。')
        else:
            flash('タイトル、日付、種別を入力してください。')
    
    return render_template('edit_calendar_event.html', event=event)

@app.route('/delete_calendar_event/<int:event_id>')
@login_required
def delete_calendar_event(event_id):
    if current_user.role != 'admin':
        flash('アクセス権がありません。')
        return redirect(url_for('company_calendar'))
    
    event = CompanyCalendar.query.get_or_404(event_id)
    
    try:
        db.session.delete(event)
        db.session.commit()
        flash(f'{"休日" if event.event_type == "holiday" else "イベント"}「{event.title}」を削除しました。')
    except Exception:
        flash('削除中にエラーが発生しました。')
        db.session.rollback()
    
    return redirect(url_for('company_calendar'))

@app.route('/calendar_settings', methods=['POST'])
@login_required
def calendar_settings():
    if current_user.role != 'admin':
        flash('アクセス権がありません。')
        return redirect(url_for('company_calendar'))
    
    if request.method == 'POST':
        start_month = request.form.get('start_month')
        
        if start_month and start_month.isdigit():
            start_month = int(start_month)
            if 1 <= start_month <= 12:
                set_calendar_setting('start_month', str(start_month), 'カレンダー表示開始月')
                flash(f'カレンダー開始月を{start_month}月に設定しました。')
            else:
                flash('開始月は1-12の範囲で指定してください。')
        else:
            flash('正しい月を選択してください。')
    
    return redirect(url_for('company_calendar'))

@app.route('/employee_detail/<int:employee_id>')
@login_required
def employee_detail(employee_id):
    if current_user.role != 'admin':
        flash('アクセス権がありません。')
        return redirect(url_for('login'))
    
    employee = Employee.query.get_or_404(employee_id)
    
    # 年休付与履歴
    leave_credits = LeaveCredit.query.filter_by(employee_id=employee_id).order_by(LeaveCredit.date_credited.desc()).all()
    
    # 年休取得履歴
    leave_records = LeaveRecord.query.filter_by(employee_id=employee_id).order_by(LeaveRecord.date_taken.desc()).all()
    
    # 年休合計計算
    total_credited = db.session.query(db.func.sum(LeaveCredit.days_credited))\
        .filter(LeaveCredit.employee_id == employee_id)\
        .scalar() or 0
    
    total_taken = db.session.query(db.func.sum(LeaveRecord.days_taken))\
        .filter(LeaveRecord.employee_id == employee_id)\
        .scalar() or 0
    
    remaining_leave = total_credited - total_taken
    
    # 法律に基づく付与日数
    if employee.join_date:
        legal_leave_days = calculate_annual_leave_days(employee.join_date)
    else:
        legal_leave_days = 0
    
    return render_template('employee_detail.html', 
                         employee=employee,
                         leave_credits=leave_credits,
                         leave_records=leave_records,
                         total_credited=total_credited,
                         total_taken=total_taken,
                         remaining_leave=remaining_leave,
                         legal_leave_days=legal_leave_days)

@app.route('/employee_pdf/<int:employee_id>')
@login_required
def employee_pdf(employee_id):
    if current_user.role != 'admin':
        flash('アクセス権がありません。')
        return redirect(url_for('login'))
    
    employee = Employee.query.get_or_404(employee_id)
    
    try:
        pdf_buffer = create_employee_pdf(employee)
        
        pdf_data = pdf_buffer.read()
        response = make_response(pdf_data)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=employee_{employee.id}_{date.today().strftime("%Y%m%d")}.pdf'
        
        return response
        
    except Exception as e:
        print(f"PDF generation error: {str(e)}")
        flash(f'PDFの生成中にエラーが発生しました: {str(e)}')
        return redirect(url_for('employee_detail', employee_id=employee_id))

def create_employee_excel_data(employee):
    """従業員情報をExcel形式で生成"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    
    # 年休データを計算
    total_credited = db.session.query(db.func.sum(LeaveCredit.days_credited))\
                               .filter_by(employee_id=employee.id).scalar() or 0
    total_taken = db.session.query(db.func.sum(LeaveRecord.days_taken))\
                            .filter_by(employee_id=employee.id).scalar() or 0
    remaining_leave = total_credited - total_taken
    
    # 法定付与日数の計算
    years_employed = (date.today() - employee.join_date).days / 365.25 if employee.join_date else 0
    if years_employed < 0.5:
        legal_leave_days = 0
    elif years_employed < 1.5:
        legal_leave_days = 10
    elif years_employed < 2.5:
        legal_leave_days = 11
    elif years_employed < 3.5:
        legal_leave_days = 12
    elif years_employed < 4.5:
        legal_leave_days = 14
    elif years_employed < 5.5:
        legal_leave_days = 16
    elif years_employed < 6.5:
        legal_leave_days = 18
    else:
        legal_leave_days = 20
    
    # 新しいワークブック作成
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"従業員情報_{employee.name}"
    
    # スタイル設定
    header_font = Font(name='メイリオ', size=14, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    title_font = Font(name='メイリオ', size=16, bold=True)
    normal_font = Font(name='メイリオ', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # タイトル行
    ws['A1'] = f"従業員情報 - {employee.name}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:B1')
    
    # 発行日
    ws['A2'] = f"発行日: {date.today().strftime('%Y年%m月%d日')}"
    ws['A2'].font = normal_font
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:B2')
    
    # 空行
    current_row = 4
    
    # 基本情報セクション
    ws[f'A{current_row}'] = "基本情報"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
    ws[f'B{current_row}'].fill = header_fill
    ws.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 1
    
    # 基本情報データ
    basic_info = [
        ('氏名', employee.name),
        ('生年月日', employee.birth_date.strftime('%Y年%m月%d日') if employee.birth_date else '未設定'),
        ('性別', employee.gender or '未設定'),
        ('入社年月日', employee.join_date.strftime('%Y年%m月%d日') if employee.join_date else '未設定'),
        ('電話番号', employee.phone_number or '未設定'),
        ('住所', employee.address or '未設定'),
        ('国籍', employee.nationality or '未設定'),
        ('在留カード期限', employee.residence_card_expiry.strftime('%Y年%m月%d日') if employee.residence_card_expiry else '未設定'),
        ('自動車保険満了日', employee.car_insurance_expiry.strftime('%Y年%m月%d日') if employee.car_insurance_expiry else '未設定'),
        ('在籍状況', employee.status or '未設定'),
    ]
    
    for label, value in basic_info:
        ws[f'A{current_row}'] = label
        ws[f'B{current_row}'] = value
        
        # スタイル適用
        ws[f'A{current_row}'].font = Font(name='メイリオ', size=11, bold=True)
        ws[f'B{current_row}'].font = normal_font
        ws[f'A{current_row}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        # ボーダー
        ws[f'A{current_row}'].border = border
        ws[f'B{current_row}'].border = border
        
        current_row += 1
    
    # 空行
    current_row += 1
    
    # 年次有給休暇情報
    ws[f'A{current_row}'] = "年次有給休暇情報"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = PatternFill(start_color='28a745', end_color='28a745', fill_type='solid')
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
    ws[f'B{current_row}'].fill = PatternFill(start_color='28a745', end_color='28a745', fill_type='solid')
    ws.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 1
    
    leave_info = [
        ('付与日数合計', f"{total_credited}日"),
        ('取得日数合計', f"{total_taken}日"),
        ('残日数', f"{remaining_leave}日"),
        ('法定付与日数', f"{legal_leave_days}日"),
    ]
    
    for label, value in leave_info:
        ws[f'A{current_row}'] = label
        ws[f'B{current_row}'] = value
        
        # スタイル適用
        ws[f'A{current_row}'].font = Font(name='メイリオ', size=11, bold=True)
        ws[f'B{current_row}'].font = normal_font
        ws[f'A{current_row}'].fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
        
        # ボーダー
        ws[f'A{current_row}'].border = border
        ws[f'B{current_row}'].border = border
        
        current_row += 1
    
    # 列幅調整
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    
    # BytesIOに保存
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer

@app.route('/employee_excel/<int:employee_id>')
@login_required
def employee_excel(employee_id):
    """従業員情報のExcelファイルを生成してダウンロード"""
    if current_user.role != 'admin':
        flash('アクセス権がありません。')
        return redirect(url_for('login'))
    
    employee = Employee.query.get_or_404(employee_id)
    
    try:
        excel_buffer = create_employee_excel_data(employee)
        
        excel_data = excel_buffer.read()
        response = make_response(excel_data)
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=employee_{employee.id}_{date.today().strftime("%Y%m%d")}.xlsx'
        
        return response
        
    except Exception as e:
        print(f"Excel generation error: {str(e)}")
        flash(f'Excelファイルの生成中にエラーが発生しました: {str(e)}')
        return redirect(url_for('employee_detail', employee_id=employee_id))

@app.route('/calendar_pdf')
@login_required  
def calendar_pdf():
    if current_user.role not in ['admin', 'employee']:
        flash('アクセス権がありません。')
        return redirect(url_for('login'))
    
    try:
        current_year = datetime.now().year
        pdf_buffer = create_calendar_pdf()
        
        response = make_response(pdf_buffer.read())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=company_calendar_{current_year}.pdf'
        
        return response
        
    except Exception as e:
        flash('PDFの生成中にエラーが発生しました。')
        return redirect(url_for('company_calendar'))

# 他の従業員管理機能（追加、編集、削除）はここに追加していきます。

# --- 人事評価機能 ---

@app.route('/performance_evaluation')
@login_required
def performance_evaluation():
    if current_user.role not in ['admin', 'hr_affairs']:
        flash('アクセス権がありません。')
        return redirect(url_for('dashboard'))
    
    # フィルタリングパラメータ
    period = request.args.get('period', '')
    department = request.args.get('department', '')
    
    # 評価データの取得
    query = PerformanceEvaluation.query.join(Employee)
    
    if period:
        query = query.filter(PerformanceEvaluation.evaluation_period == period)
    if department:
        query = query.filter(Employee.department == department)
    
    evaluations = query.order_by(PerformanceEvaluation.created_at.desc()).all()
    
    # 部署一覧の取得（フィルター用）
    departments = db.session.query(Employee.department).filter(Employee.department.isnot(None)).distinct().all()
    departments = [dept[0] for dept in departments]
    
    # 統計計算
    avg_rating = None
    approved_count = 0
    draft_count = 0
    
    if evaluations:
        ratings = [e.overall_rating for e in evaluations if e.overall_rating]
        if ratings:
            avg_rating = sum(ratings) / len(ratings)
        
        approved_count = len([e for e in evaluations if e.status == 'approved'])
        draft_count = len([e for e in evaluations if e.status == 'draft'])
    
    return render_template('performance_evaluation.html', 
                         evaluations=evaluations, 
                         departments=departments,
                         avg_rating=avg_rating,
                         approved_count=approved_count,
                         draft_count=draft_count)

@app.route('/add_evaluation', methods=['GET', 'POST'])
@login_required
def add_evaluation():
    if current_user.role != 'admin':
        flash('アクセス権がありません。')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        try:
            # フォームデータを取得
            employee_id = request.form.get('employee_id')
            evaluation_period = request.form.get('evaluation_period')
            performance_score = request.form.get('performance_score')
            behavior_score = request.form.get('behavior_score')
            potential_score = request.form.get('potential_score')
            strengths = request.form.get('strengths')
            areas_for_improvement = request.form.get('areas_for_improvement')
            goals_next_period = request.form.get('goals_next_period')
            evaluator_comment = request.form.get('evaluator_comment')
            status = request.form.get('status', 'draft')
            
            # 総合評価を計算
            scores = []
            if performance_score:
                scores.append(float(performance_score))
            if behavior_score:
                scores.append(float(behavior_score))
            if potential_score:
                scores.append(float(potential_score))
            
            overall_rating = sum(scores) / len(scores) if scores else None
            
            # 新しい評価を作成
            evaluation = PerformanceEvaluation(
                employee_id=employee_id,
                evaluation_period=evaluation_period,
                overall_rating=overall_rating,
                performance_score=float(performance_score) if performance_score else None,
                behavior_score=float(behavior_score) if behavior_score else None,
                potential_score=float(potential_score) if potential_score else None,
                strengths=strengths,
                areas_for_improvement=areas_for_improvement,
                goals_next_period=goals_next_period,
                evaluator_comment=evaluator_comment,
                evaluator_id=current_user.id,
                status=status
            )
            
            db.session.add(evaluation)
            db.session.commit()
            
            flash('人事評価を作成しました。', 'success')
            return redirect(url_for('performance_evaluation'))
            
        except Exception as e:
            db.session.rollback()
            flash('評価作成中にエラーが発生しました。', 'error')
    
    # GETリクエストの場合、作成フォームを表示
    employees = Employee.query.filter_by(status='在籍中').all()
    return render_template('add_evaluation.html', employees=employees)

@app.route('/view_evaluation/<int:evaluation_id>')
@login_required
def view_evaluation(evaluation_id):
    if current_user.role != 'admin':
        flash('アクセス権がありません。')
        return redirect(url_for('dashboard'))
    
    evaluation = PerformanceEvaluation.query.get_or_404(evaluation_id)
    return render_template('view_evaluation.html', evaluation=evaluation)

@app.route('/edit_evaluation/<int:evaluation_id>', methods=['GET', 'POST'])
@login_required
def edit_evaluation(evaluation_id):
    if current_user.role != 'admin':
        flash('アクセス権がありません。')
        return redirect(url_for('dashboard'))
    
    evaluation = PerformanceEvaluation.query.get_or_404(evaluation_id)
    
    if request.method == 'POST':
        try:
            # フォームデータを更新
            evaluation.evaluation_period = request.form.get('evaluation_period')
            evaluation.performance_score = float(request.form.get('performance_score')) if request.form.get('performance_score') else None
            evaluation.behavior_score = float(request.form.get('behavior_score')) if request.form.get('behavior_score') else None
            evaluation.potential_score = float(request.form.get('potential_score')) if request.form.get('potential_score') else None
            evaluation.strengths = request.form.get('strengths')
            evaluation.areas_for_improvement = request.form.get('areas_for_improvement')
            evaluation.goals_next_period = request.form.get('goals_next_period')
            evaluation.evaluator_comment = request.form.get('evaluator_comment')
            evaluation.status = request.form.get('status', 'draft')
            
            # 総合評価を再計算
            scores = []
            if evaluation.performance_score:
                scores.append(evaluation.performance_score)
            if evaluation.behavior_score:
                scores.append(evaluation.behavior_score)
            if evaluation.potential_score:
                scores.append(evaluation.potential_score)
            
            evaluation.overall_rating = sum(scores) / len(scores) if scores else None
            evaluation.updated_at = datetime.now()
            
            db.session.commit()
            
            flash('人事評価を更新しました。', 'success')
            return redirect(url_for('performance_evaluation'))
            
        except Exception as e:
            db.session.rollback()
            flash('評価更新中にエラーが発生しました。', 'error')
    
    employees = Employee.query.filter_by(status='在籍中').all()
    return render_template('edit_evaluation.html', evaluation=evaluation, employees=employees)

@app.route('/delete_evaluation/<int:evaluation_id>', methods=['POST'])
@login_required
def delete_evaluation(evaluation_id):
    if current_user.role != 'admin':
        flash('アクセス権がありません。')
        return redirect(url_for('dashboard'))
    
    try:
        evaluation = PerformanceEvaluation.query.get_or_404(evaluation_id)
        
        # 承認済みの評価は削除不可
        if evaluation.status == 'approved':
            flash('承認済みの評価は削除できません。', 'error')
            return redirect(url_for('performance_evaluation'))
        
        db.session.delete(evaluation)
        db.session.commit()
        
        flash('人事評価を削除しました。', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash('評価削除中にエラーが発生しました。', 'error')
    
    return redirect(url_for('performance_evaluation'))

# --- 人事評価承認機能 ---

@app.route('/hr_approve_evaluation/<int:evaluation_id>', methods=['POST'])
@login_required
def hr_approve_evaluation(evaluation_id):
    if not current_user.can_hr_approve:
        flash('人事承認権限がありません。', 'error')
        return redirect(url_for('performance_evaluation'))
    
    try:
        evaluation = PerformanceEvaluation.query.get_or_404(evaluation_id)
        
        # 提出済みの評価のみ承認可能
        if evaluation.status != 'submitted':
            flash('この評価は承認できません。', 'error')
            return redirect(url_for('performance_evaluation'))
        
        # 人事承認を記録
        evaluation.hr_approved_by = current_user.id
        evaluation.hr_approved_at = datetime.now()
        evaluation.hr_approval_comment = request.form.get('approval_comment', '')
        evaluation.status = 'hr_approved'
        evaluation.updated_at = datetime.now()
        
        db.session.commit()
        flash('人事承認が完了しました。取締役承認をお待ちください。', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash('承認処理中にエラーが発生しました。', 'error')
    
    return redirect(url_for('performance_evaluation'))

@app.route('/director_approve_evaluation/<int:evaluation_id>', methods=['POST'])
@login_required
def director_approve_evaluation(evaluation_id):
    if not current_user.can_director_approve:
        flash('取締役承認権限がありません。', 'error')
        return redirect(url_for('performance_evaluation'))
    
    try:
        evaluation = PerformanceEvaluation.query.get_or_404(evaluation_id)
        
        # 人事承認済みの評価のみ承認可能
        if evaluation.status != 'hr_approved':
            flash('この評価は承認できません。人事承認が必要です。', 'error')
            return redirect(url_for('performance_evaluation'))
        
        # 取締役承認を記録
        evaluation.director_approved_by = current_user.id
        evaluation.director_approved_at = datetime.now()
        evaluation.director_approval_comment = request.form.get('approval_comment', '')
        evaluation.status = 'approved'  # 最終承認完了
        evaluation.updated_at = datetime.now()
        
        db.session.commit()
        flash('取締役承認が完了しました。評価が正式に確定されました。', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash('承認処理中にエラーが発生しました。', 'error')
    
    return redirect(url_for('performance_evaluation'))

@app.route('/reject_evaluation/<int:evaluation_id>', methods=['POST'])
@login_required
def reject_evaluation(evaluation_id):
    if not (current_user.can_hr_approve or current_user.can_director_approve):
        flash('承認権限がありません。', 'error')
        return redirect(url_for('performance_evaluation'))
    
    try:
        evaluation = PerformanceEvaluation.query.get_or_404(evaluation_id)
        rejection_comment = request.form.get('rejection_comment', '')
        rejection_reason = f"承認者コメント: {rejection_comment}"
        
        # 評価を差し戻し（下書きに戻す）
        evaluation.status = 'draft'
        evaluation.evaluator_comment = f"{evaluation.evaluator_comment or ''}\n\n【差し戻し理由】\n{rejection_reason}"
        evaluation.updated_at = datetime.now()
        
        # 承認情報をクリア
        evaluation.hr_approved_by = None
        evaluation.hr_approved_at = None
        evaluation.hr_approval_comment = None
        evaluation.director_approved_by = None
        evaluation.director_approved_at = None
        evaluation.director_approval_comment = None
        
        db.session.commit()
        flash('評価を差し戻しました。評価者による修正をお待ちください。', 'warning')
        
    except Exception as e:
        db.session.rollback()
        flash('差し戻し処理中にエラーが発生しました。', 'error')
    
    return redirect(url_for('performance_evaluation'))

# --- 組織図機能 ---

@app.route('/organization_chart')
@login_required
def organization_chart():
    if current_user.role not in ['admin', 'general_affairs']:
        flash('アクセス権がありません。')
        return redirect(url_for('dashboard'))
    
    # 従業員データを取得
    employees = Employee.query.filter_by(status='在籍中').all()
    
    # 統計データの計算
    dept_stats = {}
    position_stats = {}
    hire_type_stats = {}
    
    for emp in employees:
        # 部署別統計
        dept = emp.department or '未設定'
        dept_stats[dept] = dept_stats.get(dept, 0) + 1
        
        # 役職別統計
        pos = emp.position or '未設定'
        position_stats[pos] = position_stats.get(pos, 0) + 1
        
        # 雇用形態別統計
        hire = emp.hire_type or '未設定'
        hire_type_stats[hire] = hire_type_stats.get(hire, 0) + 1
    
    # 組織図用データ（JSON形式）
    org_data = []
    for emp in employees:
        org_data.append({
            'id': emp.id,
            'name': emp.name,
            'position': emp.position or '',
            'department': emp.department or '',
            'manager_id': emp.manager_id
        })
    
    import json
    org_data_json = json.dumps(org_data, ensure_ascii=False)
    
    return render_template('organization_chart.html',
                         employees=employees,
                         dept_stats=dept_stats,
                         position_stats=position_stats,
                         hire_type_stats=hire_type_stats,
                         org_data=org_data_json)

@app.route('/organization_chart_pdf')
@login_required
def organization_chart_pdf():
    if current_user.role != 'admin':
        flash('アクセス権がありません。')
        return redirect(url_for('dashboard'))
    
    # 組織図PDF出力（簡易実装）
    flash('組織図PDF出力機能は開発中です。')
    return redirect(url_for('organization_chart'))

# --- システム管理者機能 ---

@app.route('/system_admin_login', methods=['GET', 'POST'])
def system_admin_login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        
        user = User.query.filter_by(email=email, role='system_admin').first()
        if user and check_password_hash(user.password, password):
            login_user(user)
            return redirect(url_for('system_admin_dashboard'))
        else:
            flash('メールアドレスまたはパスワードが間違っています。')
    
    return render_template('system_admin_login.html')

@app.route('/system_admin_dashboard')
@login_required
def system_admin_dashboard():
    if current_user.role != 'system_admin':
        flash('アクセス権がありません。')
        return redirect(url_for('login'))
    
    admin_users = User.query.filter(User.role.in_(['admin', 'system_admin'])).all()
    return render_template('system_admin_dashboard.html', admin_users=admin_users)

@app.route('/add_admin_user', methods=['POST'])
@login_required
def add_admin_user():
    if current_user.role != 'system_admin':
        flash('アクセス権がありません。')
        return redirect(url_for('login'))
    
    email = request.form['email']
    password = request.form['password']
    role = request.form['role']
    
    # メールアドレスの重複チェック
    existing_user = User.query.filter_by(email=email).first()
    if existing_user:
        flash('このメールアドレスは既に使用されています。')
        return redirect(url_for('system_admin_dashboard'))
    
    # 新しいユーザーを作成
    new_user = User(
        email=email,
        password=generate_password_hash(password),
        role=role
    )
    
    db.session.add(new_user)
    db.session.commit()
    
    flash(f'{role}アカウントが正常に追加されました。')
    return redirect(url_for('system_admin_dashboard'))

@app.route('/edit_admin_user/<int:user_id>', methods=['GET', 'POST'])
@login_required
def edit_admin_user(user_id):
    if current_user.role != 'system_admin':
        flash('アクセス権がありません。')
        return redirect(url_for('login'))
    
    user = User.query.get_or_404(user_id)
    
    if request.method == 'POST':
        user.email = request.form['email']
        user.role = request.form['role']
        
        # パスワードが入力された場合のみ更新
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        
        if new_password:
            if new_password != confirm_password:
                flash('パスワードが一致しません。')
                return render_template('edit_admin_user.html', user=user)
            user.password = generate_password_hash(new_password)
        
        db.session.commit()
        flash('ユーザー情報が正常に更新されました。')
        return redirect(url_for('system_admin_dashboard'))
    
    return render_template('edit_admin_user.html', user=user)

@app.route('/delete_admin_user/<int:user_id>')
@login_required
def delete_admin_user(user_id):
    if current_user.role != 'system_admin':
        flash('アクセス権がありません。')
        return redirect(url_for('login'))
    
    user = User.query.get_or_404(user_id)
    
    # システム管理者が最後の1人の場合は削除を防ぐ
    if user.role == 'system_admin':
        system_admin_count = User.query.filter_by(role='system_admin').count()
        if system_admin_count <= 1:
            flash('システム管理者は最低1人必要です。')
            return redirect(url_for('system_admin_dashboard'))
    
    db.session.delete(user)
    db.session.commit()
    
    flash('ユーザーアカウントが削除されました。')
    return redirect(url_for('system_admin_dashboard'))

@app.route('/accounting_dashboard')
@login_required
def accounting_dashboard():
    if current_user.role != 'accounting':
        flash('アクセス権限がありません。')
        return redirect(url_for('index'))
    
    # 統計データ取得
    total_employees = Employee.query.filter_by(status='在籍中').count()
    
    # 今月の有給申請数
    today = datetime.now()
    start_of_month = today.replace(day=1)
    monthly_leave_requests = LeaveRequest.query.filter(
        LeaveRequest.created_at >= start_of_month
    ).count()
    
    # 人事評価総数
    total_evaluations = PerformanceEvaluation.query.count()
    
    # 今月の休日数
    monthly_holidays = CompanyCalendar.query.filter(
        CompanyCalendar.event_date >= start_of_month,
        CompanyCalendar.event_date < (start_of_month + timedelta(days=32)).replace(day=1),
        CompanyCalendar.event_type == 'holiday'
    ).count()
    
    # 最近の活動（サンプルデータ）
    recent_activities = [
        {
            'date': '2025年08月27日',
            'type': '有給申請',
            'target': '田中 太郎',
            'status': '承認待ち',
            'status_color': 'warning'
        },
        {
            'date': '2025年08月26日',
            'type': '人事評価',
            'target': '佐藤 花子',
            'status': '完了',
            'status_color': 'success'
        }
    ]
    
    return render_template('accounting_dashboard.html', 
                         total_employees=total_employees,
                         monthly_leave_requests=monthly_leave_requests,
                         total_evaluations=total_evaluations,
                         monthly_holidays=monthly_holidays,
                         recent_activities=recent_activities)

@app.route('/working_time_input', methods=['GET', 'POST'])
@login_required
def working_time_input():
    if current_user.role != 'accounting':
        flash('アクセス権限がありません。')
        return redirect(url_for('index'))
    
    employees = Employee.query.filter_by(status='在籍中').all()
    current_year = datetime.now().year
    years = list(range(current_year - 2, current_year + 2))
    
    # フィルター取得
    selected_employee_id = request.args.get('employee_id') or request.form.get('employee_id')
    selected_year = int(request.args.get('year') or request.form.get('year') or current_year)
    selected_month = int(request.args.get('month') or request.form.get('month') or datetime.now().month)
    
    selected_employee = None
    calendar_days = []
    
    if selected_employee_id:
        selected_employee = Employee.query.get(selected_employee_id)
        if selected_employee:
            # カレンダー生成
            from calendar import monthrange
            _, days_in_month = monthrange(selected_year, selected_month)
            
            # 既存の勤怠データを取得
            existing_records = {}
            records = WorkingTimeRecord.query.filter(
                WorkingTimeRecord.employee_id == selected_employee.id,
                db.extract('year', WorkingTimeRecord.work_date) == selected_year,
                db.extract('month', WorkingTimeRecord.work_date) == selected_month
            ).all()
            
            for record in records:
                existing_records[record.work_date.day] = record
            
            # 会社カレンダーの祝日を取得
            holidays = set()
            company_holidays = CompanyCalendar.query.filter(
                db.extract('year', CompanyCalendar.event_date) == selected_year,
                db.extract('month', CompanyCalendar.event_date) == selected_month,
                CompanyCalendar.event_type == 'holiday'
            ).all()
            for holiday in company_holidays:
                holidays.add(holiday.event_date.day)
            
            # 法定休日設定を取得
            holiday_settings = LegalHolidaySettings.query.first()
            
            # カレンダーデータ生成
            import calendar
            for day in range(1, days_in_month + 1):
                work_date = date(selected_year, selected_month, day)
                weekday = work_date.weekday()
                weekday_names = ['月', '火', '水', '木', '金', '土', '日']
                
                record = existing_records.get(day)
                # 時分を分離して表示
                start_hour, start_minute = '', ''
                end_hour, end_minute = '', ''
                if record and record.start_time:
                    start_hour = record.start_time.hour
                    start_minute = record.start_time.minute
                if record and record.end_time:
                    end_hour = record.end_time.hour
                    end_minute = record.end_time.minute
                
                # 法定休日判定
                is_legal_holiday = False
                is_company_holiday = False
                
                # 会社カレンダーの祝日チェック
                if day in holidays:
                    # 会社カレンダーの祝日は設定に基づいて判定
                    if holiday_settings and hasattr(holiday_settings, 'specific_date_legal'):
                        is_legal_holiday = holiday_settings.specific_date_legal
                    else:
                        is_legal_holiday = True  # デフォルトでは法定休日
                else:
                    # 曜日別法定休日設定をチェック
                    if holiday_settings:
                        weekday_legal_flags = [
                            holiday_settings.monday_legal_holiday,    # 0: 月曜日
                            holiday_settings.tuesday_legal_holiday,   # 1: 火曜日  
                            holiday_settings.wednesday_legal_holiday, # 2: 水曜日
                            holiday_settings.thursday_legal_holiday,  # 3: 木曜日
                            holiday_settings.friday_legal_holiday,    # 4: 金曜日
                            holiday_settings.saturday_legal_holiday,  # 5: 土曜日
                            holiday_settings.sunday_legal_holiday     # 6: 日曜日
                        ]
                        
                        if weekday < len(weekday_legal_flags):
                            is_legal_holiday = weekday_legal_flags[weekday]
                        
                        # 法定休日でない場合は会社休日として扱わない（設定に基づいて判定のみ）
                        if not is_legal_holiday:
                            is_company_holiday = False
                    else:
                        # 設定がない場合はデフォルト
                        if weekday == 6:  # 日曜日
                            is_legal_holiday = True
                        elif weekday == 5:  # 土曜日
                            is_company_holiday = True  # 法定外休日
                
                calendar_days.append({
                    'day': day,
                    'weekday': weekday,
                    'weekday_name': weekday_names[weekday],
                    'is_holiday': is_legal_holiday or is_company_holiday,
                    'is_legal_holiday': is_legal_holiday,
                    'is_company_holiday': is_company_holiday,
                    'start_hour': start_hour,
                    'start_minute': start_minute,
                    'end_hour': end_hour,
                    'end_minute': end_minute,
                    'break_time': record.break_time_minutes if record else 0,
                    'is_paid_leave': record.is_paid_leave if record else False,
                    'is_special_leave': record.is_special_leave if record else False,
                    'is_absence': record.is_absence if record else False,
                    'is_company_closure': record.is_company_closure if record else False,
                    'remarks': ''
                })
    
    if request.method == 'POST' and selected_employee:
        action = request.form.get('action')
        
        # データ保存処理
        for day in range(1, len(calendar_days) + 1):
            work_date = date(selected_year, selected_month, day)
            
            # 既存レコードを取得または作成
            record = WorkingTimeRecord.query.filter(
                WorkingTimeRecord.employee_id == selected_employee.id,
                WorkingTimeRecord.work_date == work_date
            ).first()
            
            if not record:
                record = WorkingTimeRecord(
                    employee_id=selected_employee.id,
                    work_date=work_date,
                    input_by=current_user.id
                )
                db.session.add(record)
            
            # フォームデータから更新（時分を個別に取得）
            start_hour = request.form.get(f'start_hour_{day}')
            start_minute = request.form.get(f'start_minute_{day}')
            end_hour = request.form.get(f'end_hour_{day}')
            end_minute = request.form.get(f'end_minute_{day}')
            
            # 時分が入力された場合に時刻を設定（時のみまたは分のみでも可）
            if start_hour is not None and start_hour != '':
                try:
                    hour = int(start_hour)
                    minute = int(start_minute) if start_minute and start_minute != '' else 0
                    if 0 <= hour <= 23 and 0 <= minute <= 59:
                        record.start_time = datetime.strptime(f'{hour:02d}:{minute:02d}', '%H:%M').time()
                    else:
                        record.start_time = None
                except (ValueError, TypeError):
                    record.start_time = None
            else:
                record.start_time = None
                
            if end_hour is not None and end_hour != '':
                try:
                    hour = int(end_hour)
                    minute = int(end_minute) if end_minute and end_minute != '' else 0
                    if 0 <= hour <= 23 and 0 <= minute <= 59:
                        record.end_time = datetime.strptime(f'{hour:02d}:{minute:02d}', '%H:%M').time()
                    else:
                        record.end_time = None
                except (ValueError, TypeError):
                    record.end_time = None
            else:
                record.end_time = None
            record.break_time_minutes = int(request.form.get(f'break_time_{day}') or 0)
            record.is_paid_leave = bool(request.form.get(f'paid_leave_{day}'))
            record.is_special_leave = bool(request.form.get(f'special_leave_{day}'))
            record.is_absence = bool(request.form.get(f'absence_{day}'))
            record.is_company_closure = bool(request.form.get(f'company_closure_{day}'))
            record.updated_at = datetime.now()
            
            # 労働時間計算（日本労働基準法準拠）
            if record.start_time and record.end_time and not (record.is_paid_leave or record.is_special_leave or record.is_absence):
                start_datetime = datetime.combine(work_date, record.start_time)
                end_datetime = datetime.combine(work_date, record.end_time)
                
                # 日をまたぐ場合の処理
                if end_datetime <= start_datetime:
                    end_datetime = end_datetime + timedelta(days=1)
                
                total_minutes = int((end_datetime - start_datetime).total_seconds() / 60) - record.break_time_minutes
                
                if total_minutes > 0:
                    # 週の始まりを取得（月曜日を0とする）
                    weekday = work_date.weekday()
                    
                    # 法定休日設定を取得
                    holiday_settings = LegalHolidaySettings.query.first()
                    
                    # 法定休日判定（設定に基づく）
                    is_legal_holiday = False
                    is_company_holiday = False
                    
                    # 会社カレンダーの祝日チェック
                    company_holiday = CompanyCalendar.query.filter(
                        CompanyCalendar.event_date == work_date,
                        CompanyCalendar.event_type == 'holiday'
                    ).first()
                    if company_holiday:
                        # 会社カレンダーの祝日は設定に基づいて判定
                        if holiday_settings and hasattr(holiday_settings, 'specific_date_legal'):
                            is_legal_holiday = holiday_settings.specific_date_legal
                        else:
                            is_legal_holiday = True  # デフォルトでは法定休日
                    
                    # 曜日別法定休日設定をチェック（会社カレンダーがない場合）
                    if not company_holiday:
                        if holiday_settings:
                            weekday_legal_flags = [
                                holiday_settings.monday_legal_holiday,    # 0: 月曜日
                                holiday_settings.tuesday_legal_holiday,   # 1: 火曜日  
                                holiday_settings.wednesday_legal_holiday, # 2: 水曜日
                                holiday_settings.thursday_legal_holiday,  # 3: 木曜日
                                holiday_settings.friday_legal_holiday,    # 4: 金曜日
                                holiday_settings.saturday_legal_holiday,  # 5: 土曜日
                                holiday_settings.sunday_legal_holiday     # 6: 日曜日
                            ]
                            
                            if weekday < len(weekday_legal_flags):
                                is_legal_holiday = weekday_legal_flags[weekday]
                            
                            # 法定休日でない場合は会社休日として扱う
                            if not is_legal_holiday:
                                is_company_holiday = False  # 設定に基づいて判定のみ
                        else:
                            # 設定がない場合はデフォルト
                            if weekday == 6:  # 日曜日
                                is_legal_holiday = True
                            elif weekday == 5:  # 土曜日
                                is_company_holiday = True  # 法定外休日
                    
                    # 労働基準法準拠の労働時間分類
                    if is_legal_holiday:
                        # 法定休日労働（35%割増）- 全て法定休日労働時間
                        record.legal_holiday_minutes = total_minutes
                        record.regular_working_minutes = 0
                        record.legal_overtime_minutes = 0
                        record.overtime_minutes = 0
                        record.holiday_minutes = 0
                    else:
                        # 平日・法定外休日の労働時間計算（週40時間制適用前の仮分類）
                        # ここでは全て通常労働時間として記録し、後で週40時間制で再分類する
                        record.regular_working_minutes = total_minutes
                        record.legal_overtime_minutes = 0
                        record.overtime_minutes = 0
                        record.holiday_minutes = 0
                        record.legal_holiday_minutes = 0
                    
                    # 深夜労働時間計算（22:00-5:00）
                    record.night_working_minutes = calculate_night_work_minutes(start_datetime, end_datetime, work_date)
            else:
                # 休暇・欠勤の場合は労働時間を0にリセット
                record.regular_working_minutes = 0
                record.legal_overtime_minutes = 0
                record.overtime_minutes = 0
                record.legal_holiday_minutes = 0
                record.holiday_minutes = 0
                record.night_working_minutes = 0
        
        # 週40時間制限に基づく労働時間再計算
        calculate_weekly_overtime(selected_employee.id, selected_year, selected_month)
        
        db.session.commit()
        flash('労働時間データを保存しました。')
        
        # 給与計算実行
        if action == 'calculate':
            try:
                calculate_monthly_payroll(selected_employee.id, selected_year, selected_month)
                flash('給与計算を実行しました。')
            except Exception as e:
                db.session.rollback()  # セッションをロールバック
                flash(f'給与計算でエラーが発生しました: {str(e)}')
        
        # セッションから切り離されたIDを使用
        employee_id = selected_employee.id
        return redirect(url_for('working_time_input', 
                               employee_id=employee_id, 
                               year=selected_year, 
                               month=selected_month))
    
    # 環境設定を取得
    holiday_settings = LegalHolidaySettings.query.first()
    
    return render_template('working_time_input.html',
                         employees=employees,
                         years=years,
                         selected_employee=selected_employee,
                         selected_year=selected_year,
                         selected_month=selected_month,
                         calendar_days=calendar_days,
                         holiday_settings=holiday_settings)

def calculate_night_work_minutes(start_datetime, end_datetime, work_date):
    """深夜労働時間計算（22:00-5:00）"""
    night_minutes = 0
    
    # 当日の深夜時間帯（22:00-24:00）
    night_start_today = datetime.combine(work_date, datetime.strptime('22:00', '%H:%M').time())
    night_end_today = datetime.combine(work_date + timedelta(days=1), datetime.strptime('00:00', '%H:%M').time())
    
    # 翌日の深夜時間帯（0:00-5:00）
    night_start_tomorrow = datetime.combine(work_date + timedelta(days=1), datetime.strptime('00:00', '%H:%M').time())
    night_end_tomorrow = datetime.combine(work_date + timedelta(days=1), datetime.strptime('05:00', '%H:%M').time())
    
    # 当日22:00-24:00の計算
    if start_datetime < night_end_today and end_datetime > night_start_today:
        actual_start = max(start_datetime, night_start_today)
        actual_end = min(end_datetime, night_end_today)
        if actual_end > actual_start:
            night_minutes += int((actual_end - actual_start).total_seconds() / 60)
    
    # 翌日0:00-5:00の計算
    if start_datetime < night_end_tomorrow and end_datetime > night_start_tomorrow:
        actual_start = max(start_datetime, night_start_tomorrow)
        actual_end = min(end_datetime, night_end_tomorrow)
        if actual_end > actual_start:
            night_minutes += int((actual_end - actual_start).total_seconds() / 60)
    
    return night_minutes

def calculate_weekly_overtime(employee_id, year, month):
    """週40時間制限に基づく労働時間再計算（月曜日リセット・クロスマンス対応）
    
    日本の労働基準法第32条に基づき、週40時間を超える労働時間を
    法定外労働時間（25%割増）として再分類します。
    
    週の起算日は月曜日とし、毎週月曜日に法定内労働時間がリセットされます。
    月をまたぐ週についても適切に処理します。
    
    Args:
        employee_id: 従業員ID
        year: 年
        month: 月
    """
    from calendar import monthrange
    from datetime import date, timedelta
    
    print(f"🔄 週40時間制限計算開始（月曜日リセット）: {year}年{month}月 (従業員ID: {employee_id})")
    
    # 月の開始日と終了日を取得
    _, days_in_month = monthrange(year, month)
    month_start = date(year, month, 1)
    month_end = date(year, month, days_in_month)
    
    # 対象月に含まれる週を特定（月曜日起算）
    weeks_to_process = set()
    
    # 月の各日について、その日が属する週の月曜日を記録
    current_date = month_start
    while current_date <= month_end:
        # その日が属する週の月曜日を計算
        # weekday(): 月=0, 火=1, 水=2, 木=3, 金=4, 土=5, 日=6
        days_since_monday = current_date.weekday()  # 月曜日からの日数
        week_monday = current_date - timedelta(days=days_since_monday)
        weeks_to_process.add(week_monday)
        current_date += timedelta(days=1)
    
    print(f"📅 処理対象の週: {len(weeks_to_process)}週（月曜日起算）")
    for week_start in sorted(weeks_to_process):
        week_end = week_start + timedelta(days=6)
        print(f"  {week_start}(月) 〜 {week_end}(日)")
    
    # 各週について40時間制限を適用
    for week_monday in weeks_to_process:
        week_sunday = week_monday + timedelta(days=6)
        
        print(f"\n⏰ 週処理中: {week_monday}(月) 〜 {week_sunday}(日)")
        
        # その週の全レコードを取得（クロスマンス対応）
        all_week_records = WorkingTimeRecord.query.filter(
            WorkingTimeRecord.employee_id == employee_id,
            WorkingTimeRecord.work_date >= week_monday,
            WorkingTimeRecord.work_date <= week_sunday
        ).order_by(WorkingTimeRecord.work_date).all()
        
        if not all_week_records:
            print(f"   📋 この週にはレコードがありません")
            continue
        
        print(f"   📋 この週のレコード数: {len(all_week_records)}")
        
        # 法定休日を除いた労働時間を計算
        workday_records = []
        total_work_minutes = 0
        
        for record in all_week_records:
            day_name = ['月', '火', '水', '木', '金', '土', '日'][record.work_date.weekday()]
            
            # 法定休日労働は週40時間計算から除外
            if record.holiday_minutes and record.holiday_minutes > 0:
                print(f"   {record.work_date} ({day_name}): 法定休日労働 {record.holiday_minutes}分 - 除外")
                continue
            
            # 平日・法定外休日の労働時間
            daily_work_minutes = (record.regular_working_minutes or 0) + (record.overtime_minutes or 0)
            
            if daily_work_minutes > 0:
                total_work_minutes += daily_work_minutes
                workday_records.append((record, daily_work_minutes))
                print(f"   {record.work_date} ({day_name}): {daily_work_minutes}分")
        
        print(f"   📊 週合計労働時間: {total_work_minutes}分 ({total_work_minutes/60:.1f}時間)")
        
        # 週40時間制限を適用
        WEEKLY_LIMIT_MINUTES = 40 * 60  # 2400分
        
        if total_work_minutes > 0:
            # 法定内労働時間と法定外労働時間を計算
            legal_regular_minutes = min(total_work_minutes, WEEKLY_LIMIT_MINUTES)
            legal_overtime_minutes = max(0, total_work_minutes - WEEKLY_LIMIT_MINUTES)
            
            print(f"   ⚖️  週40時間制限適用:")
            print(f"      法定内労働時間: {legal_regular_minutes}分 ({legal_regular_minutes/60:.1f}時間)")
            print(f"      法定外労働時間: {legal_overtime_minutes}分 ({legal_overtime_minutes/60:.1f}時間)")
            
            # 対象月のレコードのみを更新（クロスマンス考慮）
            target_month_records = [
                (record, daily_minutes) for record, daily_minutes in workday_records 
                if record.work_date.year == year and record.work_date.month == month
            ]
            
            if target_month_records:
                print(f"   🎯 {year}年{month}月のレコード更新: {len(target_month_records)}件")
                
                # 対象月の総労働時間を計算
                target_month_total = sum(daily_minutes for _, daily_minutes in target_month_records)
                print(f"      対象月労働時間: {target_month_total}分 ({target_month_total/60:.1f}時間)")
                
                # 比例配分で法定内・法定外を分配
                if target_month_total > 0:
                    # まず全レコードの労働時間をリセット
                    for record, _ in target_month_records:
                        record.regular_working_minutes = 0
                        record.overtime_minutes = 0
                    
                    # 対象月の労働時間を週の比例で法定内・法定外に分配
                    # 対象月の労働時間のうち、週の比例で法定内・法定外を決定
                    month_regular_ratio = min(1.0, legal_regular_minutes / total_work_minutes) if total_work_minutes > 0 else 0
                    month_overtime_ratio = max(0.0, legal_overtime_minutes / total_work_minutes) if total_work_minutes > 0 else 0
                    
                    print(f"      週比例: 法定内{month_regular_ratio:.3f}, 法定外{month_overtime_ratio:.3f}")
                    
                    # 各日に比例配分
                    remaining_month_regular = int(target_month_total * month_regular_ratio)
                    remaining_month_overtime = target_month_total - remaining_month_regular
                    
                    print(f"      対象月分配: 法定内{remaining_month_regular}分, 法定外{remaining_month_overtime}分")
                    
                    for record, daily_minutes in sorted(target_month_records, key=lambda x: x[0].work_date):
                        if daily_minutes > 0:
                            # この日に割り当てる法定内労働時間
                            daily_regular = min(daily_minutes, remaining_month_regular)
                            remaining_month_regular -= daily_regular
                            
                            # 残りは法定外労働時間
                            daily_overtime = daily_minutes - daily_regular
                            
                            record.regular_working_minutes = daily_regular
                            record.overtime_minutes = daily_overtime
                            
                            day_name = ['月', '火', '水', '木', '金', '土', '日'][record.work_date.weekday()]
                            print(f"      {record.work_date} ({day_name}): "
                                  f"法定内 {daily_regular}分, 法定外 {daily_overtime}分")
                
                # 更新日時を設定
                for record, _ in target_month_records:
                    record.updated_at = datetime.now()
            else:
                print(f"   ⏭️  この週の{year}年{month}月レコードはありません")
    
    # データベースに変更をコミット
    db.session.commit()
    print(f"✅ 週40時間制限計算完了: {year}年{month}月")

def calculate_weekly_overtime_adjustment(employee, year, month, weekly_data, week_start_day):
    """週40時間超過分の時間外労働調整計算（土曜日の労働時間を週40時間基準で再分類）
    
    Args:
        employee: Employee 従業員オブジェクト
        year: int 計算対象年
        month: int 計算対象月
        weekly_data: dict 週別労働時間データ
        week_start_day: int 週起算日（0=月曜日、6=日曜日）
    
    Returns:
        dict: 週単位時間外労働調整データ
    """
    adjustment_data = {
        'weekly_overtime_minutes': 0,
        'adjusted_records': [],
        'saturday_adjustments': []
    }
    
    # 週40時間 = 2400分
    WEEKLY_OVERTIME_THRESHOLD = 2400
    
    for week_start, week_info in weekly_data.items():
        # 該当週の全レコードを取得（クロスマンス対応）
        week_end = week_start + timedelta(days=6)
        week_records = []
        
        # 対象月のレコード
        for record in week_info['records']:
            week_records.append(record)
        
        # クロスマンスの場合は前月・翌月のデータも取得
        if week_start.month != month or week_end.month != month:
            # 前月データ
            if week_start.month != month:
                prev_month = month - 1 if month > 1 else 12
                prev_year = year if month > 1 else year - 1
                
                prev_records = WorkingTimeRecord.query.filter(
                    WorkingTimeRecord.employee_id == employee.id,
                    db.extract('year', WorkingTimeRecord.work_date) == prev_year,
                    db.extract('month', WorkingTimeRecord.work_date) == prev_month,
                    WorkingTimeRecord.work_date >= week_start,
                    WorkingTimeRecord.work_date <= week_end
                ).all()
                week_records.extend(prev_records)
            
            # 翌月データ
            if week_end.month != month:
                next_month = month + 1 if month < 12 else 1
                next_year = year if month < 12 else year + 1
                
                next_records = WorkingTimeRecord.query.filter(
                    WorkingTimeRecord.employee_id == employee.id,
                    db.extract('year', WorkingTimeRecord.work_date) == next_year,
                    db.extract('month', WorkingTimeRecord.work_date) == next_month,
                    WorkingTimeRecord.work_date >= week_start,
                    WorkingTimeRecord.work_date <= week_end
                ).all()
                week_records.extend(next_records)
        
        # 週の総労働時間を計算
        week_total_minutes = 0
        saturday_records = []
        
        for record in week_records:
            daily_total = (
                (record.regular_working_minutes or 0) +
                (record.legal_overtime_minutes or 0) +
                (record.overtime_minutes or 0)
            )
            week_total_minutes += daily_total
            
            # 土曜日のレコードを特定
            if record.work_date.weekday() == 5:  # 土曜日
                saturday_records.append(record)
        
        # 週40時間を超過している場合
        if week_total_minutes > WEEKLY_OVERTIME_THRESHOLD:
            weekly_overtime = week_total_minutes - WEEKLY_OVERTIME_THRESHOLD
            adjustment_data['weekly_overtime_minutes'] += weekly_overtime
            
            # 土曜日の労働時間を週40時間基準で再分類
            for saturday_record in saturday_records:
                if saturday_record.work_date.month == month:  # 対象月のみ調整
                    saturday_total = (
                        (saturday_record.regular_working_minutes or 0) +
                        (saturday_record.legal_overtime_minutes or 0) +
                        (saturday_record.overtime_minutes or 0)
                    )
                    
                    # 週40時間を超過した分のみ時間外労働とする
                    # 土曜日の労働時間を週40時間以内は法定外休日労働、超過分は時間外労働に分類
                    weekly_overtime_portion = min(saturday_total, weekly_overtime)
                    holiday_portion = saturday_total - weekly_overtime_portion
                    
                    # 土曜日の労働時間を再分類
                    saturday_record.regular_working_minutes = 0
                    saturday_record.legal_overtime_minutes = 0
                    saturday_record.overtime_minutes = weekly_overtime_portion
                    saturday_record.holiday_minutes = holiday_portion
                    
                    adjustment_data['saturday_adjustments'].append({
                        'record': saturday_record,
                        'original_total': saturday_total,
                        'holiday_minutes': holiday_portion,
                        'overtime_minutes': weekly_overtime_portion,
                        'week_start': week_start,
                        'week_total': week_total_minutes
                    })
                    
                    # 調整した分を週間時間外労働から差し引く
                    weekly_overtime -= weekly_overtime_portion
                    if weekly_overtime <= 0:
                        break
        
        else:
            # 週40時間以内の場合、土曜日は法定外休日労働として扱う
            for saturday_record in saturday_records:
                if saturday_record.work_date.month == month:  # 対象月のみ調整
                    saturday_total = (
                        (saturday_record.regular_working_minutes or 0) +
                        (saturday_record.legal_overtime_minutes or 0) +
                        (saturday_record.overtime_minutes or 0)
                    )
                    
                    # 全て法定外休日労働に分類
                    saturday_record.regular_working_minutes = 0
                    saturday_record.legal_overtime_minutes = 0
                    saturday_record.overtime_minutes = 0
                    saturday_record.holiday_minutes = saturday_total
                    
                    adjustment_data['saturday_adjustments'].append({
                        'record': saturday_record,
                        'original_total': saturday_total,
                        'holiday_minutes': saturday_total,
                        'overtime_minutes': 0,
                        'week_start': week_start,
                        'week_total': week_total_minutes
                    })
    
    return adjustment_data

def calculate_annual_working_hours(year):
    """年間所定労働時間を計算する"""
    import calendar
    
    # 年間日数
    days_in_year = 366 if calendar.isleap(year) else 365
    
    # 土日の数を計算
    weekends = 0
    for month in range(1, 13):
        for day in range(1, calendar.monthrange(year, month)[1] + 1):
            weekday = calendar.weekday(year, month, day)
            if weekday in [5, 6]:  # 土曜日=5, 日曜日=6
                weekends += 1
    
    # 祝日数（日本の祝日概算）
    holidays = 16  # 基本的な祝日数
    
    # 年間労働日数 = 年間日数 - 土日 - 祝日
    working_days = days_in_year - weekends - holidays
    
    # 年間所定労働時間 = 労働日数 × 8時間
    annual_working_hours = working_days * 8
    
    return annual_working_hours, working_days

def apply_weekly_40hour_rule(records, weekly_data, week_start_day, year, month):
    """週40時間制を正しく適用する関数"""
    from datetime import datetime, timedelta
    
    WEEKLY_LIMIT_MINUTES = 40 * 60  # 2400分 = 40時間
    
    
    for week_start, week_info in weekly_data.items():
        # その週の法定休日以外の労働記録を取得
        workday_records = []
        total_workday_minutes = 0
        
        for record in week_info['records']:
            # 法定休日労働は週40時間計算から除外
            if record.legal_holiday_minutes and record.legal_holiday_minutes > 0:
                continue
            
            # 平日・法定外休日の労働時間を集計
            daily_minutes = record.regular_working_minutes or 0
            if daily_minutes > 0:
                workday_records.append((record, daily_minutes))
                total_workday_minutes += daily_minutes
        
        # 週40時間を超過している場合の再分類
        if total_workday_minutes > WEEKLY_LIMIT_MINUTES:
            overtime_minutes = total_workday_minutes - WEEKLY_LIMIT_MINUTES
            # 週の労働日を時系列で並び替え（月曜日から日曜日順）
            workday_records.sort(key=lambda x: x[0].work_date)
            
            # 週40時間超過分を後半の労働日から時間外労働に分類
            remaining_overtime = overtime_minutes
            
            # 逆順（日曜日から）で超過時間を時間外労働に分類
            for record, daily_minutes in reversed(workday_records):
                if remaining_overtime <= 0:
                    break
                
                # この日の労働時間を時間外労働に分類
                overtime_for_this_day = min(remaining_overtime, daily_minutes)
                regular_for_this_day = daily_minutes - overtime_for_this_day
                
                # 労働時間を再分類
                record.regular_working_minutes = regular_for_this_day
                record.overtime_minutes = overtime_for_this_day
                record.updated_at = datetime.now()
                
                remaining_overtime -= overtime_for_this_day
    
    # データベースに変更をコミット
    from app import db
    db.session.commit()

def calculate_monthly_payroll(employee_id, year, month):
    """月次給与計算処理"""
    employee = Employee.query.get(employee_id)
    if not employee:
        raise ValueError("従業員が見つかりません")
    
    # 給与設定を取得（最新の有効な設定を取得）
    payroll_settings = EmployeePayrollSettings.query.filter(
        EmployeePayrollSettings.employee_id == employee_id,
        EmployeePayrollSettings.effective_from <= date(year, month, 1)
    ).filter(
        db.or_(
            EmployeePayrollSettings.effective_until.is_(None),
            EmployeePayrollSettings.effective_until >= date(year, month, 1)
        )
    ).order_by(EmployeePayrollSettings.effective_from.desc()).first()
    
    if not payroll_settings:
        raise ValueError(f"有効な給与設定が見つかりません。先に給与設定を登録してください。")
    
    # 年間所定労働時間を計算
    annual_working_hours, annual_working_days = calculate_annual_working_hours(year)
    
    # 既存の計算結果と関連する給与明細書を削除
    existing = PayrollCalculation.query.filter(
        PayrollCalculation.employee_id == employee_id,
        PayrollCalculation.year == year,
        PayrollCalculation.month == month
    ).first()
    
    if existing:
        # 関連する給与明細書を先に削除
        existing_slips = PayrollSlip.query.filter(
            PayrollSlip.payroll_calculation_id == existing.id
        ).all()
        for slip in existing_slips:
            db.session.delete(slip)
        
        # 次に給与計算を削除
        db.session.delete(existing)
    
    # その月の勤怠データを取得
    records = WorkingTimeRecord.query.filter(
        WorkingTimeRecord.employee_id == employee_id,
        db.extract('year', WorkingTimeRecord.work_date) == year,
        db.extract('month', WorkingTimeRecord.work_date) == month
    ).all()
    
    # 労働時間集計（日本労働基準法準拠）
    total_working_minutes = 0
    regular_working_minutes = 0
    legal_overtime_minutes = 0
    overtime_minutes = 0
    legal_holiday_minutes = 0
    holiday_minutes = 0
    night_working_minutes = 0
    
    paid_leave_days = 0
    special_leave_days = 0
    absence_days = 0
    company_closure_days = 0
    
    # 法定休日設定から週起算日を取得
    holiday_settings = LegalHolidaySettings.query.first()
    week_start_day = 6  # デフォルトは日曜日（0=月曜日、6=日曜日）
    if holiday_settings and holiday_settings.week_start_day is not None:
        week_start_day = holiday_settings.week_start_day
    
    # 週40時間制チェック用
    weekly_data = {}  # {week_start_date: {'total_minutes': int, 'records': [record, ...]}}
    
    for record in records:
        if record.is_paid_leave:
            paid_leave_days += 1
        elif record.is_special_leave:
            special_leave_days += 1
        elif record.is_absence:
            absence_days += 1
        elif record.is_company_closure:
            company_closure_days += 1
        else:
            # 労働時間集計
            daily_total = (record.regular_working_minutes + record.legal_overtime_minutes + 
                          record.overtime_minutes + record.legal_holiday_minutes + record.holiday_minutes)
            total_working_minutes += daily_total
            regular_working_minutes += record.regular_working_minutes
            legal_overtime_minutes += record.legal_overtime_minutes
            overtime_minutes += record.overtime_minutes
            legal_holiday_minutes += record.legal_holiday_minutes
            holiday_minutes += record.holiday_minutes
            night_working_minutes += record.night_working_minutes
            
            # 週の開始日を計算
            work_date = record.work_date
            days_from_week_start = (work_date.weekday() - week_start_day) % 7
            week_start = work_date - timedelta(days=days_from_week_start)
            
            if week_start not in weekly_data:
                weekly_data[week_start] = {'total_minutes': 0, 'records': []}
            weekly_data[week_start]['total_minutes'] += daily_total
            weekly_data[week_start]['records'].append(record)
    
    # 週40時間制の正しい適用
    apply_weekly_40hour_rule(records, weekly_data, week_start_day, year, month)
    
    # 週40時間ルール適用後の労働時間を再集計
    regular_working_minutes = 0
    legal_overtime_minutes = 0
    overtime_minutes = 0
    legal_holiday_minutes = 0
    holiday_minutes = 0
    night_working_minutes = 0
    
    for record in records:
        regular_working_minutes += record.regular_working_minutes or 0
        legal_overtime_minutes += record.legal_overtime_minutes or 0
        overtime_minutes += record.overtime_minutes or 0
        legal_holiday_minutes += record.legal_holiday_minutes or 0
        holiday_minutes += record.holiday_minutes or 0
        night_working_minutes += record.night_working_minutes or 0
    
    # 給与形態別基本給計算
    wage_type = payroll_settings.wage_type or 'monthly'
    
    if wage_type == 'hourly':
        # 時給制
        hourly_rate = payroll_settings.hourly_rate or 0
        regular_working_pay = int(regular_working_minutes / 60 * hourly_rate)
        legal_overtime_pay = int(legal_overtime_minutes / 60 * hourly_rate)  # 法定内残業は通常賃金
        overtime_pay = int(overtime_minutes / 60 * hourly_rate * 1.25)  # 法定外残業は25%増し
        legal_holiday_pay = int(legal_holiday_minutes / 60 * hourly_rate * 1.35)  # 法定休日は35%増し
        holiday_pay = int(holiday_minutes / 60 * hourly_rate)  # 法定外休日は通常賃金
        base_salary = regular_working_pay
        
    elif wage_type == 'daily':
        # 日給制
        daily_rate = payroll_settings.daily_rate or 0
        working_days = len([r for r in records if r.regular_working_minutes > 0 or r.legal_overtime_minutes > 0 or r.overtime_minutes > 0])
        regular_working_pay = working_days * daily_rate
        
        # 日給を時給換算（1日8時間で計算）
        hourly_rate = daily_rate / 8 if daily_rate > 0 else 0
        legal_overtime_pay = int(legal_overtime_minutes / 60 * hourly_rate)
        overtime_pay = int(overtime_minutes / 60 * hourly_rate * 1.25)
        legal_holiday_pay = int(legal_holiday_minutes / 60 * hourly_rate * 1.35)
        holiday_pay = int(holiday_minutes / 60 * hourly_rate)
        base_salary = regular_working_pay
        
    else:
        # 月給制（デフォルト）
        base_salary = payroll_settings.base_salary or 0
        regular_working_pay = 0  # 月給に含まれる
        
        # 年間所定労働時間から時給を算出
        if base_salary > 0 and annual_working_hours > 0:
            # 月給 × 12ヶ月 ÷ 年間所定労働時間
            annual_salary = base_salary * 12
            hourly_rate = annual_salary / annual_working_hours
            
            # 月給制では所定労働時間超過分のみ法定内残業として支払い
            # legal_overtime_minutesには所定時間超過〜法定時間内の残業が含まれる
            legal_overtime_pay = int(legal_overtime_minutes / 60 * hourly_rate)
            overtime_pay = int(overtime_minutes / 60 * hourly_rate * 1.25)
            legal_holiday_pay = int(legal_holiday_minutes / 60 * hourly_rate * 1.35)
            holiday_pay = int(holiday_minutes / 60 * hourly_rate)
        else:
            legal_overtime_pay = 0
            overtime_pay = 0
            legal_holiday_pay = 0
            holiday_pay = 0
    
    # 深夜労働手当（25%増し）- 労働基準法第37条第4項
    if employee.wage_type == 'hourly':
        night_working_pay = int(night_working_minutes / 60 * base_salary * 0.25)
    elif employee.wage_type == 'daily':
        hourly_rate = base_salary / 8
        night_working_pay = int(night_working_minutes / 60 * hourly_rate * 0.25)
    else:
        monthly_working_hours = employee.standard_working_hours * employee.standard_working_days * 4.33
        hourly_rate = base_salary / monthly_working_hours if monthly_working_hours > 0 else 0
        night_working_pay = int(night_working_minutes / 60 * hourly_rate * 0.25)
    
    # 週40時間超過分は既にovertime_minutesに分類済み（新ロジック）
    weekly_overtime_pay = 0
    
    # 休業補償（60%）- 労働基準法第26条
    closure_compensation = 0
    if company_closure_days > 0:
        if employee.wage_type == 'hourly':
            # 時給制の場合：標準労働時間×時給×60%×日数
            daily_compensation = employee.standard_working_hours * base_salary * 0.6
            closure_compensation = int(company_closure_days * daily_compensation)
        elif employee.wage_type == 'daily':
            # 日給制の場合：日給×60%×日数
            closure_compensation = int(company_closure_days * base_salary * 0.6)
        else:
            # 月給制の場合：月給÷30日×60%×日数
            daily_rate = base_salary / 30
            closure_compensation = int(company_closure_days * daily_rate * 0.6)
    
    # 欠勤控除（ノーワーク・ノーペイの原則）
    absence_deduction = 0
    if absence_days > 0:
        if wage_type == 'hourly':
            # 時給制の場合：8時間×時給×日数
            absence_deduction = int(absence_days * 8 * hourly_rate)
        elif wage_type == 'daily':
            # 日給制の場合：日給×日数
            absence_deduction = int(absence_days * (payroll_settings.daily_rate or 0))
        else:
            # 月給制の場合：年間所定労働日数から月平均労働日数を算出
            monthly_working_days = annual_working_days / 12
            daily_rate = base_salary / monthly_working_days if monthly_working_days > 0 else 0
            absence_deduction = int(absence_days * daily_rate)
    
    # 総支給額計算
    gross_salary = (base_salary + regular_working_pay + legal_overtime_pay + overtime_pay + 
                   legal_holiday_pay + holiday_pay + night_working_pay + weekly_overtime_pay + 
                   closure_compensation - absence_deduction)
    
    # 給与計算結果を保存
    payroll = PayrollCalculation(
        employee_id=employee_id,
        year=year,
        month=month,
        wage_type=wage_type,
        base_salary=base_salary,
        regular_working_minutes=regular_working_minutes,
        legal_overtime_minutes=legal_overtime_minutes,
        overtime_minutes=overtime_minutes,
        legal_holiday_minutes=legal_holiday_minutes,
        holiday_minutes=holiday_minutes,
        night_working_minutes=night_working_minutes,
        paid_leave_days=paid_leave_days,
        special_leave_days=special_leave_days,
        absence_days=absence_days,
        overtime_allowance=overtime_pay + legal_overtime_pay + weekly_overtime_pay,
        night_allowance=night_working_pay,
        holiday_allowance=holiday_pay + legal_holiday_pay,
        gross_salary=gross_salary,
        net_salary=gross_salary,  # 控除計算は後で実装
        calculated_by=current_user.id if current_user and not current_user.is_anonymous else None
    )
    
    db.session.add(payroll)
    db.session.commit()
    
    # 賃金台帳を自動更新
    try:
        from wage_register_manager import WageRegisterManager
        wage_manager = WageRegisterManager()
        
        # 給与計算データから賃金台帳用データを構築
        payroll_data = {
            'base_salary': base_salary,
            'overtime_allowance': payroll.overtime_allowance,
            'holiday_allowance': payroll.holiday_allowance,
            'night_allowance': payroll.night_allowance,
            'position_allowance': payroll_settings.position_allowance or 0,
            'transportation_allowance': payroll_settings.transportation_allowance or 0,
            'housing_allowance': payroll_settings.housing_allowance or 0,
            'family_allowance': payroll_settings.family_allowance or 0,
            'other_allowances': payroll.other_allowances,
            'health_insurance': payroll.health_insurance,
            'pension_insurance': payroll.pension_insurance,
            'employment_insurance': payroll.employment_insurance,
            'income_tax': payroll.income_tax,
            'resident_tax': payroll.resident_tax,
            'other_deductions': payroll.other_deductions,
            'gross_salary': payroll.gross_salary,
            'total_deductions': payroll.total_deductions,
            'net_salary': payroll.net_salary,
            'working_days': len([r for r in records if r.regular_working_minutes > 0 or r.overtime_minutes > 0]),
            'overtime_hours': (overtime_minutes + legal_overtime_minutes) / 60.0,
            'paid_leave_days': paid_leave_days,
            'absence_days': absence_days
        }
        
        wage_manager.update_wage_register(employee_id, year, month, payroll_data)
        print(f"✅ 賃金台帳を更新しました: 従業員{employee_id}, {year}年{month}月")
        
    except Exception as e:
        print(f"⚠️ 賃金台帳更新エラー: {e}")
        # エラーが発生しても給与計算処理は続行
    
    return payroll

@app.route('/payroll_dashboard', methods=['GET', 'POST'])
@login_required
def payroll_dashboard():
    """給与計算ダッシュボード"""
    if current_user.role != 'accounting':
        flash('アクセス権限がありません。')
        return redirect(url_for('index'))
    
    employees = Employee.query.filter_by(status='在籍中').all()
    current_year = datetime.now().year
    current_month = datetime.now().month
    years = list(range(current_year - 2, current_year + 2))
    months = list(range(1, 13))
    
    selected_employee = None
    selected_year = current_year
    selected_month = current_month
    payroll_data = None
    
    if request.method == 'POST':
        employee_id = request.form.get('employee_id')
        selected_year = int(request.form.get('year', current_year))
        selected_month = int(request.form.get('month', current_month))
        
        if employee_id:
            selected_employee = Employee.query.get(employee_id)
            
            # 既存の給与計算結果を取得
            existing_calculation = PayrollCalculation.query.filter(
                PayrollCalculation.employee_id == employee_id,
                PayrollCalculation.year == selected_year,
                PayrollCalculation.month == selected_month
            ).first()
            
            # 従業員の給与設定を取得（最新の有効な設定を取得）
            target_date = date(selected_year, selected_month, 1)
            print(f"[DEBUG] 給与設定を検索: 従業員ID={employee_id}, 対象日={target_date}")
            
            payroll_settings = EmployeePayrollSettings.query.filter(
                EmployeePayrollSettings.employee_id == employee_id,
                EmployeePayrollSettings.effective_from <= target_date
            ).filter(
                db.or_(
                    EmployeePayrollSettings.effective_until.is_(None),
                    EmployeePayrollSettings.effective_until >= target_date
                )
            ).order_by(EmployeePayrollSettings.effective_from.desc()).first()
            
            if payroll_settings:
                print(f"[DEBUG] 給与設定発見: ID={payroll_settings.id}, 基本給={payroll_settings.base_salary}, 適用期間={payroll_settings.effective_from}〜{payroll_settings.effective_until}")
            else:
                print(f"[DEBUG] 給与設定が見つかりません")
                # 全設定を確認
                all_settings = EmployeePayrollSettings.query.filter(
                    EmployeePayrollSettings.employee_id == employee_id
                ).all()
                print(f"[DEBUG] 該当従業員の全設定数: {len(all_settings)}")
                for setting in all_settings:
                    print(f"[DEBUG] 設定: ID={setting.id}, 適用期間={setting.effective_from}〜{setting.effective_until}, 基本給={setting.base_salary}")
            
            # 勤怠データを取得
            working_records = WorkingTimeRecord.query.filter(
                WorkingTimeRecord.employee_id == employee_id,
                db.extract('year', WorkingTimeRecord.work_date) == selected_year,
                db.extract('month', WorkingTimeRecord.work_date) == selected_month
            ).all()
            
            # 勤怠データの合計を事前に計算
            total_regular = 0
            total_legal_overtime = 0
            total_overtime = 0
            total_legal_holiday = 0
            total_holiday = 0
            total_night = 0
            
            for record in working_records:
                total_regular += record.regular_working_minutes or 0
                total_legal_overtime += record.legal_overtime_minutes or 0
                total_overtime += record.overtime_minutes or 0
                total_legal_holiday += record.legal_holiday_minutes or 0
                total_holiday += record.holiday_minutes or 0
                total_night += record.night_working_minutes or 0
            
            # 給与設定がNoneの場合の対処
            if not payroll_settings:
                # デフォルトの給与設定を作成（表示のみ）
                class DefaultPayrollSettings:
                    def __init__(self):
                        self.wage_type = 'monthly'
                        self.base_salary = 0
                        self.transportation_allowance = 0
                        self.position_allowance = 0
                
                payroll_settings = DefaultPayrollSettings()
            
            payroll_data = {
                'employee': selected_employee,
                'calculation': existing_calculation,
                'settings': payroll_settings,
                'records': working_records,
                'year': selected_year,
                'month': selected_month,
                'total_regular': total_regular,
                'total_legal_overtime': total_legal_overtime,
                'total_overtime': total_overtime,
                'total_legal_holiday': total_legal_holiday,
                'total_holiday': total_holiday,
                'total_night': total_night
            }
    
    return render_template('payroll_dashboard.html',
                         employees=employees,
                         years=years,
                         months=months,
                         selected_employee=selected_employee,
                         selected_year=selected_year,
                         selected_month=selected_month,
                         payroll_data=payroll_data)

@app.route('/api/calculate_payroll', methods=['POST'])
@login_required
def api_calculate_payroll():
    """給与計算API"""
    if current_user.role != 'accounting':
        return jsonify({'success': False, 'error': 'アクセス権限がありません'})
    
    try:
        data = request.get_json()
        employee_id = data.get('employee_id')
        year = data.get('year')
        month = data.get('month')
        
        if not all([employee_id, year, month]):
            return jsonify({'success': False, 'error': '必要な情報が不足しています'})
        
        # 給与計算実行
        calculate_monthly_payroll(employee_id, year, month)
        
        return jsonify({'success': True, 'message': '給与計算が完了しました'})
        
    except ValueError as e:
        error_msg = str(e)
        if "給与設定が見つかりません" in error_msg:
            return jsonify({
                'success': False, 
                'error': error_msg,
                'redirect': url_for('employee_payroll_settings', employee_id=employee_id)
            })
        return jsonify({'success': False, 'error': error_msg})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/employee_payroll_settings/<int:employee_id>', methods=['GET', 'POST'])
@login_required
def employee_payroll_settings(employee_id):
    """従業員給与設定"""
    if current_user.role not in ['accounting', 'admin']:
        flash('アクセス権限がありません。')
        return redirect(url_for('index'))
    
    employee = Employee.query.get_or_404(employee_id)
    
    # 現在の給与設定を取得
    print(f"[DEBUG] 給与設定画面: 従業員ID={employee_id}, 今日={date.today()}")
    current_settings = EmployeePayrollSettings.query.filter(
        EmployeePayrollSettings.employee_id == employee_id,
        EmployeePayrollSettings.effective_from <= date.today()
    ).filter(
        db.or_(
            EmployeePayrollSettings.effective_until.is_(None),
            EmployeePayrollSettings.effective_until >= date.today()
        )
    ).first()
    
    if current_settings:
        print(f"[DEBUG] 現在の給与設定: ID={current_settings.id}, 基本給={current_settings.base_salary}, 適用期間={current_settings.effective_from}〜{current_settings.effective_until}")
    else:
        print(f"[DEBUG] 現在の給与設定がありません")
    
    if request.method == 'POST':
        try:
            # 既存設定の終了日を昨日に設定（新しい設定との重複を避けるため）
            if current_settings:
                from datetime import timedelta
                current_settings.effective_until = date.today() - timedelta(days=1)
                print(f"[DEBUG] 既存設定の終了日を設定: {current_settings.effective_until}")
            
            # 新しい設定を作成
            new_settings = EmployeePayrollSettings(
                employee_id=employee_id,
                wage_type=request.form.get('wage_type', 'monthly'),
                base_salary=int(request.form.get('base_salary', 0)),
                hourly_rate=int(request.form.get('hourly_rate', 0)),
                daily_rate=int(request.form.get('daily_rate', 0)),
                position_allowance=int(request.form.get('position_allowance', 0)),
                family_allowance=int(request.form.get('family_allowance', 0)),
                transportation_allowance=int(request.form.get('transportation_allowance', 0)),
                housing_allowance=int(request.form.get('housing_allowance', 0)),
                meal_allowance=int(request.form.get('meal_allowance', 0)),
                skill_allowance=int(request.form.get('skill_allowance', 0)),
                health_insurance_rate=float(request.form.get('health_insurance_rate', 4.95)),
                pension_insurance_rate=float(request.form.get('pension_insurance_rate', 9.15)),
                employment_insurance_rate=float(request.form.get('employment_insurance_rate', 0.3)),
                long_term_care_insurance_rate=float(request.form.get('long_term_care_insurance_rate', 0.58)),
                income_tax_type=request.form.get('income_tax_type', 'automatic'),
                resident_tax=int(request.form.get('resident_tax', 0)),
                union_fee=int(request.form.get('union_fee', 0)),
                parking_fee=int(request.form.get('parking_fee', 0)),
                uniform_fee=int(request.form.get('uniform_fee', 0)),
                effective_from=date.today()
            )
            
            print(f"[DEBUG] 新しい設定を作成: 従業員ID={employee_id}, 基本給={new_settings.base_salary}, 適用開始日={new_settings.effective_from}")
            
            db.session.add(new_settings)
            db.session.commit()
            
            print(f"[DEBUG] 給与設定保存完了: ID={new_settings.id}")
            flash('給与設定を保存しました。')
            return redirect(url_for('payroll_dashboard'))
            
        except Exception as e:
            db.session.rollback()
            print(f"[ERROR] 給与設定保存エラー: {e}")
            import traceback
            traceback.print_exc()
            flash(f'給与設定の保存でエラーが発生しました: {str(e)}')
    
    return render_template('employee_payroll_settings.html',
                         employee=employee,
                         settings=current_settings)

@app.route('/create_payroll_slip/<int:employee_id>/<int:year>/<int:month>', methods=['GET', 'POST'])
@login_required
def create_payroll_slip(employee_id, year, month):
    """給与明細書作成"""
    if current_user.role != 'accounting':
        flash('アクセス権限がありません。')
        return redirect(url_for('index'))
    
    employee = Employee.query.get_or_404(employee_id)
    
    # 給与計算結果を取得
    payroll_calculation = PayrollCalculation.query.filter(
        PayrollCalculation.employee_id == employee_id,
        PayrollCalculation.year == year,
        PayrollCalculation.month == month
    ).first()
    
    if not payroll_calculation:
        flash('先に給与計算を実行してください。')
        return redirect(url_for('payroll_dashboard'))
    
    # 給与設定を取得（最新の有効な設定を取得）
    payroll_settings = EmployeePayrollSettings.query.filter(
        EmployeePayrollSettings.employee_id == employee_id,
        EmployeePayrollSettings.effective_from <= date(year, month, 1)
    ).filter(
        db.or_(
            EmployeePayrollSettings.effective_until.is_(None),
            EmployeePayrollSettings.effective_until >= date(year, month, 1)
        )
    ).order_by(EmployeePayrollSettings.effective_from.desc()).first()
    
    # 既存の給与明細書を取得
    existing_slip = PayrollSlip.query.filter(
        PayrollSlip.employee_id == employee_id,
        PayrollSlip.slip_year == year,
        PayrollSlip.slip_month == month
    ).first()
    
    if request.method == 'POST':
        try:
            # 社会保険料計算
            gross_salary = payroll_calculation.gross_salary
            health_insurance = int(gross_salary * (payroll_settings.health_insurance_rate / 100)) if payroll_settings else 0
            pension_insurance = int(gross_salary * (payroll_settings.pension_insurance_rate / 100)) if payroll_settings else 0
            employment_insurance = int(gross_salary * (payroll_settings.employment_insurance_rate / 100)) if payroll_settings else 0
            
            # 年齢に応じて介護保険料を計算
            long_term_care_insurance = 0
            if employee.birth_date:
                age = date.today().year - employee.birth_date.year
                if age >= 40:
                    long_term_care_insurance = int(gross_salary * (payroll_settings.long_term_care_insurance_rate / 100)) if payroll_settings else 0
            
            # 所得税計算
            if payroll_settings and payroll_settings.income_tax_type == 'automatic':
                # 簡易所得税計算（総支給額ベース）
                if gross_salary <= 88000:
                    income_tax = 0
                elif gross_salary <= 162500:
                    income_tax = int((gross_salary - 88000) * 0.05)
                elif gross_salary <= 275000:
                    income_tax = int(3725 + (gross_salary - 162500) * 0.10)
                elif gross_salary <= 383333:
                    income_tax = int(15000 + (gross_salary - 275000) * 0.20)
                else:
                    income_tax = int(36666 + (gross_salary - 383333) * 0.23)
            else:
                # フォームから手動入力
                income_tax = int(request.form.get('income_tax', 0))
            
            if existing_slip:
                slip = existing_slip
                # 既存のスリップでもpayroll_calculation_idを確実に設定
                slip.payroll_calculation_id = payroll_calculation.id
            else:
                slip = PayrollSlip(
                    employee_id=employee_id,
                    payroll_calculation_id=payroll_calculation.id,
                    slip_year=year,
                    slip_month=month
                )
                db.session.add(slip)
            
            # 基本給与項目
            slip.base_salary = payroll_calculation.base_salary
            slip.overtime_allowance = payroll_calculation.overtime_allowance or 0
            slip.holiday_allowance = payroll_calculation.holiday_allowance or 0
            slip.night_allowance = payroll_calculation.night_allowance or 0
            
            # 各種手当
            if payroll_settings:
                slip.position_allowance = payroll_settings.position_allowance or 0
                slip.family_allowance = payroll_settings.family_allowance or 0
                slip.transportation_allowance = payroll_settings.transportation_allowance or 0
                slip.housing_allowance = payroll_settings.housing_allowance or 0
                slip.meal_allowance = payroll_settings.meal_allowance or 0
                slip.skill_allowance = payroll_settings.skill_allowance or 0
            else:
                slip.position_allowance = 0
                slip.family_allowance = 0
                slip.transportation_allowance = 0
                slip.housing_allowance = 0
                slip.meal_allowance = 0
                slip.skill_allowance = 0
            
            # 新しい支給項目3つ
            temporary_closure_compensation = int(request.form.get('temporary_closure_compensation', 0))
            salary_payment = int(request.form.get('salary_payment', 0))
            bonus_payment = int(request.form.get('bonus_payment', 0))
            
            # 新しい支給項目をデータベースに保存
            slip.temporary_closure_compensation = temporary_closure_compensation
            slip.salary_payment = salary_payment
            slip.bonus_payment = bonus_payment
            
            # その他手当5つの合計
            other_allowances = []
            for i in range(1, 6):
                amount = int(request.form.get(f'other_allowance_{i}', 0))
                name = request.form.get(f'other_allowance_{i}_name', '')
                if amount > 0 or name:
                    other_allowances.append({'name': name, 'amount': amount})
            
            slip.other_allowance = sum(item['amount'] for item in other_allowances)
            # 手当名と金額の情報をJSONで保存
            import json
            slip.other_allowances_json = json.dumps(other_allowances) if other_allowances else None
            # PDF生成用の一時的な属性
            slip.other_allowances_detail = other_allowances
            
            # 総支給額
            slip.gross_salary = (slip.base_salary + slip.overtime_allowance + slip.holiday_allowance + 
                               slip.night_allowance + slip.position_allowance + slip.family_allowance + 
                               slip.transportation_allowance + slip.housing_allowance + slip.meal_allowance + 
                               slip.skill_allowance + temporary_closure_compensation + salary_payment + 
                               bonus_payment + slip.other_allowance)
            
            # 法定控除
            slip.health_insurance = health_insurance
            slip.pension_insurance = pension_insurance
            slip.employment_insurance = employment_insurance
            slip.long_term_care_insurance = long_term_care_insurance
            slip.income_tax = income_tax
            slip.resident_tax = int(request.form.get('resident_tax', payroll_settings.resident_tax if payroll_settings else 0))
            
            # 法定外控除
            if payroll_settings:
                slip.union_fee = payroll_settings.union_fee or 0
                slip.parking_fee = payroll_settings.parking_fee or 0
                slip.uniform_fee = payroll_settings.uniform_fee or 0
            else:
                slip.union_fee = 0
                slip.parking_fee = 0
                slip.uniform_fee = 0
            # その他控除2つの合計
            other_deductions = []
            for i in range(1, 3):
                amount = int(request.form.get(f'other_deduction_{i}', 0))
                name = request.form.get(f'other_deduction_{i}_name', '')
                if amount > 0 or name:
                    other_deductions.append({'name': name, 'amount': amount})
            
            slip.other_deduction = sum(item['amount'] for item in other_deductions)
            # 控除名と金額の情報をJSONで保存
            slip.other_deductions_json = json.dumps(other_deductions) if other_deductions else None
            # PDF生成用の一時的な属性
            slip.other_deductions_detail = other_deductions
            
            # 総控除額・手取額
            slip.total_deduction = (slip.health_insurance + slip.pension_insurance + slip.employment_insurance +
                                  slip.long_term_care_insurance + slip.income_tax + slip.resident_tax +
                                  slip.union_fee + slip.parking_fee + slip.uniform_fee + slip.other_deduction)
            slip.net_salary = slip.gross_salary - slip.total_deduction

            # ===== データ整合性検証 =====
            # 基本給与の整合性チェック
            calc_overtime_allowance = payroll_calculation.overtime_allowance or 0
            if abs(slip.overtime_allowance - calc_overtime_allowance) > 0:
                print(f"[WARNING] 時間外手当の差異: 明細書={slip.overtime_allowance}, 計算結果={calc_overtime_allowance}")

            calc_night_allowance = payroll_calculation.night_allowance or 0
            if abs(slip.night_allowance - calc_night_allowance) > 0:
                print(f"[WARNING] 深夜手当の差異: 明細書={slip.night_allowance}, 計算結果={calc_night_allowance}")

            # 基本給が変更されていた場合のログ
            if slip.base_salary != payroll_calculation.base_salary:
                print(f"[INFO] 基本給修正: 計算結果={payroll_calculation.base_salary} → 明細書={slip.base_salary}")

            # 総支給額の再計算検証
            calc_gross_salary = (slip.base_salary + slip.overtime_allowance + slip.holiday_allowance +
                               slip.night_allowance + slip.position_allowance + slip.family_allowance +
                               slip.transportation_allowance + slip.housing_allowance + slip.meal_allowance +
                               slip.skill_allowance + temporary_closure_compensation + salary_payment +
                               bonus_payment + slip.other_allowance)
            if slip.gross_salary != calc_gross_salary:
                print(f"[ERROR] 総支給額計算不整合: 保存値={slip.gross_salary}, 再計算値={calc_gross_salary}")
                slip.gross_salary = calc_gross_salary  # 正しい値で上書き
                slip.net_salary = slip.gross_salary - slip.total_deduction  # 手取額も再計算
            
            # 勤怠情報
            slip.working_days = len([r for r in payroll_calculation.employee.working_time_records 
                                   if r.work_date.year == year and r.work_date.month == month 
                                   and (r.regular_working_minutes or 0) > 0])
            slip.absence_days = payroll_calculation.absence_days or 0
            slip.paid_leave_days = payroll_calculation.paid_leave_days or 0
            slip.overtime_hours = (payroll_calculation.overtime_minutes or 0) / 60.0
            
            slip.remarks = request.form.get('remarks', '')
            slip.created_by = current_user.id if current_user and not current_user.is_anonymous else None
            slip.issued_at = datetime.now()
            
            db.session.commit()
            
            # PDF生成とダウンロード
            try:
                pdf_buffer = create_payroll_slip_pdf(slip, employee, payroll_calculation, payroll_settings)
                pdf_data = pdf_buffer.read()
                
                response = make_response(pdf_data)
                response.headers['Content-Type'] = 'application/pdf'
                filename = f"payroll_slip_emp{employee.id}_{year}_{month:02d}.pdf"
                response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
                
                flash('給与明細書を作成しました。PDFファイルをダウンロードしてください。')
                return response
                
            except Exception as pdf_e:
                flash(f'給与明細書は保存されましたが、PDFの生成でエラーが発生しました: {str(pdf_e)}')
                return redirect(url_for('payroll_dashboard'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'給与明細書の作成でエラーが発生しました: {str(e)}')
    
    return render_template('create_payroll_slip.html',
                         employee=employee,
                         year=year,
                         month=month,
                         payroll_calculation=payroll_calculation,
                         payroll_settings=payroll_settings,
                         existing_slip=existing_slip)

@app.route('/payroll_results', methods=['GET'])
@login_required
def payroll_results():
    if current_user.role != 'accounting':
        flash('アクセス権限がありません。')
        return redirect(url_for('index'))
    
    employees = Employee.query.filter_by(status='在籍中').all()
    current_year = datetime.now().year
    years = list(range(current_year - 2, current_year + 2))
    
    # フィルター取得
    selected_employee_id = request.args.get('employee_id')
    selected_year = request.args.get('year')
    selected_month = request.args.get('month')
    
    # 基本クエリ
    query = PayrollCalculation.query.join(Employee)
    
    # フィルター適用
    if selected_employee_id:
        query = query.filter(PayrollCalculation.employee_id == selected_employee_id)
    if selected_year:
        query = query.filter(PayrollCalculation.year == int(selected_year))
    if selected_month:
        query = query.filter(PayrollCalculation.month == int(selected_month))
    
    # 結果取得（新しい順）
    payroll_results = query.order_by(
        PayrollCalculation.year.desc(),
        PayrollCalculation.month.desc(),
        PayrollCalculation.calculated_at.desc()
    ).all()
    
    # 詳細表示用（単一の結果が選択された場合）
    selected_result = None
    if selected_employee_id and selected_year and selected_month:
        selected_result = PayrollCalculation.query.filter(
            PayrollCalculation.employee_id == selected_employee_id,
            PayrollCalculation.year == int(selected_year),
            PayrollCalculation.month == int(selected_month)
        ).first()
    
    return render_template('payroll_results.html',
                         employees=employees,
                         years=years,
                         payroll_results=payroll_results,
                         selected_result=selected_result)

@app.route('/bulk_issue_payroll_slips', methods=['POST'])
@login_required
def bulk_issue_payroll_slips():
    """一括給与明細書発行"""
    if current_user.role != 'accounting':
        flash('アクセス権限がありません。')
        return redirect(url_for('index'))
    
    try:
        year = int(request.form.get('year'))
        month = int(request.form.get('month'))
        employee_scope = request.form.get('employee_scope')
        
        # 対象の従業員を取得
        if employee_scope == 'selected':
            employee_ids = request.form.getlist('employee_ids')
            if not employee_ids:
                flash('従業員を選択してください。')
                return redirect(url_for('payroll_results'))
            employee_ids = [int(emp_id) for emp_id in employee_ids]
        else:
            # 全従業員の場合、保存済み明細データがある従業員のIDを取得
            payroll_slips = PayrollSlip.query.filter(
                PayrollSlip.slip_year == year,
                PayrollSlip.slip_month == month
            ).all()
            employee_ids = [slip.employee_id for slip in payroll_slips]
        
        if not employee_ids:
            flash(f'{year}年{month}月の保存済み給与明細データが見つかりませんでした。')
            return redirect(url_for('payroll_results'))
        
        # 複数のPDFを生成してZIPファイルで返す
        import zipfile
        from io import BytesIO
        import os
        
        # メモリ上でZIPファイルを作成
        zip_buffer = BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            generated_count = 0
            
            for employee_id in employee_ids:
                # 保存された給与明細データを取得
                payroll_slip = PayrollSlip.query.filter(
                    PayrollSlip.employee_id == employee_id,
                    PayrollSlip.slip_year == year,
                    PayrollSlip.slip_month == month
                ).first()
                
                if not payroll_slip:
                    continue
                
                # 従業員情報を取得
                employee = Employee.query.get(employee_id)
                if not employee:
                    continue
                
                # 給与計算結果を取得
                payroll_calculation = PayrollCalculation.query.get(payroll_slip.payroll_calculation_id)
                if not payroll_calculation:
                    continue
                
                # 給与設定を取得
                payroll_settings = EmployeePayrollSettings.query.filter(
                    EmployeePayrollSettings.employee_id == employee_id,
                    EmployeePayrollSettings.effective_from <= date(year, month, 1)
                ).filter(
                    db.or_(
                        EmployeePayrollSettings.effective_until.is_(None),
                        EmployeePayrollSettings.effective_until >= date(year, month, 1)
                    )
                ).first()
                
                # その他手当の詳細を復元
                if payroll_slip.other_allowances_json:
                    import json
                    payroll_slip.other_allowances_detail = json.loads(payroll_slip.other_allowances_json)
                else:
                    payroll_slip.other_allowances_detail = []
                
                # その他控除の詳細を復元
                if payroll_slip.other_deductions_json:
                    import json
                    payroll_slip.other_deductions_detail = json.loads(payroll_slip.other_deductions_json)
                else:
                    payroll_slip.other_deductions_detail = []
                
                try:
                    # PDFを生成
                    pdf_buffer = create_payroll_slip_pdf(payroll_slip, employee, payroll_calculation, payroll_settings)
                    pdf_data = pdf_buffer.read()
                    
                    # ZIPファイルにPDFを追加
                    filename = f"{year}年{month}月_{employee.name}_給与明細書.pdf"
                    zip_file.writestr(filename, pdf_data)
                    generated_count += 1
                    
                except Exception as e:
                    print(f"PDF生成エラー（従業員ID: {employee_id}）: {e}")
                    continue
        
        if generated_count == 0:
            flash('給与明細PDFを生成できませんでした。')
            return redirect(url_for('payroll_results'))
        
        zip_buffer.seek(0)
        
        # ZIPファイルとしてダウンロード
        response = make_response(zip_buffer.getvalue())
        response.headers['Content-Type'] = 'application/zip'
        # 日本語文字を含むファイル名をUTF-8でエンコード
        filename = f"{year}年{month}月_給与明細書一括_{generated_count}名.zip"
        encoded_filename = urllib.parse.quote(filename.encode('utf-8'))
        response.headers['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"
        
        flash(f'{generated_count}名の給与明細PDFを一括生成しました。')
        return response
        
    except Exception as e:
        flash(f'一括発行でエラーが発生しました: {str(e)}')
        return redirect(url_for('payroll_results'))

# --- 賃金台帳管理 ---
@app.route('/wage_ledger')
@login_required
def wage_ledger():
    """賃金台帳管理画面"""
    if current_user.role != 'accounting':
        flash('アクセス権限がありません。')
        return redirect(url_for('index'))
    
    employees = Employee.query.filter_by(status='在籍中').order_by(Employee.name).all()
    current_year = datetime.now().year
    years = list(range(current_year - 2, current_year + 2))
    
    return render_template('wage_ledger.html', 
                         employees=employees,
                         years=years)

@app.route('/create_wage_ledger_pdf', methods=['POST'])
@login_required
def create_wage_ledger_pdf():
    """賃金台帳PDF生成"""
    if current_user.role != 'accounting':
        flash('アクセス権限がありません。')
        return redirect(url_for('index'))
    
    try:
        employee_id = request.form.get('employee_id')
        year = request.form.get('year')
        
        if not employee_id or not year:
            flash('従業員と年度を選択してください。', 'error')
            return redirect(url_for('wage_ledger'))
        
        employee = Employee.query.get_or_404(employee_id)
        year = int(year)
        
        # 給与明細書データから賃金台帳データを生成
        wage_data = _generate_wage_ledger_from_payroll_slips(employee_id, year)
        
        if not wage_data:
            flash(f'{employee.name}の{year}年度給与明細データが見つかりません。先に給与明細を作成してください。', 'error')
            return redirect(url_for('wage_ledger'))
        
        # 従業員データ準備
        employee_data = {
            'id': employee.id,
            'name': employee.name,
            'employee_number': f'EMP{employee.id:03d}'  # IDベースで従業員番号を生成
        }
        
        # PDF生成
        from wage_ledger_pdf_generator import WageLedgerPDFGenerator
        generator = WageLedgerPDFGenerator()
        
        # 一時ファイルパス作成
        import tempfile
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp_file:
            temp_path = tmp_file.name
        
        # PDF生成
        success = generator.generate_wage_ledger_pdf(employee_data, wage_data, year, temp_path)
        
        if not success:
            flash('賃金台帳PDFの生成に失敗しました。', 'error')
            return redirect(url_for('wage_ledger'))
        
        # PDFファイルを読み込んでレスポンス作成
        with open(temp_path, 'rb') as pdf_file:
            pdf_data = pdf_file.read()
        
        # 一時ファイル削除
        os.unlink(temp_path)
        
        # レスポンス作成
        response = make_response(pdf_data)
        response.headers['Content-Type'] = 'application/pdf'
        encoded_filename = urllib.parse.quote(f'{year}年度_賃金台帳_{employee.name}.pdf', safe='')
        response.headers['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_filename}'
        
        flash(f'{employee.name}の{year}年度賃金台帳PDFを作成しました。', 'success')
        return response
        
    except Exception as e:
        flash(f'賃金台帳PDF作成中にエラーが発生しました: {str(e)}', 'error')
        return redirect(url_for('wage_ledger'))

def _generate_wage_ledger_from_payroll_slips(employee_id: int, year: int) -> dict:
    """給与明細書データから賃金台帳データを生成する"""
    import json
    
    # 該当年の全ての給与明細データを取得
    payroll_slips = PayrollSlip.query.filter(
        PayrollSlip.employee_id == employee_id,
        PayrollSlip.slip_year == year
    ).order_by(PayrollSlip.slip_month).all()
    
    if not payroll_slips:
        return None
    
    # 月次データ辞書を初期化
    monthly_data = {}
    annual_data = {}
    
    # 各項目の月次データを初期化
    fields = [
        'base_salary', 'overtime_allowance', 'holiday_allowance', 'night_allowance',
        'position_allowance', 'family_allowance', 'transportation_allowance', 
        'housing_allowance', 'meal_allowance', 'skill_allowance',
        'temporary_closure_compensation', 'salary_payment', 'bonus_payment',
        'other_allowance', 'gross_salary',
        'health_insurance', 'pension_insurance', 'employment_insurance', 
        'long_term_care_insurance', 'income_tax', 'resident_tax',
        'union_fee', 'parking_fee', 'uniform_fee', 'other_deduction',
        'total_deduction', 'net_salary',
        'working_days', 'absence_days', 'paid_leave_days', 'overtime_hours'
    ]
    
    for field in fields:
        monthly_data[f'monthly_{field}'] = {}
        annual_data[f'annual_{field}'] = 0
    
    # 手当項目の詳細データ用（5項目に対応）
    for i in range(1, 6):
        monthly_data[f'monthly_allowance{i}'] = {}
        annual_data[f'annual_allowance{i}'] = 0
    
    # 控除項目の詳細データ用（2項目に対応）
    for i in range(1, 3):
        monthly_data[f'monthly_other_deduction{i}'] = {}
        annual_data[f'annual_other_deduction{i}'] = 0
    
    # 各月のデータを集計
    for slip in payroll_slips:
        month_str = str(slip.slip_month)
        
        # 基本項目のデータを設定
        for field in fields:
            value = getattr(slip, field, 0) or 0
            if field == 'overtime_hours':
                # 時間データは時間単位で保存
                monthly_data[f'monthly_{field}'][month_str] = float(value)
            else:
                monthly_data[f'monthly_{field}'][month_str] = int(value)
            annual_data[f'annual_{field}'] += (float(value) if field == 'overtime_hours' else int(value))
        
        # その他手当の詳細データを展開（5項目）
        if slip.other_allowances_json:
            try:
                allowances = json.loads(slip.other_allowances_json)
                for i, allowance in enumerate(allowances[:5], 1):  # 最大5項目
                    amount = allowance.get('amount', 0)
                    monthly_data[f'monthly_allowance{i}'][month_str] = int(amount)
                    annual_data[f'annual_allowance{i}'] += int(amount)
            except:
                pass
        
        # その他控除の詳細データを展開（2項目）
        if slip.other_deductions_json:
            try:
                deductions = json.loads(slip.other_deductions_json)
                for i, deduction in enumerate(deductions[:2], 1):  # 最大2項目
                    amount = deduction.get('amount', 0)
                    monthly_data[f'monthly_other_deduction{i}'][month_str] = int(amount)
                    annual_data[f'annual_other_deduction{i}'] += int(amount)
            except:
                pass
    
    # JSON文字列に変換
    wage_data = {}
    for key, value in monthly_data.items():
        wage_data[key] = json.dumps(value)
    
    for key, value in annual_data.items():
        wage_data[key] = value
    
    return wage_data

# --- 各種届出管理 ---
@app.route('/company_submissions')
@login_required
def company_submissions():
    if current_user.role not in ['admin', 'general_affairs', 'hr_affairs']:
        flash('アクセス権がありません。')
        return redirect(url_for('dashboard'))
    
    return render_template('company_submissions.html')

# --- 企業情報管理 ---
@app.route('/company_info')
@login_required
def company_info():
    if current_user.role not in ['admin', 'general_affairs']:
        flash('アクセス権がありません。')
        return redirect(url_for('dashboard'))
    
    company_settings = CompanySettings.query.first()
    
    return render_template('company_info.html', company_settings=company_settings)

@app.route('/update_company_info', methods=['POST'])
@login_required
def update_company_info():
    if current_user.role not in ['admin', 'general_affairs']:
        flash('アクセス権がありません。')
        return redirect(url_for('dashboard'))
    
    try:
        company_settings = CompanySettings.query.first()
        
        if not company_settings:
            company_settings = CompanySettings()
            db.session.add(company_settings)
        
        company_settings.company_name = request.form.get('company_name')
        company_settings.company_code = request.form.get('company_code')
        company_settings.address = request.form.get('address')
        company_settings.phone = request.form.get('phone')
        company_settings.fax = request.form.get('fax')
        company_settings.email = request.form.get('email')
        company_settings.representative = request.form.get('representative')
        
        establishment_date = request.form.get('establishment_date')
        if establishment_date:
            company_settings.establishment_date = datetime.strptime(establishment_date, '%Y-%m-%d').date()
        
        capital = request.form.get('capital')
        if capital:
            company_settings.capital = int(capital)
        
        employee_count = request.form.get('employee_count')
        if employee_count:
            company_settings.employee_count = int(employee_count)
        
        company_settings.business_type = request.form.get('business_type')
        
        # 会計年度設定の更新
        fiscal_year_start_month = request.form.get('fiscal_year_start_month')
        if fiscal_year_start_month:
            company_settings.fiscal_year_start_month = int(fiscal_year_start_month)
        
        fiscal_year_start_day = request.form.get('fiscal_year_start_day')
        if fiscal_year_start_day:
            company_settings.fiscal_year_start_day = int(fiscal_year_start_day)
        
        company_settings.updated_at = datetime.now()
        
        db.session.commit()
        flash('企業情報を更新しました。', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash('企業情報の更新中にエラーが発生しました。', 'error')
    
    return redirect(url_for('company_info'))

# --- 年度繰越処理 ---
@app.route('/accounting_period_management')
@login_required
def accounting_period_management():
    """会計年度管理画面"""
    if current_user.role != 'accounting':
        flash('アクセス権限がありません。')
        return redirect(url_for('index'))
    
    # 会計年度一覧取得
    periods = AccountingPeriod.query.order_by(AccountingPeriod.fiscal_year.desc()).all()
    
    # 企業設定から会計年度開始日を取得
    company_settings = CompanySettings.query.first()
    fiscal_start_month = company_settings.fiscal_year_start_month if company_settings else 4
    fiscal_start_day = company_settings.fiscal_year_start_day if company_settings else 1
    
    return render_template('accounting_period_management.html', 
                         periods=periods,
                         fiscal_start_month=fiscal_start_month,
                         fiscal_start_day=fiscal_start_day,
                         current_year=datetime.now().year)

@app.route('/create_accounting_period', methods=['POST'])
@login_required
def create_accounting_period():
    """新規会計年度作成"""
    if current_user.role != 'accounting':
        flash('アクセス権限がありません。')
        return redirect(url_for('index'))
    
    try:
        fiscal_year = int(request.form.get('fiscal_year'))
        
        # 既存期間のチェック
        existing = AccountingPeriod.query.filter_by(fiscal_year=fiscal_year).first()
        if existing:
            flash(f'{fiscal_year}年度は既に作成されています。')
            return redirect(url_for('accounting_period_management'))
        
        # 企業設定から会計年度開始日を取得
        company_settings = CompanySettings.query.first()
        start_month = company_settings.fiscal_year_start_month if company_settings else 4
        start_day = company_settings.fiscal_year_start_day if company_settings else 1
        
        # 期間設定
        start_date = date(fiscal_year, start_month, start_day)
        if start_month == 1:
            end_date = date(fiscal_year, 12, 31)
        else:
            end_date = date(fiscal_year + 1, start_month - 1, 
                          calendar.monthrange(fiscal_year + 1, start_month - 1)[1])
        
        # 新規期間作成
        new_period = AccountingPeriod(
            fiscal_year=fiscal_year,
            start_date=start_date,
            end_date=end_date
        )
        
        db.session.add(new_period)
        db.session.commit()
        
        flash(f'{fiscal_year}年度の会計期間を作成しました。')
        
    except Exception as e:
        db.session.rollback()
        flash(f'会計期間の作成でエラーが発生しました: {str(e)}')
    
    return redirect(url_for('accounting_period_management'))

@app.route('/carryover_balances', methods=['POST'])
@login_required
def carryover_balances():
    """期末残高の繰越処理"""
    if current_user.role != 'accounting':
        flash('アクセス権限がありません。')
        return redirect(url_for('index'))
    
    try:
        from_year = int(request.form.get('from_year'))
        to_year = int(request.form.get('to_year'))
        
        # 繰越元の期間チェック
        from_period = AccountingPeriod.query.filter_by(fiscal_year=from_year).first()
        if not from_period:
            flash(f'{from_year}年度の会計期間が作成されていません。')
            return redirect(url_for('accounting_period_management'))
        
        # 繰越先の期間チェック
        to_period = AccountingPeriod.query.filter_by(fiscal_year=to_year).first()
        if not to_period:
            flash(f'{to_year}年度の会計期間が作成されていません。')
            return redirect(url_for('accounting_period_management'))
        
        # 資産・負債・純資産科目の期末残高を計算して繰越
        carryover_accounts = AccountingAccount.query.filter(
            AccountingAccount.account_type.in_(['資産', '負債', '純資産'])
        ).all()
        
        carryover_count = 0
        
        for account in carryover_accounts:
            # 繰越元年度の期末残高を計算
            balance_query = db.session.query(
                db.func.sum(JournalEntryDetail.debit_amount - JournalEntryDetail.credit_amount)
            ).join(JournalEntry).filter(
                JournalEntryDetail.account_id == account.id,
                JournalEntry.entry_date >= from_period.start_date,
                JournalEntry.entry_date <= from_period.end_date
            )
            
            period_balance = balance_query.scalar() or 0
            
            # 既存の期首残高があるかチェック
            existing_opening = OpeningBalance.query.filter_by(
                fiscal_year=to_year, 
                account_id=account.id
            ).first()
            
            if existing_opening:
                # 既存の期首残高を更新
                existing_opening.opening_balance = period_balance
                existing_opening.source_type = 'carryover'
                existing_opening.updated_at = datetime.now()
            else:
                # 新規期首残高を作成
                opening_balance = OpeningBalance(
                    fiscal_year=to_year,
                    account_id=account.id,
                    opening_balance=period_balance,
                    source_type='carryover'
                )
                db.session.add(opening_balance)
            
            if period_balance != 0:  # 残高がある場合のみカウント
                carryover_count += 1
        
        # 繰越元期間を締め済みに設定
        from_period.is_closed = True
        from_period.closing_date = datetime.now()
        
        db.session.commit()
        
        flash(f'{from_year}年度から{to_year}年度へ{carryover_count}科目の残高を繰越しました。')
        
    except Exception as e:
        db.session.rollback()
        flash(f'繰越処理でエラーが発生しました: {str(e)}')
    
    return redirect(url_for('accounting_period_management'))

# --- 雇用契約書作成 ---
@app.route('/employment_contract')
@login_required
def employment_contract():
    if current_user.role not in ['admin', 'general_affairs', 'hr_affairs']:
        flash('アクセス権がありません。')
        return redirect(url_for('dashboard'))
    
    employees = Employee.query.filter_by(status='在籍中').order_by(Employee.name).all()
    company_settings = CompanySettings.query.first()
    
    return render_template('employment_contract.html', 
                         employees=employees,
                         company_settings=company_settings)

@app.route('/create_employment_contract', methods=['POST'])
@login_required
def create_employment_contract():
    if current_user.role not in ['admin', 'general_affairs', 'hr_affairs']:
        flash('アクセス権がありません。')
        return redirect(url_for('dashboard'))
    
    try:
        company_settings = CompanySettings.query.first()
        
        # 雇用契約書データの収集
        # 新規雇用の場合は従業員情報を動的に作成
        if request.form.get('employee_id'):
            employee = Employee.query.get_or_404(request.form.get('employee_id'))
            employee_name = employee.name
        else:
            # 新規雇用の場合
            employee = None
            employee_name = request.form.get('employee_name')
        
        contract_data = {
            'employee': employee,
            'employee_name': employee_name,
            'employee_birth_date': request.form.get('birth_date'),
            'employee_address': request.form.get('address'),
            'employee_phone': request.form.get('phone_number'),
            'company': company_settings,
            'contract_type': request.form.get('contract_type'),
            'contract_period_type': request.form.get('contract_period_type'),
            'start_date': request.form.get('start_date'),
            'end_date': request.form.get('end_date'),
            'contract_renewal': request.form.get('contract_renewal'),
            'renewal_criteria': request.form.get('renewal_criteria'),
            'work_location': request.form.get('work_location'),
            'work_location_change': request.form.get('work_location_change'),
            'position': request.form.get('position'),
            'department': request.form.get('department'),
            'job_description': request.form.get('job_description'),
            'work_start_time': request.form.get('work_start_time'),
            'work_end_time': request.form.get('work_end_time'),
            'break_time': request.form.get('break_time'),
            'scheduled_working_hours': request.form.get('scheduled_working_hours'),
            'shift_work': request.form.get('shift_work'),
            'work_days': request.form.getlist('work_days'),
            'holidays': request.form.get('holidays'),
            'overtime_work': request.form.get('overtime_work'),
            'salary_type': request.form.get('salary_type'),
            'base_salary': request.form.get('base_salary'),
            'wage_calculation_method': request.form.get('wage_calculation_method'),
            'salary_closing_date': request.form.get('salary_closing_date'),
            'payment_date': request.form.get('payment_date'),
            'payment_method': request.form.get('payment_method'),
            'allowances': request.form.get('allowances'),
            'bonus_payment': request.form.get('bonus_payment'),
            'bonus_details': request.form.get('bonus_details'),
            'trial_period': request.form.get('trial_period'),
            'social_insurance': request.form.getlist('social_insurance'),
            'retirement_allowance': request.form.get('retirement_allowance'),
            'retirement_age': request.form.get('retirement_age'),
            'termination_conditions': request.form.get('termination_conditions'),
            'dismissal_reasons': request.form.get('dismissal_reasons'),
            'special_conditions': request.form.get('special_conditions')
        }
        
        # PDF生成
        pdf_buffer = generate_employment_contract_pdf(contract_data)
        
        # レスポンスの作成
        response = make_response(pdf_buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        # 日本語ファイル名をエンコードしてHTTPヘッダーに対応
        import urllib.parse
        encoded_filename = urllib.parse.quote(f'雇用契約書_{employee_name}.pdf', safe='')
        response.headers['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_filename}'
        
        flash(f'{employee_name}の雇用契約書を作成しました。', 'success')
        return response
        
    except Exception as e:
        flash('雇用契約書の作成中にエラーが発生しました。', 'error')
        return redirect(url_for('employment_contract'))

def format_japanese_date(date_str):
    """日付を日本語形式（年月日）に変換"""
    if not date_str:
        return ""
    try:
        from datetime import datetime
        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        return f"{date_obj.year}年{date_obj.month}月{date_obj.day}日"
    except:
        return date_str

def generate_employment_contract_pdf(contract_data):
    """雇用契約書PDFを生成（1ページに収まるよう最適化）"""
    buffer = BytesIO()
    # マージンを調整してページ領域を最大化（上部マージンを元に戻す）
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4,
        topMargin=15*mm,    # 10mm → 15mmに戻す
        bottomMargin=15*mm,
        leftMargin=15*mm,
        rightMargin=15*mm
    )
    styles = getSampleStyleSheet()
    story = []
    
    # 既存のフォント設定関数を使用
    japanese_font, japanese_font_bold = setup_japanese_font()
    
    # スタイル設定（コンパクト化）
    japanese_style = ParagraphStyle(
        'Japanese',
        parent=styles['Normal'],
        fontName=japanese_font,
        fontSize=8,
        leading=10
    )
    title_style = ParagraphStyle(
        'JapaneseTitle',
        parent=styles['Title'],
        fontName=japanese_font_bold,
        fontSize=14,
        leading=16,
        alignment=1  # Center
    )
    section_style = ParagraphStyle(
        'JapaneseSection',
        parent=styles['Normal'],
        fontName=japanese_font_bold,
        fontSize=9,
        leading=11
    )
    
    # タイトル
    story.append(Paragraph("雇用契約書", title_style))
    story.append(Spacer(1, 10))
    
    # 基本情報テーブルを削除して、変数のみ定義
    if contract_data['company']:
        company_name = contract_data['company'].company_name or "会社名未設定"
        representative = contract_data['company'].representative_name or "代表者名未設定"
    else:
        company_name = "会社名未設定"
        representative = "代表者名未設定"
    
    employee = contract_data['employee']
    employee_name = contract_data['employee_name']
    
    # 契約内容
    story.append(Paragraph("【労働契約の内容】", section_style))
    story.append(Spacer(1, 5))
    
    # テーブル形式で契約内容を表示（統一された2列構造に修正）
    contract_info = [
        ["契約タイプ", contract_data['contract_type'] or "未設定"],
        ["労働契約の期間", contract_data['contract_period_type'] or "未設定"],
        ["契約開始日", format_japanese_date(contract_data['start_date']) or "未設定"]
    ]
    
    # 有期契約の場合の追加情報
    if contract_data['end_date']:
        contract_info.append(["契約終了日", format_japanese_date(contract_data['end_date'])])
    if contract_data['contract_renewal']:
        contract_info.append(["契約の更新", contract_data['contract_renewal']])
    if contract_data['renewal_criteria']:
        contract_info.append(["契約更新の判断基準", contract_data['renewal_criteria']])
    
    # 就業場所・就業場所の変更（同じ行に統合）
    work_location_combined = f"{contract_data['work_location'] or '未設定'} / 変更: {contract_data['work_location_change'] or 'なし'}"
    contract_info.append(["就業場所・変更有無", work_location_combined])
    
    # 従事する業務の内容
    contract_info.append(["従事する業務の内容", contract_data['job_description'] or "未設定"])
    
    # 職種・役職・所属部署（同じ行に統合）
    position_combined = f"職種: {contract_data['position'] or '未設定'} / 役職: {contract_data.get('role', '未設定')} / 部署: {contract_data['department'] or '未設定'}"
    contract_info.append(["職種・役職・所属部署", position_combined])
    
    # 始業・終業・休憩時間（同じマスに統合）- 全角コロン使用
    work_time = f"{contract_data['work_start_time'].replace(':', '：') if contract_data['work_start_time'] else '未設定'}〜{contract_data['work_end_time'].replace(':', '：') if contract_data['work_end_time'] else '未設定'} (休憩{contract_data['break_time'] or '未設定'}分)"
    contract_info.append(["労働時間", work_time])
    
    if contract_data['scheduled_working_hours']:
        contract_info.append(["所定労働時間", contract_data['scheduled_working_hours']])
    
    if contract_data['shift_work'] and contract_data['shift_work'] != 'なし':
        contract_info.append(["交替勤務", contract_data['shift_work']])
    
    if contract_data['holidays']:
        contract_info.append(["休日・休暇", contract_data['holidays']])
    
    if contract_data['overtime_work'] and contract_data['overtime_work'] != 'なし':
        contract_info.append(["所定時間外労働", contract_data['overtime_work']])
    
    # 賃金関連
    contract_info.extend([
        ["賃金形態", contract_data['salary_type'] or "未設定"],
        ["基本給", f"{int(contract_data['base_salary']):,}円" if contract_data['base_salary'] else "未設定"],
        ["賃金の決定・計算方法", contract_data['wage_calculation_method'] or "未設定"]
    ])
    
    # 賃金締切日・支払日（同じ行に統合）
    payment_combined = f"締切日: {contract_data['salary_closing_date'] or '未設定'} / 支払日: {contract_data['payment_date'] or '未設定'}"
    contract_info.append(["賃金締切日・支払日", payment_combined])
    
    contract_info.append(["賃金支払方法", contract_data['payment_method'] or "未設定"])
    
    # 追加項目
    if contract_data['trial_period']:
        contract_info.append(["試用期間", contract_data['trial_period']])
    
    if contract_data['bonus_payment'] and contract_data['bonus_payment'] != 'なし':
        bonus_info = contract_data['bonus_details'] if contract_data['bonus_details'] else "あり"
        contract_info.append(["賞与", bonus_info])
    
    if contract_data['retirement_allowance'] and contract_data['retirement_allowance'] != 'なし':
        contract_info.append(["退職金", contract_data['retirement_allowance']])
    
    if contract_data['retirement_age'] and contract_data['retirement_age'] != 'なし':
        contract_info.append(["定年", contract_data['retirement_age']])
    
    # 文字列の長さに応じて手動で改行を挿入する方法で文字折り返しを実現
    def wrap_text(text, max_length=50):
        """テキストが表の幅を超える場合のみ改行する（より長い判定基準）"""
        if not isinstance(text, str) or len(text) <= max_length:
            return text
        
        # 表の幅を超える場合のみ改行
        words = []
        current_word = ""
        for char in text:
            current_word += char
            if len(current_word) >= max_length:
                words.append(current_word)
                current_word = ""
        if current_word:
            words.append(current_word)
        
        return "\n".join(words)
    
    # contract_infoの内容に改行を挿入
    wrapped_contract_info = []
    for label, content in contract_info:
        if isinstance(content, str):
            wrapped_content = wrap_text(content)
        else:
            wrapped_content = content
        wrapped_contract_info.append([label, wrapped_content])
    
    # 統一された2列テーブル（A4幅210mm - 左右15mm - 右15mm = 180mm有効幅）
    # 180mm = 約510ptで、項目列120pt + 内容列390ptで配分
    table = Table(wrapped_contract_info, colWidths=[120, 390])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, -1), japanese_font),  # 全列に日本語フォント適用
        ('FONTSIZE', (0, 0), (-1, -1), 8),              # 全列にフォントサイズ適用
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),             # ヘッダー列（項目名）は左揃え
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),             # 内容列も左揃え
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),            # 上揃え（複数行表示時の配置）
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('LEFTPADDING', (0, 0), (-1, -1), 15),          # 左余白15mmに調整
        ('RIGHTPADDING', (0, 0), (-1, -1), 15),         # 右余白15mmに調整
        ('TOPPADDING', (0, 0), (-1, -1), 3),           
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3)         
    ]))
    
    story.append(table)
    story.append(Spacer(1, 10))
    
    # 追加情報
    if contract_data['allowances']:
        story.append(Paragraph("【諸手当の詳細】", section_style))
        story.append(Paragraph(contract_data['allowances'], japanese_style))
        story.append(Spacer(1, 5))
    
    if contract_data['social_insurance']:
        story.append(Paragraph("【社会保険】", section_style))
        story.append(Paragraph("、".join(contract_data['social_insurance']), japanese_style))
        story.append(Spacer(1, 5))
    
    # 退職に関する事項（必須）
    if contract_data['termination_conditions']:
        story.append(Paragraph("【退職に関する事項】", section_style))
        story.append(Paragraph(contract_data['termination_conditions'], japanese_style))
        story.append(Spacer(1, 5))
    
    # 解雇事由（必須）
    if contract_data['dismissal_reasons']:
        story.append(Paragraph("【解雇事由】", section_style))
        story.append(Paragraph(contract_data['dismissal_reasons'], japanese_style))
        story.append(Spacer(1, 5))
    
    if contract_data['special_conditions']:
        story.append(Paragraph("【特記事項】", section_style))
        story.append(Paragraph(contract_data['special_conditions'], japanese_style))
        story.append(Spacer(1, 5))
    
    # 署名欄（コンパクト版）
    story.append(Spacer(1, 15))
    story.append(Paragraph("上記の条件により雇用契約を締結します。", japanese_style))
    story.append(Spacer(1, 10))
    
    today = date.today().strftime('%Y年%m月%d日')
    story.append(Paragraph(f"契約日: {today}", japanese_style))
    story.append(Spacer(1, 10))
    
    # スタッフ詳細PDFと同じフォント設定を使用
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    
    # CIDフォント候補を順番に試行（スタッフ詳細PDFと同じ順序）
    cid_fonts = [
        ('HeiseiKakuGo-W5', 'HeiseiKakuGo-W5'),  # 日本語ゴシック
        ('HeiseiMin-W3', 'HeiseiMin-W3'),        # 日本語明朝
    ]
    
    signature_font = japanese_font  # デフォルトは既存のフォント
    for regular_name, bold_name in cid_fonts:
        try:
            pdfmetrics.registerFont(UnicodeCIDFont(regular_name))
            signature_font = regular_name
            break
        except Exception:
            continue
    
    # 署名欄用のスタイル（CIDフォント、1段階小さくして10pt、左揃え）
    signature_style = ParagraphStyle(
        'SignatureStyle',
        parent=styles['Normal'],
        fontName=signature_font,
        fontSize=10,  # 12 → 10に変更
        alignment=0,  # Left
        leading=14   # 行間も少し狭める
    )
    
    # 印の文字を薄く表示するためのスタイル設定（CIDフォント、小さなフォントサイズ）
    light_style = ParagraphStyle(
        'LightText',
        parent=styles['Normal'],
        fontName=signature_font,
        fontSize=10,  # 12 → 10に変更
        textColor=colors.Color(0.7, 0.7, 0.7, 1),  # グレー色で薄く表示
        alignment=1  # Center（印は中央揃えのまま）
    )
    
    # 代表者の役職・氏名を取得（会社名追加、行間調整）
    if contract_data['company'] and contract_data['company'].representative_name:
        employer_text = f"{company_name}<br/><br/>代表取締役　{contract_data['company'].representative_name}"
    else:
        employer_text = f"{company_name}<br/><br/>代表取締役"
    
    # 被雇用者の氏名を雇用者の高さに合わせるため改行を追加
    employee_text = f"<br/><br/>{employee_name}"
    
    # ヘッダー用のスタイル（左揃え）
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontName=signature_font,
        fontSize=10,  # 12 → 10に変更
        alignment=0  # Left
    )
    
    # 印の位置を氏名の右横、高さは氏名に合わせる（有効幅510ptで配分）
    signature_info = [
        [Paragraph("雇用者", header_style), "", Paragraph("被雇用者", header_style), ""],
        [Paragraph(employer_text, signature_style), Paragraph("印", light_style), Paragraph(employee_text, signature_style), Paragraph("印", light_style)]
    ]
    
    signature_table = Table(signature_info, colWidths=[200, 55, 200, 55])
    signature_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),     # 雇用者ヘッダーは左揃え
        ('ALIGN', (2, 0), (2, 0), 'LEFT'),     # 被雇用者ヘッダーは左揃え
        ('ALIGN', (0, 1), (0, 1), 'LEFT'),     # 雇用者本文は左揃え
        ('ALIGN', (2, 1), (2, 1), 'LEFT'),     # 被雇用者本文は左揃え
        ('ALIGN', (1, 1), (1, 1), 'LEFT'),     # 印を左揃えに変更（氏名の右横に配置）
        ('ALIGN', (3, 1), (3, 1), 'LEFT'),     # 印を左揃えに変更（氏名の右横に配置）
        ('VALIGN', (0, 0), (-1, -1), 'BOTTOM'), # 下揃えに変更（氏名の行の高さに合わせる）
        # 罫線なし
    ]))
    
    story.append(signature_table)
    
    # PDF生成
    doc.build(story)
    buffer.seek(0)
    return buffer

# --- 労働条件変更機能 ---
@app.route('/working_conditions_change')
@login_required
def working_conditions_change():
    """労働条件変更ページ"""
    if current_user.role not in ['admin', 'general_affairs', 'hr_affairs']:
        flash('アクセス権がありません。')
        return redirect(url_for('dashboard'))
    
    employees = Employee.query.filter_by(status='在籍中').order_by(Employee.name).all()
    company_settings = CompanySettings.query.first()
    
    return render_template('working_conditions_change_mhlw.html', 
                         employees=employees,
                         company_settings=company_settings)

@app.route('/social_insurance_acquisition')
@login_required
def social_insurance_acquisition():
    """社会保険資格取得届作成画面"""
    if current_user.role not in ['admin', 'general_affairs', 'hr_affairs']:
        flash('アクセス権がありません。')
        return redirect(url_for('dashboard'))
    
    employees = Employee.query.filter_by(status='在籍中').order_by(Employee.name).all()
    company_settings = CompanySettings.query.first()
    
    return render_template('social_insurance_acquisition.html', 
                         employees=employees,
                         company_settings=company_settings)

@app.route('/social_insurance_loss')
@login_required
def social_insurance_loss():
    """社会保険資格喪失届作成画面"""
    if current_user.role not in ['admin', 'general_affairs', 'hr_affairs']:
        flash('アクセス権がありません。')
        return redirect(url_for('dashboard'))
    
    employees = Employee.query.filter_by(status='在籍中').order_by(Employee.name).all()
    company_settings = CompanySettings.query.first()
    
    return render_template('social_insurance_loss.html', 
                         employees=employees,
                         company_settings=company_settings)

@app.route('/social_insurance_change')
@login_required
def social_insurance_change():
    """社会保険変更届作成画面"""
    if current_user.role not in ['admin', 'general_affairs', 'hr_affairs']:
        flash('アクセス権がありません。')
        return redirect(url_for('dashboard'))
    
    employees = Employee.query.filter_by(status='在籍中').order_by(Employee.name).all()
    company_settings = CompanySettings.query.first()
    
    return render_template('social_insurance_change.html', 
                         employees=employees,
                         company_settings=company_settings)

@app.route('/create_social_insurance_acquisition', methods=['POST'])
@login_required
def create_social_insurance_acquisition():
    """社会保険資格取得届作成"""
    if current_user.role not in ['admin', 'general_affairs', 'hr_affairs']:
        flash('アクセス権がありません。')
        return redirect(url_for('dashboard'))
    
    try:
        # フォームデータの取得
        employee_id = request.form.get('employee_id')
        acquisition_date = request.form.get('acquisition_date')
        employee_name = request.form.get('employee_name')
        employee_kana = request.form.get('employee_kana')
        my_number = request.form.get('my_number')
        birth_date = request.form.get('birth_date')
        gender = request.form.get('gender')
        employment_type = request.form.get('employment_type')
        monthly_salary = request.form.get('monthly_salary')
        job_description = request.form.get('job_description')
        has_dependents = request.form.get('has_dependents') == 'on'
        
        # 必須項目チェック
        if not all([employee_id, acquisition_date, my_number, birth_date, gender, employment_type, monthly_salary]):
            flash('必須項目をすべて入力してください。', 'error')
            return redirect(url_for('social_insurance_acquisition'))
        
        # 従業員情報の取得
        employee = Employee.query.get_or_404(employee_id)
        company_settings = CompanySettings.query.first()
        
        # 資格取得届データ
        acquisition_data = {
            'employee': employee,
            'employee_name': employee_name,
            'employee_kana': employee_kana,
            'my_number': my_number,
            'birth_date': birth_date,
            'gender': gender,
            'acquisition_date': acquisition_date,
            'employment_type': employment_type,
            'monthly_salary': int(monthly_salary),
            'job_description': job_description,
            'has_dependents': has_dependents,
            'company': company_settings,
            'created_by': current_user.email,
            'created_date': datetime.now().strftime('%Y年%m月%d日')
        }
        
        # PDF生成
        pdf_buffer = generate_social_insurance_acquisition_pdf(acquisition_data)
        
        # PDFレスポンスの作成
        response = make_response(pdf_buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{urllib.parse.quote(f"社会保険資格取得届_{employee_name}.pdf", safe="")}'
        
        flash(f'{employee_name}の社会保険資格取得届を作成しました。', 'success')
        return response
        
    except Exception as e:
        flash(f'社会保険資格取得届の作成に失敗しました: {str(e)}', 'error')
        return redirect(url_for('social_insurance_acquisition'))

@app.route('/create_working_conditions_change', methods=['POST'])
@login_required
def create_working_conditions_change():
    """労働条件変更通知書作成"""
    if current_user.role not in ['admin', 'general_affairs', 'hr_affairs']:
        flash('アクセス権がありません。')
        return redirect(url_for('dashboard'))
    
    try:
        # フォームデータの取得
        employee_id = request.form.get('employee_id')
        change_date = request.form.get('change_date')
        change_reason = request.form.get('change_reason')
        
        # 変更項目の取得（厚労省モデルに基づく全15項目）
        changes = {}
        change_fields = [
            'contract_period', 'contract_renewal', 'workplace', 'job_duties', 
            'work_hours', 'break_time', 'overtime', 'holidays', 'vacation', 
            'base_salary', 'allowances', 'payment_terms', 'payment_method',
            'retirement', 'dismissal', 'social_insurance', 'employment_insurance'
        ]
        
        for field in change_fields:
            change_type = request.form.get(f'change_type_{field}')
            if change_type == 'change':
                # 契約期間の特別処理
                if field == 'contract_period':
                    old_period_type = request.form.get('old_contract_period_type')
                    new_period_type = request.form.get('new_contract_period_type')
                    
                    old_value = old_period_type
                    if old_period_type == '期間の定めあり':
                        old_start = request.form.get('old_contract_start_date')
                        old_end = request.form.get('old_contract_end_date')
                        if old_start and old_end:
                            old_value += f"（{format_japanese_date(old_start)}～{format_japanese_date(old_end)}）"
                    
                    new_value = new_period_type
                    if new_period_type == '期間の定めあり':
                        new_start = request.form.get('new_contract_start_date')
                        new_end = request.form.get('new_contract_end_date')
                        if new_start and new_end:
                            new_value += f"（{format_japanese_date(new_start)}～{format_japanese_date(new_end)}）"
                else:
                    old_value = request.form.get(f'old_{field}')
                    new_value = request.form.get(f'new_{field}')
                
                if old_value or new_value:
                    # 賃金関連の値はカンマ区切りでフォーマット
                    if field == 'base_salary':
                        formatted_old = format_salary(old_value) if old_value else '未設定'
                        formatted_new = format_salary(new_value) if new_value else '未設定'
                    # 時刻表示フィールドの全角コロン変換
                    elif field in ['work_hours', 'break_time']:
                        formatted_old = format_time_display(old_value) if old_value else '未設定'
                        formatted_new = format_time_display(new_value) if new_value else '未設定'
                    else:
                        formatted_old = old_value or '未設定'
                        formatted_new = new_value or '未設定'
                    
                    changes[field] = {
                        'old_value': formatted_old,
                        'new_value': formatted_new,
                        'label': get_change_field_label(field),
                        'change_type': 'change'
                    }
            elif change_type == 'no_change':
                # 契約期間の特別処理
                if field == 'contract_period':
                    current_period_type = request.form.get('current_contract_period_type')
                    current_value = current_period_type
                    if current_period_type == '期間の定めあり':
                        current_start = request.form.get('current_contract_start_date')
                        current_end = request.form.get('current_contract_end_date')
                        if current_start and current_end:
                            current_value += f"（{format_japanese_date(current_start)}～{format_japanese_date(current_end)}）"
                else:
                    current_value = request.form.get(f'current_{field}')
                
                if current_value:
                    # 賃金関連の値はカンマ区切りでフォーマット
                    if field == 'base_salary':
                        formatted_current = format_salary(current_value)
                    # 時刻表示フィールドの全角コロン変換
                    elif field in ['work_hours', 'break_time']:
                        formatted_current = format_time_display(current_value)
                    else:
                        formatted_current = current_value
                    
                    changes[field] = {
                        'old_value': formatted_current,
                        'new_value': formatted_current,
                        'label': get_change_field_label(field),
                        'change_type': 'no_change'
                    }
        
        if not changes:
            flash('変更項目を入力してください。', 'error')
            return redirect(url_for('working_conditions_change'))
        
        # 従業員情報の取得
        if employee_id:
            employee = Employee.query.get_or_404(employee_id)
            employee_name = employee.name
        else:
            employee_name = request.form.get('employee_name', '従業員名未設定')
            employee = None
        
        # 会社情報の取得
        company_settings = CompanySettings.query.first()
        
        # 労働条件変更通知書データ
        change_data = {
            'employee': employee,
            'employee_name': employee_name,
            'company': company_settings,
            'change_date': change_date,
            'change_reason': change_reason,
            'changes': changes,
            'created_by': current_user.email,
            'created_date': datetime.now().strftime('%Y年%m月%d日')
        }
        
        # PDF生成
        pdf_buffer = generate_working_conditions_change_pdf(change_data)
        
        # PDFレスポンスの作成
        response = make_response(pdf_buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{urllib.parse.quote(f"労働条件変更通知書_{employee_name}.pdf", safe="")}'
        
        flash(f'{employee_name}の労働条件変更通知書を作成しました。', 'success')
        return response
        
    except Exception as e:
        flash(f'労働条件変更通知書の作成に失敗しました: {str(e)}', 'error')
        return redirect(url_for('working_conditions_change'))

def format_japanese_date(date_str):
    """YYYY-MM-DD形式を年月日形式に変換"""
    if not date_str or len(date_str) != 10:
        return date_str
    try:
        from datetime import datetime
        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        return f"{date_obj.year}年{date_obj.month}月{date_obj.day}日"
    except:
        return date_str

def format_salary(salary_str):
    """賃金をカンマ区切りでフォーマット"""
    if not salary_str:
        return salary_str
    try:
        # 数値に変換してカンマ区切りでフォーマット
        salary_num = int(salary_str)
        return f"{salary_num:,}円"
    except (ValueError, TypeError):
        return salary_str

def format_time_display(text):
    """時刻表示のコロンを半角から全角に変換し、ハイフンの前後にスペースを追加"""
    if not text:
        return text
    # 半角コロン(:)を全角コロン(：)に置換
    formatted = text.replace(':', '：')
    # ハイフン(-)の前後にスペースを追加（既にスペースがある場合は重複を避ける）
    formatted = formatted.replace(' - ', '-').replace('-', ' - ')
    return formatted

def get_change_field_label(field):
    """変更項目のラベルを取得（厚労省完全準拠）"""
    labels = {
        'contract_period': '1. 契約期間',
        'contract_renewal': '1-1. 契約の更新の有無',
        'workplace': '2. 就業の場所',
        'job_duties': '3. 従事すべき業務',
        'work_hours': '4. 始業、終業時間',
        'break_time': '5. 休憩時間',
        'overtime': '6. 所定時間外労働の有無',
        'holidays': '7. 休日',
        'vacation': '8. 休暇',
        'base_salary': '9. 賃金',
        'allowances': '10. 諸手当',
        'payment_terms': '11. 賃金締切日、支払日',
        'payment_method': '12. 賃金支払方法',
        'retirement': '13. 退職に関する事項',
        'dismissal': '14. 解雇の事由及び手続',
        'social_insurance': '15. 社会保険加入状況',
        'employment_insurance': '16. 雇用保険の適用'
    }
    return labels.get(field, field)

def generate_working_conditions_change_pdf(change_data):
    """労働条件変更通知書PDFを生成（1ページ最適化・厚労省完全準拠）"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4,
        topMargin=10*mm,    # 1ページ最適化のため上部マージンを縮小
        bottomMargin=10*mm,
        leftMargin=12*mm,   # テーブル幅拡張のため左右マージンを縮小
        rightMargin=12*mm
    )
    
    story = []
    
    # 雇用契約書と同じフォント設定関数を使用
    japanese_font, japanese_font_bold = setup_japanese_font()
    
    # 雇用契約書と同じCIDフォント設定
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    
    cid_fonts = [
        ('HeiseiKakuGo-W5', 'HeiseiKakuGo-W5'),  # 日本語ゴシック
        ('HeiseiMin-W3', 'HeiseiMin-W3'),        # 日本語明朝
    ]
    
    cid_font = japanese_font  # デフォルトは既存のフォント
    for regular_name, bold_name in cid_fonts:
        try:
            pdfmetrics.registerFont(UnicodeCIDFont(regular_name))
            cid_font = regular_name
            break
        except Exception:
            continue
    
    # 雇用契約書と同じスタイル設定
    styles = getSampleStyleSheet()
    # 1ページ最適化のためフォントサイズを調整
    title_style = ParagraphStyle(
        'JapaneseTitle',
        parent=styles['Title'],
        fontName=japanese_font_bold,
        fontSize=12,        # 14 → 12に縮小
        leading=14,         # 16 → 14に縮小
        alignment=1,  # Center
        spaceAfter=8        # 15 → 8に縮小
    )
    
    normal_style = ParagraphStyle(
        'Japanese',
        parent=styles['Normal'],
        fontName=cid_font,
        fontSize=9,         # 8 → 9に変更
        leading=11          # 10 → 11に変更
    )
    
    section_style = ParagraphStyle(
        'JapaneseSection',
        parent=styles['Normal'],
        fontName=japanese_font_bold,
        fontSize=8,         # 9 → 8に縮小
        leading=10          # 11 → 10に縮小
    )
    
    # 発信者用スタイル（右寄せ左揃え）
    sender_style = ParagraphStyle(
        'SenderStyle',
        parent=styles['Normal'],
        fontName=cid_font,
        fontSize=8,
        leading=10,
        alignment=2,        # Right（右揃え）
        leftIndent=300      # 左インデントで右寄せ効果を作成
    )
    
    # タイトル
    story.append(Paragraph("労働条件変更通知書", title_style))
    story.append(Spacer(1, 8))       # 15 → 8に縮小
    
    # 基本情報
    if change_data['company']:
        company_name = change_data['company'].company_name or "会社名未設定"
        representative = change_data['company'].representative_name or "代表者名未設定"
    else:
        company_name = "会社名未設定"
        representative = "代表者名未設定"
    
    # 日付と宛先をコンパクトに配置
    date_info = f"通知日：{change_data['created_date']}"
    story.append(Paragraph(date_info, normal_style))
    story.append(Spacer(1, 6))       # 15 → 6に縮小
    
    # 宛先（10ptスタイル作成）
    addressee_style = ParagraphStyle(
        'AddresseeStyle',
        parent=styles['Normal'],
        fontName=cid_font,
        fontSize=10,
        leading=13
    )
    story.append(Paragraph(f"{change_data['employee_name']} 殿", addressee_style))
    story.append(Spacer(1, 6))       # 15 → 6に縮小
    
    # 発信者（右寄せ左揃え・10ptスタイル作成）
    sender_10pt_style = ParagraphStyle(
        'Sender10ptStyle',
        parent=styles['Normal'],
        fontName=cid_font,
        fontSize=10,
        leading=13,
        alignment=2,        # Right（右揃え）
        leftIndent=300      # 左インデントで右寄せ効果を作成
    )
    story.append(Paragraph(f"{company_name}", sender_10pt_style))
    story.append(Paragraph(f"代表取締役　{representative}", sender_10pt_style))
    story.append(Spacer(1, 10))      # 20 → 10に縮小
    
    # 本文
    story.append(Paragraph("下記のとおり労働条件を変更いたしますのでお知らせします。", normal_style))
    story.append(Spacer(1, 8))       # 15 → 8に縮小
    
    # 変更日と理由
    change_info_data = []
    if change_data.get('change_date'):
        # 日付を日本語形式に変換
        try:
            if isinstance(change_data['change_date'], str) and len(change_data['change_date']) == 10:
                # YYYY-MM-DD形式の場合
                from datetime import datetime
                date_obj = datetime.strptime(change_data['change_date'], '%Y-%m-%d')
                formatted_date = f"{date_obj.year}年{date_obj.month}月{date_obj.day}日"
            else:
                formatted_date = change_data['change_date']
        except:
            formatted_date = change_data['change_date']
        
        change_info_data.append(["変更実施日", formatted_date])
    
    if change_data.get('change_reason'):
        change_info_data.append(["変更理由", change_data['change_reason']])
    
    if change_info_data:
        info_table = Table(change_info_data, colWidths=[80, 400])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), cid_font),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3)
        ]))
        story.append(info_table)
        story.append(Spacer(1, 15))
    
    # 変更内容テーブル（雇用契約書と同じスタイル）
    story.append(Paragraph("【変更内容】", section_style))
    story.append(Spacer(1, 5))
    
    change_table_data = [["変更項目", "変更前", "変更後", "備考"]]
    
    for field_key, change_info in change_data['changes'].items():
        if change_info.get('change_type') == 'no_change':
            remark = '変更なし'
            new_value_display = change_info['new_value']
        else:
            remark = '変更'
            new_value_display = change_info['new_value'] or '-'
        
        change_table_data.append([
            change_info['label'],
            change_info['old_value'] or '-',
            new_value_display,
            remark
        ])
    
    # 1ページ最適化テーブル設定（4列・変更前後欄を拡張）
    change_table = Table(change_table_data, colWidths=[90, 200, 200, 25])  # 変更前後の列幅を180から200に拡張
    change_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, -1), cid_font),
        ('FONTSIZE', (0, 0), (-1, -1), 7.5),        # 8 → 7.5に変更
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),       # ヘッダー行のみ中央揃え
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),        # データ行は左揃え
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),       # 15 → 3に縮小
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),      # 15 → 3に縮小
        ('TOPPADDING', (0, 0), (-1, -1), 2),        # 3 → 2に縮小
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2)      # 3 → 2に縮小
    ]))
    
    story.append(change_table)
    story.append(Spacer(1, 8))                      # 20 → 8に縮小
    
    # 注意事項（コンパクト版）
    story.append(Paragraph("【備考】", section_style))
    story.append(Paragraph("※ 本通知書は労働基準法第15条に基づく労働条件の明示です。", normal_style))
    story.append(Paragraph("※ ご不明な点がございましたら、総務部までお問い合わせください。", normal_style))
    story.append(Spacer(1, 10))
    
    # 署名・押印欄（雇用契約書と同様のスタイル）
    story.append(Paragraph("上記の労働条件変更について、当事者間で合意したことを確認し、署名・押印いたします。", normal_style))
    story.append(Spacer(1, 8))
    
    
    # 雇用者と被雇用者の情報（改行を適切に処理）
    if change_data['company']:
        company_name = change_data['company'].company_name or "会社名未設定"
        representative = change_data['company'].representative_name or "代表者名未設定"
        employer_text = f"{company_name}\n代表取締役　{representative}"
    else:
        employer_text = "雇用者名\n代表取締役　代表者名"
    
    employee_text = f"{change_data['employee_name']}"
    
    # 署名テーブル（印の文字を削除し2列構造に変更）
    signature_info = [
        ["雇用者", "被雇用者"],
        [employer_text, employee_text]
    ]
    
    signature_table = Table(signature_info, colWidths=[250, 250])   # 印の列を削除し2列構造に変更
    signature_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),     # 雇用者ヘッダーを左揃え
        ('ALIGN', (1, 0), (1, 0), 'LEFT'),     # 被雇用者ヘッダーを左揃え
        ('ALIGN', (0, 1), (0, 1), 'LEFT'),     # 雇用者本文を左揃え
        ('ALIGN', (1, 1), (1, 1), 'LEFT'),     # 被雇用者本文を左揃え
        ('VALIGN', (0, 0), (-1, -1), 'BOTTOM'), # 下揃え
        ('FONTNAME', (0, 0), (-1, -1), cid_font),
    ]))
    
    story.append(signature_table)
    story.append(Spacer(1, 5))
    
    # PDF生成
    doc.build(story)
    buffer.seek(0)
    return buffer

def generate_social_insurance_acquisition_pdf(acquisition_data):
    """社会保険資格取得届PDFを生成（厚生労働省準拠）"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=15*mm,
        bottomMargin=15*mm,
        leftMargin=15*mm,
        rightMargin=15*mm
    )
    
    story = []
    
    # フォント設定
    japanese_font, japanese_font_bold = setup_japanese_font()
    
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    cid_fonts = [
        ('HeiseiKakuGo-W5', 'HeiseiKakuGo-W5'),
        ('HeiseiMin-W3', 'HeiseiMin-W3'),
    ]
    
    cid_font = japanese_font
    for regular_name, bold_name in cid_fonts:
        try:
            pdfmetrics.registerFont(UnicodeCIDFont(regular_name))
            cid_font = regular_name
            break
        except Exception:
            continue
    
    # スタイル設定
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'JapaneseTitle',
        parent=styles['Title'],
        fontName=cid_font,
        fontSize=16,
        leading=20,
        alignment=1,
        spaceAfter=20
    )
    
    normal_style = ParagraphStyle(
        'Japanese',
        parent=styles['Normal'],
        fontName=cid_font,
        fontSize=10,
        leading=14
    )
    
    # タイトル
    story.append(Paragraph("健康保険・厚生年金保険被保険者資格取得届", title_style))
    story.append(Spacer(1, 10))
    
    # 会社情報
    if acquisition_data['company']:
        company_name = acquisition_data['company'].company_name or "会社名未設定"
        company_address = acquisition_data['company'].company_address or "住所未設定"
    else:
        company_name = "会社名未設定"
        company_address = "住所未設定"
    
    story.append(Paragraph(f"提出日：{acquisition_data['created_date']}", normal_style))
    story.append(Paragraph(f"事業所名称：{company_name}", normal_style))
    story.append(Paragraph(f"事業所所在地：{company_address}", normal_style))
    story.append(Spacer(1, 15))
    
    # 基本情報テーブル
    basic_info_data = [
        ["項目", "内容"],
        ["氏名", acquisition_data['employee_name']],
        ["フリガナ", acquisition_data['employee_kana']],
        ["個人番号", acquisition_data['my_number']],
        ["生年月日", format_japanese_date(acquisition_data['birth_date'])],
        ["性別", acquisition_data['gender']],
        ["資格取得年月日", format_japanese_date(acquisition_data['acquisition_date'])],
        ["雇用形態", acquisition_data['employment_type']],
        ["標準報酬月額", f"{acquisition_data['monthly_salary']:,}円"],
        ["従事する業務の内容", acquisition_data['job_description'] or "未記載"],
        ["扶養家族", "有" if acquisition_data['has_dependents'] else "無"]
    ]
    
    basic_info_table = Table(basic_info_data, colWidths=[120, 300])
    basic_info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, -1), cid_font),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4)
    ]))
    
    story.append(basic_info_table)
    story.append(Spacer(1, 20))
    
    # 注意事項
    story.append(Paragraph("【注意事項】", ParagraphStyle('注意', parent=normal_style, fontName=japanese_font_bold)))
    story.append(Paragraph("1. 本届出は雇用から5日以内に年金事務所へ提出してください。", normal_style))
    story.append(Paragraph("2. 年金手帳またはマイナンバーカードの写しを添付してください。", normal_style))
    if acquisition_data['has_dependents']:
        story.append(Paragraph("3. 扶養家族がいる場合は「健康保険被扶養者異動届」も併せて提出してください。", normal_style))
    story.append(Spacer(1, 20))
    
    # 作成者情報
    story.append(Paragraph(f"作成者：{acquisition_data['created_by']}", normal_style))
    story.append(Paragraph(f"作成日：{acquisition_data['created_date']}", normal_style))
    
    # PDF生成
    doc.build(story)
    buffer.seek(0)
    return buffer

# --- 会計機能 ---
@app.route('/journal_entries')
@login_required
def journal_entries():
    """仕訳入力画面"""
    if current_user.role != 'accounting':
        flash('アクセス権限がありません。')
        return redirect(url_for('index'))
    
    # 会計科目一覧を取得
    accounts = AccountingAccount.query.filter_by(is_active=True).order_by(AccountingAccount.account_code).all()
    
    # 最近の仕訳を取得
    recent_entries = JournalEntry.query.order_by(JournalEntry.created_at.desc()).limit(10).all()
    
    return render_template('journal_entries.html', accounts=accounts, recent_entries=recent_entries)

@app.route('/create_journal_entry', methods=['POST'])
@login_required
def create_journal_entry():
    """仕訳登録"""
    if current_user.role != 'accounting':
        flash('アクセス権限がありません。')
        return redirect(url_for('index'))
    
    try:
        entry_date = datetime.strptime(request.form.get('entry_date'), '%Y-%m-%d').date()
        description = request.form.get('description')
        reference_number = request.form.get('reference_number')
        
        # 仕訳明細の取得
        debit_accounts = request.form.getlist('debit_account')
        debit_amounts = request.form.getlist('debit_amount')
        credit_accounts = request.form.getlist('credit_account')
        credit_amounts = request.form.getlist('credit_amount')
        
        # 借方・貸方の合計チェック
        total_debit = sum(int(amount) for amount in debit_amounts if amount)
        total_credit = sum(int(amount) for amount in credit_amounts if amount)
        
        if total_debit != total_credit:
            flash('借方と貸方の金額が一致しません。')
            return redirect(url_for('journal_entries'))
        
        # 仕訳ヘッダーを作成
        journal_entry = JournalEntry(
            entry_date=entry_date,
            description=description,
            reference_number=reference_number,
            total_amount=total_debit,
            created_by=current_user.id
        )
        db.session.add(journal_entry)
        db.session.flush()  # IDを取得するためにflush
        
        # 借方明細を作成
        for i, (account_id, amount) in enumerate(zip(debit_accounts, debit_amounts)):
            if account_id and amount:
                detail = JournalEntryDetail(
                    journal_entry_id=journal_entry.id,
                    account_id=int(account_id),
                    debit_amount=int(amount),
                    credit_amount=0
                )
                db.session.add(detail)
        
        # 貸方明細を作成
        for i, (account_id, amount) in enumerate(zip(credit_accounts, credit_amounts)):
            if account_id and amount:
                detail = JournalEntryDetail(
                    journal_entry_id=journal_entry.id,
                    account_id=int(account_id),
                    debit_amount=0,
                    credit_amount=int(amount)
                )
                db.session.add(detail)
        
        db.session.commit()
        flash('仕訳を登録しました。')
        
    except Exception as e:
        db.session.rollback()
        flash(f'仕訳登録でエラーが発生しました: {str(e)}')
    
    return redirect(url_for('journal_entries'))

@app.route('/accounting_ledger')
@login_required
def accounting_ledger():
    """総勘定元帳画面"""
    if current_user.role != 'accounting':
        flash('アクセス権限がありません。')
        return redirect(url_for('index'))
    
    # クエリパラメータ取得
    account_id = request.args.get('account_id', type=int)
    year = request.args.get('year', type=int, default=datetime.now().year)
    month = request.args.get('month', type=int)
    
    # 会計科目一覧
    accounts = AccountingAccount.query.filter_by(is_active=True).order_by(AccountingAccount.account_code).all()
    
    ledger_data = []
    if account_id:
        # 指定された科目の取引明細を取得
        query = JournalEntryDetail.query.join(JournalEntry).filter(
            JournalEntryDetail.account_id == account_id,
            db.extract('year', JournalEntry.entry_date) == year
        )
        
        if month:
            query = query.filter(db.extract('month', JournalEntry.entry_date) == month)
        
        details = query.order_by(JournalEntry.entry_date).all()
        
        # 期首残高を取得
        opening_balance_obj = OpeningBalance.query.filter_by(
            fiscal_year=year,
            account_id=account_id
        ).first()
        opening_balance = opening_balance_obj.opening_balance if opening_balance_obj else 0
        
        # 各明細に相手科目の情報と累計差引金額を追加
        running_balance = opening_balance  # 期首残高から開始
        for detail in details:
            # 同じ仕訳の他の科目を取得（相手科目）
            opposite_details = JournalEntryDetail.query.filter(
                JournalEntryDetail.journal_entry_id == detail.journal_entry_id,
                JournalEntryDetail.id != detail.id
            ).all()
            
            # 相手科目名を取得（複数ある場合は最初の一つ）
            if opposite_details:
                opposite_account = AccountingAccount.query.get(opposite_details[0].account_id)
                detail.opposite_account_name = opposite_account.account_name if opposite_account else '-'
            else:
                detail.opposite_account_name = '-'
            
            # 累計差引金額を計算
            running_balance += detail.debit_amount - detail.credit_amount
            detail.running_balance = running_balance
                
        ledger_data = details
    
    years = list(range(datetime.now().year - 2, datetime.now().year + 2))
    
    # 期首残高情報も渡す
    opening_balance = 0
    if account_id:
        opening_balance_obj = OpeningBalance.query.filter_by(
            fiscal_year=year,
            account_id=account_id
        ).first()
        opening_balance = opening_balance_obj.opening_balance if opening_balance_obj else 0
    
    return render_template('accounting_ledger.html', 
                         accounts=accounts, 
                         ledger_data=ledger_data,
                         selected_account_id=account_id,
                         selected_year=year,
                         selected_month=month,
                         opening_balance=opening_balance,
                         years=years)

@app.route('/export_ledger_excel')
@login_required
def export_ledger_excel():
    """総勘定元帳のExcelエクスポート"""
    try:
        if current_user.role != 'accounting':
            flash('アクセス権限がありません。')
            return redirect(url_for('index'))
        
        # クエリパラメータ取得
        account_id = request.args.get('account_id', type=int)
        year = request.args.get('year', type=int, default=datetime.now().year)
        month = request.args.get('month', type=int)
        
        if not account_id:
            flash('勘定科目が選択されていません。')
            return redirect(url_for('accounting_ledger'))
        
        # 勘定科目情報取得
        account = AccountingAccount.query.get_or_404(account_id)
        
        # 元帳データ取得（制限付き - 最大1000件）
        query = JournalEntryDetail.query.join(JournalEntry).filter(
            JournalEntryDetail.account_id == account_id,
            db.extract('year', JournalEntry.entry_date) == year
        )
        
        if month:
            query = query.filter(db.extract('month', JournalEntry.entry_date) == month)
        
        details = query.order_by(JournalEntry.entry_date).limit(1000).all()
        
        # 期首残高を取得
        opening_balance_obj = OpeningBalance.query.filter_by(
            fiscal_year=year,
            account_id=account_id
        ).first()
        opening_balance = opening_balance_obj.opening_balance if opening_balance_obj else 0
        
        # 各明細に相手科目の情報と累計差引金額を追加
        running_balance = opening_balance  # 期首残高から開始
        for detail in details:
            try:
                # 相手科目取得
                opposite_details = JournalEntryDetail.query.filter(
                    JournalEntryDetail.journal_entry_id == detail.journal_entry_id,
                    JournalEntryDetail.id != detail.id
                ).first()  # firstに変更してパフォーマンス向上
                
                if opposite_details:
                    opposite_account = AccountingAccount.query.get(opposite_details.account_id)
                    detail.opposite_account_name = opposite_account.account_name if opposite_account else '-'
                else:
                    detail.opposite_account_name = '-'
                
                # 累計差引金額を計算
                running_balance += detail.debit_amount - detail.credit_amount
                detail.running_balance = running_balance
            except Exception as e:
                print(f"Error processing detail {detail.id}: {e}")
                detail.opposite_account_name = '-'
                detail.running_balance = running_balance
        
        # Excelワークブック作成
        wb = Workbook()
        ws = wb.active
        ws.title = "総勘定元帳"
        
        # ヘッダー情報
        period_str = f"{year}年"
        if month:
            period_str += f"{month}月"
        
        ws['A1'] = "総勘定元帳"
        ws['A3'] = f"{account.account_code}　{account.account_name}"
        
        # ヘッダースタイル（明朝体を指定）
        header_font = Font(name='ＭＳ 明朝', bold=True, size=14)
        center_alignment = Alignment(horizontal='center')
        double_underline_border = Border(bottom=Side(style='double'))
        
        # 総勘定元帳のセルのみ下線二重線（文字部分のみ）
        ws['A1'].font = header_font
        ws['A1'].alignment = center_alignment
        ws['A1'].border = double_underline_border  # 総勘定元帳に下線二重線（文字部分のみ）
        
        # 総勘定元帳を中央表示するためにマージ
        ws.merge_cells('A1:G1')
        
        ws['A3'].font = Font(name='ＭＳ 明朝', bold=True, size=11)
        ws['A3'].alignment = center_alignment
        
        # 科目情報を中央表示
        ws.merge_cells('A3:G3')
        
        # A4サイズ設定
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT  # 縦向き
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1  # A4縦に収める
        
        # 印刷設定
        ws.print_options.horizontalCentered = True
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        
        # 罫線スタイル定義
        thick_border = Side(style='medium')  # 外枠用中太線（少し細くする）
        solid_border = Side(style='thin')   # 縦線用実線
        dashed_border = Side(style='dashed') # 横線用破線
        
        # テーブルヘッダー
        headers = ['伝票No.', '日付', '相手科目', '摘要', '借方', '貸方', '差引金額']
        start_row = 5  # 1行空けたので5行目から開始
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = Font(name='ＭＳ 明朝', bold=True, size=9)  # ヘッダーも明朝体9ptに統一
            cell.alignment = Alignment(horizontal='center')
            
            # ヘッダー行の罫線（外枠太線）
            cell.border = Border(
                top=thick_border,
                bottom=thick_border,  # 項目名の下も同じ太線にする
                left=thick_border if col == 1 else solid_border,
                right=thick_border if col == len(headers) else solid_border
            )
        
        # A4縦に収まる行数を計算（行数を増やす）
        max_rows = 45
        data_rows = len(details)
        current_row = start_row + 1
        
        # 期首残高行を追加（期首残高がある場合）
        if opening_balance != 0:
            debit_value = opening_balance if opening_balance > 0 else 0
            credit_value = -opening_balance if opening_balance < 0 else 0
            
            cells = [
                ws.cell(row=current_row, column=1, value='期首残高'),
                ws.cell(row=current_row, column=2, value=f'{year % 100}.4.1'),
                ws.cell(row=current_row, column=3, value='-'),
                ws.cell(row=current_row, column=4, value='期首残高'),
                ws.cell(row=current_row, column=5, value=debit_value),
                ws.cell(row=current_row, column=6, value=credit_value),
                ws.cell(row=current_row, column=7, value=opening_balance)
            ]
            
            # 期首残高行のスタイル設定
            for col, cell in enumerate(cells, 1):
                cell.font = Font(name='ＭＳ 明朝', bold=True, size=9)  # 太字にする
                
                if col >= 5:  # 借方、貸方、差引金額
                    cell.alignment = Alignment(horizontal='right', shrinkToFit=True)
                    if cell.value > 0:
                        cell.number_format = '#,##0'
                else:
                    cell.alignment = Alignment(horizontal='left' if col > 2 else 'center', shrinkToFit=True)
                
                # 罫線設定
                cell.border = Border(
                    top=dashed_border,
                    bottom=dashed_border,
                    left=thick_border if col == 1 else solid_border,
                    right=thick_border if col == len(headers) else solid_border
                )
            
            current_row += 1
            max_rows -= 1  # 期首残高行の分を差し引く
        
        # データ行またはダミー行を作成
        for row_idx in range(current_row, current_row + max_rows):
            try:
                detail_idx = row_idx - current_row
                is_last_row = (row_idx == current_row + max_rows - 1)
                
                if detail_idx < data_rows:
                    # 実際のデータ行
                    detail = details[detail_idx]
                    # 日付フォーマット変更: 25.9.12形式
                    date_str = detail.journal_entry.entry_date.strftime('%y.%-m.%-d')
                    
                    cells = [
                        ws.cell(row=row_idx, column=1, value=detail.journal_entry.reference_number or '-'),
                        ws.cell(row=row_idx, column=2, value=date_str),
                        ws.cell(row=row_idx, column=3, value=detail.opposite_account_name),
                        ws.cell(row=row_idx, column=4, value=detail.journal_entry.description),
                        ws.cell(row=row_idx, column=5, value=detail.debit_amount if detail.debit_amount > 0 else 0),
                        ws.cell(row=row_idx, column=6, value=detail.credit_amount if detail.credit_amount > 0 else 0),
                        ws.cell(row=row_idx, column=7, value=detail.running_balance)
                    ]
                else:
                    # 空白行
                    cells = [
                        ws.cell(row=row_idx, column=1, value=''),
                        ws.cell(row=row_idx, column=2, value=''),
                        ws.cell(row=row_idx, column=3, value=''),
                        ws.cell(row=row_idx, column=4, value=''),
                        ws.cell(row=row_idx, column=5, value=''),
                        ws.cell(row=row_idx, column=6, value=''),
                        ws.cell(row=row_idx, column=7, value='')
                    ]
                
                # 各セルに罫線とフォントサイズを設定
                for col, cell in enumerate(cells, 1):
                    # 明細部分のフォントサイズを小さく（明朝体）
                    cell.font = Font(name='ＭＳ 明朝', size=9)  # 明朝体9pt
                    
                    # 数値セルは右揃え、テキストセルは自動縮小設定
                    if col >= 5:  # 借方、貸方、差引金額
                        cell.alignment = Alignment(horizontal='right', shrinkToFit=True)
                        if detail_idx < data_rows:  # データがある場合のみ数値フォーマット
                            cell.number_format = '#,##0'
                    else:
                        # テキストセルは自動縮小を有効化
                        cell.alignment = Alignment(horizontal='left' if col > 2 else 'center', shrinkToFit=True)
                    
                    # 罫線設定（格子状）
                    cell.border = Border(
                        top=dashed_border,  # 横線は破線
                        bottom=thick_border if is_last_row else dashed_border,  # 最終行は太線
                        left=thick_border if col == 1 else solid_border,  # 左端は太線、縦線は実線
                        right=thick_border if col == len(headers) else solid_border  # 右端は太線、縦線は実線
                    )
                
            except Exception as e:
                print(f"Error writing row {row_idx}: {e}")
                continue
        
        # 件数を右下に表示（明細の直後の行）
        count_row = current_row + max_rows + 1
        count_cell = ws.cell(row=count_row, column=7, value=f"件数: {len(details)}件")
        count_cell.font = Font(name='ＭＳ 明朝', bold=True, size=10)
        count_cell.alignment = Alignment(horizontal='right')
        
        # 注記を件数の下の行に左寄せで追加
        note_row = count_row + 1
        note_cell = ws.cell(row=note_row, column=1, value="注)軽印は軽減税率対象　☆印は80%控除対象")
        note_cell.font = Font(name='ＭＳ 明朝', size=9)
        note_cell.alignment = Alignment(horizontal='left')
        
        # 列幅調整（摘要を42に調整）
        column_widths = [6, 7, 12, 42, 9, 9, 9]  # 伝票No.: 6, 日付: 7, 相手科目: 12, 摘要: 42, 借方: 9, 貸方: 9, 差引: 9
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
            # 各列の幅に合わせてテキストを自動縮小
            ws.column_dimensions[get_column_letter(col)].bestFit = True
        
        # レスポンス作成
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"ledger_{account.account_code}_{year}.xlsx"
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        print(f"Excel export error: {e}")
        flash(f'Excelエクスポートでエラーが発生しました: {str(e)}')
        return redirect(url_for('accounting_ledger', account_id=account_id, year=year, month=month))

@app.route('/test_excel')
@login_required
def test_excel():
    """Excel生成テスト"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "テスト"
        
        ws['A1'] = "テストファイル"
        ws['A2'] = "正常に生成されました"
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename="test.xlsx"'
        
        return response
    except Exception as e:
        return f"Error: {e}"

def create_cash_flow_statement(year):
    """キャッシュフロー計算書データ作成"""
    try:
        from types import SimpleNamespace
        
        # 現金科目（現金、預金等）の期首・期末残高を取得
        cash_accounts = ['現金', '当座預金', '普通預金']
        
        # 期首現金残高
        beginning_cash = 0
        for account_name in cash_accounts:
            account = AccountingAccount.query.filter_by(account_name=account_name).first()
            if account:
                opening_balance = OpeningBalance.query.filter_by(
                    fiscal_year=year,
                    account_id=account.id
                ).first()
                if opening_balance:
                    beginning_cash += opening_balance.opening_balance
        
        # 期末現金残高（期首＋当期取引）
        ending_cash = beginning_cash
        for account_name in cash_accounts:
            account = AccountingAccount.query.filter_by(account_name=account_name).first()
            if account:
                transactions = db.session.query(
                    db.func.sum(JournalEntryDetail.debit_amount - JournalEntryDetail.credit_amount)
                ).join(JournalEntry).filter(
                    JournalEntryDetail.account_id == account.id,
                    db.extract('year', JournalEntry.entry_date) == year
                ).scalar() or 0
                ending_cash += transactions
        
        # 当期純利益計算
        revenues = db.session.query(
            db.func.sum(JournalEntryDetail.credit_amount - JournalEntryDetail.debit_amount)
        ).join(JournalEntry).join(AccountingAccount).filter(
            AccountingAccount.account_type == '収益',
            db.extract('year', JournalEntry.entry_date) == year
        ).scalar() or 0
        
        expenses = db.session.query(
            db.func.sum(JournalEntryDetail.debit_amount - JournalEntryDetail.credit_amount)
        ).join(JournalEntry).join(AccountingAccount).filter(
            AccountingAccount.account_type == '費用',
            db.extract('year', JournalEntry.entry_date) == year
        ).scalar() or 0
        
        net_income = revenues - expenses
        
        # 簡易的なキャッシュフロー計算（実際はより複雑な調整が必要）
        depreciation = 0  # 減価償却費（簡易版では0）
        receivables_change = 0  # 売上債権の増減（簡易版では0）
        payables_change = 0  # 仕入債務の増減（簡易版では0）
        
        operating_cf = net_income + depreciation + receivables_change + payables_change
        
        # 投資・財務活動は簡易版では0
        investing_cf = 0
        financing_cf = 0
        
        return SimpleNamespace(
            operating=SimpleNamespace(
                pre_tax_income=net_income,
                depreciation=depreciation,
                receivables_change=receivables_change,
                payables_change=payables_change,
                total=operating_cf
            ),
            investing=SimpleNamespace(
                asset_purchase=0,
                asset_sale=0,
                securities_purchase=0,
                total=investing_cf
            ),
            financing=SimpleNamespace(
                loan_increase=0,
                loan_repayment=0,
                dividend_payment=0,
                total=financing_cf
            ),
            beginning_cash=beginning_cash,
            ending_cash=ending_cash
        )
    except Exception:
        return None

def create_equity_change_statement(year):
    """株主資本等変動計算書データ作成"""
    try:
        from types import SimpleNamespace
        
        # 資本金、剰余金等の期首・期末残高（簡易版）
        capital = 1000000  # 資本金（固定値）
        capital_surplus = 0  # 資本剰余金
        
        # 前期末利益剰余金
        prev_retained_earnings = 0
        
        # 当期純利益計算
        revenues = db.session.query(
            db.func.sum(JournalEntryDetail.credit_amount - JournalEntryDetail.debit_amount)
        ).join(JournalEntry).join(AccountingAccount).filter(
            AccountingAccount.account_type == '収益',
            db.extract('year', JournalEntry.entry_date) == year
        ).scalar() or 0
        
        expenses = db.session.query(
            db.func.sum(JournalEntryDetail.debit_amount - JournalEntryDetail.credit_amount)
        ).join(JournalEntry).join(AccountingAccount).filter(
            AccountingAccount.account_type == '費用',
            db.extract('year', JournalEntry.entry_date) == year
        ).scalar() or 0
        
        net_income = revenues - expenses
        dividend = 0  # 配当金（簡易版では0）
        
        # 当期末利益剰余金
        ending_retained_earnings = prev_retained_earnings + net_income - dividend
        
        return SimpleNamespace(
            beginning=SimpleNamespace(
                capital=capital,
                capital_surplus=capital_surplus,
                retained_earnings=prev_retained_earnings,
                treasury_stock=0,
                total=capital + capital_surplus + prev_retained_earnings
            ),
            changes=SimpleNamespace(
                dividend=-dividend,
                net_income=net_income,
                capital_change=0,
                surplus_change=0,
                earnings_change=net_income - dividend,
                treasury_change=0,
                total=net_income - dividend
            ),
            ending=SimpleNamespace(
                capital=capital,
                capital_surplus=capital_surplus,
                retained_earnings=ending_retained_earnings,
                treasury_stock=0,
                total=capital + capital_surplus + ending_retained_earnings
            )
        )
    except Exception:
        return None

def create_fixed_assets_schedule(year):
    """有形固定資産等明細書データ作成"""
    try:
        # 簡易版：固定資産科目の残高変動を取得
        fixed_asset_types = ['建物', '機械装置', '車両運搬具', '工具器具備品']
        assets = []
        
        for asset_type in fixed_asset_types:
            account = AccountingAccount.query.filter_by(account_name=asset_type).first()
            if account:
                # 期首残高
                opening_balance = OpeningBalance.query.filter_by(
                    fiscal_year=year,
                    account_id=account.id
                ).first()
                beginning_value = opening_balance.opening_balance if opening_balance else 0
                
                # 当期増減
                transactions = db.session.query(
                    db.func.sum(JournalEntryDetail.debit_amount - JournalEntryDetail.credit_amount)
                ).join(JournalEntry).filter(
                    JournalEntryDetail.account_id == account.id,
                    db.extract('year', JournalEntry.entry_date) == year
                ).scalar() or 0
                
                increase = max(transactions, 0)
                decrease = max(-transactions, 0)
                ending_value = beginning_value + increase - decrease
                
                from types import SimpleNamespace
                assets.append(SimpleNamespace(
                    asset_type=asset_type,
                    beginning_book_value=beginning_value,
                    increase=increase,
                    decrease=decrease,
                    ending_book_value=ending_value,
                    accumulated_depreciation=0,  # 簡易版では0
                    acquisition_cost=ending_value
                ))
        
        return assets
    except Exception:
        return []

def create_bonds_schedule(year):
    """社債明細書データ作成"""
    try:
        # 簡易版：社債科目の残高変動を取得
        bonds = []
        account = AccountingAccount.query.filter_by(account_name='社債').first()
        if account:
            opening_balance = OpeningBalance.query.filter_by(
                fiscal_year=year,
                account_id=account.id
            ).first()
            beginning_value = opening_balance.opening_balance if opening_balance else 0
            
            transactions = db.session.query(
                db.func.sum(JournalEntryDetail.credit_amount - JournalEntryDetail.debit_amount)
            ).join(JournalEntry).filter(
                JournalEntryDetail.account_id == account.id,
                db.extract('year', JournalEntry.entry_date) == year
            ).scalar() or 0
            
            ending_value = beginning_value + transactions
            
            if beginning_value != 0 or transactions != 0:
                from types import SimpleNamespace
                bonds.append(SimpleNamespace(
                    name='第1回無担保社債',
                    beginning_balance=beginning_value,
                    change=transactions,
                    ending_balance=ending_value
                ))
        
        return bonds
    except Exception:
        return []

def create_loans_schedule(year):
    """借入金明細書データ作成"""
    try:
        loans = []
        loan_accounts = ['短期借入金', '長期借入金']
        
        for account_name in loan_accounts:
            account = AccountingAccount.query.filter_by(account_name=account_name).first()
            if account:
                opening_balance = OpeningBalance.query.filter_by(
                    fiscal_year=year,
                    account_id=account.id
                ).first()
                beginning_value = opening_balance.opening_balance if opening_balance else 0
                
                transactions = db.session.query(
                    db.func.sum(JournalEntryDetail.credit_amount - JournalEntryDetail.debit_amount)
                ).join(JournalEntry).filter(
                    JournalEntryDetail.account_id == account.id,
                    db.extract('year', JournalEntry.entry_date) == year
                ).scalar() or 0
                
                ending_value = beginning_value + transactions
                
                if beginning_value != 0 or transactions != 0:
                    from types import SimpleNamespace
                    loans.append(SimpleNamespace(
                        lender='○○銀行' if '短期' in account_name else '△△銀行',
                        beginning_balance=beginning_value,
                        change=transactions,
                        ending_balance=ending_value
                    ))
        
        return loans
    except Exception:
        return []

def create_reserves_schedule(year):
    """引当金明細書データ作成"""
    try:
        reserves = []
        reserve_accounts = ['貸倒引当金', '賞与引当金', '退職給付引当金']
        
        for account_name in reserve_accounts:
            account = AccountingAccount.query.filter_by(account_name=account_name).first()
            if account:
                opening_balance = OpeningBalance.query.filter_by(
                    fiscal_year=year,
                    account_id=account.id
                ).first()
                beginning_value = opening_balance.opening_balance if opening_balance else 0
                
                transactions = db.session.query(
                    db.func.sum(JournalEntryDetail.credit_amount - JournalEntryDetail.debit_amount)
                ).join(JournalEntry).filter(
                    JournalEntryDetail.account_id == account.id,
                    db.extract('year', JournalEntry.entry_date) == year
                ).scalar() or 0
                
                ending_value = beginning_value + transactions
                
                if beginning_value != 0 or transactions != 0:
                    from types import SimpleNamespace
                    reserves.append(SimpleNamespace(
                        account_name=account_name,
                        beginning_balance=beginning_value,
                        increase=max(transactions, 0),
                        purpose_decrease=0,  # 簡易版では0
                        other_decrease=max(-transactions, 0),
                        ending_balance=ending_value
                    ))
        
        return reserves
    except Exception:
        return []

def generate_financial_statements_excel(assets, liabilities, revenues, expenses, cash_flow, equity_change,
                                       fixed_assets, bonds, loans, reserves, year, report_type):
    """財務諸表のExcel生成（データベース連動版）"""
    # データベース連動財務諸表生成機能を使用
    try:
        from database_integrated_financial_generator import generate_mixed_orientation_financial_statements_from_db

        # 会社名をサンプルに設定（実際のシステムでは動的に取得）
        company_name = "株式会社サンプル"

        # 実際のデータベースデータを辞書形式に変換
        assets_data = []
        liabilities_data = []
        revenues_data = []
        expenses_data = []

        # 資産データ変換
        for asset in assets:
            assets_data.append({
                'account_name': asset.account_name,
                'total_balance': float(asset.balance) if asset.balance else 0
            })

        # 負債データ変換
        for liability in liabilities:
            liabilities_data.append({
                'account_name': liability.account_name,
                'total_balance': float(liability.balance) if liability.balance else 0
            })

        # 収益データ変換
        for revenue in revenues:
            revenues_data.append({
                'account_name': revenue.account_name,
                'total_amount': float(revenue.balance) if revenue.balance else 0
            })

        # 費用データ変換
        for expense in expenses:
            expenses_data.append({
                'account_name': expense.account_name,
                'total_amount': float(expense.balance) if expense.balance else 0
            })

        # データベース連動Excel生成
        return generate_mixed_orientation_financial_statements_from_db(
            company_name, year, assets_data, liabilities_data, revenues_data, expenses_data
        )

    except ImportError:
        # フォールバック：混合レイアウト版を使用
        try:
            from mixed_orientation_financial_generator import generate_mixed_orientation_financial_statements
            company_name = "株式会社サンプル"
            return generate_mixed_orientation_financial_statements(company_name, year)
        except ImportError:
            # フォールバック：改良版を使用
            try:
                from improved_financial_excel_generator import create_improved_financial_statements_excel
                return create_improved_financial_statements_excel(
                    assets, liabilities, revenues, expenses, cash_flow, equity_change,
                    fixed_assets, bonds, loans, reserves, year, report_type
                )
            except ImportError:
                # 最後のフォールバック：従来版を使用
                from io import BytesIO
                import openpyxl
                from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
                from openpyxl.utils import get_column_letter

                # ワークブック作成
                wb = openpyxl.Workbook()
    
    # スタイル定義
    title_font = Font(name='ＭＳ 明朝', size=16, bold=True)
    header_font = Font(name='ＭＳ 明朝', size=11, bold=True)
    normal_font = Font(name='ＭＳ 明朝', size=10)
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    right_alignment = Alignment(horizontal='right', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    thick_border = Border(
        left=Side(style='thick'), right=Side(style='thick'),
        top=Side(style='thick'), bottom=Side(style='thick')
    )
    
    # 全ての帳票または貸借対照表
    if report_type in ['all', 'balance_sheet']:
        ws_bs = wb.active if report_type == 'balance_sheet' else wb.create_sheet("貸借対照表")
        ws_bs.title = "貸借対照表"
        create_balance_sheet_excel(ws_bs, assets, liabilities, year, title_font, header_font, normal_font, 
                                 center_alignment, left_alignment, right_alignment, thin_border)
    
    # 損益計算書
    if report_type in ['all', 'income_statement']:
        ws_pl = wb.create_sheet("損益計算書")
        create_income_statement_excel(ws_pl, revenues, expenses, year, title_font, header_font, normal_font,
                                    center_alignment, left_alignment, right_alignment, thin_border)
    
    # キャッシュフロー計算書
    if report_type in ['all', 'cash_flow']:
        ws_cf = wb.create_sheet("キャッシュフロー計算書")
        create_cash_flow_excel(ws_cf, cash_flow, year, title_font, header_font, normal_font,
                             center_alignment, left_alignment, right_alignment, thin_border)
    
    # 株主資本等変動計算書
    if report_type in ['all', 'equity_change']:
        ws_eq = wb.create_sheet("株主資本等変動計算書")
        create_equity_change_excel(ws_eq, equity_change, year, title_font, header_font, normal_font,
                                 center_alignment, left_alignment, right_alignment, thin_border)
    
    # 附属明細書
    if report_type in ['all', 'notes']:
        ws_notes = wb.create_sheet("附属明細書")
        create_notes_excel(ws_notes, fixed_assets, bonds, loans, reserves, year, title_font, header_font, normal_font,
                          center_alignment, left_alignment, right_alignment, thin_border)
    
    # デフォルトシートを削除（allの場合）
    if report_type == 'all' and 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # バイナリストリームに保存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

def create_balance_sheet_excel(ws, assets, liabilities, year, title_font, header_font, normal_font,
                              center_alignment, left_alignment, right_alignment, thin_border):
    """貸借対照表シート作成（標準アカウント式フォーマット）"""
    
    # タイトル設定
    ws['A1'] = '貸借対照表'
    ws['A1'].font = title_font
    ws['A1'].alignment = center_alignment
    ws.merge_cells('A1:H1')
    
    # 期間表示
    ws['A3'] = f'{year+1}年 3月31日現在'
    ws['A3'].font = header_font
    ws['A3'].alignment = center_alignment
    ws.merge_cells('A3:H3')
    
    # 単位表示
    ws['H2'] = '(単位：円)'
    ws['H2'].font = normal_font
    ws['H2'].alignment = right_alignment
    
    # 列幅設定（アカウント式レイアウト）
    ws.column_dimensions['A'].width = 5   # 番号列
    ws.column_dimensions['B'].width = 25  # 資産科目
    ws.column_dimensions['C'].width = 15  # 資産金額
    ws.column_dimensions['D'].width = 5   # 空白
    ws.column_dimensions['E'].width = 5   # 番号列
    ws.column_dimensions['F'].width = 25  # 負債・純資産科目
    ws.column_dimensions['G'].width = 15  # 負債・純資産金額
    ws.column_dimensions['H'].width = 5   # 空白
    
    # ヘッダー行設定
    row = 5
    # 左側（資産の部）
    ws[f'B{row}'] = '資産の部'
    ws[f'C{row}'] = '金額'
    # 右側（負債・純資産の部）
    ws[f'F{row}'] = '負債・純資産の部'
    ws[f'G{row}'] = '金額'
    
    # ヘッダー行のスタイル設定
    for col in ['B', 'C', 'F', 'G']:
        ws[f'{col}{row}'].font = header_font
        ws[f'{col}{row}'].alignment = center_alignment
        ws[f'{col}{row}'].border = thin_border
    
    row += 1
    
    # 左側：資産の部
    asset_row = row
    
    # 流動資産
    ws[f'B{asset_row}'] = '流動資産'
    ws[f'B{asset_row}'].font = header_font
    ws[f'B{asset_row}'].alignment = left_alignment
    asset_row += 1
    
    current_assets = 0
    for asset in assets:
        # 流動資産の判定（簡略化）
        if any(keyword in asset.account_name for keyword in ['現金', '預金', '売掛金', '未収金', '在庫', '商品']):
            ws[f'B{asset_row}'] = f'　{asset.account_name}'
            ws[f'C{asset_row}'] = f"{asset.balance:,}" if asset.balance != 0 else "-"
            ws[f'B{asset_row}'].font = normal_font
            ws[f'C{asset_row}'].font = normal_font
            ws[f'C{asset_row}'].alignment = right_alignment
            current_assets += asset.balance
            asset_row += 1
    
    # 流動資産計
    ws[f'B{asset_row}'] = '流動資産計'
    ws[f'C{asset_row}'] = f"{current_assets:,}"
    ws[f'B{asset_row}'].font = header_font
    ws[f'C{asset_row}'].font = header_font
    ws[f'C{asset_row}'].alignment = right_alignment
    for col in ['B', 'C']:
        ws[f'{col}{asset_row}'].border = thin_border
    asset_row += 1
    
    # 固定資産
    ws[f'B{asset_row}'] = '固定資産'
    ws[f'B{asset_row}'].font = header_font
    ws[f'B{asset_row}'].alignment = left_alignment
    asset_row += 1
    
    fixed_assets = 0
    for asset in assets:
        # 固定資産の判定
        if not any(keyword in asset.account_name for keyword in ['現金', '預金', '売掛金', '未収金', '在庫', '商品']):
            ws[f'B{asset_row}'] = f'　{asset.account_name}'
            ws[f'C{asset_row}'] = f"{asset.balance:,}" if asset.balance != 0 else "-"
            ws[f'B{asset_row}'].font = normal_font
            ws[f'C{asset_row}'].font = normal_font
            ws[f'C{asset_row}'].alignment = right_alignment
            fixed_assets += asset.balance
            asset_row += 1
    
    # 固定資産計
    ws[f'B{asset_row}'] = '固定資産計'
    ws[f'C{asset_row}'] = f"{fixed_assets:,}"
    ws[f'B{asset_row}'].font = header_font
    ws[f'C{asset_row}'].font = header_font
    ws[f'C{asset_row}'].alignment = right_alignment
    for col in ['B', 'C']:
        ws[f'{col}{asset_row}'].border = thin_border
    asset_row += 1
    
    # 資産合計
    total_assets = current_assets + fixed_assets
    ws[f'B{asset_row}'] = '資産合計'
    ws[f'C{asset_row}'] = f"{total_assets:,}"
    ws[f'B{asset_row}'].font = title_font
    ws[f'C{asset_row}'].font = title_font
    ws[f'C{asset_row}'].alignment = right_alignment
    for col in ['B', 'C']:
        ws[f'{col}{asset_row}'].border = thin_border
    
    # 右側：負債・純資産の部
    liability_row = row
    
    # 流動負債
    ws[f'F{liability_row}'] = '流動負債'
    ws[f'F{liability_row}'].font = header_font
    ws[f'F{liability_row}'].alignment = left_alignment
    liability_row += 1
    
    current_liabilities = 0
    for liability in liabilities:
        # 流動負債の判定
        if any(keyword in liability.account_name for keyword in ['買掛金', '未払金', '短期借入金', '預り金']):
            ws[f'F{liability_row}'] = f'　{liability.account_name}'
            ws[f'G{liability_row}'] = f"{liability.balance:,}" if liability.balance != 0 else "-"
            ws[f'F{liability_row}'].font = normal_font
            ws[f'G{liability_row}'].font = normal_font
            ws[f'G{liability_row}'].alignment = right_alignment
            current_liabilities += liability.balance
            liability_row += 1
    
    # 流動負債計
    ws[f'F{liability_row}'] = '流動負債計'
    ws[f'G{liability_row}'] = f"{current_liabilities:,}"
    ws[f'F{liability_row}'].font = header_font
    ws[f'G{liability_row}'].font = header_font
    ws[f'G{liability_row}'].alignment = right_alignment
    for col in ['F', 'G']:
        ws[f'{col}{liability_row}'].border = thin_border
    liability_row += 1
    
    # 固定負債
    ws[f'F{liability_row}'] = '固定負債'
    ws[f'F{liability_row}'].font = header_font
    ws[f'F{liability_row}'].alignment = left_alignment
    liability_row += 1
    
    fixed_liabilities = 0
    for liability in liabilities:
        # 固定負債の判定
        if not any(keyword in liability.account_name for keyword in ['買掛金', '未払金', '短期借入金', '預り金']):
            ws[f'F{liability_row}'] = f'　{liability.account_name}'
            ws[f'G{liability_row}'] = f"{liability.balance:,}" if liability.balance != 0 else "-"
            ws[f'F{liability_row}'].font = normal_font
            ws[f'G{liability_row}'].font = normal_font
            ws[f'G{liability_row}'].alignment = right_alignment
            fixed_liabilities += liability.balance
            liability_row += 1
    
    # 固定負債計
    ws[f'F{liability_row}'] = '固定負債計'
    ws[f'G{liability_row}'] = f"{fixed_liabilities:,}"
    ws[f'F{liability_row}'].font = header_font
    ws[f'G{liability_row}'].font = header_font
    ws[f'G{liability_row}'].alignment = right_alignment
    for col in ['F', 'G']:
        ws[f'{col}{liability_row}'].border = thin_border
    liability_row += 1
    
    # 負債合計
    total_liabilities = current_liabilities + fixed_liabilities
    ws[f'F{liability_row}'] = '負債合計'
    ws[f'G{liability_row}'] = f"{total_liabilities:,}"
    ws[f'F{liability_row}'].font = header_font
    ws[f'G{liability_row}'].font = header_font
    ws[f'G{liability_row}'].alignment = right_alignment
    for col in ['F', 'G']:
        ws[f'{col}{liability_row}'].border = thin_border
    liability_row += 1
    
    # 純資産の部
    liability_row += 1
    ws[f'F{liability_row}'] = '純資産の部'
    ws[f'F{liability_row}'].font = header_font
    ws[f'F{liability_row}'].alignment = left_alignment
    liability_row += 1
    
    # 資本金
    capital = 1000000  # 仮の資本金
    ws[f'F{liability_row}'] = '　資本金'
    ws[f'G{liability_row}'] = f"{capital:,}"
    ws[f'F{liability_row}'].font = normal_font
    ws[f'G{liability_row}'].font = normal_font
    ws[f'G{liability_row}'].alignment = right_alignment
    liability_row += 1
    
    # 利益剰余金
    retained_earnings = total_assets - total_liabilities - capital
    ws[f'F{liability_row}'] = '　利益剰余金'
    ws[f'G{liability_row}'] = f"{retained_earnings:,}"
    ws[f'F{liability_row}'].font = normal_font
    ws[f'G{liability_row}'].font = normal_font
    ws[f'G{liability_row}'].alignment = right_alignment
    liability_row += 1
    
    # 純資産合計
    total_equity = capital + retained_earnings
    ws[f'F{liability_row}'] = '純資産合計'
    ws[f'G{liability_row}'] = f"{total_equity:,}"
    ws[f'F{liability_row}'].font = header_font
    ws[f'G{liability_row}'].font = header_font
    ws[f'G{liability_row}'].alignment = right_alignment
    for col in ['F', 'G']:
        ws[f'{col}{liability_row}'].border = thin_border
    liability_row += 1
    
    # 負債・純資産合計
    total_liabilities_equity = total_liabilities + total_equity
    ws[f'F{liability_row}'] = '負債・純資産合計'
    ws[f'G{liability_row}'] = f"{total_liabilities_equity:,}"
    ws[f'F{liability_row}'].font = title_font
    ws[f'G{liability_row}'].font = title_font
    ws[f'G{liability_row}'].alignment = right_alignment
    for col in ['F', 'G']:
        ws[f'{col}{liability_row}'].border = thin_border

def create_income_statement_excel(ws, revenues, expenses, year, title_font, header_font, normal_font,
                                center_alignment, left_alignment, right_alignment, thin_border):
    """損益計算書シート作成（標準日本会計基準フォーマット）"""
    
    # タイトル設定
    ws['A1'] = '損益計算書'
    ws['A1'].font = title_font
    ws['A1'].alignment = center_alignment
    ws.merge_cells('A1:E1')
    
    # 期間表示
    ws['A3'] = f'自　{year}年 4月 1日　至　{year+1}年 3月31日'
    ws['A3'].font = header_font
    ws['A3'].alignment = center_alignment
    ws.merge_cells('A3:E3')
    
    # 単位表示
    ws['E2'] = '(単位：円)'
    ws['E2'].font = normal_font
    ws['E2'].alignment = right_alignment
    
    # 列幅設定
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # ヘッダー行設定
    row = 5
    ws[f'B{row}'] = '科目'
    ws[f'D{row}'] = '注記'
    ws[f'E{row}'] = '金額'
    
    # ヘッダー行のスタイル設定
    for col in ['B', 'D', 'E']:
        ws[f'{col}{row}'].font = header_font
        ws[f'{col}{row}'].alignment = center_alignment
        ws[f'{col}{row}'].border = thin_border
    
    row += 1
    
    # 売上高
    sales_total = 0
    for revenue in revenues:
        if '売上' in revenue.account_name:
            ws[f'B{row}'] = revenue.account_name
            ws[f'E{row}'] = f"{revenue.balance:,}" if revenue.balance != 0 else "-"
            ws[f'B{row}'].font = normal_font
            ws[f'E{row}'].font = normal_font
            ws[f'E{row}'].alignment = right_alignment
            sales_total += revenue.balance
            row += 1
    
    # 売上高計
    if sales_total > 0:
        ws[f'B{row}'] = '売上高'
        ws[f'E{row}'] = f"{sales_total:,}"
        ws[f'B{row}'].font = header_font
        ws[f'E{row}'].font = header_font
        ws[f'E{row}'].alignment = right_alignment
        # 罫線追加
        for col in ['B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].border = thin_border
        row += 1
    
    # 売上原価
    cost_of_sales = 0
    for expense in expenses:
        if any(keyword in expense.account_name for keyword in ['材料費', '商品仕入', '製造', '原価']):
            ws[f'B{row}'] = f'　{expense.account_name}'
            ws[f'E{row}'] = f"{expense.balance:,}" if expense.balance != 0 else "-"
            ws[f'B{row}'].font = normal_font
            ws[f'E{row}'].font = normal_font
            ws[f'E{row}'].alignment = right_alignment
            cost_of_sales += expense.balance
            row += 1
    
    # 売上原価計
    ws[f'B{row}'] = '売上原価'
    ws[f'E{row}'] = f"{cost_of_sales:,}"
    ws[f'B{row}'].font = header_font
    ws[f'E{row}'].font = header_font
    ws[f'E{row}'].alignment = right_alignment
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].border = thin_border
    row += 1
    
    # 売上総利益
    gross_profit = sales_total - cost_of_sales
    ws[f'B{row}'] = '売上総利益'
    ws[f'E{row}'] = f"{gross_profit:,}"
    ws[f'B{row}'].font = header_font
    ws[f'E{row}'].font = header_font
    ws[f'E{row}'].alignment = right_alignment
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].border = thin_border
    row += 1
    
    # 販売費及び一般管理費
    sg_a_expenses = 0
    for expense in expenses:
        if not any(keyword in expense.account_name for keyword in ['材料費', '商品仕入', '製造', '原価', '営業外', '特別']):
            ws[f'B{row}'] = f'　{expense.account_name}'
            ws[f'E{row}'] = f"{expense.balance:,}" if expense.balance != 0 else "-"
            ws[f'B{row}'].font = normal_font
            ws[f'E{row}'].font = normal_font
            ws[f'E{row}'].alignment = right_alignment
            sg_a_expenses += expense.balance
            row += 1
    
    # 販売費及び一般管理費計
    ws[f'B{row}'] = '販売費及び一般管理費'
    ws[f'E{row}'] = f"{sg_a_expenses:,}"
    ws[f'B{row}'].font = header_font
    ws[f'E{row}'].font = header_font
    ws[f'E{row}'].alignment = right_alignment
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].border = thin_border
    row += 1
    
    # 営業利益
    operating_income = gross_profit - sg_a_expenses
    ws[f'B{row}'] = '営業利益'
    ws[f'E{row}'] = f"{operating_income:,}"
    ws[f'B{row}'].font = header_font
    ws[f'E{row}'].font = header_font
    ws[f'E{row}'].alignment = right_alignment
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].border = thin_border
    row += 1
    
    # 営業外収益
    non_operating_income = 0
    for revenue in revenues:
        if '売上' not in revenue.account_name:
            ws[f'B{row}'] = f'　{revenue.account_name}'
            ws[f'E{row}'] = f"{revenue.balance:,}" if revenue.balance != 0 else "-"
            ws[f'B{row}'].font = normal_font
            ws[f'E{row}'].font = normal_font
            ws[f'E{row}'].alignment = right_alignment
            non_operating_income += revenue.balance
            row += 1
    
    # 営業外費用
    non_operating_expense = 0
    for expense in expenses:
        if '営業外' in expense.account_name:
            ws[f'B{row}'] = f'　{expense.account_name}'
            ws[f'E{row}'] = f"{expense.balance:,}" if expense.balance != 0 else "-"
            ws[f'B{row}'].font = normal_font
            ws[f'E{row}'].font = normal_font
            ws[f'E{row}'].alignment = right_alignment
            non_operating_expense += expense.balance
            row += 1
    
    # 経常利益
    ordinary_income = operating_income + non_operating_income - non_operating_expense
    ws[f'B{row}'] = '経常利益'
    ws[f'E{row}'] = f"{ordinary_income:,}"
    ws[f'B{row}'].font = header_font
    ws[f'E{row}'].font = header_font
    ws[f'E{row}'].alignment = right_alignment
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].border = thin_border
    row += 1
    
    # 特別利益・特別損失（必要に応じて追加）
    
    # 税引前当期純利益
    ws[f'B{row}'] = '税引前当期純利益'
    ws[f'E{row}'] = f"{ordinary_income:,}"
    ws[f'B{row}'].font = header_font
    ws[f'E{row}'].font = header_font
    ws[f'E{row}'].alignment = right_alignment
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].border = thin_border
    row += 1
    
    # 法人税等（簡略化）
    tax_rate = 0.3  # 仮定
    tax_expense = ordinary_income * tax_rate
    ws[f'B{row}'] = '法人税等'
    ws[f'E{row}'] = f"{tax_expense:,.0f}"
    ws[f'B{row}'].font = normal_font
    ws[f'E{row}'].font = normal_font
    ws[f'E{row}'].alignment = right_alignment
    row += 1
    
    # 当期純利益
    net_income = ordinary_income - tax_expense
    ws[f'B{row}'] = '当期純利益'
    ws[f'E{row}'] = f"{net_income:,.0f}"
    ws[f'B{row}'].font = title_font
    ws[f'E{row}'].font = title_font
    ws[f'E{row}'].alignment = right_alignment
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].border = thin_border
    
    # 罫線適用
    for row_num in range(1, row + 1):
        for col_num in range(1, 4):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border

def create_cash_flow_excel(ws, cash_flow, year, title_font, header_font, normal_font,
                          center_alignment, left_alignment, right_alignment, thin_border):
    """キャッシュフロー計算書シート作成"""
    
    # タイトル
    ws['A1'] = 'キャッシュフロー計算書'
    ws['A1'].font = title_font
    ws['A1'].alignment = center_alignment
    ws.merge_cells('A1:C1')
    
    # 期間表示
    ws['A2'] = f'{year}年4月1日から{year+1}年3月31日まで'
    ws['A2'].font = header_font
    ws['A2'].alignment = center_alignment
    ws.merge_cells('A2:C2')
    
    # 列幅設定
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    
    row = 4
    
    if cash_flow:
        # 営業活動によるキャッシュフロー
        ws[f'A{row}'] = '営業活動によるキャッシュフロー'
        ws[f'A{row}'].font = header_font
        row += 1
        
        ws[f'A{row}'] = '  税引前当期純利益'
        ws[f'B{row}'] = f"{cash_flow.operating.pre_tax_income:,}"
        ws[f'A{row}'].font = normal_font
        ws[f'B{row}'].font = normal_font
        ws[f'B{row}'].alignment = right_alignment
        row += 1
        
        ws[f'A{row}'] = '  減価償却費'
        ws[f'B{row}'] = f"{cash_flow.operating.depreciation:,}"
        ws[f'A{row}'].font = normal_font
        ws[f'B{row}'].font = normal_font
        ws[f'B{row}'].alignment = right_alignment
        row += 1
        
        ws[f'A{row}'] = '営業活動によるキャッシュフロー'
        ws[f'B{row}'] = f"{cash_flow.operating.total:,}"
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].font = header_font
        ws[f'B{row}'].alignment = right_alignment
        row += 2
        
        # 投資活動によるキャッシュフロー
        ws[f'A{row}'] = '投資活動によるキャッシュフロー'
        ws[f'A{row}'].font = header_font
        row += 1
        
        ws[f'A{row}'] = '投資活動によるキャッシュフロー'
        ws[f'B{row}'] = f"{cash_flow.investing.total:,}"
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].font = header_font
        ws[f'B{row}'].alignment = right_alignment
        row += 2
        
        # 財務活動によるキャッシュフロー
        ws[f'A{row}'] = '財務活動によるキャッシュフロー'
        ws[f'A{row}'].font = header_font
        row += 1
        
        ws[f'A{row}'] = '財務活動によるキャッシュフロー'
        ws[f'B{row}'] = f"{cash_flow.financing.total:,}"
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].font = header_font
        ws[f'B{row}'].alignment = right_alignment
        row += 2
        
        # 現金および現金同等物の増減額
        cash_change = cash_flow.operating.total + cash_flow.investing.total + cash_flow.financing.total
        ws[f'A{row}'] = '現金および現金同等物の増減額'
        ws[f'B{row}'] = f"{cash_change:,}"
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].font = header_font
        ws[f'B{row}'].alignment = right_alignment
        row += 1
        
        # 期首現金および現金同等物
        ws[f'A{row}'] = '期首現金および現金同等物'
        ws[f'B{row}'] = f"{cash_flow.beginning_cash:,}"
        ws[f'A{row}'].font = normal_font
        ws[f'B{row}'].font = normal_font
        ws[f'B{row}'].alignment = right_alignment
        row += 1
        
        # 期末現金および現金同等物
        ws[f'A{row}'] = '期末現金および現金同等物'
        ws[f'B{row}'] = f"{cash_flow.ending_cash:,}"
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].font = header_font
        ws[f'B{row}'].alignment = right_alignment
    else:
        ws[f'A{row}'] = 'データがありません'
        ws[f'A{row}'].font = normal_font
    
    # 罫線適用
    for row_num in range(1, row + 1):
        for col_num in range(1, 3):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border

def create_equity_change_excel(ws, equity_change, year, title_font, header_font, normal_font,
                              center_alignment, left_alignment, right_alignment, thin_border):
    """株主資本等変動計算書シート作成"""
    
    # タイトル
    ws['A1'] = '株主資本等変動計算書'
    ws['A1'].font = title_font
    ws['A1'].alignment = center_alignment
    ws.merge_cells('A1:F1')
    
    # 期間表示
    ws['A2'] = f'{year}年4月1日から{year+1}年3月31日まで'
    ws['A2'].font = header_font
    ws['A2'].alignment = center_alignment
    ws.merge_cells('A2:F2')
    
    # 列幅設定
    ws.column_dimensions['A'].width = 20
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 12
    
    row = 4
    
    # ヘッダー
    headers = ['', '資本金', '資本剰余金', '利益剰余金', '自己株式', '株主資本合計']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    row += 1
    
    if equity_change:
        # 当期首残高
        ws[f'A{row}'] = '当期首残高'
        ws[f'B{row}'] = f"{equity_change.beginning.capital:,}"
        ws[f'C{row}'] = f"{equity_change.beginning.capital_surplus:,}"
        ws[f'D{row}'] = f"{equity_change.beginning.retained_earnings:,}"
        ws[f'E{row}'] = f"{equity_change.beginning.treasury_stock:,}"
        ws[f'F{row}'] = f"{equity_change.beginning.total:,}"
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.font = normal_font
            cell.alignment = right_alignment if col > 1 else left_alignment
            cell.border = thin_border
        row += 1
        
        # 当期純利益
        ws[f'A{row}'] = '当期純利益'
        ws[f'D{row}'] = f"{equity_change.changes.net_income:,}"
        ws[f'F{row}'] = f"{equity_change.changes.net_income:,}"
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.font = normal_font
            cell.alignment = right_alignment if col > 1 else left_alignment
            cell.border = thin_border
        row += 1
        
        # 当期末残高
        ws[f'A{row}'] = '当期末残高'
        ws[f'B{row}'] = f"{equity_change.ending.capital:,}"
        ws[f'C{row}'] = f"{equity_change.ending.capital_surplus:,}"
        ws[f'D{row}'] = f"{equity_change.ending.retained_earnings:,}"
        ws[f'E{row}'] = f"{equity_change.ending.treasury_stock:,}"
        ws[f'F{row}'] = f"{equity_change.ending.total:,}"
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.alignment = right_alignment if col > 1 else left_alignment
            cell.border = thin_border
    else:
        ws[f'A{row}'] = 'データがありません'
        ws[f'A{row}'].font = normal_font

def create_notes_excel(ws, fixed_assets, bonds, loans, reserves, year, title_font, header_font, normal_font,
                      center_alignment, left_alignment, right_alignment, thin_border):
    """附属明細書シート作成"""
    
    # タイトル
    ws['A1'] = '附属明細書'
    ws['A1'].font = title_font
    ws['A1'].alignment = center_alignment
    ws.merge_cells('A1:G1')
    
    # 期間表示
    ws['A2'] = f'{year}年4月1日から{year+1}年3月31日まで'
    ws['A2'].font = header_font
    ws['A2'].alignment = center_alignment
    ws.merge_cells('A2:G2')
    
    row = 4
    
    # 有形固定資産等明細書
    ws[f'A{row}'] = '有形固定資産等明細書'
    ws[f'A{row}'].font = header_font
    row += 1
    
    # ヘッダー
    headers = ['資産の種類', '期首帳簿価額', '当期増加額', '当期減少額', '期末帳簿価額', '減価償却累計額', '期末取得価額']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    row += 1
    
    if fixed_assets:
        for asset in fixed_assets:
            ws[f'A{row}'] = asset.asset_type
            ws[f'B{row}'] = f"{asset.beginning_book_value:,}"
            ws[f'C{row}'] = f"{asset.increase:,}"
            ws[f'D{row}'] = f"{asset.decrease:,}"
            ws[f'E{row}'] = f"{asset.ending_book_value:,}"
            ws[f'F{row}'] = f"{asset.accumulated_depreciation:,}"
            ws[f'G{row}'] = f"{asset.acquisition_cost:,}"
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.font = normal_font
                cell.alignment = right_alignment if col > 1 else left_alignment
                cell.border = thin_border
            row += 1
    else:
        ws[f'A{row}'] = 'データがありません'
        ws[f'A{row}'].font = normal_font
        row += 1
    
    row += 2
    
    # 借入金明細書
    ws[f'A{row}'] = '借入金明細書'
    ws[f'A{row}'].font = header_font
    row += 1
    
    # ヘッダー
    headers = ['借入先', '期首残高', '当期増減', '期末残高']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    row += 1
    
    if loans:
        for loan in loans:
            ws[f'A{row}'] = loan.lender
            ws[f'B{row}'] = f"{loan.beginning_balance:,}"
            ws[f'C{row}'] = f"{loan.change:,}"
            ws[f'D{row}'] = f"{loan.ending_balance:,}"
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.font = normal_font
                cell.alignment = right_alignment if col > 1 else left_alignment
                cell.border = thin_border
            row += 1
    else:
        ws[f'A{row}'] = 'データがありません'
        ws[f'A{row}'].font = normal_font
    
    # 列幅設定
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col_letter].width = 15

@app.route('/financial_statements')
@login_required
def financial_statements():
    """財務諸表画面"""
    if current_user.role != 'accounting':
        flash('アクセス権限がありません。')
        return redirect(url_for('index'))
    
    year = request.args.get('year', type=int, default=datetime.now().year)
    
    # 簡易的な財務諸表データを作成
    # 実際の実装では、より複雑な集計処理が必要
    
    # 資産科目の残高
    assets = db.session.query(
        AccountingAccount.account_name,
        db.func.sum(JournalEntryDetail.debit_amount - JournalEntryDetail.credit_amount).label('balance')
    ).join(JournalEntryDetail).join(JournalEntry).filter(
        AccountingAccount.account_type == '資産',
        db.extract('year', JournalEntry.entry_date) == year
    ).group_by(AccountingAccount.id, AccountingAccount.account_name).all()
    
    # 負債科目の残高
    liabilities = db.session.query(
        AccountingAccount.account_name,
        db.func.sum(JournalEntryDetail.credit_amount - JournalEntryDetail.debit_amount).label('balance')
    ).join(JournalEntryDetail).join(JournalEntry).filter(
        AccountingAccount.account_type == '負債',
        db.extract('year', JournalEntry.entry_date) == year
    ).group_by(AccountingAccount.id, AccountingAccount.account_name).all()
    
    # 収益科目の残高
    revenues = db.session.query(
        AccountingAccount.account_name,
        db.func.sum(JournalEntryDetail.credit_amount - JournalEntryDetail.debit_amount).label('balance')
    ).join(JournalEntryDetail).join(JournalEntry).filter(
        AccountingAccount.account_type == '収益',
        db.extract('year', JournalEntry.entry_date) == year
    ).group_by(AccountingAccount.id, AccountingAccount.account_name).all()
    
    # 費用科目の残高
    expenses = db.session.query(
        AccountingAccount.account_name,
        db.func.sum(JournalEntryDetail.debit_amount - JournalEntryDetail.credit_amount).label('balance')
    ).join(JournalEntryDetail).join(JournalEntry).filter(
        AccountingAccount.account_type == '費用',
        db.extract('year', JournalEntry.entry_date) == year
    ).group_by(AccountingAccount.id, AccountingAccount.account_name).all()
    
    years = list(range(datetime.now().year - 2, datetime.now().year + 2))
    
    # キャッシュフロー計算書データ作成
    cash_flow = create_cash_flow_statement(year)
    
    # 株主資本等変動計算書データ作成
    equity_change = create_equity_change_statement(year)
    
    # 附属明細書データ作成
    fixed_assets = create_fixed_assets_schedule(year)
    bonds = create_bonds_schedule(year)
    loans = create_loans_schedule(year)
    reserves = create_reserves_schedule(year)
    
    return render_template('financial_statements.html',
                         assets=assets,
                         liabilities=liabilities, 
                         revenues=revenues,
                         expenses=expenses,
                         cash_flow=cash_flow,
                         equity_change=equity_change,
                         fixed_assets=fixed_assets,
                         bonds=bonds,
                         loans=loans,
                         reserves=reserves,
                         selected_year=year,
                         years=years)

@app.route('/export_financial_statements_excel')
@login_required
def export_financial_statements_excel():
    """財務諸表のExcelエクスポート"""
    if current_user.role != 'accounting':
        flash('アクセス権限がありません。')
        return redirect(url_for('index'))
    
    try:
        year = request.args.get('year', type=int, default=datetime.now().year)
        report_type = request.args.get('type', default='all')  # all, balance_sheet, income_statement, cash_flow, equity_change, notes
        
        # 財務諸表データを再取得
        # 資産科目の残高
        assets = db.session.query(
            AccountingAccount.account_name,
            db.func.sum(JournalEntryDetail.debit_amount - JournalEntryDetail.credit_amount).label('balance')
        ).join(JournalEntryDetail).join(JournalEntry).filter(
            AccountingAccount.account_type == '資産',
            db.extract('year', JournalEntry.entry_date) == year
        ).group_by(AccountingAccount.id, AccountingAccount.account_name).all()
        
        # 負債科目の残高
        liabilities = db.session.query(
            AccountingAccount.account_name,
            db.func.sum(JournalEntryDetail.credit_amount - JournalEntryDetail.debit_amount).label('balance')
        ).join(JournalEntryDetail).join(JournalEntry).filter(
            AccountingAccount.account_type == '負債',
            db.extract('year', JournalEntry.entry_date) == year
        ).group_by(AccountingAccount.id, AccountingAccount.account_name).all()
        
        # 収益科目の残高
        revenues = db.session.query(
            AccountingAccount.account_name,
            db.func.sum(JournalEntryDetail.credit_amount - JournalEntryDetail.debit_amount).label('balance')
        ).join(JournalEntryDetail).join(JournalEntry).filter(
            AccountingAccount.account_type == '収益',
            db.extract('year', JournalEntry.entry_date) == year
        ).group_by(AccountingAccount.id, AccountingAccount.account_name).all()
        
        # 費用科目の残高
        expenses = db.session.query(
            AccountingAccount.account_name,
            db.func.sum(JournalEntryDetail.debit_amount - JournalEntryDetail.credit_amount).label('balance')
        ).join(JournalEntry).filter(
            AccountingAccount.account_type == '費用',
            db.extract('year', JournalEntry.entry_date) == year
        ).group_by(AccountingAccount.id, AccountingAccount.account_name).all()
        
        # 新しい財務諸表データを作成
        cash_flow = create_cash_flow_statement(year)
        equity_change = create_equity_change_statement(year)
        fixed_assets = create_fixed_assets_schedule(year)
        bonds = create_bonds_schedule(year)
        loans = create_loans_schedule(year)
        reserves = create_reserves_schedule(year)
        
        # Excel生成
        output = generate_financial_statements_excel(
            assets, liabilities, revenues, expenses, cash_flow, equity_change,
            fixed_assets, bonds, loans, reserves, year, report_type
        )
        
        # レスポンス作成
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        filename = f"financial_statements_{year}.xlsx"
        # Unicode文字対応
        from urllib.parse import quote
        filename_utf8 = quote(filename.encode('utf-8'))
        response.headers['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{filename_utf8}'
        
        return response
        
    except Exception as e:
        flash(f'Excel出力でエラーが発生しました: {str(e)}')
        return redirect(url_for('financial_statements'))

# 勘定科目管理
@app.route('/account_management')
@login_required
def account_management():
    """勘定科目管理画面"""
    if current_user.role != 'accounting':
        flash('アクセス権限がありません。')
        return redirect(url_for('index'))
    
    accounts = AccountingAccount.query.order_by(AccountingAccount.account_code).all()
    return render_template('account_management.html', accounts=accounts)

@app.route('/create_account', methods=['POST'])
@login_required
def create_account():
    """新規勘定科目作成"""
    if current_user.role != 'accounting':
        flash('アクセス権限がありません。')
        return redirect(url_for('index'))
    
    try:
        account_code = request.form.get('account_code')
        account_name = request.form.get('account_name')
        account_type = request.form.get('account_type')
        bank_name = request.form.get('bank_name')
        branch_name = request.form.get('branch_name')
        
        # 勘定科目コードの重複チェック
        existing_account = AccountingAccount.query.filter_by(account_code=account_code).first()
        if existing_account:
            flash('この勘定科目コードは既に使用されています。')
            return redirect(url_for('account_management'))
        
        # 新規勘定科目を作成
        new_account = AccountingAccount(
            account_code=account_code,
            account_name=account_name,
            account_type=account_type,
            bank_name=bank_name if bank_name else None,
            branch_name=branch_name if branch_name else None,
            is_active=True
        )
        
        db.session.add(new_account)
        db.session.commit()
        
        flash(f'勘定科目「{account_name}」を作成しました。')
        return redirect(url_for('account_management'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'勘定科目の作成に失敗しました: {str(e)}')
        return redirect(url_for('account_management'))

@app.route('/edit_account/<int:account_id>', methods=['POST'])
@login_required
def edit_account(account_id):
    """勘定科目編集"""
    if current_user.role != 'accounting':
        flash('アクセス権限がありません。')
        return redirect(url_for('index'))
    
    try:
        account = AccountingAccount.query.get_or_404(account_id)
        
        account.account_name = request.form.get('account_name')
        account.account_type = request.form.get('account_type')
        account.bank_name = request.form.get('bank_name') if request.form.get('bank_name') else None
        account.branch_name = request.form.get('branch_name') if request.form.get('branch_name') else None
        account.is_active = request.form.get('is_active') == 'on'
        
        db.session.commit()
        
        flash(f'勘定科目「{account.account_name}」を更新しました。')
        return redirect(url_for('account_management'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'勘定科目の更新に失敗しました: {str(e)}')
        return redirect(url_for('account_management'))

@app.route('/delete_account/<int:account_id>', methods=['POST'])
@login_required
def delete_account(account_id):
    """勘定科目削除"""
    if current_user.role != 'accounting':
        flash('アクセス権限がありません。')
        return redirect(url_for('index'))
    
    try:
        account = AccountingAccount.query.get_or_404(account_id)
        
        # 仕訳で使用されているかチェック
        used_in_journal = JournalEntryDetail.query.filter_by(account_id=account_id).first()
        if used_in_journal:
            flash('この勘定科目は仕訳で使用されているため削除できません。')
            return redirect(url_for('account_management'))
        
        db.session.delete(account)
        db.session.commit()
        
        flash(f'勘定科目「{account.account_name}」を削除しました。')
        return redirect(url_for('account_management'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'勘定科目の削除に失敗しました: {str(e)}')
        return redirect(url_for('account_management'))

# 簡単仕訳入力
@app.route('/simple_journal_entry')
@login_required
def simple_journal_entry():
    """簡単仕訳入力画面"""
    if current_user.role != 'accounting':
        flash('アクセス権限がありません。')
        return redirect(url_for('index'))
    
    # 取引パターンをカテゴリ別に取得
    patterns = TransactionPattern.query.filter_by(is_active=True).order_by(TransactionPattern.pattern_name).all()
    
    # カテゴリの表示順序を定義
    category_order = [
        '資産関係',
        '経費関係', 
        '仕入関係',
        '売上関係',
        '給与・報酬',
        '負債関係',
        'その他'
    ]
    
    # カテゴリ別にグループ化
    patterns_by_category = {}
    for pattern in patterns:
        if pattern.category not in patterns_by_category:
            patterns_by_category[pattern.category] = []
        patterns_by_category[pattern.category].append(pattern)
    
    # 定義された順序で辞書を作成
    ordered_patterns_by_category = {}
    for category in category_order:
        if category in patterns_by_category:
            ordered_patterns_by_category[category] = patterns_by_category[category]
    
    # 現金・預金口座を取得
    cash_accounts = AccountingAccount.query.filter(
        AccountingAccount.account_type == '資産',
        AccountingAccount.account_code.in_(['101', '102', '103', '104', '105', '106', '107', '108', '109', '110', '113']),
        AccountingAccount.is_active == True
    ).order_by(AccountingAccount.account_code).all()
    
    # 最近の仕訳を取得
    recent_entries = JournalEntry.query.order_by(JournalEntry.created_at.desc()).limit(5).all()
    
    # 取引先を取得
    partners = BusinessPartner.query.filter_by(is_active=True).order_by(BusinessPartner.partner_name).all()
    
    return render_template('simple_journal_entry.html', 
                         patterns_by_category=ordered_patterns_by_category,
                         cash_accounts=cash_accounts,
                         recent_entries=recent_entries,
                         partners=partners,
                         today=date.today())

@app.route('/create_simple_journal_entry', methods=['POST'])
@login_required
def create_simple_journal_entry():
    """簡単仕訳入力からの仕訳登録"""
    if current_user.role != 'accounting':
        flash('アクセス権限がありません。')
        return redirect(url_for('index'))
    
    try:
        pattern_id = request.form.get('pattern_id')
        amount = int(request.form.get('amount'))
        cash_account_id = request.form.get('cash_account_id')
        partner_id = request.form.get('partner_id') if request.form.get('partner_id') else None
        description = request.form.get('description', '')
        entry_date = datetime.strptime(request.form.get('entry_date'), '%Y-%m-%d').date()
        reference_number = request.form.get('reference_number', '')
        
        # 取引パターンを取得
        pattern = TransactionPattern.query.get_or_404(pattern_id)
        
        # 現金・預金口座を取得
        cash_account = AccountingAccount.query.get_or_404(cash_account_id)
        
        # 取引先情報を取得
        partner_name = ''
        if partner_id:
            partner = BusinessPartner.query.get(partner_id)
            if partner:
                partner_name = partner.partner_name
        
        # 摘要を生成
        if description:
            full_description = f"{pattern.pattern_name}（{description}）"
        else:
            full_description = pattern.pattern_name
            
        if partner_name:
            full_description += f" - {partner_name}"
        
        # 仕訳エントリを作成
        journal_entry = JournalEntry(
            entry_date=entry_date,
            description=full_description,
            reference_number=reference_number if reference_number else None,
            total_amount=amount,
            partner_id=partner_id,
            created_by=current_user.id
        )
        
        db.session.add(journal_entry)
        db.session.flush()  # IDを取得するためのflush
        
        # 仕訳明細を作成
        details_to_add = []
        
        if pattern.main_account_side == 'cash_debit':
            # 現金・預金が借方の場合（入金）
            cash_detail = JournalEntryDetail(
                journal_entry_id=journal_entry.id,
                account_id=cash_account.id,
                debit_amount=amount,
                credit_amount=0
            )
            details_to_add.append(cash_detail)
            
            # 相手科目を取得
            if pattern.credit_account_code:
                credit_account = AccountingAccount.query.filter_by(account_code=pattern.credit_account_code).first()
                if credit_account:
                    opposite_detail = JournalEntryDetail(
                        journal_entry_id=journal_entry.id,
                        account_id=credit_account.id,
                        debit_amount=0,
                        credit_amount=amount
                    )
                    details_to_add.append(opposite_detail)
            
        else:
            # 現金・預金が貸方の場合（出金）
            cash_detail = JournalEntryDetail(
                journal_entry_id=journal_entry.id,
                account_id=cash_account.id,
                debit_amount=0,
                credit_amount=amount
            )
            details_to_add.append(cash_detail)
            
            # 相手科目を取得
            if pattern.debit_account_code:
                debit_account = AccountingAccount.query.filter_by(account_code=pattern.debit_account_code).first()
                if debit_account:
                    opposite_detail = JournalEntryDetail(
                        journal_entry_id=journal_entry.id,
                        account_id=debit_account.id,
                        debit_amount=amount,
                        credit_amount=0
                    )
                    details_to_add.append(opposite_detail)
        
        # 全ての明細を追加
        for detail in details_to_add:
            db.session.add(detail)
        db.session.commit()
        
        flash(f'仕訳「{full_description}」を登録しました。')
        return redirect(url_for('simple_journal_entry'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'仕訳の登録に失敗しました: {str(e)}')
        return redirect(url_for('simple_journal_entry'))

# === 取引先管理機能 ===

@app.route('/partner_management')
@login_required
def partner_management():
    """取引先管理画面"""
    if current_user.role != 'accounting':
        flash('アクセス権限がありません。')
        return redirect(url_for('index'))
    
    partners = BusinessPartner.query.filter_by(is_active=True).order_by(BusinessPartner.partner_code).all()
    return render_template('partner_management.html', partners=partners)

@app.route('/create_partner', methods=['POST'])
@login_required
def create_partner():
    """取引先新規登録"""
    if current_user.role != 'accounting':
        flash('アクセス権限がありません。')
        return redirect(url_for('index'))
    
    try:
        partner = BusinessPartner(
            partner_code=request.form.get('partner_code'),
            partner_name=request.form.get('partner_name'),
            partner_type=request.form.get('partner_type'),
            postal_code=request.form.get('postal_code'),
            address=request.form.get('address'),
            phone=request.form.get('phone'),
            fax=request.form.get('fax'),
            email=request.form.get('email'),
            contact_person=request.form.get('contact_person'),
            notes=request.form.get('notes')
        )
        
        db.session.add(partner)
        db.session.commit()
        
        flash(f'取引先「{partner.partner_name}」を登録しました。')
        
    except Exception as e:
        db.session.rollback()
        flash(f'取引先の登録に失敗しました: {str(e)}')
    
    return redirect(url_for('partner_management'))

@app.route('/edit_partner/<int:partner_id>', methods=['POST'])
@login_required
def edit_partner(partner_id):
    """取引先編集"""
    if current_user.role != 'accounting':
        flash('アクセス権限がありません。')
        return redirect(url_for('index'))
    
    try:
        partner = BusinessPartner.query.get_or_404(partner_id)
        
        partner.partner_code = request.form.get('partner_code')
        partner.partner_name = request.form.get('partner_name')
        partner.partner_type = request.form.get('partner_type')
        partner.postal_code = request.form.get('postal_code')
        partner.address = request.form.get('address')
        partner.phone = request.form.get('phone')
        partner.fax = request.form.get('fax')
        partner.email = request.form.get('email')
        partner.contact_person = request.form.get('contact_person')
        partner.notes = request.form.get('notes')
        
        db.session.commit()
        
        flash(f'取引先「{partner.partner_name}」を更新しました。')
        
    except Exception as e:
        db.session.rollback()
        flash(f'取引先の更新に失敗しました: {str(e)}')
    
    return redirect(url_for('partner_management'))

@app.route('/delete_partner/<int:partner_id>', methods=['POST'])
@login_required
def delete_partner(partner_id):
    """取引先削除（無効化）"""
    if current_user.role != 'accounting':
        flash('アクセス権限がありません。')
        return redirect(url_for('index'))
    
    try:
        partner = BusinessPartner.query.get_or_404(partner_id)
        partner.is_active = False
        
        db.session.commit()
        
        flash(f'取引先「{partner.partner_name}」を削除しました。')
        
    except Exception as e:
        db.session.rollback()
        flash(f'取引先の削除に失敗しました: {str(e)}')
    
    return redirect(url_for('partner_management'))

# --- タイムカード発行 ---

@app.route('/timecard_issuance')
@login_required
def timecard_issuance():
    """タイムカード発行画面"""
    if current_user.role != 'accounting':
        flash('アクセス権限がありません。')
        return redirect(url_for('index'))

    # 従業員一覧取得
    employees = Employee.query.filter_by(is_active=True).order_by(Employee.employee_id).all()

    # 年月選択用のデータ
    current_date = datetime.now()
    years = list(range(current_date.year - 2, current_date.year + 2))
    months = list(range(1, 13))

    return render_template('timecard_issuance.html',
                         employees=employees,
                         years=years,
                         months=months,
                         current_year=current_date.year,
                         current_month=current_date.month)

@app.route('/generate_timecard_pdf', methods=['POST'])
@login_required
def generate_timecard_pdf():
    """タイムカードPDF生成"""
    if current_user.role != 'accounting':
        flash('アクセス権限がありません。')
        return redirect(url_for('index'))

    try:
        employee_id = request.form.get('employee_id')
        year = int(request.form.get('year'))
        month = int(request.form.get('month'))

        # 従業員情報取得
        employee = Employee.query.get_or_404(employee_id)

        # 指定月の労働時間データ取得
        working_time_records = WorkingTimeRecord.query.filter(
            WorkingTimeRecord.employee_id == employee_id,
            WorkingTimeRecord.year == year,
            WorkingTimeRecord.month == month
        ).order_by(WorkingTimeRecord.day).all()

        # PDFレスポンス生成
        response = make_response(create_timecard_pdf(employee, working_time_records, year, month))
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'inline; filename="timecard_{employee.employee_id}_{year}{month:02d}.pdf"'

        return response

    except Exception as e:
        flash(f'タイムカードの生成に失敗しました: {str(e)}')
        return redirect(url_for('timecard_issuance'))

# --- 起動と初期設定のためのコマンド ---

if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5001))
    debug = os.environ.get('FLASK_ENV') != 'production'
    app.run(debug=debug, host='0.0.0.0', port=port)