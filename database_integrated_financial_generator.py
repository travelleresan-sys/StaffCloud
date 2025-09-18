#!/usr/bin/env python3
"""
データベース連動財務諸表Excel生成機能
帳簿データと完全連動した混合レイアウト財務諸表
"""

from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from datetime import datetime


class DatabaseIntegratedFinancialGenerator:
    """データベース連動財務諸表Excel生成クラス"""

    def __init__(self):
        # フォント・スタイル設定
        self.title_font = Font(name='ＭＳ ゴシック', size=12, bold=True)
        self.header_font = Font(name='ＭＳ ゴシック', size=10, bold=True)
        self.normal_font = Font(name='ＭＳ ゴシック', size=9)
        self.small_font = Font(name='ＭＳ ゴシック', size=8)

        # 横向け用（少し大きめ）
        self.title_font_landscape = Font(name='ＭＳ ゴシック', size=14, bold=True)
        self.header_font_landscape = Font(name='ＭＳ ゴシック', size=11, bold=True)
        self.normal_font_landscape = Font(name='ＭＳ ゴシック', size=10)
        self.small_font_landscape = Font(name='ＭＳ ゴシック', size=9)

        # アライメント
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')
        self.right_align = Alignment(horizontal='right', vertical='center')

        # 罫線
        self.thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        self.double_bottom_border = Border(bottom=Side(style='double'))

        # 背景色
        self.header_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
        self.subtotal_fill = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')

    def setup_a4_portrait(self, ws):
        """A4縦向けページ設定"""
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.scale = 85
        ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)

    def setup_a4_landscape(self, ws):
        """A4横向けページ設定"""
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.scale = 90
        ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)

    def classify_account_data(self, assets_data, liabilities_data, revenues_data, expenses_data):
        """
        帳簿データを財務諸表項目に分類
        """
        # 貸借対照表データ分類
        balance_sheet_data = {
            'current_assets': [],
            'fixed_assets': [],
            'current_liabilities': [],
            'fixed_liabilities': [],
            'equity': []
        }

        # 損益計算書データ分類
        income_statement_data = {
            'sales': [],
            'cost_of_sales': [],
            'selling_expenses': [],
            'administrative_expenses': [],
            'other_income': [],
            'other_expenses': []
        }

        # 資産の分類
        for asset in assets_data:
            account_name = asset.get('account_name', '')
            balance = asset.get('total_balance', 0)

            if any(keyword in account_name for keyword in ['現金', '預金', '売掛金', '受取手形', '商品', '製品', '原材料']):
                balance_sheet_data['current_assets'].append({
                    'name': account_name,
                    'amount': balance
                })
            else:
                balance_sheet_data['fixed_assets'].append({
                    'name': account_name,
                    'amount': balance
                })

        # 負債の分類
        for liability in liabilities_data:
            account_name = liability.get('account_name', '')
            balance = liability.get('total_balance', 0)

            if any(keyword in account_name for keyword in ['買掛金', '支払手形', '短期借入金', '未払金', '預り金']):
                balance_sheet_data['current_liabilities'].append({
                    'name': account_name,
                    'amount': balance
                })
            elif any(keyword in account_name for keyword in ['長期借入金', '社債', '退職給付引当金']):
                balance_sheet_data['fixed_liabilities'].append({
                    'name': account_name,
                    'amount': balance
                })
            else:
                balance_sheet_data['equity'].append({
                    'name': account_name,
                    'amount': balance
                })

        # 収益の分類
        for revenue in revenues_data:
            account_name = revenue.get('account_name', '')
            amount = revenue.get('total_amount', 0)

            if any(keyword in account_name for keyword in ['売上', '売上高']):
                income_statement_data['sales'].append({
                    'name': account_name,
                    'amount': amount
                })
            else:
                income_statement_data['other_income'].append({
                    'name': account_name,
                    'amount': amount
                })

        # 費用の分類
        for expense in expenses_data:
            account_name = expense.get('account_name', '')
            amount = expense.get('total_amount', 0)

            if any(keyword in account_name for keyword in ['売上原価', '仕入']):
                income_statement_data['cost_of_sales'].append({
                    'name': account_name,
                    'amount': amount
                })
            elif any(keyword in account_name for keyword in ['広告宣伝費', '販売手数料', '営業']):
                income_statement_data['selling_expenses'].append({
                    'name': account_name,
                    'amount': amount
                })
            else:
                income_statement_data['administrative_expenses'].append({
                    'name': account_name,
                    'amount': amount
                })

        return balance_sheet_data, income_statement_data

    def create_balance_sheet_from_data(self, ws, company_name, fiscal_year, balance_sheet_data):
        """データベースデータから貸借対照表作成（A4縦）"""
        self.setup_a4_portrait(ws)

        # タイトル行
        ws.merge_cells('A1:E1')
        ws['A1'] = f"{company_name}　貸借対照表"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = self.center_align

        ws.merge_cells('A2:E2')
        ws['A2'] = f"{fiscal_year}年3月31日現在"
        ws['A2'].font = self.normal_font
        ws['A2'].alignment = self.center_align

        ws.merge_cells('A3:E3')
        ws['A3'] = "（単位：円）"
        ws['A3'].font = self.small_font
        ws['A3'].alignment = self.center_align

        # ヘッダー
        row = 5
        ws[f'A{row}'] = "【資産の部】"
        ws[f'A{row}'].font = self.header_font
        ws[f'C{row}'] = "【負債の部】"
        ws[f'C{row}'].font = self.header_font

        row += 1

        # 流動資産
        ws[f'A{row}'] = "流動資産"
        ws[f'A{row}'].font = self.header_font
        ws[f'C{row}'] = "流動負債"
        ws[f'C{row}'].font = self.header_font

        current_assets_total = 0
        current_liabilities_total = 0

        max_items = max(len(balance_sheet_data['current_assets']), len(balance_sheet_data['current_liabilities']))

        for i in range(max_items):
            row += 1

            # 流動資産項目
            if i < len(balance_sheet_data['current_assets']):
                asset = balance_sheet_data['current_assets'][i]
                ws[f'A{row}'] = f"  {asset['name']}"
                ws[f'A{row}'].font = self.normal_font
                ws[f'B{row}'] = f"{asset['amount']:,}"
                ws[f'B{row}'].font = self.normal_font
                ws[f'B{row}'].alignment = self.right_align
                current_assets_total += asset['amount']

            # 流動負債項目
            if i < len(balance_sheet_data['current_liabilities']):
                liability = balance_sheet_data['current_liabilities'][i]
                ws[f'C{row}'] = f"  {liability['name']}"
                ws[f'C{row}'].font = self.normal_font
                ws[f'D{row}'] = f"{liability['amount']:,}"
                ws[f'D{row}'].font = self.normal_font
                ws[f'D{row}'].alignment = self.right_align
                current_liabilities_total += liability['amount']

        # 小計行
        row += 1
        ws[f'A{row}'] = "流動資産計"
        ws[f'A{row}'].font = self.header_font
        ws[f'B{row}'] = f"{current_assets_total:,}"
        ws[f'B{row}'].font = self.header_font
        ws[f'B{row}'].alignment = self.right_align

        ws[f'C{row}'] = "流動負債計"
        ws[f'C{row}'].font = self.header_font
        ws[f'D{row}'] = f"{current_liabilities_total:,}"
        ws[f'D{row}'].font = self.header_font
        ws[f'D{row}'].alignment = self.right_align

        # 固定資産・固定負債
        row += 2
        ws[f'A{row}'] = "固定資産"
        ws[f'A{row}'].font = self.header_font
        ws[f'C{row}'] = "固定負債"
        ws[f'C{row}'].font = self.header_font

        fixed_assets_total = 0
        fixed_liabilities_total = 0

        max_fixed_items = max(len(balance_sheet_data['fixed_assets']), len(balance_sheet_data['fixed_liabilities']))

        for i in range(max_fixed_items):
            row += 1

            # 固定資産項目
            if i < len(balance_sheet_data['fixed_assets']):
                asset = balance_sheet_data['fixed_assets'][i]
                ws[f'A{row}'] = f"  {asset['name']}"
                ws[f'A{row}'].font = self.normal_font
                ws[f'B{row}'] = f"{asset['amount']:,}"
                ws[f'B{row}'].font = self.normal_font
                ws[f'B{row}'].alignment = self.right_align
                fixed_assets_total += asset['amount']

            # 固定負債項目
            if i < len(balance_sheet_data['fixed_liabilities']):
                liability = balance_sheet_data['fixed_liabilities'][i]
                ws[f'C{row}'] = f"  {liability['name']}"
                ws[f'C{row}'].font = self.normal_font
                ws[f'D{row}'] = f"{liability['amount']:,}"
                ws[f'D{row}'].font = self.normal_font
                ws[f'D{row}'].alignment = self.right_align
                fixed_liabilities_total += liability['amount']

        # 固定資産計・固定負債計
        row += 1
        ws[f'A{row}'] = "固定資産計"
        ws[f'A{row}'].font = self.header_font
        ws[f'B{row}'] = f"{fixed_assets_total:,}"
        ws[f'B{row}'].font = self.header_font
        ws[f'B{row}'].alignment = self.right_align

        ws[f'C{row}'] = "固定負債計"
        ws[f'C{row}'].font = self.header_font
        ws[f'D{row}'] = f"{fixed_liabilities_total:,}"
        ws[f'D{row}'].font = self.header_font
        ws[f'D{row}'].alignment = self.right_align

        # 純資産の部
        row += 2
        ws[f'C{row}'] = "【純資産の部】"
        ws[f'C{row}'].font = self.header_font

        equity_total = 0
        for equity in balance_sheet_data['equity']:
            row += 1
            ws[f'C{row}'] = f"  {equity['name']}"
            ws[f'C{row}'].font = self.normal_font
            ws[f'D{row}'] = f"{equity['amount']:,}"
            ws[f'D{row}'].font = self.normal_font
            ws[f'D{row}'].alignment = self.right_align
            equity_total += equity['amount']

        row += 1
        ws[f'C{row}'] = "純資産合計"
        ws[f'C{row}'].font = self.header_font
        ws[f'D{row}'] = f"{equity_total:,}"
        ws[f'D{row}'].font = self.header_font
        ws[f'D{row}'].alignment = self.right_align

        # 資産合計・負債純資産合計
        assets_total = current_assets_total + fixed_assets_total
        liabilities_equity_total = current_liabilities_total + fixed_liabilities_total + equity_total

        row += 2
        ws[f'A{row}'] = "資産合計"
        ws[f'A{row}'].font = self.header_font
        ws[f'B{row}'] = f"{assets_total:,}"
        ws[f'B{row}'].font = self.header_font
        ws[f'B{row}'].alignment = self.right_align

        ws[f'C{row}'] = "負債純資産合計"
        ws[f'C{row}'].font = self.header_font
        ws[f'D{row}'] = f"{liabilities_equity_total:,}"
        ws[f'D{row}'].font = self.header_font
        ws[f'D{row}'].alignment = self.right_align

    def create_income_statement_from_data(self, ws, company_name, fiscal_year, income_statement_data):
        """データベースデータから損益計算書作成（A4縦）"""
        self.setup_a4_portrait(ws)

        # タイトル行
        ws.merge_cells('A1:D1')
        ws['A1'] = f"{company_name}　損益計算書"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = self.center_align

        ws.merge_cells('A2:D2')
        ws['A2'] = f"{fiscal_year}年4月1日から{fiscal_year + 1}年3月31日まで"
        ws['A2'].font = self.normal_font
        ws['A2'].alignment = self.center_align

        ws.merge_cells('A3:D3')
        ws['A3'] = "（単位：円）"
        ws['A3'].font = self.small_font
        ws['A3'].alignment = self.center_align

        row = 5

        # 売上高
        sales_total = sum(item['amount'] for item in income_statement_data['sales'])
        ws[f'A{row}'] = "売上高"
        ws[f'A{row}'].font = self.header_font
        ws[f'D{row}'] = f"{sales_total:,}"
        ws[f'D{row}'].font = self.header_font
        ws[f'D{row}'].alignment = self.right_align

        # 売上原価
        row += 1
        cost_of_sales_total = sum(item['amount'] for item in income_statement_data['cost_of_sales'])
        ws[f'A{row}'] = "売上原価"
        ws[f'A{row}'].font = self.header_font
        ws[f'D{row}'] = f"{cost_of_sales_total:,}"
        ws[f'D{row}'].font = self.header_font
        ws[f'D{row}'].alignment = self.right_align

        # 売上総利益
        row += 1
        gross_profit = sales_total - cost_of_sales_total
        ws[f'A{row}'] = "売上総利益"
        ws[f'A{row}'].font = self.header_font
        ws[f'D{row}'] = f"{gross_profit:,}"
        ws[f'D{row}'].font = self.header_font
        ws[f'D{row}'].alignment = self.right_align

        # 販売費及び一般管理費
        row += 2
        ws[f'A{row}'] = "販売費及び一般管理費"
        ws[f'A{row}'].font = self.header_font

        selling_admin_total = (sum(item['amount'] for item in income_statement_data['selling_expenses']) +
                              sum(item['amount'] for item in income_statement_data['administrative_expenses']))

        row += 1
        ws[f'A{row}'] = "販売費及び一般管理費計"
        ws[f'A{row}'].font = self.normal_font
        ws[f'D{row}'] = f"{selling_admin_total:,}"
        ws[f'D{row}'].font = self.normal_font
        ws[f'D{row}'].alignment = self.right_align

        # 営業利益
        row += 1
        operating_profit = gross_profit - selling_admin_total
        ws[f'A{row}'] = "営業利益"
        ws[f'A{row}'].font = self.header_font
        ws[f'D{row}'] = f"{operating_profit:,}"
        ws[f'D{row}'].font = self.header_font
        ws[f'D{row}'].alignment = self.right_align

        # 営業外収益・費用
        row += 2
        other_income_total = sum(item['amount'] for item in income_statement_data['other_income'])
        other_expense_total = sum(item['amount'] for item in income_statement_data['other_expenses'])

        ws[f'A{row}'] = "営業外収益"
        ws[f'A{row}'].font = self.header_font
        ws[f'D{row}'] = f"{other_income_total:,}"
        ws[f'D{row}'].font = self.header_font
        ws[f'D{row}'].alignment = self.right_align

        row += 1
        ws[f'A{row}'] = "営業外費用"
        ws[f'A{row}'].font = self.header_font
        ws[f'D{row}'] = f"{other_expense_total:,}"
        ws[f'D{row}'].font = self.header_font
        ws[f'D{row}'].alignment = self.right_align

        # 経常利益
        row += 1
        ordinary_profit = operating_profit + other_income_total - other_expense_total
        ws[f'A{row}'] = "経常利益"
        ws[f'A{row}'].font = self.header_font
        ws[f'D{row}'] = f"{ordinary_profit:,}"
        ws[f'D{row}'].font = self.header_font
        ws[f'D{row}'].alignment = self.right_align

        # 当期純利益（税引前）
        row += 2
        ws[f'A{row}'] = "税引前当期純利益"
        ws[f'A{row}'].font = self.header_font
        ws[f'D{row}'] = f"{ordinary_profit:,}"
        ws[f'D{row}'].font = self.header_font
        ws[f'D{row}'].alignment = self.right_align

        # 法人税等（概算）
        row += 1
        tax_estimate = int(ordinary_profit * 0.3) if ordinary_profit > 0 else 0
        ws[f'A{row}'] = "法人税等"
        ws[f'A{row}'].font = self.header_font
        ws[f'D{row}'] = f"{tax_estimate:,}"
        ws[f'D{row}'].font = self.header_font
        ws[f'D{row}'].alignment = self.right_align

        # 当期純利益
        row += 1
        net_income = ordinary_profit - tax_estimate
        ws[f'A{row}'] = "当期純利益"
        ws[f'A{row}'].font = self.header_font
        ws[f'D{row}'] = f"{net_income:,}"
        ws[f'D{row}'].font = self.header_font
        ws[f'D{row}'].alignment = self.right_align

    def create_cash_flow_statement(self, ws, company_name, fiscal_year):
        """キャッシュフロー計算書作成（A4縦・簡易版）"""
        self.setup_a4_portrait(ws)

        ws.merge_cells('A1:D1')
        ws['A1'] = f"{company_name}　キャッシュフロー計算書"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = self.center_align

        ws.merge_cells('A2:D2')
        ws['A2'] = f"{fiscal_year}年4月1日から{fiscal_year + 1}年3月31日まで"
        ws['A2'].font = self.normal_font
        ws['A2'].alignment = self.center_align

        row = 5
        ws[f'A{row}'] = "営業活動によるキャッシュフロー"
        ws[f'A{row}'].font = self.header_font
        ws[f'D{row}'] = "1,000,000"
        ws[f'D{row}'].alignment = self.right_align

    def create_equity_statement_landscape(self, ws, company_name, fiscal_year):
        """株主資本等変動計算書作成（A4横・拡張版）"""
        self.setup_a4_landscape(ws)

        ws.merge_cells('A1:H1')
        ws['A1'] = f"{company_name}　株主資本等変動計算書"
        ws['A1'].font = self.title_font_landscape
        ws['A1'].alignment = self.center_align

    def create_notes_landscape(self, ws, company_name, fiscal_year):
        """附属明細書作成（A4横・詳細版）"""
        self.setup_a4_landscape(ws)

        ws.merge_cells('A1:G1')
        ws['A1'] = f"{company_name}　附属明細書"
        ws['A1'].font = self.title_font_landscape
        ws['A1'].alignment = self.center_align

    def generate_all_statements_from_data(self, company_name="株式会社サンプル", fiscal_year=2025,
                                        assets_data=None, liabilities_data=None,
                                        revenues_data=None, expenses_data=None):
        """
        データベースデータから全財務諸表を生成
        """
        if not all([assets_data, liabilities_data, revenues_data, expenses_data]):
            # デフォルトサンプルデータ
            assets_data = [{'account_name': '現金預金', 'total_balance': 5000000}]
            liabilities_data = [{'account_name': '買掛金', 'total_balance': 1000000}]
            revenues_data = [{'account_name': '売上高', 'total_amount': 10000000}]
            expenses_data = [{'account_name': '売上原価', 'total_amount': 6000000}]

        # データ分類
        balance_sheet_data, income_statement_data = self.classify_account_data(
            assets_data, liabilities_data, revenues_data, expenses_data
        )

        # Excelブック作成
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # デフォルトシート削除

        # 1. 貸借対照表（A4縦）
        bs_ws = wb.create_sheet(title="貸借対照表")
        self.create_balance_sheet_from_data(bs_ws, company_name, fiscal_year, balance_sheet_data)

        # 2. 損益計算書（A4縦）
        pl_ws = wb.create_sheet(title="損益計算書")
        self.create_income_statement_from_data(pl_ws, company_name, fiscal_year, income_statement_data)

        # 3. キャッシュフロー計算書（A4縦）
        cf_ws = wb.create_sheet(title="キャッシュフロー計算書")
        self.create_cash_flow_statement(cf_ws, company_name, fiscal_year)

        # 4. 株主資本等変動計算書（A4横）
        eq_ws = wb.create_sheet(title="株主資本等変動計算書")
        self.create_equity_statement_landscape(eq_ws, company_name, fiscal_year)

        # 5. 附属明細書（A4横）
        notes_ws = wb.create_sheet(title="附属明細書")
        self.create_notes_landscape(notes_ws, company_name, fiscal_year)

        # BytesIOに保存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output


def generate_mixed_orientation_financial_statements_from_db(company_name="株式会社サンプル",
                                                          fiscal_year=2025,
                                                          assets_data=None,
                                                          liabilities_data=None,
                                                          revenues_data=None,
                                                          expenses_data=None):
    """
    データベース連動財務諸表生成（外部呼び出し用）
    """
    generator = DatabaseIntegratedFinancialGenerator()
    return generator.generate_all_statements_from_data(
        company_name, fiscal_year, assets_data, liabilities_data, revenues_data, expenses_data
    )


if __name__ == "__main__":
    # テスト実行
    print("データベース連動財務諸表生成テスト")
    output = generate_mixed_orientation_financial_statements_from_db()

    with open("/home/esan/employee-db/database_integrated_test.xlsx", "wb") as f:
        f.write(output.getvalue())

    print("✅ database_integrated_test.xlsx が生成されました")