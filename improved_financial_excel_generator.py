#!/usr/bin/env python3
"""
改良版財務諸表Excel出力機能
日本の会計基準に準拠したフォーマット
"""

from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime


def create_improved_financial_statements_excel(assets, liabilities, revenues, expenses, cash_flow, equity_change,
                                             fixed_assets, bonds, loans, reserves, year, report_type):
    """改良版財務諸表Excel生成機能"""

    wb = openpyxl.Workbook()

    # 日本語フォント設定
    title_font = Font(name='ＭＳ ゴシック', size=14, bold=True)
    subtitle_font = Font(name='ＭＳ ゴシック', size=12, bold=True)
    header_font = Font(name='ＭＳ ゴシック', size=11, bold=True)
    normal_font = Font(name='ＭＳ ゴシック', size=10)
    small_font = Font(name='ＭＳ ゴシック', size=9)

    # アライメント設定
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_alignment = Alignment(horizontal='right', vertical='center')

    # 罫線設定
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    thick_border = Border(
        left=Side(style='thick'), right=Side(style='thick'),
        top=Side(style='thick'), bottom=Side(style='thick')
    )
    double_border = Border(
        top=Side(style='double'), bottom=Side(style='double')
    )

    # 背景色設定
    header_fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
    subtotal_fill = PatternFill(start_color='F0F8FF', end_color='F0F8FF', fill_type='solid')
    total_fill = PatternFill(start_color='FFF8DC', end_color='FFF8DC', fill_type='solid')

    # 貸借対照表作成
    if report_type in ['all', 'balance_sheet']:
        ws_bs = wb.active if report_type == 'balance_sheet' else wb.create_sheet("貸借対照表")
        ws_bs.title = "貸借対照表"
        create_improved_balance_sheet(ws_bs, assets, liabilities, year,
                                    title_font, header_font, normal_font, small_font,
                                    center_alignment, left_alignment, right_alignment,
                                    thin_border, thick_border, double_border,
                                    header_fill, subtotal_fill, total_fill)

    # 損益計算書作成
    if report_type in ['all', 'income_statement']:
        ws_pl = wb.create_sheet("損益計算書")
        create_improved_income_statement(ws_pl, revenues, expenses, year,
                                       title_font, header_font, normal_font, small_font,
                                       center_alignment, left_alignment, right_alignment,
                                       thin_border, thick_border, double_border,
                                       header_fill, subtotal_fill, total_fill)

    # キャッシュフロー計算書作成
    if report_type in ['all', 'cash_flow']:
        ws_cf = wb.create_sheet("キャッシュフロー計算書")
        create_improved_cash_flow(ws_cf, cash_flow, year,
                                title_font, header_font, normal_font, small_font,
                                center_alignment, left_alignment, right_alignment,
                                thin_border, thick_border, double_border,
                                header_fill, subtotal_fill, total_fill)

    # 株主資本等変動計算書作成
    if report_type in ['all', 'equity_change']:
        ws_eq = wb.create_sheet("株主資本等変動計算書")
        create_improved_equity_change(ws_eq, equity_change, year,
                                    title_font, header_font, normal_font, small_font,
                                    center_alignment, left_alignment, right_alignment,
                                    thin_border, thick_border, double_border,
                                    header_fill, subtotal_fill, total_fill)

    # 附属明細書作成
    if report_type in ['all', 'notes']:
        ws_notes = wb.create_sheet("附属明細書")
        create_improved_notes(ws_notes, fixed_assets, bonds, loans, reserves, year,
                            title_font, header_font, normal_font, small_font,
                            center_alignment, left_alignment, right_alignment,
                            thin_border, thick_border, double_border,
                            header_fill, subtotal_fill, total_fill)

    # 最初のワークシートを削除（空のワークシートがある場合）
    if len(wb.sheetnames) > 1 and wb['Sheet']:
        wb.remove(wb['Sheet'])

    # BytesIOオブジェクトに保存
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def create_improved_balance_sheet(ws, assets, liabilities, year,
                                title_font, header_font, normal_font, small_font,
                                center_alignment, left_alignment, right_alignment,
                                thin_border, thick_border, double_border,
                                header_fill, subtotal_fill, total_fill):
    """改良版貸借対照表作成"""

    # 列幅設定
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 3
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 15

    # タイトル部分
    ws.merge_cells('A1:F1')
    ws['A1'] = "貸借対照表"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_alignment

    ws.merge_cells('A2:F2')
    ws['A2'] = f"{year}年12月31日現在"
    ws['A2'].font = header_font
    ws['A2'].alignment = center_alignment

    ws.merge_cells('A3:F3')
    ws['A3'] = "（単位：円）"
    ws['A3'].font = small_font
    ws['A3'].alignment = center_alignment

    # ヘッダー行
    row = 5
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = "資産の部"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].alignment = center_alignment
    ws[f'A{row}'].fill = header_fill

    ws.merge_cells(f'D{row}:F{row}')
    ws[f'D{row}'] = "負債及び純資産の部"
    ws[f'D{row}'].font = header_font
    ws[f'D{row}'].alignment = center_alignment
    ws[f'D{row}'].fill = header_fill

    # 罫線設定
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{row}'].border = thin_border

    row += 1

    # 資産の部
    ws[f'B{row}'] = "科目"
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].alignment = center_alignment
    ws[f'B{row}'].fill = subtotal_fill
    ws[f'C{row}'] = "金額"
    ws[f'C{row}'].font = header_font
    ws[f'C{row}'].alignment = center_alignment
    ws[f'C{row}'].fill = subtotal_fill

    # 負債の部
    ws[f'E{row}'] = "科目"
    ws[f'E{row}'].font = header_font
    ws[f'E{row}'].alignment = center_alignment
    ws[f'E{row}'].fill = subtotal_fill
    ws[f'F{row}'] = "金額"
    ws[f'F{row}'].font = header_font
    ws[f'F{row}'].alignment = center_alignment
    ws[f'F{row}'].fill = subtotal_fill

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{row}'].border = thin_border

    row += 1

    # 資産データ
    asset_total = 0
    start_row = row

    # 流動資産
    ws[f'B{row}'] = "Ⅰ　流動資産"
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].alignment = left_alignment
    row += 1

    for asset in assets[:3]:  # 流動資産として最初の3つ
        ws[f'B{row}'] = f"　　{asset.account_name}"
        ws[f'B{row}'].font = normal_font
        ws[f'B{row}'].alignment = left_alignment
        ws[f'C{row}'] = f"{asset.balance:,}"
        ws[f'C{row}'].font = normal_font
        ws[f'C{row}'].alignment = right_alignment
        asset_total += asset.balance
        row += 1

    # 流動資産小計
    ws[f'B{row}'] = "　流動資産合計"
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].alignment = left_alignment
    ws[f'B{row}'].fill = subtotal_fill
    current_assets_total = sum(asset.balance for asset in assets[:3])
    ws[f'C{row}'] = f"{current_assets_total:,}"
    ws[f'C{row}'].font = header_font
    ws[f'C{row}'].alignment = right_alignment
    ws[f'C{row}'].fill = subtotal_fill
    row += 1

    # 固定資産
    ws[f'B{row}'] = "Ⅱ　固定資産"
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].alignment = left_alignment
    row += 1

    for asset in assets[3:]:  # 固定資産として残り
        ws[f'B{row}'] = f"　　{asset.account_name}"
        ws[f'B{row}'].font = normal_font
        ws[f'B{row}'].alignment = left_alignment
        ws[f'C{row}'] = f"{asset.balance:,}"
        ws[f'C{row}'].font = normal_font
        ws[f'C{row}'].alignment = right_alignment
        row += 1

    # 固定資産小計
    ws[f'B{row}'] = "　固定資産合計"
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].alignment = left_alignment
    ws[f'B{row}'].fill = subtotal_fill
    fixed_assets_total = sum(asset.balance for asset in assets[3:])
    ws[f'C{row}'] = f"{fixed_assets_total:,}"
    ws[f'C{row}'].font = header_font
    ws[f'C{row}'].alignment = right_alignment
    ws[f'C{row}'].fill = subtotal_fill
    row += 1

    # 資産合計
    ws[f'B{row}'] = "資産合計"
    ws[f'B{row}'].font = title_font
    ws[f'B{row}'].alignment = left_alignment
    ws[f'B{row}'].fill = total_fill
    total_assets = sum(asset.balance for asset in assets)
    ws[f'C{row}'] = f"{total_assets:,}"
    ws[f'C{row}'].font = title_font
    ws[f'C{row}'].alignment = right_alignment
    ws[f'C{row}'].fill = total_fill
    ws[f'C{row}'].border = double_border

    # 負債の部（右側）
    liability_row = start_row

    # 流動負債
    ws[f'E{liability_row}'] = "Ⅰ　流動負債"
    ws[f'E{liability_row}'].font = header_font
    ws[f'E{liability_row}'].alignment = left_alignment
    liability_row += 1

    for liability in liabilities[:2]:  # 流動負債として最初の2つ
        ws[f'E{liability_row}'] = f"　　{liability.account_name}"
        ws[f'E{liability_row}'].font = normal_font
        ws[f'E{liability_row}'].alignment = left_alignment
        ws[f'F{liability_row}'] = f"{liability.balance:,}"
        ws[f'F{liability_row}'].font = normal_font
        ws[f'F{liability_row}'].alignment = right_alignment
        liability_row += 1

    # 流動負債小計
    ws[f'E{liability_row}'] = "　流動負債合計"
    ws[f'E{liability_row}'].font = header_font
    ws[f'E{liability_row}'].alignment = left_alignment
    ws[f'E{liability_row}'].fill = subtotal_fill
    current_liabilities_total = sum(liability.balance for liability in liabilities[:2])
    ws[f'F{liability_row}'] = f"{current_liabilities_total:,}"
    ws[f'F{liability_row}'].font = header_font
    ws[f'F{liability_row}'].alignment = right_alignment
    ws[f'F{liability_row}'].fill = subtotal_fill
    liability_row += 1

    # 固定負債
    ws[f'E{liability_row}'] = "Ⅱ　固定負債"
    ws[f'E{liability_row}'].font = header_font
    ws[f'E{liability_row}'].alignment = left_alignment
    liability_row += 1

    for liability in liabilities[2:]:  # 固定負債として残り
        ws[f'E{liability_row}'] = f"　　{liability.account_name}"
        ws[f'E{liability_row}'].font = normal_font
        ws[f'E{liability_row}'].alignment = left_alignment
        ws[f'F{liability_row}'] = f"{liability.balance:,}"
        ws[f'F{liability_row}'].font = normal_font
        ws[f'F{liability_row}'].alignment = right_alignment
        liability_row += 1

    # 固定負債小計
    ws[f'E{liability_row}'] = "　固定負債合計"
    ws[f'E{liability_row}'].font = header_font
    ws[f'E{liability_row}'].alignment = left_alignment
    ws[f'E{liability_row}'].fill = subtotal_fill
    fixed_liabilities_total = sum(liability.balance for liability in liabilities[2:])
    ws[f'F{liability_row}'] = f"{fixed_liabilities_total:,}"
    ws[f'F{liability_row}'].font = header_font
    ws[f'F{liability_row}'].alignment = right_alignment
    ws[f'F{liability_row}'].fill = subtotal_fill
    liability_row += 1

    # 負債合計
    ws[f'E{liability_row}'] = "負債合計"
    ws[f'E{liability_row}'].font = header_font
    ws[f'E{liability_row}'].alignment = left_alignment
    ws[f'E{liability_row}'].fill = subtotal_fill
    total_liabilities = sum(liability.balance for liability in liabilities)
    ws[f'F{liability_row}'] = f"{total_liabilities:,}"
    ws[f'F{liability_row}'].font = header_font
    ws[f'F{liability_row}'].alignment = right_alignment
    ws[f'F{liability_row}'].fill = subtotal_fill
    liability_row += 1

    # 純資産の部
    ws[f'E{liability_row}'] = "Ⅲ　純資産の部"
    ws[f'E{liability_row}'].font = header_font
    ws[f'E{liability_row}'].alignment = left_alignment
    liability_row += 1

    # 資本金
    ws[f'E{liability_row}'] = "　　資本金"
    ws[f'E{liability_row}'].font = normal_font
    ws[f'E{liability_row}'].alignment = left_alignment
    capital = 1000000  # サンプル値
    ws[f'F{liability_row}'] = f"{capital:,}"
    ws[f'F{liability_row}'].font = normal_font
    ws[f'F{liability_row}'].alignment = right_alignment
    liability_row += 1

    # 利益剰余金
    ws[f'E{liability_row}'] = "　　利益剰余金"
    ws[f'E{liability_row}'].font = normal_font
    ws[f'E{liability_row}'].alignment = left_alignment
    retained_earnings = total_assets - total_liabilities - capital
    ws[f'F{liability_row}'] = f"{retained_earnings:,}"
    ws[f'F{liability_row}'].font = normal_font
    ws[f'F{liability_row}'].alignment = right_alignment
    liability_row += 1

    # 純資産合計
    ws[f'E{liability_row}'] = "　純資産合計"
    ws[f'E{liability_row}'].font = header_font
    ws[f'E{liability_row}'].alignment = left_alignment
    ws[f'E{liability_row}'].fill = subtotal_fill
    total_equity = capital + retained_earnings
    ws[f'F{liability_row}'] = f"{total_equity:,}"
    ws[f'F{liability_row}'].font = header_font
    ws[f'F{liability_row}'].alignment = right_alignment
    ws[f'F{liability_row}'].fill = subtotal_fill
    liability_row += 1

    # 負債及び純資産合計
    ws[f'E{liability_row}'] = "負債及び純資産合計"
    ws[f'E{liability_row}'].font = title_font
    ws[f'E{liability_row}'].alignment = left_alignment
    ws[f'E{liability_row}'].fill = total_fill
    total_liab_equity = total_liabilities + total_equity
    ws[f'F{liability_row}'] = f"{total_liab_equity:,}"
    ws[f'F{liability_row}'].font = title_font
    ws[f'F{liability_row}'].alignment = right_alignment
    ws[f'F{liability_row}'].fill = total_fill
    ws[f'F{liability_row}'].border = double_border

    # 全体に罫線を適用
    max_row = max(row, liability_row)
    for r in range(5, max_row + 1):
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            if ws[f'{col}{r}'].value is not None or r <= max_row:
                ws[f'{col}{r}'].border = thin_border


def create_improved_income_statement(ws, revenues, expenses, year,
                                   title_font, header_font, normal_font, small_font,
                                   center_alignment, left_alignment, right_alignment,
                                   thin_border, thick_border, double_border,
                                   header_fill, subtotal_fill, total_fill):
    """改良版損益計算書作成"""

    # 列幅設定
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    # タイトル部分
    ws.merge_cells('A1:D1')
    ws['A1'] = "損益計算書"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_alignment

    ws.merge_cells('A2:D2')
    ws['A2'] = f"{year}年1月1日から{year}年12月31日まで"
    ws['A2'].font = header_font
    ws['A2'].alignment = center_alignment

    ws.merge_cells('A3:D3')
    ws['A3'] = "（単位：円）"
    ws['A3'].font = small_font
    ws['A3'].alignment = center_alignment

    row = 5

    # ヘッダー
    ws['B5'] = "科目"
    ws['B5'].font = header_font
    ws['B5'].alignment = center_alignment
    ws['B5'].fill = header_fill
    ws['C5'] = "当期"
    ws['C5'].font = header_font
    ws['C5'].alignment = center_alignment
    ws['C5'].fill = header_fill
    ws['D5'] = "前期"
    ws['D5'].font = header_font
    ws['D5'].alignment = center_alignment
    ws['D5'].fill = header_fill

    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}5'].border = thin_border

    row = 6

    # 売上高
    ws[f'B{row}'] = "Ⅰ　売上高"
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].alignment = left_alignment
    row += 1

    total_revenue = 0
    for revenue in revenues:
        ws[f'B{row}'] = f"　　{revenue.account_name}"
        ws[f'B{row}'].font = normal_font
        ws[f'B{row}'].alignment = left_alignment
        ws[f'C{row}'] = f"{revenue.balance:,}"
        ws[f'C{row}'].font = normal_font
        ws[f'C{row}'].alignment = right_alignment
        total_revenue += revenue.balance
        row += 1

    # 売上高計
    ws[f'B{row}'] = "　売上高計"
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].alignment = left_alignment
    ws[f'B{row}'].fill = subtotal_fill
    ws[f'C{row}'] = f"{total_revenue:,}"
    ws[f'C{row}'].font = header_font
    ws[f'C{row}'].alignment = right_alignment
    ws[f'C{row}'].fill = subtotal_fill
    row += 1

    # 売上原価
    ws[f'B{row}'] = "Ⅱ　売上原価"
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].alignment = left_alignment
    row += 1

    cost_of_sales = sum(expense.balance for expense in expenses[:2])  # 最初の2つを売上原価とする
    ws[f'B{row}'] = "　　売上原価"
    ws[f'B{row}'].font = normal_font
    ws[f'B{row}'].alignment = left_alignment
    ws[f'C{row}'] = f"{cost_of_sales:,}"
    ws[f'C{row}'].font = normal_font
    ws[f'C{row}'].alignment = right_alignment
    row += 1

    # 売上総利益
    gross_profit = total_revenue - cost_of_sales
    ws[f'B{row}'] = "　売上総利益"
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].alignment = left_alignment
    ws[f'B{row}'].fill = subtotal_fill
    ws[f'C{row}'] = f"{gross_profit:,}"
    ws[f'C{row}'].font = header_font
    ws[f'C{row}'].alignment = right_alignment
    ws[f'C{row}'].fill = subtotal_fill
    row += 1

    # 販売費及び一般管理費
    ws[f'B{row}'] = "Ⅲ　販売費及び一般管理費"
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].alignment = left_alignment
    row += 1

    sg_a_expenses = 0
    for expense in expenses[2:]:  # 残りを販管費とする
        ws[f'B{row}'] = f"　　{expense.account_name}"
        ws[f'B{row}'].font = normal_font
        ws[f'B{row}'].alignment = left_alignment
        ws[f'C{row}'] = f"{expense.balance:,}"
        ws[f'C{row}'].font = normal_font
        ws[f'C{row}'].alignment = right_alignment
        sg_a_expenses += expense.balance
        row += 1

    # 販管費計
    ws[f'B{row}'] = "　販売費及び一般管理費計"
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].alignment = left_alignment
    ws[f'B{row}'].fill = subtotal_fill
    ws[f'C{row}'] = f"{sg_a_expenses:,}"
    ws[f'C{row}'].font = header_font
    ws[f'C{row}'].alignment = right_alignment
    ws[f'C{row}'].fill = subtotal_fill
    row += 1

    # 営業利益
    operating_profit = gross_profit - sg_a_expenses
    ws[f'B{row}'] = "　営業利益"
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].alignment = left_alignment
    ws[f'B{row}'].fill = subtotal_fill
    ws[f'C{row}'] = f"{operating_profit:,}"
    ws[f'C{row}'].font = header_font
    ws[f'C{row}'].alignment = right_alignment
    ws[f'C{row}'].fill = subtotal_fill
    row += 1

    # 当期純利益
    net_income = operating_profit
    ws[f'B{row}'] = "　当期純利益"
    ws[f'B{row}'].font = title_font
    ws[f'B{row}'].alignment = left_alignment
    ws[f'B{row}'].fill = total_fill
    ws[f'C{row}'] = f"{net_income:,}"
    ws[f'C{row}'].font = title_font
    ws[f'C{row}'].alignment = right_alignment
    ws[f'C{row}'].fill = total_fill
    ws[f'C{row}'].border = double_border

    # 全体に罫線を適用
    for r in range(5, row + 1):
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{r}'].border = thin_border


def create_improved_cash_flow(ws, cash_flow, year,
                            title_font, header_font, normal_font, small_font,
                            center_alignment, left_alignment, right_alignment,
                            thin_border, thick_border, double_border,
                            header_fill, subtotal_fill, total_fill):
    """改良版キャッシュフロー計算書作成"""

    # 列幅設定
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15

    # タイトル部分
    ws.merge_cells('A1:C1')
    ws['A1'] = "キャッシュ・フロー計算書"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_alignment

    ws.merge_cells('A2:C2')
    ws['A2'] = f"{year}年1月1日から{year}年12月31日まで"
    ws['A2'].font = header_font
    ws['A2'].alignment = center_alignment

    ws.merge_cells('A3:C3')
    ws['A3'] = "（単位：円）"
    ws['A3'].font = small_font
    ws['A3'].alignment = center_alignment

    row = 5

    # ヘッダー
    ws['B5'] = "科目"
    ws['B5'].font = header_font
    ws['B5'].alignment = center_alignment
    ws['B5'].fill = header_fill
    ws['C5'] = "金額"
    ws['C5'].font = header_font
    ws['C5'].alignment = center_alignment
    ws['C5'].fill = header_fill

    for col in ['A', 'B', 'C']:
        ws[f'{col}5'].border = thin_border

    row = 6

    # 営業活動によるキャッシュフロー
    ws[f'B{row}'] = "Ⅰ　営業活動によるキャッシュ・フロー"
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].alignment = left_alignment
    row += 1

    # 税引前当期純利益
    ws[f'B{row}'] = "　　税引前当期純利益"
    ws[f'B{row}'].font = normal_font
    ws[f'B{row}'].alignment = left_alignment
    pre_tax_income = cash_flow.operating.pre_tax_income if cash_flow else 300000
    ws[f'C{row}'] = f"{pre_tax_income:,}"
    ws[f'C{row}'].font = normal_font
    ws[f'C{row}'].alignment = right_alignment
    row += 1

    # 減価償却費
    ws[f'B{row}'] = "　　減価償却費"
    ws[f'B{row}'].font = normal_font
    ws[f'B{row}'].alignment = left_alignment
    depreciation = cash_flow.operating.depreciation if cash_flow else 50000
    ws[f'C{row}'] = f"{depreciation:,}"
    ws[f'C{row}'].font = normal_font
    ws[f'C{row}'].alignment = right_alignment
    row += 1

    # 営業活動によるキャッシュフロー小計
    operating_cf = pre_tax_income + depreciation
    ws[f'B{row}'] = "　営業活動によるキャッシュ・フロー"
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].alignment = left_alignment
    ws[f'B{row}'].fill = subtotal_fill
    ws[f'C{row}'] = f"{operating_cf:,}"
    ws[f'C{row}'].font = header_font
    ws[f'C{row}'].alignment = right_alignment
    ws[f'C{row}'].fill = subtotal_fill
    row += 1

    # 投資活動によるキャッシュフロー
    ws[f'B{row}'] = "Ⅱ　投資活動によるキャッシュ・フロー"
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].alignment = left_alignment
    row += 1

    # 有形固定資産の取得による支出
    ws[f'B{row}'] = "　　有形固定資産の取得による支出"
    ws[f'B{row}'].font = normal_font
    ws[f'B{row}'].alignment = left_alignment
    asset_purchase = cash_flow.investing.asset_purchase if cash_flow else -100000
    ws[f'C{row}'] = f"{asset_purchase:,}"
    ws[f'C{row}'].font = normal_font
    ws[f'C{row}'].alignment = right_alignment
    row += 1

    # 投資活動によるキャッシュフロー小計
    investing_cf = asset_purchase
    ws[f'B{row}'] = "　投資活動によるキャッシュ・フロー"
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].alignment = left_alignment
    ws[f'B{row}'].fill = subtotal_fill
    ws[f'C{row}'] = f"{investing_cf:,}"
    ws[f'C{row}'].font = header_font
    ws[f'C{row}'].alignment = right_alignment
    ws[f'C{row}'].fill = subtotal_fill
    row += 1

    # 財務活動によるキャッシュフロー
    ws[f'B{row}'] = "Ⅲ　財務活動によるキャッシュ・フロー"
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].alignment = left_alignment
    row += 1

    # 借入による収入
    ws[f'B{row}'] = "　　長期借入金の借入による収入"
    ws[f'B{row}'].font = normal_font
    ws[f'B{row}'].alignment = left_alignment
    loan_increase = cash_flow.financing.loan_increase if cash_flow else 200000
    ws[f'C{row}'] = f"{loan_increase:,}"
    ws[f'C{row}'].font = normal_font
    ws[f'C{row}'].alignment = right_alignment
    row += 1

    # 財務活動によるキャッシュフロー小計
    financing_cf = loan_increase
    ws[f'B{row}'] = "　財務活動によるキャッシュ・フロー"
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].alignment = left_alignment
    ws[f'B{row}'].fill = subtotal_fill
    ws[f'C{row}'] = f"{financing_cf:,}"
    ws[f'C{row}'].font = header_font
    ws[f'C{row}'].alignment = right_alignment
    ws[f'C{row}'].fill = subtotal_fill
    row += 1

    # 現金及び現金同等物の増減額
    cash_change = operating_cf + investing_cf + financing_cf
    ws[f'B{row}'] = "現金及び現金同等物の増減額"
    ws[f'B{row}'].font = title_font
    ws[f'B{row}'].alignment = left_alignment
    ws[f'B{row}'].fill = total_fill
    ws[f'C{row}'] = f"{cash_change:,}"
    ws[f'C{row}'].font = title_font
    ws[f'C{row}'].alignment = right_alignment
    ws[f'C{row}'].fill = total_fill
    ws[f'C{row}'].border = double_border

    # 全体に罫線を適用
    for r in range(5, row + 1):
        for col in ['A', 'B', 'C']:
            ws[f'{col}{r}'].border = thin_border


def create_improved_equity_change(ws, equity_change, year,
                                title_font, header_font, normal_font, small_font,
                                center_alignment, left_alignment, right_alignment,
                                thin_border, thick_border, double_border,
                                header_fill, subtotal_fill, total_fill):
    """改良版株主資本等変動計算書作成"""

    # 列幅設定
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    # タイトル部分
    ws.merge_cells('A1:F1')
    ws['A1'] = "株主資本等変動計算書"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_alignment

    ws.merge_cells('A2:F2')
    ws['A2'] = f"{year}年1月1日から{year}年12月31日まで"
    ws['A2'].font = header_font
    ws['A2'].alignment = center_alignment

    ws.merge_cells('A3:F3')
    ws['A3'] = "（単位：円）"
    ws['A3'].font = small_font
    ws['A3'].alignment = center_alignment

    row = 5

    # ヘッダー行
    ws.merge_cells(f'B{row}:E{row}')
    ws[f'B{row}'] = "株主資本"
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].alignment = center_alignment
    ws[f'B{row}'].fill = header_fill

    ws[f'F{row}'] = "株主資本合計"
    ws[f'F{row}'].font = header_font
    ws[f'F{row}'].alignment = center_alignment
    ws[f'F{row}'].fill = header_fill

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{row}'].border = thin_border

    row += 1

    # サブヘッダー
    ws['A6'] = ""
    ws['B6'] = "資本金"
    ws['C6'] = "資本剰余金"
    ws['D6'] = "利益剰余金"
    ws['E6'] = "自己株式"
    ws['F6'] = "合計"

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}6'].font = header_font
        ws[f'{col}6'].alignment = center_alignment
        ws[f'{col}6'].fill = subtotal_fill
        ws[f'{col}6'].border = thin_border

    row = 7

    # 当期首残高
    ws[f'A{row}'] = "当期首残高"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].alignment = left_alignment

    capital = 1000000
    capital_surplus = 0
    retained_earnings = 500000
    treasury_stock = 0
    beginning_total = capital + capital_surplus + retained_earnings - treasury_stock

    ws[f'B{row}'] = f"{capital:,}"
    ws[f'C{row}'] = f"{capital_surplus:,}"
    ws[f'D{row}'] = f"{retained_earnings:,}"
    ws[f'E{row}'] = f"{treasury_stock:,}"
    ws[f'F{row}'] = f"{beginning_total:,}"

    for col in ['B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{row}'].font = normal_font
        ws[f'{col}{row}'].alignment = right_alignment
        ws[f'{col}{row}'].border = thin_border

    row += 1

    # 当期変動額見出し
    ws[f'A{row}'] = "当期変動額"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].alignment = left_alignment
    ws[f'A{row}'].fill = subtotal_fill

    for col in ['B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{row}'].fill = subtotal_fill
        ws[f'{col}{row}'].border = thin_border

    row += 1

    # 当期純利益
    ws[f'A{row}'] = "　当期純利益"
    ws[f'A{row}'].font = normal_font
    ws[f'A{row}'].alignment = left_alignment

    net_income = 300000
    ws[f'B{row}'] = "-"
    ws[f'C{row}'] = "-"
    ws[f'D{row}'] = f"{net_income:,}"
    ws[f'E{row}'] = "-"
    ws[f'F{row}'] = f"{net_income:,}"

    for col in ['B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{row}'].font = normal_font
        ws[f'{col}{row}'].alignment = right_alignment
        ws[f'{col}{row}'].border = thin_border

    row += 1

    # 当期変動額合計
    ws[f'A{row}'] = "当期変動額合計"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].alignment = left_alignment
    ws[f'A{row}'].fill = subtotal_fill

    ws[f'B{row}'] = "0"
    ws[f'C{row}'] = "0"
    ws[f'D{row}'] = f"{net_income:,}"
    ws[f'E{row}'] = "0"
    ws[f'F{row}'] = f"{net_income:,}"

    for col in ['B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{row}'].font = header_font
        ws[f'{col}{row}'].alignment = right_alignment
        ws[f'{col}{row}'].fill = subtotal_fill
        ws[f'{col}{row}'].border = thin_border

    row += 1

    # 当期末残高
    ws[f'A{row}'] = "当期末残高"
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].alignment = left_alignment
    ws[f'A{row}'].fill = total_fill

    ending_retained_earnings = retained_earnings + net_income
    ending_total = capital + capital_surplus + ending_retained_earnings - treasury_stock

    ws[f'B{row}'] = f"{capital:,}"
    ws[f'C{row}'] = f"{capital_surplus:,}"
    ws[f'D{row}'] = f"{ending_retained_earnings:,}"
    ws[f'E{row}'] = f"{treasury_stock:,}"
    ws[f'F{row}'] = f"{ending_total:,}"

    for col in ['B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{row}'].font = title_font
        ws[f'{col}{row}'].alignment = right_alignment
        ws[f'{col}{row}'].fill = total_fill
        ws[f'{col}{row}'].border = double_border

    # 全体に罫線を適用
    for r in range(5, row + 1):
        ws[f'A{r}'].border = thin_border


def create_improved_notes(ws, fixed_assets, bonds, loans, reserves, year,
                        title_font, header_font, normal_font, small_font,
                        center_alignment, left_alignment, right_alignment,
                        thin_border, thick_border, double_border,
                        header_fill, subtotal_fill, total_fill):
    """改良版附属明細書作成"""

    # 列幅設定
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15

    # タイトル部分
    ws.merge_cells('A1:G1')
    ws['A1'] = "附属明細書"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_alignment

    ws.merge_cells('A2:G2')
    ws['A2'] = f"{year}年12月31日現在"
    ws['A2'].font = header_font
    ws['A2'].alignment = center_alignment

    ws.merge_cells('A3:G3')
    ws['A3'] = "（単位：円）"
    ws['A3'].font = small_font
    ws['A3'].alignment = center_alignment

    row = 5

    # 有形固定資産等明細書
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = "有形固定資産等明細書"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].alignment = center_alignment
    ws[f'A{row}'].fill = header_fill

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws[f'{col}{row}'].border = thin_border

    row += 1

    # ヘッダー
    headers = ["資産の種類", "期首帳簿価額", "当期増加額", "当期減少額", "期末帳簿価額", "減価償却累計額", "期末取得価額"]
    for i, header in enumerate(headers):
        col = chr(ord('A') + i)
        ws[f'{col}{row}'] = header
        ws[f'{col}{row}'].font = header_font
        ws[f'{col}{row}'].alignment = center_alignment
        ws[f'{col}{row}'].fill = subtotal_fill
        ws[f'{col}{row}'].border = thin_border

    row += 1

    # サンプル固定資産データ
    sample_assets = [
        {
            'name': '建物',
            'beginning': 5000000,
            'increase': 0,
            'decrease': 0,
            'ending': 5000000,
            'accumulated_dep': 500000,
            'acquisition_cost': 5500000
        },
        {
            'name': '機械装置',
            'beginning': 2000000,
            'increase': 500000,
            'decrease': 0,
            'ending': 2500000,
            'accumulated_dep': 300000,
            'acquisition_cost': 2800000
        }
    ]

    for asset in sample_assets:
        ws[f'A{row}'] = asset['name']
        ws[f'B{row}'] = f"{asset['beginning']:,}"
        ws[f'C{row}'] = f"{asset['increase']:,}"
        ws[f'D{row}'] = f"{asset['decrease']:,}"
        ws[f'E{row}'] = f"{asset['ending']:,}"
        ws[f'F{row}'] = f"{asset['accumulated_dep']:,}"
        ws[f'G{row}'] = f"{asset['acquisition_cost']:,}"

        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = left_alignment

        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws[f'{col}{row}'].font = normal_font
            ws[f'{col}{row}'].alignment = right_alignment
            ws[f'{col}{row}'].border = thin_border

        ws[f'A{row}'].border = thin_border
        row += 1

    # 全体に罫線を適用
    for r in range(5, row):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws[f'{col}{r}'].border = thin_border


if __name__ == "__main__":
    print("改良版財務諸表Excel出力機能モジュール")
    print("app.pyから呼び出して使用してください")